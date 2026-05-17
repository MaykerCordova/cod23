# =============================================================================
# MÓDULO DE EXPORTACIÓN — Agregar al FINAL del script analisis_mcc7995_v3.py
# Genera un Excel con todas las tablas + interpretaciones automáticas
# =============================================================================

from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

RUTA_EXCEL_OUT = r"C:\ruta\salida\analisis_mcc7995_resultado.xlsx"  # ← AJUSTAR

# ── Colores institucionales ────────────────────────────────────────────────────
FILL_HEADER   = PatternFill("solid", fgColor="1F3864")   # azul oscuro
FILL_SUBHEAD  = PatternFill("solid", fgColor="2E75B6")   # azul medio
FILL_ALERTA   = PatternFill("solid", fgColor="C00000")   # rojo fraude
FILL_OK       = PatternFill("solid", fgColor="375623")   # verde ok
FILL_FILA_A   = PatternFill("solid", fgColor="DEEAF1")   # celda alternada
FILL_AMARILLO = PatternFill("solid", fgColor="FFF2CC")   # highlight interpretación
FONT_HEADER   = Font(color="FFFFFF", bold=True, size=10)
FONT_BOLD     = Font(bold=True, size=10)
FONT_NORMAL   = Font(size=10)
FONT_INTERP   = Font(italic=True, size=9, color="1F3864")
BORDER_THIN   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def estilizar_header(ws, fila, n_cols, texto, fill=FILL_HEADER):
    """Fila de encabezado fusionada."""
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila,   end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.fill      = fill
    c.font      = FONT_HEADER
    c.alignment = ALIGN_CENTER
    c.border    = BORDER_THIN

def escribir_df(ws, df, fila_inicio, col_inicio=1, alternar=True):
    """Escribe un DataFrame con encabezados y filas alternadas."""
    # Encabezados de columna
    for j, col in enumerate(df.columns, start=col_inicio):
        c = ws.cell(row=fila_inicio, column=j, value=col)
        c.fill = FILL_SUBHEAD; c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    # Índice (si tiene)
    if df.index.name or str(df.index.dtype) != "int64":
        idx_col = col_inicio - 1
        ws.cell(row=fila_inicio, column=idx_col,
                value=df.index.name or "").fill = FILL_SUBHEAD

    # Datos
    for i, (idx, row) in enumerate(df.iterrows(), start=1):
        fill = FILL_FILA_A if (alternar and i % 2 == 0) else PatternFill()
        # índice
        if df.index.name or str(df.index.dtype) != "int64":
            c = ws.cell(row=fila_inicio+i, column=col_inicio-1, value=str(idx))
            c.fill = fill; c.font = FONT_BOLD
            c.alignment = ALIGN_LEFT; c.border = BORDER_THIN
        # valores
        for j, val in enumerate(row, start=col_inicio):
            c = ws.cell(row=fila_inicio+i, column=j, value=val)
            c.fill = fill; c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    return fila_inicio + len(df) + 1

def agregar_interpretacion(ws, fila, n_cols, texto):
    """Fila de interpretación en amarillo."""
    ws.merge_cells(start_row=fila, start_column=1,
                   end_row=fila,   end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=f"📌 {texto}")
    c.fill      = FILL_AMARILLO
    c.font      = FONT_INTERP
    c.alignment = ALIGN_LEFT
    c.border    = BORDER_THIN
    ws.row_dimensions[fila].height = 30

def ajustar_columnas(ws):
    for col in ws.columns:
        max_len = 0
        for c in col:
            try:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
            except: pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

# =============================================================================
# CONSTRUIR TODAS LAS TABLAS
# =============================================================================

# ── 1. RESUMEN EJECUTIVO ──────────────────────────────────────────────────────
n_clientes   = df["ID_CLIENTE"].nunique() if "ID_CLIENTE" in df.columns else 0
n_ap         = (df["ESTADO"] == "APROBADA").sum()
monto_ap     = df_ap["MONTO"].sum()
ticket_ap    = df_ap["MONTO"].mean()
n_fraudes    = df["ES_FRAUDE"].sum()
monto_fraude = df_f["MONTO"].sum() if len(df_f) > 0 else 0
ticket_f     = df_f["MONTO"].mean() if len(df_f) > 0 else 0
ratio_trx_f  = n_fraudes / n_ap if n_ap > 0 else 0
ratio_monto_f= monto_fraude / monto_ap if monto_ap > 0 else 0

df_resumen_ejecutivo = pd.DataFrame({
    "Indicador": [
        "Total trx aprobadas",
        "Total monto aprobado (S/)",
        "Ticket promedio por trx (S/)",
        "N° clientes únicos",
        "Trx fraudulentas (F)",
        "Total monto fraude (S/)",
        "Ticket promedio fraude (S/)",
        "Ratio trx fraudulentas",
        "Ratio fraude en soles",
    ],
    "Valor": [
        f"{n_ap:,}",
        f"S/ {monto_ap:,.2f}",
        f"S/ {ticket_ap:,.2f}",
        f"{n_clientes:,}",
        f"{n_fraudes:,}",
        f"S/ {monto_fraude:,.2f}",
        f"S/ {ticket_f:,.2f}",
        f"{ratio_trx_f*100:.2f}%",
        f"{ratio_monto_f*100:.2f}%",
    ]
})

# Umbral de control
nivel_riesgo = "ALTO ≥5%" if ratio_trx_f >= 0.05 else ("MEDIO [1%-5%]" if ratio_trx_f >= 0.01 else "REGULAR <1%")
df_umbral = pd.DataFrame({
    "Umbral":  ["Regular", "Medio",      "Alto"],
    "Rango":   ["< 1%",    "[1% – 5%]",  "≥ 5%"],
    "Nivel actual": [
        "◀ AQUÍ" if ratio_trx_f < 0.01 else "",
        "◀ AQUÍ" if 0.01 <= ratio_trx_f < 0.05 else "",
        "◀ AQUÍ" if ratio_trx_f >= 0.05 else "",
    ]
})

# ── 2. TOP COMERCIOS CON FRAUDE ───────────────────────────────────────────────
df_top_comercios = (
    df_ap
    .groupby("COMERCIO")
    .agg(
        Tarjetas    =("ID_CLIENTE", "nunique"),
        Trx         =("MONTO",      "count"),
        Fraudes     =("ES_FRAUDE",  "sum"),
        Monto_fraude=("MONTO",      lambda x: x[df_ap.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_total =("MONTO",      "sum"),
        FR_pct      =("ES_FRAUDE",  lambda x: x.mean()*100),
    )
    .sort_values("Monto_fraude", ascending=False)
    .head(10)
    .reset_index()
    .rename(columns={
        "COMERCIO":"Comercio", "Tarjetas":"# Tarjetas",
        "Trx":"# Trx", "Fraudes":"# Fraudes",
        "Monto_fraude":"Monto Fraude S/",
        "Monto_total":"Monto Total S/",
        "FR_pct":"Fraud Rate %"
    })
)
df_top_comercios["Monto Fraude S/"] = df_top_comercios["Monto Fraude S/"].map("S/ {:,.2f}".format)
df_top_comercios["Monto Total S/"]  = df_top_comercios["Monto Total S/"].map("S/ {:,.2f}".format)
df_top_comercios["Fraud Rate %"]    = df_top_comercios["Fraud Rate %"].map("{:.2f}%".format)

# ── 3. ESTADÍSTICAS DE MONTO Y TRX ───────────────────────────────────────────
pcts = [10, 25, 50, 75, 90, 95, 99]

stats_monto = pd.DataFrame({
    "Métrica": (
        ["Media", "Mediana", "Desv. Estándar", "Varianza", "Mínimo", "Máximo"] +
        [f"Percentil {p}" for p in pcts]
    ),
    "Total (aprobadas)": (
        [df_ap["MONTO"].mean(), df_ap["MONTO"].median(),
         df_ap["MONTO"].std(),  df_ap["MONTO"].var(),
         df_ap["MONTO"].min(),  df_ap["MONTO"].max()] +
        [df_ap["MONTO"].quantile(p/100) for p in pcts]
    ),
    "Fraudes (F)": (
        [df_f["MONTO"].mean(), df_f["MONTO"].median(),
         df_f["MONTO"].std(),  df_f["MONTO"].var(),
         df_f["MONTO"].min(),  df_f["MONTO"].max()] +
        [df_f["MONTO"].quantile(p/100) for p in pcts]
    ) if len(df_f) > 0 else [0]*13,
    "No Fraude": (
        [df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].mean(),
         df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].median(),
         df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].std(),
         df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].var(),
         df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].min(),
         df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].max()] +
        [df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].quantile(p/100) for p in pcts]
    ),
})
# Formatear
for col in ["Total (aprobadas)","Fraudes (F)","No Fraude"]:
    stats_monto[col] = stats_monto[col].apply(lambda v: f"S/ {v:,.2f}")

# ── 4. RESUMEN POR DIMENSIONES ────────────────────────────────────────────────
def tabla_dimension(col, nombre_col):
    return (
        df.groupby(col)
        .agg(
            Trx        =(col,         "count"),
            Fraudes    =("ES_FRAUDE", "sum"),
            FR_pct     =("ES_FRAUDE", lambda x: x.mean()*100),
            Monto_total=("MONTO",     "sum"),
        )
        .sort_values("Fraudes", ascending=False)
        .reset_index()
        .rename(columns={col: nombre_col, "FR_pct":"Fraud Rate %"})
        .assign(**{
            "Fraud Rate %":  lambda x: x["Fraud Rate %"].map("{:.2f}%".format),
            "Monto total":   lambda x: x["Monto_total"].map("S/ {:,.2f}".format),
        })
        .drop(columns=["Monto_total"])
    )

df_por_canal    = tabla_dimension("CANAL",        "Canal")
df_por_segmento = tabla_dimension("SEG_NOMBRE",   "Segmento")
df_por_seguro   = tabla_dimension("SEGURO",       "Seguridad Comercio")
df_por_tipo     = tabla_dimension("TIPO_PRODUCTO","Tipo Producto")
df_por_cvv      = tabla_dimension("COD_RED_LABEL","Código CVV")
df_por_pais     = tabla_dimension("PAIS",         "País").head(10)

# =============================================================================
# EXPORTAR A EXCEL
# =============================================================================
print("\nExportando a Excel...")

with pd.ExcelWriter(RUTA_EXCEL_OUT, engine="openpyxl") as writer:

    # ── HOJA 1: RESUMEN EJECUTIVO ─────────────────────────────────────────────
    sheet = "1_Resumen_Ejecutivo"
    df_resumen_ejecutivo.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]

    estilizar_header(ws, 1, 2, f"MCC 7995 — JUEGOS DE AZAR | Análisis: {datetime.today().strftime('%d/%m/%Y')}")
    estilizar_header(ws, 2, 2, "RESUMEN EJECUTIVO DE TRANSACCIONALIDAD", fill=FILL_SUBHEAD)

    # Estilizar encabezados y filas
    for row in ws.iter_rows(min_row=4, max_row=4):
        for c in row:
            c.fill = FILL_SUBHEAD; c.font = FONT_HEADER; c.alignment = ALIGN_CENTER

    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=1):
        fill = FILL_FILA_A if i % 2 == 0 else PatternFill()
        # Destacar filas de fraude en rojo suave
        if row[0].value and "fraude" in str(row[0].value).lower():
            fill = PatternFill("solid", fgColor="FCE4D6")
        for c in row:
            c.fill = fill; c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    # Umbral de control
    fila_u = ws.max_row + 2
    estilizar_header(ws, fila_u, 3, "UMBRAL DE CONTROL — Ratio Trx Fraudulentas", fill=FILL_SUBHEAD)
    df_umbral.to_excel(writer, sheet_name=sheet, index=False, startrow=fila_u)

    # Interpretación automática
    fila_interp = ws.max_row + 2
    riesgo_texto = {
        "REGULAR <1%": "El fraud rate se encuentra en nivel REGULAR (<1%). El segmento presenta bajo riesgo relativo.",
        "MEDIO [1%-5%]": "⚠ El fraud rate está en nivel MEDIO (1%–5%). Se recomienda monitoreo activo y revisión de reglas.",
        "ALTO ≥5%": "🔴 ALERTA: El fraud rate supera el 5%. El segmento requiere intervención inmediata.",
    }
    agregar_interpretacion(ws, fila_interp, 3,
        f"Nivel de riesgo actual: {nivel_riesgo}. {riesgo_texto[nivel_riesgo]}")
    agregar_interpretacion(ws, fila_interp+1, 3,
        f"El ticket promedio de fraude (S/ {ticket_f:,.2f}) es "
        f"{'MAYOR' if ticket_f > ticket_ap else 'MENOR'} al ticket general (S/ {ticket_ap:,.2f}). "
        f"{'Las transacciones fraudulentas son de mayor valor — típico de ataques dirigidos.' if ticket_f > ticket_ap else 'Los fraudes son de montos bajos — posible testeo de tarjetas.'}")

    ajustar_columnas(ws)

    # ── HOJA 2: DECILES ──────────────────────────────────────────────────────
    sheet = "2_Deciles"
    tabla_deciles.to_excel(writer, sheet_name=sheet, startrow=3)
    ws = writer.sheets[sheet]

    estilizar_header(ws, 1, len(tabla_deciles.columns)+2,
        "MCC 7995 — TABLA DE DECILES POR MONTO (D01=menor monto, D10=mayor monto)")
    estilizar_header(ws, 2, len(tabla_deciles.columns)+2,
        "Cada decil contiene ~10% de las transacciones ordenadas de menor a mayor monto",
        fill=FILL_SUBHEAD)
    for row in ws.iter_rows(min_row=4, max_row=4):
        for c in row:
            c.fill = FILL_SUBHEAD; c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=1):
        fill = FILL_FILA_A if i % 2 == 0 else PatternFill()
        for c in row:
            c.fill = fill; c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    # Interpretaciones automáticas de deciles
    decil_max_fraude = tabla_deciles["Fraude Trx Aprobado"].idxmax()
    decil_max_monto  = tabla_deciles["Fraude Monto Aprobado"].idxmax()
    fila_di = ws.max_row + 2
    agregar_interpretacion(ws, fila_di,   len(tabla_deciles.columns)+2,
        f"El decil con mayor NÚMERO de fraudes es {decil_max_fraude} "
        f"({tabla_deciles.loc[decil_max_fraude,'Fraude Trx Aprobado']:,} fraudes) — "
        f"rango de monto: {tabla_deciles.loc[decil_max_fraude,'Rango S/']}")
    agregar_interpretacion(ws, fila_di+1, len(tabla_deciles.columns)+2,
        f"El decil con mayor MONTO de fraude es {decil_max_monto} "
        f"(S/ {tabla_deciles.loc[decil_max_monto,'Fraude Monto Aprobado']:,.2f}) — "
        f"rango de monto: {tabla_deciles.loc[decil_max_monto,'Rango S/']}")
    agregar_interpretacion(ws, fila_di+2, len(tabla_deciles.columns)+2,
        "Si el fraude se concentra en D10: las transacciones de mayor monto son el principal riesgo. "
        "Si se concentra en D01–D03: hay testeo de tarjetas con montos bajos.")

    ajustar_columnas(ws)

    # ── HOJA 3: TOP COMERCIOS ─────────────────────────────────────────────────
    sheet = "3_Top_Comercios"
    df_top_comercios.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]

    estilizar_header(ws, 1, len(df_top_comercios.columns),
        "MCC 7995 — TOP 10 COMERCIOS CON MAYOR MONTO DE FRAUDE")
    estilizar_header(ws, 2, len(df_top_comercios.columns),
        "Ordenado por monto fraudulento acumulado", fill=FILL_SUBHEAD)
    for row in ws.iter_rows(min_row=4, max_row=4):
        for c in row:
            c.fill = FILL_SUBHEAD; c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=1):
        fill = FILL_FILA_A if i % 2 == 0 else PatternFill()
        for c in row:
            c.fill = fill; c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    top1 = df_top_comercios.iloc[0]["Comercio"] if len(df_top_comercios) > 0 else "N/A"
    fila_tc = ws.max_row + 2
    agregar_interpretacion(ws, fila_tc, len(df_top_comercios.columns),
        f"El comercio con mayor concentración de fraude en monto es '{top1}'. "
        f"Si un comercio concentra >50% del monto total de fraude, es candidato prioritario para regla específica.")
    ajustar_columnas(ws)

    # ── HOJA 4: ESTADÍSTICAS DE MONTO ────────────────────────────────────────
    sheet = "4_Estadisticas_Monto"
    stats_monto.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]

    estilizar_header(ws, 1, len(stats_monto.columns),
        "MCC 7995 — ESTADÍSTICAS DE MONTO: FRAUDE vs NO FRAUDE")
    estilizar_header(ws, 2, len(stats_monto.columns),
        "Comparación de distribución de montos entre transacciones fraudulentas y legítimas",
        fill=FILL_SUBHEAD)
    for row in ws.iter_rows(min_row=4, max_row=4):
        for c in row:
            c.fill = FILL_SUBHEAD; c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=1):
        fill = FILL_FILA_A if i % 2 == 0 else PatternFill()
        for c in row:
            c.fill = fill; c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    med_f   = df_f["MONTO"].median() if len(df_f) > 0 else 0
    med_nof = df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].median()
    p90_f   = df_f["MONTO"].quantile(0.90) if len(df_f) > 0 else 0
    p90_nof = df_ap[df_ap["ES_FRAUDE"]==0]["MONTO"].quantile(0.90)

    fila_st = ws.max_row + 2
    agregar_interpretacion(ws, fila_st, len(stats_monto.columns),
        f"La mediana de monto en fraude (S/ {med_f:,.2f}) es "
        f"{'MAYOR' if med_f > med_nof else 'MENOR'} a la mediana legítima (S/ {med_nof:,.2f}). "
        f"Diferencia: S/ {abs(med_f - med_nof):,.2f}.")
    agregar_interpretacion(ws, fila_st+1, len(stats_monto.columns),
        f"El P90 del fraude (S/ {p90_f:,.2f}) vs P90 legítimo (S/ {p90_nof:,.2f}). "
        f"Si el P90 del fraude es mucho mayor, los ataques se dan en la cola alta de montos — "
        f"un umbral de monto puede ser efectivo como regla de control.")
    agregar_interpretacion(ws, fila_st+2, len(stats_monto.columns),
        "La desviación estándar alta en fraude indica heterogeneidad — "
        "no todos los fraudes siguen el mismo patrón de monto. "
        "Combinar umbral de monto con otras variables (canal, CVV, velocidad) mejora la precisión.")
    ajustar_columnas(ws)

    # ── HOJA 5: DIMENSIONES ───────────────────────────────────────────────────
    sheet = "5_Por_Dimension"
    ws5 = writer.book.create_sheet(sheet)
    writer.sheets[sheet] = ws5
    fila_actual = 1

    dims_export = [
        (df_por_canal,    "Por Canal",             "El canal con mayor fraude concentrado indica el vector de ataque principal."),
        (df_por_segmento, "Por Segmento Cliente",  "Si el fraude se concentra en un segmento, la regla puede ser más quirúrgica."),
        (df_por_seguro,   "Por Seguridad Comercio","'No Seguro' indica ausencia de 3DSecure — vector de ataque frecuente en e-commerce."),
        (df_por_tipo,     "Por Tipo Producto",     "TD vs TC puede tener distinto perfil de fraude y distintas reglas aplicables."),
        (df_por_cvv,      "Por Código CVV/Red",    "CVV Dinámico es más seguro. Alta concentración en Estático o No Match es señal de alerta."),
        (df_por_pais,     "Por País Origen (Top10)","Transacciones desde países distintos al Perú son señal de alto riesgo en este MCC."),
    ]

    for df_dim, titulo, interp in dims_export:
        estilizar_header(ws5, fila_actual, len(df_dim.columns)+1, f"MCC 7995 — {titulo.upper()}")
        fila_actual += 1
        # Encabezados
        for j, col in enumerate(df_dim.columns, start=1):
            c = ws5.cell(row=fila_actual, column=j, value=col)
            c.fill = FILL_SUBHEAD; c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN
        fila_actual += 1
        # Datos
        for i, row in df_dim.iterrows():
            fill = FILL_FILA_A if i % 2 == 0 else PatternFill()
            for j, val in enumerate(row, start=1):
                c = ws5.cell(row=fila_actual, column=j, value=val)
                c.fill = fill; c.font = FONT_NORMAL
                c.alignment = ALIGN_CENTER; c.border = BORDER_THIN
            fila_actual += 1
        # Interpretación
        ws5.merge_cells(start_row=fila_actual, start_column=1,
                        end_row=fila_actual,   end_column=len(df_dim.columns)+1)
        c = ws5.cell(row=fila_actual, column=1, value=f"📌 {interp}")
        c.fill = FILL_AMARILLO; c.font = FONT_INTERP
        c.alignment = ALIGN_LEFT; c.border = BORDER_THIN
        ws5.row_dimensions[fila_actual].height = 30
        fila_actual += 3  # espacio entre tablas

    ajustar_columnas(ws5)

    # ── HOJA 6: VELOCIDAD ─────────────────────────────────────────────────────
    sheet = "6_Velocidad_Intervalo"
    tabla_vel.reset_index().to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]

    estilizar_header(ws, 1, len(tabla_vel.columns)+1,
        "MCC 7995 — VELOCIDAD: INTERVALO ENTRE TRANSACCIONES DEL MISMO CLIENTE")
    estilizar_header(ws, 2, len(tabla_vel.columns)+1,
        "¿En qué ventana de tiempo se concentran los ataques fraudulentos?",
        fill=FILL_SUBHEAD)
    for row in ws.iter_rows(min_row=4, max_row=4):
        for c in row:
            c.fill = FILL_SUBHEAD; c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN
    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=1):
        fill = FILL_FILA_A if i % 2 == 0 else PatternFill()
        for c in row:
            c.fill = fill; c.font = FONT_NORMAL
            c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

    bucket_max = tabla_vel["FR_pct"].idxmax() if len(tabla_vel) > 0 else "N/A"
    pct_vel    = tabla_vel.loc[bucket_max,"Pct_fraudes_total"] if len(tabla_vel) > 0 else 0
    fila_v = ws.max_row + 2
    agregar_interpretacion(ws, fila_v, len(tabla_vel.columns)+1,
        f"El intervalo con mayor fraud rate es '{bucket_max}' "
        f"({pct_vel:.1f}% del total de fraudes). "
        f"Si este bucket es ≤5 min, indica ataques automatizados (bots) sobre el mismo cliente.")
    agregar_interpretacion(ws, fila_v+1, len(tabla_vel.columns)+1,
        "Una regla de velocidad efectiva bloquea al cliente si hace N transacciones "
        "en el intervalo de mayor concentración de fraude detectado en esta tabla.")
    ajustar_columnas(ws)

print(f"\n✅ Excel exportado exitosamente: {RUTA_EXCEL_OUT}")
print("   Hojas generadas:")
print("   1_Resumen_Ejecutivo | 2_Deciles | 3_Top_Comercios")
print("   4_Estadisticas_Monto | 5_Por_Dimension | 6_Velocidad_Intervalo")
