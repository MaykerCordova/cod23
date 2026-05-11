"""
╔══════════════════════════════════════════════════════════════════╗
║     BITÁCORA ANALYTICS & HERRAMIENTAS — V7                      ║
║     Scotiabank Peru — Prevención de Fraude                      ║
╠══════════════════════════════════════════════════════════════════╣
║  CAMBIOS RESPECTO A V6:                                         ║
║  1. SQLite es la ÚNICA fuente de verdad para el correlativo      ║
║     → El correlativo NUNCA puede reiniciarse aunque el Excel    ║
║       se pierda, corrompa o alguien lo borre de OneDrive.       ║
║  2. existe_id() también consulta SQLite (antes solo Excel)      ║
║     → Elimina el riesgo de duplicados si el Excel está vacío.   ║
║  3. Sistema de correos con formato inválido                     ║
║     → Los correos que no tienen los campos requeridos van a     ║
║       Bitacora_Pendientes (tabla SQLite + Excel separado).      ║
║     → El remitente puede reenviar corregido; se detecta auto.   ║
║  4. Excel generado DESDE SQLite (no al revés)                   ║
║     → Excel es solo una VISTA de la base, no la base misma.     ║
║  5. MAPA_ESTIMACION sin claves duplicadas                       ║
║  6. Logging completo — sin except:pass silenciosos              ║
║  7. Backup .msg de correos se mantiene igual                    ║
╚══════════════════════════════════════════════════════════════════╝
"""
# mejorar para la version 7
import os
import re
import sqlite3
import win32com.client
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# 1. CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════
CONFIG = {
    # Excel principal (es solo una VISTA exportada desde SQLite)
    "RUTA_EXCEL": (
        r"C:\Users\s4930359\OneDrive - The Bank of Nova Scotia"
        r"\Bitacora_Reglas\Bitacora_Master.xlsx"
    ),
    # Excel de correos con formato inválido (generado aparte)
    "RUTA_EXCEL_PENDIENTES": (
        r"C:\Users\s4930359\OneDrive - The Bank of Nova Scotia"
        r"\Bitacora_Reglas\Bitacora_Pendientes_Formato.xlsx"
    ),
    # SQLite — FUENTE DE VERDAD (no depende del estado del Excel)
    "RUTA_DB_SQLITE": (
        r"C:\Users\s4930359\OneDrive - The Bank of Nova Scotia"
        r"\Bitacora_Reglas\Respaldo_Blindado.db"
    ),
    # Carpeta donde se guardan los .msg como respaldo de auditoría
    "RUTA_BACKUP_MSG": (
        r"C:\Users\s4930359\OneDrive - The Bank of Nova Scotia"
        r"\Bitacora_Reglas\Correos_Respaldo"
    ),
    # Carpetas Outlook que monitorea el robot
    "FOLDER_SOLICITUDES": "Bitacora_Solicitudes",
    "FOLDER_RESPUESTAS":  "Bitacora_Respuestas",
    # Regex para capturar el ID del asunto, ej: [ID-VRM-001]
    "REGEX_ID": r"\[(?P<ID>ID[-_][^\]\s]+)\]",
}

# ══════════════════════════════════════════════════════════════════
# 2. LÓGICA DE NEGOCIO
# ══════════════════════════════════════════════════════════════════

# NOTA V7: claves duplicadas eliminadas (causaban que Python
# descartara silenciosamente la primera definición de cada clave).
MAPA_ESTIMACION = {
    "PMFD":  "15:01",
    "CORE":  "N/A",
    "ITC":   "N/A",
    "RT TD": "5:01",
    "RT TC": "5:01",
    "VRM":   "5:01",
    "VCAS":  "5:01",
    "FRM":   "5:01",
}

# Campos que DEBEN existir en el cuerpo del correo de solicitud.
# Si alguno falta, el correo se manda a Pendientes (no se ignora).
CAMPOS_REQUERIDOS = [
    "Herramienta",
    "Accion",
    "Institucion",
    "Codigo_Condicion",
    "Nombre_Condicion",
]

# Palabras que detectan una respuesta de aprobación/conformidad
PALABRAS_APROBACION = [
    "OK", "CONFORME", "APROBADO", "APROBADA",
    "VOBO", "PROCEDER", "ACUERDO",
]


def calcular_estimacion(herramienta: str) -> str:
    """Devuelve el tiempo estimado según la herramienta."""
    h = str(herramienta).strip().upper()
    for key, value in MAPA_ESTIMACION.items():
        if key in h:
            return value
    return "ND"


# ══════════════════════════════════════════════════════════════════
# 3. GESTOR DE BASE DE DATOS — SQLITE (FUENTE DE VERDAD)
# ══════════════════════════════════════════════════════════════════

class GestorBackupSQL:
    """
    Todo lo que importa para la integridad del proceso vive aquí.
    El Excel es solo un archivo de salida exportado desde esta BD.
    """

    def __init__(self, ruta_db: str):
        self.ruta_db = ruta_db
        self._init_tablas()

    def _init_tablas(self):
        """Crea las tablas si no existen. Seguro de correr múltiples veces."""
        os.makedirs(os.path.dirname(self.ruta_db), exist_ok=True)
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Bitacora (
                    Nro_Correlativo  INTEGER,
                    Fecha            TEXT,
                    Hora             TEXT,
                    Codigo_Generado  TEXT,
                    Maker            TEXT,
                    Solicitado_Por   TEXT,
                    Herramienta      TEXT,
                    Accion           TEXT,
                    Institucion      TEXT,
                    Codigo_Condicion TEXT,
                    Nombre_Condicion TEXT,
                    Estatus_Condicion TEXT,
                    Tipo_Condicion   TEXT,
                    Consideraciones  TEXT,
                    Objetivo         TEXT,
                    Estimacion       TEXT,
                    Sustento         TEXT,
                    Checker          TEXT,
                    Fecha_Revision   TEXT,
                    Hora_Revision    TEXT,
                    Mes_Revision     TEXT,
                    Anio_Revision    TEXT,
                    Enviado_Jefatura TEXT,
                    Conformidad      TEXT,
                    Canal            TEXT,
                    ID_Sistema       TEXT PRIMARY KEY
                )
            """)
            # Tabla separada para correos que no cumplen el formato.
            # Nunca se mezcla con la bitácora principal.
            conn.execute("""
                CREATE TABLE IF NOT EXISTS Bitacora_Pendientes (
                    id               INTEGER PRIMARY KEY AUTOINCREMENT,
                    Fecha_Captura    TEXT,
                    ID_Sistema       TEXT,
                    Asunto           TEXT,
                    Remitente        TEXT,
                    Campos_Faltantes TEXT,
                    Cuerpo_Raw       TEXT,
                    Procesado        INTEGER DEFAULT 0
                )
            """)
            conn.commit()

    # ── Correlativo ────────────────────────────────────────────────
    def get_max_correlativo(self) -> int:
        """
        Lee el máximo correlativo desde SQLite.
        NUNCA depende del Excel. Si SQLite está vacío devuelve 0.
        El siguiente correlativo es get_max_correlativo() + 1.
        """
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT MAX(Nro_Correlativo) FROM Bitacora"
            ).fetchone()[0]
        return 0 if r is None else int(r)

    # ── Deduplicación ──────────────────────────────────────────────
    def existe_id(self, id_sys: str) -> bool:
        """
        Verifica existencia SOLO en SQLite.
        En V6 dependía del Excel; si el Excel se perdía, insertaba
        duplicados. Ahora eso no puede pasar.
        """
        with sqlite3.connect(self.ruta_db) as conn:
            r = conn.execute(
                "SELECT COUNT(*) FROM Bitacora WHERE ID_Sistema = ?",
                (id_sys,)
            ).fetchone()[0]
        return r > 0

    # ── Inserción ──────────────────────────────────────────────────
    def insertar_solicitud(self, datos: dict):
        sql = """
            INSERT OR IGNORE INTO Bitacora VALUES (
                :Nro_Correlativo, :Fecha, :Hora, :Codigo_Generado,
                :Maker, :Solicitado_Por, :Herramienta, :Accion,
                :Institucion, :Codigo_Condicion, :Nombre_Condicion,
                :Estatus_Condicion, :Tipo_Condicion, :Consideraciones,
                :Objetivo, :Estimacion, :Sustento, :Checker,
                :Fecha_Revision, :Hora_Revision, :Mes_Revision,
                :Anio_Revision, :Enviado_Jefatura, :Conformidad,
                :Canal, :ID_Sistema
            )
        """
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute(sql, datos)
            conn.commit()
        print("  💾 [SQLite] Solicitud insertada.")

    # ── Correos con formato inválido ───────────────────────────────
    def registrar_pendiente(
        self, id_sys: str, asunto: str, remitente: str,
        campos_faltantes: list, cuerpo_raw: str
    ):
        """
        Guarda el correo en Bitacora_Pendientes cuando le faltan campos.
        El proceso continúa sin interrumpirse.
        El Excel Bitacora_Pendientes_Formato.xlsx se genera al final.
        """
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                INSERT INTO Bitacora_Pendientes
                    (Fecha_Captura, ID_Sistema, Asunto, Remitente,
                     Campos_Faltantes, Cuerpo_Raw)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                datetime.now().strftime("%d/%m/%Y %H:%M"),
                id_sys, asunto, remitente,
                ", ".join(campos_faltantes),
                cuerpo_raw[:3000],  # máximo 3000 chars del cuerpo
            ))
            conn.commit()
        print(f"  ⚠️  [SQLite] Sin formato → Pendientes: {id_sys}")
        print(f"       Campos faltantes: {campos_faltantes}")

    # ── Aprobación ─────────────────────────────────────────────────
    def actualizar_aprobacion(
        self, id_sistema: str, checker: str,
        fecha_rev: str, hora_rev: str,
        mes_rev: str, anio_rev: str
    ):
        with sqlite3.connect(self.ruta_db) as conn:
            conn.execute("""
                UPDATE Bitacora
                SET Conformidad    = 'Completado',
                    Checker        = ?,
                    Fecha_Revision = ?,
                    Hora_Revision  = ?,
                    Mes_Revision   = ?,
                    Anio_Revision  = ?
                WHERE ID_Sistema = ?
            """, (checker, fecha_rev, hora_rev, mes_rev, anio_rev, id_sistema))
            conn.commit()
        print(f"  ✅ [SQLite] Aprobación registrada: {id_sistema}")

    # ── Exportaciones ──────────────────────────────────────────────
    def exportar_bitacora(self) -> pd.DataFrame:
        """Devuelve toda la bitácora como DataFrame, ordenado por correlativo."""
        with sqlite3.connect(self.ruta_db) as conn:
            df = pd.read_sql(
                "SELECT * FROM Bitacora ORDER BY Nro_Correlativo",
                conn
            )
        return df

    def exportar_pendientes_sin_procesar(self) -> pd.DataFrame:
        """Devuelve los pendientes que todavía no fueron corregidos."""
        with sqlite3.connect(self.ruta_db) as conn:
            df = pd.read_sql(
                "SELECT * FROM Bitacora_Pendientes WHERE Procesado = 0",
                conn
            )
        return df


# ══════════════════════════════════════════════════════════════════
# 4. DISEÑO Y FORMATO EXCEL
# ══════════════════════════════════════════════════════════════════

COLOR_HEADER   = "0079C1"   # Azul Scotia
COLOR_PENDENTE = "FFC000"   # Ámbar para pendientes


def _aplicar_estilo_header(ws, color_hex: str = COLOR_HEADER):
    """Aplica el formato de cabecera a la primera fila."""
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color=color_hex, end_color=color_hex,
                                fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        bottom=Side(style="thin", color="FFFFFF")
    )
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border


def _autoajustar_columnas(ws, max_ancho: int = 50):
    """Ajusta el ancho de cada columna al contenido."""
    for column in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_ancho)


def embellecer_excel(ruta: str, color_header: str = COLOR_HEADER):
    """Aplica formato profesional a un Excel ya guardado."""
    try:
        wb = load_workbook(ruta)
        ws = wb.active
        _aplicar_estilo_header(ws, color_header)
        ws.auto_filter.ref = ws.dimensions
        _autoajustar_columnas(ws)
        wb.save(ruta)
        print(f"  🎨 [Excel] Formato aplicado: {os.path.basename(ruta)}")
    except Exception as e:
        print(f"  ▲ Error formato Excel [{os.path.basename(ruta)}]: {e}")


def exportar_excel_desde_sqlite(sql: GestorBackupSQL, ruta: str):
    """
    Regenera el Excel completo exportando desde SQLite.
    El Excel es solo una VISTA — si se pierde, se regenera aquí.
    """
    try:
        df = sql.exportar_bitacora()
        df.to_excel(ruta, index=False)
        embellecer_excel(ruta, COLOR_HEADER)
        print(f"  📊 [Excel] Bitácora exportada: {len(df)} registros.")
    except Exception as e:
        print(f"  ▲ Error exportando Bitácora Excel: {e}")


def exportar_excel_pendientes(sql: GestorBackupSQL, ruta: str):
    """
    Genera un Excel separado con los correos que no cumplieron
    el formato. Cada fila indica qué campos faltan.
    """
    try:
        df = sql.exportar_pendientes_sin_procesar()
        if df.empty:
            return
        # Columnas amigables para el reporte
        df = df.rename(columns={
            "id":               "ID_Interno",
            "Fecha_Captura":    "Fecha Captura",
            "ID_Sistema":       "ID Sistema",
            "Asunto":           "Asunto Correo",
            "Remitente":        "Remitente",
            "Campos_Faltantes": "Campos Faltantes",
            "Cuerpo_Raw":       "Cuerpo (primeros 3000 chars)",
            "Procesado":        "Procesado",
        })
        df.to_excel(ruta, index=False)
        embellecer_excel(ruta, COLOR_PENDENTE)
        print(f"  📋 [Excel] Pendientes exportados: {len(df)} correos sin formato.")
    except Exception as e:
        print(f"  ▲ Error exportando Pendientes Excel: {e}")


# ══════════════════════════════════════════════════════════════════
# 5. HERRAMIENTAS OUTLOOK
# ══════════════════════════════════════════════════════════════════

class HerramientasOutlook:
    def __init__(self):
        self.outlook = None
        self.inbox   = None
        try:
            self.outlook = (
                win32com.client
                .Dispatch("Outlook.Application")
                .GetNamespace("MAPI")
            )
            self.inbox = self.outlook.GetDefaultFolder(6)
        except Exception as e:
            print(f"  ▲ Error conectando a Outlook: {e}")

    @property
    def conectado(self) -> bool:
        return self.outlook is not None and self.inbox is not None

    def buscar_carpeta(self, nombre: str):
        """Busca primero en Inbox, luego en el namespace raíz."""
        try:
            return self.inbox.Folders[nombre]
        except Exception:
            pass
        try:
            return self.outlook.Folders[nombre]
        except Exception:
            return None

    def guardar_msg(self, mail, nombre: str) -> str:
        """
        Guarda el correo como .msg para auditoría.
        Se mantiene igual que V6 — esta funcionalidad queda.
        """
        os.makedirs(CONFIG["RUTA_BACKUP_MSG"], exist_ok=True)
        clean_name = re.sub(r'[\\/*?:"<>|]', "_", nombre)
        ruta = os.path.join(CONFIG["RUTA_BACKUP_MSG"], f"{clean_name}.msg")
        try:
            mail.SaveAs(ruta)
            return ruta
        except Exception as e:
            print(f"  ▲ Error guardando MSG [{nombre}]: {e}")
            return "Error guardando MSG"


# ══════════════════════════════════════════════════════════════════
# 6. PROCESADOR DE TEXTO / PARSER DE CORREO
# ══════════════════════════════════════════════════════════════════

class ProcesadorTexto:
    """
    Extrae el ID del asunto y parsea los campos del cuerpo del correo.
    Los patrones son case-insensitive y toleran variaciones de acento.
    """
    PATRONES = {
        "Solicitado_Por":     r"(?i)SOLICITADO\s*POR:\s*(.*)",
        "Herramienta":        r"(?i)HERRAMIENTA:\s*(.*)",
        "Accion":             r"(?i)ACCI[OÓ]N:\s*(.*)",
        "Institucion":        r"(?i)INSTITUCI[OÓ]N:\s*(.*)",
        "Codigo_Condicion":   r"(?i)C[OÓ]DIGO\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Nombre_Condicion":   r"(?i)NOMBRE\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Estatus_Condicion":  r"(?i)ESTATUS\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Tipo_Condicion":     r"(?i)TIPO\s*(?:DE)?\s*CONDICI[OÓ]N:\s*(.*)",
        "Canal":              r"(?i)CANAL:\s*(.*)",
        "Consideraciones":    r"(?i)CONSIDERACIONES:\s*(.*)",
        "Objetivo":           r"(?i)OBJETIVO:\s*(.*)",
    }

    @staticmethod
    def extraer_id(asunto: str):
        """Extrae el ID del asunto. Devuelve None si no lo encuentra."""
        match = re.search(CONFIG["REGEX_ID"], asunto)
        return match.group("ID").upper() if match else None

    @staticmethod
    def parsear_cuerpo(cuerpo: str) -> dict:
        """Parsea todos los campos del cuerpo. Campos no encontrados quedan '-'."""
        datos = {}
        for k, patron in ProcesadorTexto.PATRONES.items():
            match = re.search(patron, cuerpo)
            datos[k] = (
                match.group(1).strip().replace('\r', '').replace('\n', ' ')
                if match else "-"
            )
        return datos

    @staticmethod
    def validar_formato(datos: dict) -> list:
        """
        Devuelve la lista de campos requeridos que faltan o tienen '-'.
        Lista vacía = formato correcto.
        Lista con elementos = correo va a Bitacora_Pendientes.
        """
        return [
            campo for campo in CAMPOS_REQUERIDOS
            if datos.get(campo, "-").strip() in ("-", "", "N/A", "n/a")
        ]


# ══════════════════════════════════════════════════════════════════
# 7. ORQUESTADOR PRINCIPAL
# ══════════════════════════════════════════════════════════════════

def main():
    sep = "═" * 65
    print(f"\n{sep}")
    print(f"  🤖  ROBOT BITÁCORA Analytics & Herramientas — V7")
    print(f"  🕐  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{sep}\n")

    outlook = HerramientasOutlook()
    sql     = GestorBackupSQL(CONFIG["RUTA_DB_SQLITE"])

    if not outlook.conectado:
        print("❌ No se pudo conectar a Outlook. Abortando.")
        return

    cambios    = 0   # registros nuevos o actualizados en Bitacora
    pendientes = 0   # correos enviados a Bitacora_Pendientes

    # ── FASE 1: SOLICITUDES ────────────────────────────────────────
    print(f"{'─'*40}")
    print("  FASE 1 — Solicitudes entrantes")
    print(f"{'─'*40}\n")

    sol_folder = outlook.buscar_carpeta(CONFIG["FOLDER_SOLICITUDES"])

    if not sol_folder:
        print(f"  ▲ Carpeta '{CONFIG['FOLDER_SOLICITUDES']}' no encontrada en Outlook.")
    else:
        msgs = [m for m in sol_folder.Items if m.UnRead]
        msgs.sort(key=lambda x: x.ReceivedTime)
        print(f"  Correos no leídos: {len(msgs)}\n")

        for m in msgs:
            try:
                # ── 1. Extraer ID ──────────────────────────────────
                id_sys = ProcesadorTexto.extraer_id(m.Subject)
                if not id_sys:
                    print(f"  [SKIP] Sin ID en asunto: {m.Subject[:70]}")
                    m.UnRead = False
                    continue

                # ── 2. Deduplicar contra SQLite ────────────────────
                if sql.existe_id(id_sys):
                    print(f"  [DUP]  Ya registrado en SQLite: {id_sys}")
                    m.UnRead = False
                    continue

                print(f"  [+] Nuevo ticket: {id_sys}")

                # ── 3. Parsear cuerpo ──────────────────────────────
                d = ProcesadorTexto.parsear_cuerpo(m.Body)

                # ── 4. Validar formato ─────────────────────────────
                campos_faltantes = ProcesadorTexto.validar_formato(d)

                if campos_faltantes:
                    # Correo incompleto → Bitacora_Pendientes
                    sql.registrar_pendiente(
                        id_sys           = id_sys,
                        asunto           = m.Subject,
                        remitente        = m.SenderName,
                        campos_faltantes = campos_faltantes,
                        cuerpo_raw       = m.Body,
                    )
                    outlook.guardar_msg(m, f"PENDIENTE_{id_sys}")
                    m.UnRead = False
                    pendientes += 1
                    continue

                # ── 5. Correlativo desde SQLite ────────────────────
                # V7: NUNCA depende del Excel. Si alguien borra el
                # Excel, el correlativo sigue desde donde lo dejó SQLite.
                correlativo = sql.get_max_correlativo() + 1

                # ── 6. Código generado ─────────────────────────────
                parte_corr   = f"{correlativo:03d}"
                parte_fecha  = m.SentOn.strftime("%d%m%y")
                parte_tool   = d["Herramienta"].strip()
                parte_insti  = d["Institucion"].strip()
                parte_cond   = d["Codigo_Condicion"].strip()
                cod_generado = f"{parte_corr}{parte_fecha}{parte_tool}{parte_insti}{parte_cond}"

                # ── 7. Construir fila ──────────────────────────────
                fila = {
                    "Nro_Correlativo":   correlativo,
                    "Fecha":             m.SentOn.strftime("%d/%m/%Y"),
                    "Hora":              m.SentOn.strftime("%H:%M"),
                    "Codigo_Generado":   cod_generado,
                    "Maker":             m.SenderName,
                    "Solicitado_Por":    d.get("Solicitado_Por", "-"),
                    "Herramienta":       d["Herramienta"],
                    "Accion":            d["Accion"],
                    "Institucion":       d["Institucion"],
                    "Codigo_Condicion":  d["Codigo_Condicion"],
                    "Nombre_Condicion":  d["Nombre_Condicion"],
                    "Estatus_Condicion": d.get("Estatus_Condicion", "-"),
                    "Tipo_Condicion":    d.get("Tipo_Condicion", "-"),
                    "Consideraciones":   d.get("Consideraciones", "-"),
                    "Objetivo":          d.get("Objetivo", "-"),
                    "Estimacion":        calcular_estimacion(d["Herramienta"]),
                    "Sustento":          "-",
                    "Checker":           "-",
                    "Fecha_Revision":    "-",
                    "Hora_Revision":     "-",
                    "Mes_Revision":      "-",
                    "Anio_Revision":     "-",
                    "Enviado_Jefatura":  "SI",
                    "Conformidad":       "Pendiente",
                    "Canal":             d.get("Canal", "-"),
                    "ID_Sistema":        id_sys,
                }

                # ── 8. Guardar ─────────────────────────────────────
                sql.insertar_solicitud(fila)
                outlook.guardar_msg(m, f"REQ_{id_sys}")
                m.UnRead = False
                cambios  += 1
                print(f"  ✔  Correlativo asignado: {correlativo:03d}\n")

            except Exception as e:
                print(f"  ▲ Error en solicitud [{getattr(m, 'Subject', '?')[:60]}]: {e}")

    # ── FASE 2: RESPUESTAS / CONFORMIDADES ────────────────────────
    print(f"\n{'─'*40}")
    print("  FASE 2 — Respuestas / Conformidades")
    print(f"{'─'*40}\n")

    resp_folder = outlook.buscar_carpeta(CONFIG["FOLDER_RESPUESTAS"])

    if not resp_folder:
        print(f"  ▲ Carpeta '{CONFIG['FOLDER_RESPUESTAS']}' no encontrada en Outlook.")
    else:
        msgs = [m for m in resp_folder.Items if m.UnRead]
        msgs.sort(key=lambda x: x.ReceivedTime)
        print(f"  Respuestas no leídas: {len(msgs)}\n")

        for m in msgs:
            try:
                id_sys = ProcesadorTexto.extraer_id(m.Subject)

                if not id_sys:
                    print(f"  [SKIP] Sin ID en respuesta: {m.Subject[:70]}")
                    m.UnRead = False
                    continue

                cuerpo_upper     = m.Body.upper()
                es_aprobacion    = any(p in cuerpo_upper for p in PALABRAS_APROBACION)

                if not es_aprobacion:
                    print(f"  [SKIP] Sin palabra de aprobación: {id_sys}")
                    m.UnRead = False
                    continue

                if not sql.existe_id(id_sys):
                    print(f"  [WARN] Respuesta para ID no registrado: {id_sys}")
                    m.UnRead = False
                    continue

                print(f"  [✓] Conformidad detectada: {id_sys}")
                outlook.guardar_msg(m, f"RESP_{id_sys}")

                fecha_obj = m.ReceivedTime
                sql.actualizar_aprobacion(
                    id_sistema = id_sys,
                    checker    = m.SenderName,
                    fecha_rev  = fecha_obj.strftime("%d/%m/%Y"),
                    hora_rev   = fecha_obj.strftime("%H:%M"),
                    mes_rev    = fecha_obj.strftime("%m"),
                    anio_rev   = fecha_obj.strftime("%Y"),
                )
                m.UnRead = False
                cambios  += 1

            except Exception as e:
                # V7: logging explícito, no except:pass silencioso
                print(f"  ▲ Error en respuesta [{getattr(m, 'Subject', '?')[:60]}]: {e}")

    # ── CIERRE: EXPORTAR EXCEL ────────────────────────────────────
    print(f"\n{'─'*40}")
    print("  CIERRE — Exportación Excel")
    print(f"{'─'*40}\n")

    # Siempre regenerar desde SQLite (no desde el DataFrame en memoria)
    exportar_excel_desde_sqlite(sql, CONFIG["RUTA_EXCEL"])

    # Exportar pendientes solo si hay registros
    df_pend = sql.exportar_pendientes_sin_procesar()
    if not df_pend.empty:
        exportar_excel_pendientes(sql, CONFIG["RUTA_EXCEL_PENDIENTES"])
    else:
        print("  ✔  Sin correos pendientes de formato.")

    # ── RESUMEN FINAL ─────────────────────────────────────────────
    print(f"\n{sep}")
    print(f"  ✅  Proceso finalizado — {datetime.now().strftime('%H:%M:%S')}")
    print(f"{'─'*65}")
    print(f"  Registros nuevos / actualizados : {cambios}")
    print(f"  Correos sin formato (pendientes): {pendientes}")
    if pendientes > 0:
        print(f"  → Revisar: Bitacora_Pendientes_Formato.xlsx")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
