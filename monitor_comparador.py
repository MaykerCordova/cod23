"""
╔══════════════════════════════════════════════════════════════════╗
║   MONITOR COMPARADOR — Script Unificado                         ║
║   Scotiabank Perú — Prevención de Fraude                        ║
║                                                                  ║
║   PROCESO 1: Transmisiones UBA vs SBP (carpeta TRX_MONITOR_JOSE)║
║   PROCESO 2: JOY — Excel FINJOY vs Monitor JOY                  ║
║                                                                  ║
║   Fix locale peruano: usa filtro manual en Python (no Restrict) ║
║   Recuperación automática: procesa fechas pendientes acumuladas  ║
╚══════════════════════════════════════════════════════════════════╝

FLUJO GENERAL (ejecutado vía .bat):
  1. Conecta a Outlook via win32com
  2. PROCESO 1:
     - Lee carpeta TRX_MONITOR_JOSE
     - Parsea correos UBA y SBP
     - Calcula diferencias, actualiza bitácora, envía reporte
  3. PROCESO 2:
     - Lee carpeta FINJOY (Excel adjunto T-1)
     - Lee carpeta MONITOR_JOY (cuerpo del último correo del día)
     - Filtra Excel por EXISTE_EN_EL_MONITOR == SI
     - Calcula diferencias totales y por condición
     - Actualiza bitácora (hojas: DIARIO, POR_CONDICION)
     - Envía reporte con gráficos evolutivos
"""

# ─── IMPORTS ───────────────────────────────────────────────────────────────────
import win32com.client
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
import re
import os
import tempfile
import warnings
warnings.filterwarnings('ignore')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN — AJUSTAR SEGÚN ENTORNO
# ══════════════════════════════════════════════════════════════════════════════

BASE_PATH = r"C:\Users\s4930359\Data_Herramientas\data\Gold"

# Rutas de bitácoras
BITACORA_TRANS = os.path.join(BASE_PATH, "bitacora_transmisiones.xlsx")
BITACORA_JOY   = os.path.join(BASE_PATH, "bitacora_joy.xlsx")

# Carpetas Outlook — Proceso 1
CARPETA_TRANS = "TRX_MONITOR_JOSE"

# Carpetas Outlook — Proceso 2
CARPETA_FINJOY     = "FINJOY"
CARPETA_MONITOR_JOY = "MONITOR_JOY"

# Destinatarios (separados por punto y coma si hay varios)
DESTINATARIOS_TRANS = "correo1@scotiabank.com.pe; correo2@scotiabank.com.pe"
DESTINATARIOS_JOY   = "correo1@scotiabank.com.pe; correo2@scotiabank.com.pe"

# Umbral de alerta para diferencia porcentual
UMBRAL_PCT = 5.0

# Colores corporativos
C = {
    "rojo":     "#EC111A",   # Scotiabank
    "azul":     "#1B3A6B",
    "gris":     "#6B7280",
    "verde":    "#10B981",
    "naranja":  "#F59E0B",
    "fondo":    "#F9FAFB",
    "rojo_lt":  "#FEE2E2",
    "verde_lt": "#D1FAE5",
}


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES — OUTLOOK
# ══════════════════════════════════════════════════════════════════════════════

def conectar_outlook():
    """Retorna el namespace MAPI de Outlook."""
    return win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")


def get_carpeta_outlook(ns, nombre):
    """
    Busca una carpeta por nombre dentro de la Bandeja de Entrada.
    Lanza RuntimeError si no la encuentra.
    """
    inbox = ns.GetDefaultFolder(6)  # 6 = olFolderInbox
    for folder in inbox.Folders:
        if folder.Name.strip().upper() == nombre.strip().upper():
            return folder
    raise RuntimeError(
        f"Carpeta '{nombre}' no encontrada en Outlook. "
        f"Verifica que la regla esté activa y el nombre sea exacto."
    )


def get_mail_items_desde(carpeta, fecha_filtro: date) -> list:
    """
    Retorna lista de MailItems cuya fecha de recepción >= fecha_filtro.
    USA FILTRO MANUAL EN PYTHON (no Restrict) para evitar fallo por locale peruano.
    Orden ascendente (más antiguo primero).
    """
    items = carpeta.Items
    items.Sort("[ReceivedTime]", False)  # False = ascendente

    resultado = []
    for item in items:
        try:
            if item.Class != 43:          # 43 = olMail
                continue
            rt = item.ReceivedTime
            rec_date = rt.date() if hasattr(rt, "date") else rt
            if rec_date >= fecha_filtro:
                resultado.append(item)
        except Exception:
            continue

    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES — BITÁCORA EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def leer_ultima_fecha(path_excel: str, hoja: str = "DIARIO") -> date:
    """
    Lee la fecha máxima de la columna FECHA en la hoja indicada.
    Si el archivo no existe o está vacío, retorna 2026-01-01 como base.
    """
    fecha_base = date(2026, 1, 1)
    if not os.path.exists(path_excel):
        return fecha_base
    try:
        df = pd.read_excel(path_excel, sheet_name=hoja)
        if df.empty or "FECHA" not in df.columns:
            return fecha_base
        return pd.to_datetime(df["FECHA"]).max().date()
    except Exception:
        return fecha_base


def guardar_en_bitacora(path_excel: str, df_nuevo: pd.DataFrame, hoja: str):
    """
    Agrega df_nuevo al final de la hoja indicada en path_excel.
    - Crea el archivo si no existe.
    - Preserva otras hojas si el archivo existe.
    """
    if os.path.exists(path_excel):
        try:
            df_existente = pd.read_excel(path_excel, sheet_name=hoja)
            df_total = pd.concat([df_existente, df_nuevo], ignore_index=True)
        except Exception:
            df_total = df_nuevo.copy()
        mode = "a"
        if_sheet = "replace"
    else:
        df_total = df_nuevo.copy()
        mode = "w"
        if_sheet = "replace"

    with pd.ExcelWriter(path_excel, engine="openpyxl",
                        mode=mode, if_sheet_exists=if_sheet) as writer:
        df_total.to_excel(writer, sheet_name=hoja, index=False)

    print(f"  [BITACORA] {os.path.basename(path_excel)} | hoja={hoja} | +{len(df_nuevo)} fila(s)")


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES — GRÁFICOS
# ══════════════════════════════════════════════════════════════════════════════

def _base_fig(titulo: str):
    """Crea figura base con estilo corporativo. Retorna (fig, ax1, ax2)."""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(13, 7),
                                    facecolor=C["fondo"], gridspec_kw={"hspace": 0.4})
    for ax in (ax1, ax2):
        ax.set_facecolor(C["fondo"])
        ax.grid(alpha=0.25, linestyle="--")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    ax1.set_title(titulo, fontsize=13, fontweight="bold", color=C["azul"], pad=10)
    return fig, ax1, ax2


def _fmt_miles(ax):
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"{int(x):,}"))


def _guardar_fig(fig) -> str:
    """Guarda figura en archivo temporal y retorna la ruta."""
    tmp = tempfile.mktemp(suffix=".png")
    fig.savefig(tmp, dpi=150, bbox_inches="tight", facecolor=C["fondo"])
    plt.close(fig)
    return tmp


def grafico_evolutivo_diario(df: pd.DataFrame,
                              col_a: str, lbl_a: str,
                              col_b: str, lbl_b: str,
                              titulo: str) -> str:
    """Línea doble (arriba) + barras % diferencia (abajo). Últimos 30 días."""
    hoy = pd.Timestamp.today().normalize()
    df30 = df[df["FECHA"] >= (hoy - pd.Timedelta(days=30))].copy()
    if df30.empty:
        df30 = df.copy()

    fig, ax1, ax2 = _base_fig(titulo)

    x = df30["FECHA"]
    ax1.plot(x, df30[col_a], color=C["azul"],   marker="o", lw=2, label=lbl_a, ms=4)
    ax1.plot(x, df30[col_b], color=C["gris"],   marker="s", lw=2, label=lbl_b, ms=4,
             linestyle="--")
    ax1.legend(fontsize=9)
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))
    ax1.xaxis.set_major_locator(mdates.WeekdayLocator())
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right", fontsize=8)
    _fmt_miles(ax1)

    colores_barra = [C["rojo"] if abs(v) > UMBRAL_PCT else C["verde"]
                     for v in df30["PCT_DIFERENCIA"]]
    ax2.bar(x, df30["PCT_DIFERENCIA"].abs(), color=colores_barra, alpha=0.8, width=0.8)
    ax2.axhline(UMBRAL_PCT, color="red", linestyle="--", lw=1.2,
                label=f"Umbral {UMBRAL_PCT}%")
    ax2.set_ylabel("|% Diferencia|", fontsize=9)
    ax2.legend(fontsize=8)

    return _guardar_fig(fig)


def grafico_evolutivo_mensual(df: pd.DataFrame,
                               col_a: str, lbl_a: str,
                               col_b: str, lbl_b: str,
                               titulo: str) -> str:
    """Barras agrupadas mensuales (arriba) + % diferencia promedio (abajo)."""
    df = df.copy()
    df["MES"] = df["FECHA"].dt.to_period("M")
    df_men = df.groupby("MES").agg(
        A=pd.NamedAgg(column=col_a, aggfunc="sum"),
        B=pd.NamedAgg(column=col_b, aggfunc="sum"),
        PCT=pd.NamedAgg(column="PCT_DIFERENCIA", aggfunc="mean")
    ).reset_index()
    df_men["MES_STR"] = df_men["MES"].astype(str)

    fig, ax1, ax2 = _base_fig(titulo)
    x = np.arange(len(df_men))
    w = 0.35
    ax1.bar(x - w/2, df_men["A"], width=w, color=C["azul"],  alpha=0.85, label=lbl_a)
    ax1.bar(x + w/2, df_men["B"], width=w, color=C["gris"],  alpha=0.85, label=lbl_b)
    ax1.set_xticks(x)
    ax1.set_xticklabels(df_men["MES_STR"], rotation=45, ha="right", fontsize=8)
    ax1.legend(fontsize=9)
    _fmt_miles(ax1)

    colores_barra = [C["rojo"] if abs(v) > UMBRAL_PCT else C["verde"]
                     for v in df_men["PCT"]]
    ax2.bar(x, df_men["PCT"].abs(), color=colores_barra, alpha=0.8)
    ax2.set_xticks(x)
    ax2.set_xticklabels(df_men["MES_STR"], rotation=45, ha="right", fontsize=8)
    ax2.axhline(UMBRAL_PCT, color="red", linestyle="--", lw=1.2,
                label=f"Umbral {UMBRAL_PCT}%")
    ax2.set_ylabel("|% Diff promedio|", fontsize=9)
    ax2.legend(fontsize=8)

    return _guardar_fig(fig)


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES — CORREO SALIENTE
# ══════════════════════════════════════════════════════════════════════════════

def _tabla_html(df: pd.DataFrame, titulo: str) -> str:
    """Genera bloque HTML de tabla con coloreado por umbral."""
    cols = "".join(
        f"<th style='padding:7px 10px;background:{C['azul']};color:white;"
        f"border:1px solid #ccc;font-size:11px'>{c}</th>"
        for c in df.columns
    )
    filas = ""
    for _, row in df.iterrows():
        pct_val = row.get("PCT_DIFERENCIA", 0)
        fila_bg = C["rojo_lt"] if abs(pct_val) > UMBRAL_PCT else "#FFFFFF"
        celdas = ""
        for val in row:
            if isinstance(val, float):
                celdas += f"<td style='padding:5px 9px;border:1px solid #ddd;text-align:right;font-size:11px'>{val:.2f}</td>"
            elif isinstance(val, (int, np.integer)):
                celdas += f"<td style='padding:5px 9px;border:1px solid #ddd;text-align:right;font-size:11px'>{val:,}</td>"
            else:
                celdas += f"<td style='padding:5px 9px;border:1px solid #ddd;font-size:11px'>{val}</td>"
        filas += f"<tr style='background:{fila_bg}'>{celdas}</tr>"

    return f"""
    <h3 style='color:{C["azul"]};font-family:Arial;margin-top:20px'>{titulo}</h3>
    <table style='border-collapse:collapse;font-family:Arial'>
      <thead><tr>{cols}</tr></thead>
      <tbody>{filas}</tbody>
    </table>"""


def enviar_correo(asunto: str, cuerpo_html: str, destinatarios: str,
                  imagenes: list = None):
    """Envía correo vía Outlook Desktop (win32com)."""
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)  # 0 = olMailItem
    mail.Subject = asunto
    mail.HTMLBody = cuerpo_html
    mail.To = destinatarios

    if imagenes:
        for img_path in imagenes:
            if os.path.exists(img_path):
                mail.Attachments.Add(img_path)

    mail.Send()
    print(f"  [CORREO] Enviado: {asunto}")

    # Limpiar temporales
    if imagenes:
        for img_path in imagenes:
            try:
                os.remove(img_path)
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════════════
# PROCESO 1 — TRANSMISIONES (UBA vs SBP)
# ══════════════════════════════════════════════════════════════════════════════

def _parsear_numero_cuerpo(body: str) -> int | None:
    """
    Extrae el número más relevante del cuerpo del correo de transmisiones.
    Busca el primer entero grande (> 100) que aparece en el texto.
    """
    # Eliminar miles con punto o coma, buscar números
    numeros = re.findall(r'\b(\d[\d.,]+)\b', body)
    valores = []
    for n in numeros:
        try:
            limpio = int(re.sub(r'[.,]', '', n))
            if limpio > 100:      # filtrar números pequeños irrelevantes
                valores.append(limpio)
        except ValueError:
            continue
    return valores[0] if valores else None   # primer número relevante


def run_proceso_transmisiones(ns) -> bool:
    """
    Lee carpeta TRX_MONITOR_JOSE, compara UBA vs SBP,
    actualiza bitácora y envía reporte.
    """
    print("\n" + "─"*60)
    print("  PROCESO 1 — TRANSMISIONES (UBA vs SBP)")
    print("─"*60)

    try:
        carpeta = get_carpeta_outlook(ns, CARPETA_TRANS)
    except RuntimeError as e:
        print(f"  [ERROR] {e}")
        return False

    ultima_fecha = leer_ultima_fecha(BITACORA_TRANS, "DIARIO")
    fecha_filtro = ultima_fecha + timedelta(days=1)
    print(f"  Última fecha bitácora : {ultima_fecha}")
    print(f"  Buscando correos desde: {fecha_filtro}")

    items = get_mail_items_desde(carpeta, fecha_filtro)
    print(f"  Correos encontrados   : {len(items)}")

    if not items:
        print("  → Sin correos nuevos. Proceso 1 finalizado.")
        return False

    # Agrupar por fecha_datos = ReceivedTime - 1 día
    por_fecha: dict[date, dict] = {}
    for item in items:
        rt = item.ReceivedTime
        fecha_datos = (rt - timedelta(days=1)).date() if hasattr(rt, "date") else rt - timedelta(days=1)
        if fecha_datos not in por_fecha:
            por_fecha[fecha_datos] = {"uba": None, "sbp": None}
        asunto = (item.Subject or "").upper()
        body   = item.Body or ""
        if "UBA" in asunto:
            por_fecha[fecha_datos]["uba"] = _parsear_numero_cuerpo(body)
        else:
            # SBP / Monitor
            por_fecha[fecha_datos]["sbp"] = _parsear_numero_cuerpo(body)

    registros = []
    for fdate in sorted(por_fecha):
        v = por_fecha[fdate]
        uba, sbp = v["uba"], v["sbp"]
        if uba is None or sbp is None:
            print(f"  [WARN] {fdate} → par incompleto (UBA={uba}, SBP={sbp}). Se omite.")
            continue
        diff = uba - sbp
        pct  = round(diff / sbp * 100, 4) if sbp != 0 else 0.0
        registros.append({"FECHA": fdate, "TOTAL_UBA": uba,
                           "TOTAL_SBP": sbp, "DIFERENCIA": diff,
                           "PCT_DIFERENCIA": pct})
        marca = "⚠️" if abs(pct) > UMBRAL_PCT else "✅"
        print(f"  {marca} {fdate} → UBA={uba:,} | SBP={sbp:,} | Δ={diff:,} ({pct:.2f}%)")

    if not registros:
        print("  → No se generaron registros válidos. Proceso 1 finalizado.")
        return False

    df_nuevo = pd.DataFrame(registros)
    guardar_en_bitacora(BITACORA_TRANS, df_nuevo, "DIARIO")

    # Leer histórico completo para gráficos
    df_hist = pd.read_excel(BITACORA_TRANS, sheet_name="DIARIO")
    df_hist["FECHA"] = pd.to_datetime(df_hist["FECHA"])

    img_diario  = grafico_evolutivo_diario(df_hist,
                                            "TOTAL_UBA", "UBA",
                                            "TOTAL_SBP", "SBP",
                                            "Transmisiones — Evolutivo Diario (últ. 30 días)")
    img_mensual = grafico_evolutivo_mensual(df_hist,
                                             "TOTAL_UBA", "UBA",
                                             "TOTAL_SBP", "SBP",
                                             "Transmisiones — Evolutivo Mensual")

    _enviar_reporte_transmisiones(df_nuevo, [img_diario, img_mensual])
    return True


def _enviar_reporte_transmisiones(df: pd.DataFrame, imagenes: list):
    hoy_str = datetime.today().strftime("%d/%m/%Y")
    alertas = []
    for _, r in df.iterrows():
        if abs(r["PCT_DIFERENCIA"]) > UMBRAL_PCT:
            alertas.append(
                f"<li><b>{r['FECHA']}</b>: diferencia de "
                f"<span style='color:red'>{r['PCT_DIFERENCIA']:.2f}%</span>"
                f" — umbral {UMBRAL_PCT}%</li>"
            )

    bloque_alerta = (
        f"<ul style='color:red'>{''.join(alertas)}</ul>"
        if alertas else
        f"<p style='color:{C['verde']}'>✅ Todas las diferencias dentro del umbral "
        f"({UMBRAL_PCT}%).</p>"
    )

    tabla = _tabla_html(df, f"Resumen — {hoy_str}")
    cuerpo = f"""
    <html><body style='font-family:Arial;font-size:13px;color:#111'>
      <h2 style='color:{C["azul"]}'>📊 Informe de Transmisiones — {hoy_str}</h2>
      <p>Estimados,</p>
      <p>Se adjunta el resumen comparativo de transmisiones UBA vs Monitor
         correspondiente al período procesado.</p>
      {bloque_alerta}
      {tabla}
      <br>
      <p style='font-size:11px;color:{C["gris"]}'>
        Los gráficos evolutivos se encuentran adjuntos.<br>
        Reporte automático — Prevención de Fraude · Scotiabank Perú
      </p>
    </body></html>
    """
    enviar_correo(
        asunto=f"Informe Transmisiones Monitor — {hoy_str}",
        cuerpo_html=cuerpo,
        destinatarios=DESTINATARIOS_TRANS,
        imagenes=imagenes
    )


# ══════════════════════════════════════════════════════════════════════════════
# PROCESO 2 — JOY (Excel FINJOY vs Monitor JOY)
# ══════════════════════════════════════════════════════════════════════════════

# ── 2A. Extracción del Excel adjunto (FINJOY) ──────────────────────────────

def _fecha_de_nombre_archivo(nombre: str) -> date | None:
    """
    Extrae fecha de nombre como TRANSACCIONES_JOY_DDMMYYYY.xlsx
    Retorna date o None si no puede parsear.
    """
    match = re.search(r'(\d{2})(\d{2})(\d{4})', nombre)
    if match:
        d, m, y = match.groups()
        try:
            return date(int(y), int(m), int(d))
        except ValueError:
            pass
    return None


def _extraer_excel_adjunto(mail_item) -> tuple[pd.DataFrame | None, str | None]:
    """
    Extrae el primer adjunto .xlsx/.xls de un MailItem.
    Filtra filas donde EXISTE_EN_EL_MONITOR == 'SI'.
    Retorna (DataFrame, nombre_archivo) o (None, None).
    """
    for att in mail_item.Attachments:
        nombre = att.FileName or ""
        if not (nombre.lower().endswith(".xlsx") or nombre.lower().endswith(".xls")):
            continue

        tmp = os.path.join(tempfile.gettempdir(), nombre)
        try:
            att.SaveAsFile(tmp)
            df = pd.read_excel(tmp)
            # Normalizar columnas
            df.columns = [str(c).strip().upper() for c in df.columns]

            # Filtro obligatorio: solo EXISTE_EN_EL_MONITOR == SI
            if "EXISTE_EN_EL_MONITOR" in df.columns:
                df = df[
                    df["EXISTE_EN_EL_MONITOR"].astype(str).str.strip().str.upper() == "SI"
                ].copy()

            # Tipos
            if "COD_TRANSACCION" in df.columns:
                df["COD_TRANSACCION"] = df["COD_TRANSACCION"].astype(str).str.strip()
            if "TOTAL_CARGO_CUENTA" in df.columns:
                df["TOTAL_CARGO_CUENTA"] = (
                    pd.to_numeric(df["TOTAL_CARGO_CUENTA"], errors="coerce")
                    .fillna(0).astype(int)
                )

            return df, nombre
        except Exception as e:
            print(f"  [WARN] Error leyendo adjunto '{nombre}': {e}")
        finally:
            try:
                os.remove(tmp)
            except Exception:
                pass

    return None, None


# ── 2B. Parseo del cuerpo de correo Monitor JOY ────────────────────────────

def _parsear_cuerpo_monitor_joy(body: str) -> dict[str, int]:
    """
    Parsea el cuerpo del correo de Monitor JOY.

    Formato esperado (puede repetirse por código compuesto):
        (IO01) - Transferencias InterOperabilidad : 10,452
        (WJ33+WJ35+JN05) - Pagos de Servicios    : 25,699

    Retorna dict {codigo_raw: cantidad}
    Ejemplo: {"IO01": 10452, "WJ33+WJ35+JN05": 25699}
    """
    resultado: dict[str, int] = {}
    # Patrón: (CODIGO) cualquier texto : número
    patron = re.compile(r'\(([^)]+)\)[^:]+:\s*([\d,\.]+)')
    for match in patron.finditer(body):
        codigo = match.group(1).strip()
        try:
            cantidad = int(match.group(2).replace(",", "").replace(".", ""))
            resultado[codigo] = cantidad
        except ValueError:
            continue
    return resultado


# ── 2C. Cálculo de diferencias ─────────────────────────────────────────────

def _calcular_diferencias_joy(df_excel: pd.DataFrame,
                               monitor_dict: dict[str, int]) -> pd.DataFrame:
    """
    Compara Excel JOY vs dict Monitor por código.

    Para códigos compuestos en Monitor (WJ33+WJ35+JN05), suma los valores
    individuales del Excel.

    Retorna DataFrame con columnas:
      COD_TRANSACCION | DESCRIPCION | TOTAL_JOY | TOTAL_MONITOR |
      DIFERENCIA | PCT_DIFERENCIA
    """
    registros = []

    for cod_monitor, qty_monitor in monitor_dict.items():
        sub_codigos = [c.strip() for c in cod_monitor.split("+")]

        # Suma Excel para los sub-códigos encontrados
        mask = df_excel["COD_TRANSACCION"].isin(sub_codigos)
        df_match = df_excel[mask]
        qty_joy = int(df_match["TOTAL_CARGO_CUENTA"].sum())

        # Descripción (primera encontrada)
        desc = cod_monitor  # fallback
        if not df_match.empty and "DESCRIPCION_TRANSACCION" in df_match.columns:
            desc = df_match.iloc[0]["DESCRIPCION_TRANSACCION"]

        diff = qty_joy - qty_monitor
        pct  = round(diff / qty_monitor * 100, 4) if qty_monitor != 0 else 0.0

        registros.append({
            "COD_TRANSACCION":  cod_monitor,
            "DESCRIPCION":      desc,
            "TOTAL_JOY":        qty_joy,
            "TOTAL_MONITOR":    qty_monitor,
            "DIFERENCIA":       diff,
            "PCT_DIFERENCIA":   pct,
        })

    return pd.DataFrame(registros)


# ── 2D. Orquestador Proceso 2 ──────────────────────────────────────────────

def run_proceso_joy(ns) -> bool:
    """
    Lee carpetas FINJOY y MONITOR_JOY, calcula diferencias,
    actualiza bitácora (hojas DIARIO y POR_CONDICION) y envía reporte.
    """
    print("\n" + "─"*60)
    print("  PROCESO 2 — JOY (Excel FINJOY vs Monitor JOY)")
    print("─"*60)

    # Conectar carpetas
    try:
        carpeta_fin = get_carpeta_outlook(ns, CARPETA_FINJOY)
        carpeta_mon = get_carpeta_outlook(ns, CARPETA_MONITOR_JOY)
    except RuntimeError as e:
        print(f"  [ERROR] {e}")
        return False

    # Fecha filtro: procesar desde el día siguiente al último registrado
    ultima_fecha = leer_ultima_fecha(BITACORA_JOY, "DIARIO")
    # Los correos LLEGAN el día siguiente a los datos (T-1)
    # → buscamos correos recibidos desde ultima_fecha + 1 día
    fecha_filtro_recepcion = ultima_fecha + timedelta(days=1)
    print(f"  Última fecha bitácora : {ultima_fecha}")
    print(f"  Buscando correos desde: {fecha_filtro_recepcion}")

    # Obtener correos
    items_fin = get_mail_items_desde(carpeta_fin, fecha_filtro_recepcion)
    items_mon = get_mail_items_desde(carpeta_mon, fecha_filtro_recepcion)
    print(f"  Correos FINJOY       : {len(items_fin)}")
    print(f"  Correos MONITOR_JOY  : {len(items_mon)}")

    if not items_fin:
        print("  → Sin correos FINJOY nuevos. Proceso 2 finalizado.")
        return False

    # ── Agrupar Monitor JOY: tomar el ÚLTIMO correo de cada día de recepción
    #    (Monitor puede enviar 20+ duplicados por día — tomamos el más reciente)
    monitor_por_dia: dict[date, object] = {}
    for item in items_mon:
        rt = item.ReceivedTime
        rec_date = rt.date() if hasattr(rt, "date") else rt
        if (rec_date not in monitor_por_dia or
                item.ReceivedTime > monitor_por_dia[rec_date].ReceivedTime):
            monitor_por_dia[rec_date] = item

    print(f"  Días únicos Monitor  : {len(monitor_por_dia)}")

    # ── Procesar cada correo FINJOY (uno por fecha de datos)
    registros_diario    = []
    frames_condicion    = []

    for item_fin in items_fin:
        rt_fin   = item_fin.ReceivedTime
        rec_date = rt_fin.date() if hasattr(rt_fin, "date") else rt_fin

        # Extraer Excel adjunto
        df_excel, nombre_arch = _extraer_excel_adjunto(item_fin)
        if df_excel is None:
            print(f"  [WARN] {rec_date} → Sin Excel adjunto válido. Se omite.")
            continue

        # Fecha de los datos desde nombre del archivo (T-1)
        fecha_datos = _fecha_de_nombre_archivo(nombre_arch)
        if fecha_datos is None:
            fecha_datos = rec_date - timedelta(days=1)

        # Buscar correo Monitor del mismo día de recepción
        if rec_date not in monitor_por_dia:
            print(f"  [WARN] {fecha_datos} → Sin correo Monitor JOY para {rec_date}. Se omite.")
            continue

        item_mon = monitor_por_dia[rec_date]
        body_mon = item_mon.Body or ""
        monitor_dict = _parsear_cuerpo_monitor_joy(body_mon)

        if not monitor_dict:
            print(f"  [WARN] {fecha_datos} → No se parsearon datos del correo Monitor. Se omite.")
            continue

        print(f"\n  Procesando {fecha_datos}  (recibido {rec_date})")
        print(f"    Excel JOY (SI): {len(df_excel)} condiciones")
        print(f"    Monitor JOY   : {len(monitor_dict)} códigos")

        # Diferencias por condición
        df_diff = _calcular_diferencias_joy(df_excel, monitor_dict)
        df_diff.insert(0, "FECHA", fecha_datos)
        frames_condicion.append(df_diff)

        # Totales del día
        total_joy     = int(df_excel["TOTAL_CARGO_CUENTA"].sum())
        total_monitor = sum(monitor_dict.values())
        diff          = total_joy - total_monitor
        pct           = round(diff / total_monitor * 100, 4) if total_monitor != 0 else 0.0

        registros_diario.append({
            "FECHA":            fecha_datos,
            "TOTAL_JOY":        total_joy,
            "TOTAL_MONITOR":    total_monitor,
            "DIFERENCIA":       diff,
            "PCT_DIFERENCIA":   pct,
        })

        marca = "⚠️" if abs(pct) > UMBRAL_PCT else "✅"
        print(f"    {marca} JOY={total_joy:,} | Monitor={total_monitor:,} | "
              f"Δ={diff:,} ({pct:.2f}%)")

    if not registros_diario:
        print("  → No se generaron registros válidos. Proceso 2 finalizado.")
        return False

    df_diario    = pd.DataFrame(registros_diario)
    df_condicion = pd.concat(frames_condicion, ignore_index=True) if frames_condicion else pd.DataFrame()

    # Guardar en bitácora
    guardar_en_bitacora(BITACORA_JOY, df_diario, "DIARIO")
    if not df_condicion.empty:
        guardar_en_bitacora(BITACORA_JOY, df_condicion, "POR_CONDICION")

    # Leer histórico completo para gráficos
    df_hist = pd.read_excel(BITACORA_JOY, sheet_name="DIARIO")
    df_hist["FECHA"] = pd.to_datetime(df_hist["FECHA"])

    img_diario  = grafico_evolutivo_diario(df_hist,
                                            "TOTAL_JOY", "JOY (Excel)",
                                            "TOTAL_MONITOR", "Monitor",
                                            "JOY — Evolutivo Diario (últ. 30 días)")
    img_mensual = grafico_evolutivo_mensual(df_hist,
                                             "TOTAL_JOY", "JOY (Excel)",
                                             "TOTAL_MONITOR", "Monitor",
                                             "JOY — Evolutivo Mensual")

    _enviar_reporte_joy(df_diario, df_condicion, [img_diario, img_mensual])
    return True


def _enviar_reporte_joy(df_diario: pd.DataFrame,
                         df_condicion: pd.DataFrame,
                         imagenes: list):
    hoy_str = datetime.today().strftime("%d/%m/%Y")

    alertas = []
    for _, r in df_diario.iterrows():
        if abs(r["PCT_DIFERENCIA"]) > UMBRAL_PCT:
            alertas.append(
                f"<li><b>{r['FECHA']}</b>: diferencia de "
                f"<span style='color:red'>{r['PCT_DIFERENCIA']:.2f}%</span>"
                f" — umbral {UMBRAL_PCT}%</li>"
            )

    bloque_alerta = (
        f"<ul style='color:red'>{''.join(alertas)}</ul>"
        if alertas else
        f"<p style='color:{C['verde']}'>✅ Todas las diferencias dentro del umbral "
        f"({UMBRAL_PCT}%).</p>"
    )

    tabla_diario = _tabla_html(df_diario, f"Resumen Diario JOY — {hoy_str}")

    tabla_condicion = ""
    if not df_condicion.empty:
        # Mostrar top 30 por mayor diferencia absoluta
        df_top = (df_condicion
                  .assign(ABS_DIFF=lambda x: x["DIFERENCIA"].abs())
                  .sort_values("ABS_DIFF", ascending=False)
                  .drop(columns=["ABS_DIFF"])
                  .head(30))
        tabla_condicion = _tabla_html(df_top, "Detalle por Condición (top 30 por diferencia)")

    cuerpo = f"""
    <html><body style='font-family:Arial;font-size:13px;color:#111'>
      <h2 style='color:{C["azul"]}'>📊 Informe Transacciones JOY — {hoy_str}</h2>
      <p>Estimados,</p>
      <p>Se adjunta el resumen comparativo de transacciones JOY (Excel Monitor vs
         Control Monitor) correspondiente al período procesado.</p>
      {bloque_alerta}
      {tabla_diario}
      <br>
      {tabla_condicion}
      <br>
      <p style='font-size:11px;color:{C["gris"]}'>
        Los gráficos evolutivos (diario últimos 30 días y mensual) se encuentran adjuntos.<br>
        Reporte automático — Prevención de Fraude · Scotiabank Perú
      </p>
    </body></html>
    """
    enviar_correo(
        asunto=f"Informe Transacciones JOY — {hoy_str}",
        cuerpo_html=cuerpo,
        destinatarios=DESTINATARIOS_JOY,
        imagenes=imagenes
    )


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  MONITOR COMPARADOR — INICIO")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Conexión única a Outlook para ambos procesos
    try:
        ns = conectar_outlook()
        print("  [OK] Outlook conectado.")
    except Exception as e:
        print(f"  [FATAL] No se pudo conectar a Outlook: {e}")
        return

    # ── Proceso 1: Transmisiones ──
    try:
        run_proceso_transmisiones(ns)
    except Exception as e:
        print(f"\n  [ERROR] Proceso 1 (Transmisiones) falló inesperadamente: {e}")

    # ── Proceso 2: JOY ──
    try:
        run_proceso_joy(ns)
    except Exception as e:
        print(f"\n  [ERROR] Proceso 2 (JOY) falló inesperadamente: {e}")

    print("\n" + "=" * 60)
    print("  MONITOR COMPARADOR — COMPLETADO")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)


if __name__ == "__main__":
    main()
