# =============================================================================
# ANÁLISIS ECOMMERCE — Comercios de Alta Transaccionalidad
# -----------------------------------------------------------------------------
# Universo de entrada: parquet pre-filtrado en Monitor con:
#   - 1 comercio específico
#   - Solo transacciones ecommerce (Entry Mode 01)
#   - Aprobadas + Denegadas (ambas)
#
# Bloques de feature engineering:
#   A. Velocidad / frecuencia
#   B. Monto y patrones de monto
#   C. Comportamiento del cliente en el comercio
#   D. Cascada de fraude (fraudes previos del cliente)
#   E. Geográficas / IP
#   F. Motivos de rechazo (CVV fail, fondos, etc.)
# =============================================================================
import pandas as pd
import numpy as np
import polars as pl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.stats import chi2_contingency
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# =============================================================================
# AJUSTA SOLO ESTE BLOQUE
# =============================================================================
COLS = {
    "FECHA"              : "ACF-FECHA TRX",
    "HORA"               : "ACF-HORA TRX",
    "DATETIME"           : "FECHA_HORA",
    "COMERCIO"           : "ACF-NOMBRE/LOCALIZACION COMERCIO",
    "INDICADOR"          : "ACF-INDICADOR DE FRAUDE",
    "COD_RESPUESTA"      : "ACF-COD RPTA",
    "COD_MOTIVO_RECHAZO" : "ACF-COD MOTIVO RECHAZO",   # <-- columna detallada
    "RAZON_RESPUESTA"    : "ACF-RAZON RESPUESTA",       # <-- texto (CVV inv, NSF, etc.)
    "CONDICION_RT"       : "CONDICION RT",
    "CANAL"              : "ACF-CANAL",
    "ENTRY_MODE"         : "ACF-ENTRY MODE",
    "SALDO"              : "ACF-SALDO DISPONIBLE EN MONEDA TRX",
    "ID_CLIENTE"         : "ACF-ID CLIENTE",
    "ESI_UCAP"           : "ACF-ECI/UCAF",
    "PAIS"               : "ACF-PAIS ORIGEN 87519",
    "REGION"             : "ACF-REGION",                # opcional
    "CIUDAD"             : "ACF-CIUDAD",                # opcional
    "IP"                 : "ACF-IP",                    # opcional (digital)
    "COD_RED_COMERCIO"   : "ACF-COD RED COMERCIO",
    "MONTO"              : "ACF-MONTO EN MONEDA LOCAL",
    "SEGMENTO"           : "VAA-EVENTO DE COMPROMISO OTRA FUENTE",
    "TIPO_PRODUCTO"      : "ACF-TIPO PROD TC",
    "BIN"                : "ACF-BIN",
}
RUTA_PARQUET   = r"C:\ruta\al\parquet_ecommerce_comercio.parquet"
RUTA_EXCEL_OUT = r"C:\ruta\salida\analisis_ecommerce_resultado.xlsx"
# =============================================================================

SEG_NOMBRE = {"30":"Polo Direccion","99":"Polo Direccion","31":"Premium","32":"Preferente",
    "33":"Personal","34":"Estandar","5":"Inst. Financieras","21":"Corporativo",
    "2":"Mediano Empresas","15":"Sector Gobierno","16":"Otras Instituciones",
    "3":"Pequenas Empresas","4":"Negocios 2","7":"Negocios 3","8":"Negocios 1","13":"Microempresas"}
SEG_GRUPO = {"30":"Affluent","99":"Affluent","31":"Emerging Affluent","32":"Emerging Affluent",
    "33":"Top of Mass","34":"Mass","5":"Corporate","21":"Corporate","2":"Commercial",
    "15":"Commercial","16":"Commercial","3":"Small Business","4":"Small Business",
    "7":"Small Business","8":"Small Business","13":"Small Business"}
COD_RED_LABEL = {"S":"Estatico (TD)","D":"Dinamico (TD/TC)","E":"Estatico (TC)","N":"No Match / Sin CVV"}

# Motivos de rechazo agrupados — ajusta el mapping según tu data real
def clasificar_motivo(razon: str) -> str:
    r = str(razon).upper()
    if any(k in r for k in ["CVV","CVC","CODIGO SEGURIDAD","SECURITY CODE"]):
        return "CVV_FAIL"
    if any(k in r for k in ["FONDOS","SALDO","NSF","INSUFFICIENT"]):
        return "FONDOS_INSUF"
    if any(k in r for k in ["EXCEDE","LIMITE","LIMIT","EXCEED"]):
        return "EXCEDE_LIMITE"
    if any(k in r for k in ["BLOQUEAD","BLOCKED","RESTRICT"]):
        return "TARJETA_BLOQ"
    if any(k in r for k in ["EXPIR","VENCID"]):
        return "TARJETA_EXP"
    if any(k in r for k in ["PIN"]):
        return "PIN_FAIL"
    if any(k in r for k in ["3DS","AUTHEN","AUTENT"]):
        return "AUTH_FAIL"
    if any(k in r for k in ["INVALID","INVAL","NO MATCH","NO COINCID"]):
        return "DATO_INVALIDO"
    return "OTRO"

C_FRAUDE="#D9534F"; C_OK="#5B9BD5"; C_088="#F0AD4E"; C_NEU="#8EA9C1"; C_ROJO2="#8B0000"

# =============================================================================
# CARGA Y PREPARACIÓN
# =============================================================================
print("Cargando parquet...")
df_raw = pd.read_parquet(RUTA_PARQUET)
col_map = {v: k for k, v in COLS.items() if v in df_raw.columns}
df = df_raw.rename(columns=col_map).copy()

faltantes = [a for a, c in COLS.items() if c not in df_raw.columns]
if faltantes:
    print(f"  Columnas no encontradas (se omiten features dependientes): {faltantes}")

df["MONTO"]    = pd.to_numeric(df["MONTO"], errors="coerce")
df["SALDO"]    = pd.to_numeric(df.get("SALDO", pd.Series(dtype=float)), errors="coerce") if "SALDO" in df.columns else np.nan
df["DATETIME"] = pd.to_datetime(df["DATETIME"], errors="coerce")
df["FECHA"]    = df["DATETIME"].dt.normalize()
df["MES"]      = df["DATETIME"].dt.to_period("M").astype(str)

for c in ["INDICADOR","COD_RESPUESTA","COD_MOTIVO_RECHAZO","RAZON_RESPUESTA",
          "CONDICION_RT","CANAL","ESI_UCAP","COD_RED_COMERCIO","SEGMENTO",
          "TIPO_PRODUCTO","ENTRY_MODE","PAIS","REGION","CIUDAD","IP"]:
    if c in df.columns:
        df[c] = df[c].astype(str).str.strip().str.upper()
if "BIN" in df.columns:
    df["BIN"] = df["BIN"].astype(str).str.split(".").str[0].str.strip()

df["ESTADO"] = df["COD_RESPUESTA"].apply(
    lambda x: "APROBADA" if str(x).strip() in ["00","0000","000","0"] else "DENEGADA"
)
df["ES_FRAUDE"]          = (df["INDICADOR"] == "F").astype(int)
df["ES_FRAUDE_APROBADO"] = ((df["INDICADOR"] == "F") & (df["ESTADO"] == "APROBADA")).astype(int)
df["COND_088"]           = df["CONDICION_RT"].str.contains("088", na=False) if "CONDICION_RT" in df.columns else False

# SEGURO: 2, 5, 02, 05
if "ESI_UCAP" in df.columns:
    df["SEGURO"] = df["ESI_UCAP"].apply(
        lambda x: "Seguro" if str(x).strip() in ["2","5","02","05"] else "No Seguro"
    )
else:
    df["SEGURO"] = "No Seguro"

df["SEG_NOMBRE"]    = df.get("SEGMENTO", pd.Series(index=df.index)).map(SEG_NOMBRE).fillna("Otro/Sin seg") if "SEGMENTO" in df.columns else "Otro/Sin seg"
df["SEG_GRUPO"]     = df.get("SEGMENTO", pd.Series(index=df.index)).map(SEG_GRUPO).fillna("Otro/Sin seg") if "SEGMENTO" in df.columns else "Otro/Sin seg"
df["COD_RED_LABEL"] = df.get("COD_RED_COMERCIO", pd.Series(index=df.index)).map(COD_RED_LABEL).fillna("Otro") if "COD_RED_COMERCIO" in df.columns else "Otro"

# Clasificación de motivo de rechazo
if "RAZON_RESPUESTA" in df.columns:
    df["MOTIVO_RECH"] = df["RAZON_RESPUESTA"].apply(clasificar_motivo)
    # Solo aplica a denegadas; aprobadas → "N/A"
    df.loc[df["ESTADO"] == "APROBADA", "MOTIVO_RECH"] = "N/A"
else:
    df["MOTIVO_RECH"] = "N/A"

df_ap  = df[df["ESTADO"] == "APROBADA"].copy()
df_den = df[df["ESTADO"] == "DENEGADA"].copy()
df_f_ap  = df[df["ES_FRAUDE_APROBADO"] == 1].copy()
df_f_tot = df[df["ES_FRAUDE"] == 1].copy()

print(f"  Total: {len(df):,} | Ap: {len(df_ap):,} | Den: {len(df_den):,}")
print(f"  Fraude total: {df['ES_FRAUDE'].sum():,} | Fraude aprobado: {df['ES_FRAUDE_APROBADO'].sum():,}")

# Nombre del comercio (auto-detectado del parquet)
NOMBRE_COMERCIO = df["COMERCIO"].mode().iloc[0] if "COMERCIO" in df.columns and len(df) > 0 else "Comercio"
print(f"  Comercio detectado: {NOMBRE_COMERCIO}")

# =============================================================================
# INGENIERÍA DE VARIABLES CON POLARS
# =============================================================================
print("\n" + "="*65)
print("INGENIERIA DE VARIABLES — Polars")
print("="*65)

cols_feat = ["ID_CLIENTE","DATETIME","COMERCIO","MONTO",
             "ES_FRAUDE","ES_FRAUDE_APROBADO","ESTADO","INDICADOR",
             "SEG_NOMBRE","SEG_GRUPO","TIPO_PRODUCTO","SEGURO",
             "CANAL","COD_RED_LABEL","PAIS","MES","COND_088","MOTIVO_RECH"]
for opt in ["SALDO","BIN","REGION","CIUDAD","IP"]:
    if opt in df.columns:
        cols_feat.append(opt)

plf_all = pl.from_pandas(df[cols_feat].reset_index(drop=True))
plf_all = plf_all.with_columns(pl.col("DATETIME").cast(pl.Datetime("us")))
plf_all = plf_all.sort(["ID_CLIENTE","DATETIME"])

plf_ap  = plf_all.filter(pl.col("ESTADO") == "APROBADA")
plf_den = plf_all.filter(pl.col("ESTADO") == "DENEGADA")

# ── A. VELOCIDAD / FRECUENCIA ───────────────────────────────────────────────
print("\n[A] Velocidad / Frecuencia")

# A1: N transacciones del cliente en ventanas (5min, 15min, 1h, 24h)
for periodo, alias in [("5m","N_TRX_5MIN"),("15m","N_TRX_15MIN"),
                       ("1h","N_TRX_1H"),("24h","N_TRX_24H")]:
    print(f"  A1: {alias}...")
    plf_ap_sorted = plf_ap.sort("DATETIME")
    roll = (
        plf_ap_sorted
        .rolling(index_column="DATETIME", period=periodo, group_by="ID_CLIENTE")
        .agg(pl.len().alias(f"_count_{alias}"))
    )
    plf_ap = plf_ap.with_row_index("_idx_v")
    roll = roll.with_row_index("_idx_v")
    plf_ap = plf_ap.join(roll.select(["_idx_v", f"_count_{alias}"]), on="_idx_v", how="left")
    plf_ap = plf_ap.with_columns(
        (pl.col(f"_count_{alias}") - 1).clip(0, None).cast(pl.Int32).alias(alias)
    ).drop(["_idx_v", f"_count_{alias}"])

# A2: Gap en minutos desde la última transacción
print("  A2: GAP_MINUTOS desde última transacción...")
plf_ap = plf_ap.sort(["ID_CLIENTE","DATETIME"])
plf_ap = plf_ap.with_columns(
    pl.col("DATETIME").shift(1).over("ID_CLIENTE").alias("_prev_dt")
)
plf_ap = plf_ap.with_columns(
    ((pl.col("DATETIME") - pl.col("_prev_dt")).dt.total_seconds() / 60)
    .alias("GAP_MINUTOS")
).drop("_prev_dt")

# A3: ES_RAFAGA = 3+ trx en <10 min
print("  A3: ES_RAFAGA (>=3 trx en <10 min)...")
plf_ap = plf_ap.with_columns(
    (pl.col("N_TRX_15MIN") >= 2).cast(pl.Int32).alias("ES_RAFAGA")
)

# ── B. MONTO Y PATRONES DE MONTO ────────────────────────────────────────────
print("\n[B] Monto y patrones")

# B1: Monto acumulado en 2h
print("  B1: MONTO_ACUM_2H...")
plf_ap_sorted = plf_ap.sort("DATETIME")
roll_m2h = (
    plf_ap_sorted
    .rolling(index_column="DATETIME", period="2h", group_by="ID_CLIENTE")
    .agg(pl.col("MONTO").sum().alias("_macum_2h"))
)
plf_ap = plf_ap.with_row_index("_idx_m")
roll_m2h = roll_m2h.with_row_index("_idx_m")
plf_ap = plf_ap.join(roll_m2h.select(["_idx_m","_macum_2h"]), on="_idx_m", how="left")
plf_ap = plf_ap.with_columns(
    (pl.col("_macum_2h") - pl.col("MONTO")).clip(0, None).alias("MONTO_ACUM_2H")
).drop(["_idx_m","_macum_2h"])

# B2: Monto acumulado en 24h
print("  B2: MONTO_ACUM_24H...")
plf_ap_sorted = plf_ap.sort("DATETIME")
roll_m24h = (
    plf_ap_sorted
    .rolling(index_column="DATETIME", period="24h", group_by="ID_CLIENTE")
    .agg(pl.col("MONTO").sum().alias("_macum_24h"))
)
plf_ap = plf_ap.with_row_index("_idx_m24")
roll_m24h = roll_m24h.with_row_index("_idx_m24")
plf_ap = plf_ap.join(roll_m24h.select(["_idx_m24","_macum_24h"]), on="_idx_m24", how="left")
plf_ap = plf_ap.with_columns(
    (pl.col("_macum_24h") - pl.col("MONTO")).clip(0, None).alias("MONTO_ACUM_24H")
).drop(["_idx_m24","_macum_24h"])

# B3: Z-score monto vs cliente
print("  B3: ZSCORE_MONTO_CLI...")
plf_ap = plf_ap.with_columns([
    pl.col("MONTO").mean().over("ID_CLIENTE").alias("_mean_cli"),
    pl.col("MONTO").std().over("ID_CLIENTE").alias("_std_cli"),
])
plf_ap = plf_ap.with_columns(
    pl.when(pl.col("_std_cli") > 0)
    .then((pl.col("MONTO") - pl.col("_mean_cli")) / pl.col("_std_cli"))
    .otherwise(0.0)
    .alias("ZSCORE_MONTO_CLI")
)

# B4: Ratio monto / promedio histórico cliente
print("  B4: RATIO_MONTO_AVG_CLI...")
plf_ap = plf_ap.with_columns(
    pl.when(pl.col("_mean_cli") > 0)
    .then(pl.col("MONTO") / pl.col("_mean_cli"))
    .otherwise(1.0)
    .alias("RATIO_MONTO_AVG_CLI")
).drop(["_mean_cli","_std_cli"])

# B5: Ratio monto/saldo (si SALDO disponible)
if "SALDO" in plf_ap.columns:
    print("  B5: RATIO_MONTO_SALDO...")
    plf_ap = plf_ap.with_columns(
        pl.when(pl.col("SALDO") > 0)
        .then(pl.col("MONTO") / pl.col("SALDO"))
        .otherwise(None)
        .alias("RATIO_MONTO_SALDO")
    )

# B6: Monto redondo sospechoso
print("  B6: ES_MONTO_REDONDO...")
plf_ap = plf_ap.with_columns(
    ((pl.col("MONTO") % 50 == 0) & (pl.col("MONTO") >= 50)).cast(pl.Int32).alias("ES_MONTO_REDONDO")
)

# B7: Monto bajo (potencial card testing)
print("  B7: ES_MONTO_BAJO...")
plf_ap = plf_ap.with_columns(
    (pl.col("MONTO") < 20).cast(pl.Int32).alias("ES_MONTO_BAJO")
)

# ── C. COMPORTAMIENTO DEL CLIENTE EN EL COMERCIO ────────────────────────────
print("\n[C] Comportamiento del cliente en el comercio")

# C1: Primera vez del cliente en el comercio (rank=1)
print("  C1: ES_PRIMERA_VEZ_COMERCIO...")
plf_ap = plf_ap.with_columns(
    pl.lit(1).cum_sum().over("ID_CLIENTE").alias("_rank_cli")
)
plf_ap = plf_ap.with_columns(
    (pl.col("_rank_cli") == 1).cast(pl.Int32).alias("ES_PRIMERA_VEZ_COMERCIO")
)

# C2: N transacciones históricas previas del cliente
print("  C2: N_TRX_HISTORICAS_COMERCIO...")
plf_ap = plf_ap.with_columns(
    (pl.col("_rank_cli") - 1).cast(pl.Int32).alias("N_TRX_HISTORICAS_COMERCIO")
).drop("_rank_cli")

# C3: Días desde primera compra del cliente en el comercio
print("  C3: DIAS_DESDE_PRIMERA_COMPRA...")
plf_ap = plf_ap.with_columns(
    pl.col("DATETIME").min().over("ID_CLIENTE").alias("_first_dt")
)
plf_ap = plf_ap.with_columns(
    ((pl.col("DATETIME") - pl.col("_first_dt")).dt.total_seconds() / 86400)
    .alias("DIAS_DESDE_PRIMERA_COMPRA")
).drop("_first_dt")

# ── D. CASCADA DE FRAUDE ────────────────────────────────────────────────────
print("\n[D] Cascada de fraude del cliente")

# D1: HUBO_FRAUDE_PREVIO_24H, D2: 7D
print("  D1-D2: Fraudes previos 24h y 7d (rolling sum)...")
plf_ap_sorted = plf_ap.sort("DATETIME")
for periodo, alias in [("24h","HUBO_FRAUDE_PREVIO_24H"),
                       ("7d","HUBO_FRAUDE_PREVIO_7D")]:
    roll_f = (
        plf_ap_sorted
        .rolling(index_column="DATETIME", period=periodo, group_by="ID_CLIENTE")
        .agg(pl.col("ES_FRAUDE_APROBADO").sum().alias(f"_sum_{alias}"))
    )
    plf_ap = plf_ap.with_row_index("_idx_f")
    roll_f = roll_f.with_row_index("_idx_f")
    plf_ap = plf_ap.join(roll_f.select(["_idx_f", f"_sum_{alias}"]), on="_idx_f", how="left")
    plf_ap = plf_ap.with_columns(
        ((pl.col(f"_sum_{alias}") - pl.col("ES_FRAUDE_APROBADO")).clip(0, None) > 0)
        .cast(pl.Int32).alias(alias)
    ).drop(["_idx_f", f"_sum_{alias}"])

# D3: N_FRAUDES_CONSECUTIVOS — fraudes seguidos sin trx genuina en medio
print("  D3: N_FRAUDES_CONSECUTIVOS...")
plf_ap = plf_ap.sort(["ID_CLIENTE","DATETIME"])
plf_ap = plf_ap.with_columns(
    pl.col("ES_FRAUDE_APROBADO").shift(1).over("ID_CLIENTE").alias("_prev_f")
)
plf_ap = plf_ap.with_columns(
    pl.when(pl.col("_prev_f") == 1).then(1).otherwise(0).alias("PREV_FUE_FRAUDE")
).drop("_prev_f")

# D4: Minutos desde el último fraude del cliente
print("  D4: MIN_DESDE_ULTIMO_FRAUDE...")
plf_aux = plf_ap.filter(pl.col("ES_FRAUDE_APROBADO") == 1).select(
    ["ID_CLIENTE","DATETIME"]
).rename({"DATETIME":"DT_FRAUDE"}).sort(["ID_CLIENTE","DT_FRAUDE"])
if len(plf_aux) > 0:
    plf_ap_s = plf_ap.sort(["ID_CLIENTE","DATETIME"])
    merged = plf_ap_s.join_asof(
        plf_aux, left_on="DATETIME", right_on="DT_FRAUDE",
        by="ID_CLIENTE", strategy="backward"
    )
    merged = merged.with_columns(
        ((pl.col("DATETIME") - pl.col("DT_FRAUDE")).dt.total_seconds() / 60)
        .alias("MIN_DESDE_ULTIMO_FRAUDE")
    )
    # Si la trx actual ES fraude, el min_desde es 0 — lo dejamos
    plf_ap = merged.drop("DT_FRAUDE")
else:
    plf_ap = plf_ap.with_columns(pl.lit(None).cast(pl.Float64).alias("MIN_DESDE_ULTIMO_FRAUDE"))

# ── E. GEOGRÁFICAS / IP ─────────────────────────────────────────────────────
print("\n[E] Geográficas / IP")

# E1: País más frecuente del cliente
if "PAIS" in plf_ap.columns:
    print("  E1: PAIS_DISTINTO_HABITUAL...")
    pais_top = (
        plf_ap.group_by(["ID_CLIENTE","PAIS"]).agg(pl.len().alias("_n"))
        .sort(["ID_CLIENTE","_n"], descending=[False, True])
        .group_by("ID_CLIENTE").first()
        .select(["ID_CLIENTE", pl.col("PAIS").alias("PAIS_HABITUAL")])
    )
    plf_ap = plf_ap.join(pais_top, on="ID_CLIENTE", how="left")
    plf_ap = plf_ap.with_columns(
        (pl.col("PAIS") != pl.col("PAIS_HABITUAL")).cast(pl.Int32).alias("PAIS_DISTINTO_HABITUAL")
    )

    # E2: Cambio de país vs transacción anterior
    print("  E2: CAMBIO_PAIS_VS_PREV...")
    plf_ap = plf_ap.sort(["ID_CLIENTE","DATETIME"])
    plf_ap = plf_ap.with_columns(
        pl.col("PAIS").shift(1).over("ID_CLIENTE").alias("_prev_pais")
    )
    plf_ap = plf_ap.with_columns(
        pl.when(pl.col("_prev_pais").is_null()).then(0)
        .otherwise((pl.col("PAIS") != pl.col("_prev_pais")).cast(pl.Int32))
        .alias("CAMBIO_PAIS_VS_PREV")
    ).drop("_prev_pais")

    # E3: N países distintos en 24h
    print("  E3: N_PAISES_DISTINTOS_24H...")
    plf_ap_sorted = plf_ap.sort("DATETIME")
    roll_p = (
        plf_ap_sorted
        .rolling(index_column="DATETIME", period="24h", group_by="ID_CLIENTE")
        .agg(pl.col("PAIS").n_unique().alias("_npais"))
    )
    plf_ap = plf_ap.with_row_index("_idx_p")
    roll_p = roll_p.with_row_index("_idx_p")
    plf_ap = plf_ap.join(roll_p.select(["_idx_p","_npais"]), on="_idx_p", how="left")
    plf_ap = plf_ap.with_columns(
        pl.col("_npais").cast(pl.Int32).alias("N_PAISES_DISTINTOS_24H")
    ).drop(["_idx_p","_npais"])

# IP features (si disponible)
if "IP" in plf_ap.columns:
    print("  E4: IP_NUEVA_CLIENTE...")
    plf_ap = plf_ap.with_columns(
        pl.lit(1).cum_sum().over(["ID_CLIENTE","IP"]).alias("_rank_ip")
    )
    plf_ap = plf_ap.with_columns(
        (pl.col("_rank_ip") == 1).cast(pl.Int32).alias("IP_NUEVA_CLIENTE")
    ).drop("_rank_ip")

    # E5: N tarjetas distintas desde misma IP en 24h
    if "BIN" in plf_ap.columns:
        print("  E5: N_TARJETAS_MISMA_IP_24H (proxy via BIN+ID_CLIENTE)...")
        plf_ap_ip = plf_ap.sort("DATETIME")
        roll_ip = (
            plf_ap_ip
            .rolling(index_column="DATETIME", period="24h", group_by="IP")
            .agg(pl.col("ID_CLIENTE").n_unique().alias("_n_clis_ip"))
        )
        plf_ap = plf_ap.with_row_index("_idx_ip")
        roll_ip = roll_ip.with_row_index("_idx_ip")
        plf_ap = plf_ap.join(roll_ip.select(["_idx_ip","_n_clis_ip"]), on="_idx_ip", how="left")
        plf_ap = plf_ap.with_columns(
            pl.col("_n_clis_ip").cast(pl.Int32).alias("N_CLIENTES_MISMA_IP_24H")
        ).drop(["_idx_ip","_n_clis_ip"])

# ── F. MOTIVOS DE RECHAZO ────────────────────────────────────────────────────
print("\n[F] Motivos de rechazo previos")

if len(plf_den) > 0 and "MOTIVO_RECH" in plf_den.columns:
    # F1: N_RECHAZOS_24H total
    print("  F1: N_RECHAZOS_24H...")
    plf_den_cum = (
        plf_den.sort(["ID_CLIENTE","DATETIME"])
        .with_columns(pl.lit(1).cum_sum().over("ID_CLIENTE").alias("CUM_DEN"))
        .select(["ID_CLIENTE","DATETIME","CUM_DEN"])
    )
    plf_ap_r = plf_ap.sort("DATETIME").with_row_index("_idx_r")
    m_now = plf_ap_r.select(["_idx_r","ID_CLIENTE","DATETIME"]).join_asof(
        plf_den_cum.rename({"CUM_DEN":"CD_NOW"}),
        on="DATETIME", by="ID_CLIENTE", strategy="backward"
    )
    m_24h = plf_ap_r.with_columns(
        (pl.col("DATETIME") - pl.duration(hours=24)).alias("DT_24H")
    ).select(["_idx_r","ID_CLIENTE","DT_24H"])
    m_24h = m_24h.rename({"DT_24H":"DATETIME"}).join_asof(
        plf_den_cum.rename({"CUM_DEN":"CD_24H"}),
        on="DATETIME", by="ID_CLIENTE", strategy="backward"
    )
    rech = m_now.select(["_idx_r","CD_NOW"]).join(
        m_24h.select(["_idx_r","CD_24H"]), on="_idx_r", how="left"
    ).with_columns(
        (pl.col("CD_NOW").fill_null(0) - pl.col("CD_24H").fill_null(0))
        .clip(0, None).cast(pl.Int32).alias("N_RECHAZOS_24H")
    )
    plf_ap = plf_ap.with_row_index("_idx_r").join(
        rech.select(["_idx_r","N_RECHAZOS_24H"]), on="_idx_r", how="left"
    ).drop("_idx_r")

    # F2: N_CVV_FAIL_24H (rechazos específicos por CVV)
    print("  F2: N_CVV_FAIL_24H y HUBO_CVV_FAIL_PREVIO...")
    plf_cvv = plf_den.filter(pl.col("MOTIVO_RECH") == "CVV_FAIL").sort(["ID_CLIENTE","DATETIME"])
    if len(plf_cvv) > 0:
        plf_cvv_cum = plf_cvv.with_columns(
            pl.lit(1).cum_sum().over("ID_CLIENTE").alias("CUM_CVV")
        ).select(["ID_CLIENTE","DATETIME","CUM_CVV"])

        plf_ap_c = plf_ap.sort("DATETIME").with_row_index("_idx_c")
        c_now = plf_ap_c.select(["_idx_c","ID_CLIENTE","DATETIME"]).join_asof(
            plf_cvv_cum.rename({"CUM_CVV":"CC_NOW"}),
            on="DATETIME", by="ID_CLIENTE", strategy="backward"
        )
        c_24h = plf_ap_c.with_columns(
            (pl.col("DATETIME") - pl.duration(hours=24)).alias("DT_24H_C")
        ).select(["_idx_c","ID_CLIENTE","DT_24H_C"])
        c_24h = c_24h.rename({"DT_24H_C":"DATETIME"}).join_asof(
            plf_cvv_cum.rename({"CUM_CVV":"CC_24H"}),
            on="DATETIME", by="ID_CLIENTE", strategy="backward"
        )
        cvv_join = c_now.select(["_idx_c","CC_NOW"]).join(
            c_24h.select(["_idx_c","CC_24H"]), on="_idx_c", how="left"
        ).with_columns(
            (pl.col("CC_NOW").fill_null(0) - pl.col("CC_24H").fill_null(0))
            .clip(0, None).cast(pl.Int32).alias("N_CVV_FAIL_24H")
        )
        plf_ap = plf_ap.with_row_index("_idx_c").join(
            cvv_join.select(["_idx_c","N_CVV_FAIL_24H"]), on="_idx_c", how="left"
        ).drop("_idx_c")
    else:
        plf_ap = plf_ap.with_columns(pl.lit(0).cast(pl.Int32).alias("N_CVV_FAIL_24H"))
else:
    plf_ap = plf_ap.with_columns([
        pl.lit(0).cast(pl.Int32).alias("N_RECHAZOS_24H"),
        pl.lit(0).cast(pl.Int32).alias("N_CVV_FAIL_24H"),
    ])

plf_ap = plf_ap.with_columns(
    (pl.col("N_CVV_FAIL_24H") > 0).cast(pl.Int32).alias("HUBO_CVV_FAIL_PREVIO")
)

# Convertir a Pandas
print("\nConvirtiendo resultado a Pandas...")
df_feat = plf_ap.to_pandas()

# =============================================================================
# BUCKETS PARA TABLAS CRUZADAS
# =============================================================================
print("\nGenerando buckets para análisis...")

df_feat["BUCKET_N_TRX_5MIN"] = pd.cut(
    df_feat["N_TRX_5MIN"].clip(0, 20),
    bins=[-0.001, 0, 1, 2, 3, 5, 20],
    labels=["0","1","2","3","4-5","6+"], include_lowest=True
)
df_feat["BUCKET_N_TRX_24H"] = pd.cut(
    df_feat["N_TRX_24H"].clip(0, 50),
    bins=[-0.001, 0, 1, 2, 4, 9, 50],
    labels=["0","1","2","3-4","5-9","10+"], include_lowest=True
)
df_feat["BUCKET_GAP_MIN"] = pd.cut(
    df_feat["GAP_MINUTOS"].clip(0, 1440),
    bins=[-0.001, 1, 5, 15, 60, 360, 1440],
    labels=["<1min","1-5min","5-15min","15-60min","1-6h",">6h"], include_lowest=True
)
df_feat["BUCKET_MONTO_ACUM_24H"] = pd.cut(
    df_feat["MONTO_ACUM_24H"].clip(0, df_feat["MONTO_ACUM_24H"].quantile(0.99)),
    bins=5, include_lowest=True
)
df_feat["BUCKET_ZSCORE"] = pd.cut(
    df_feat["ZSCORE_MONTO_CLI"].clip(-3, 3),
    bins=[-3, -2, -1, 0, 1, 2, 3],
    labels=["<-2SD","-2a-1SD","-1a0SD","0a1SD","1a2SD",">2SD"], include_lowest=True
)
df_feat["BUCKET_RECHAZOS"] = pd.cut(
    df_feat["N_RECHAZOS_24H"].clip(0, 10),
    bins=[-0.001, 0, 1, 2, 3, 10],
    labels=["0 rech","1 rech","2 rech","3 rech","4+ rech"], include_lowest=True
)
df_feat["BUCKET_CVV_FAIL"] = pd.cut(
    df_feat["N_CVV_FAIL_24H"].clip(0, 10),
    bins=[-0.001, 0, 1, 2, 10],
    labels=["0","1","2","3+"], include_lowest=True
)
df_feat["BUCKET_DIAS_PRIMERA"] = pd.cut(
    df_feat["DIAS_DESDE_PRIMERA_COMPRA"].clip(0, 365),
    bins=[-0.001, 0, 1, 7, 30, 90, 365],
    labels=["mismo dia","1d","2-7d","8-30d","31-90d",">90d"], include_lowest=True
)

# Hora del día
df_feat["HORA_NUM"] = pd.to_datetime(df_feat["DATETIME"]).dt.hour
def franja(h):
    if h <= 5:  return "00-05 Madrugada"
    if h <= 11: return "06-11 Manana"
    if h <= 17: return "12-17 Tarde"
    if h <= 20: return "18-20 Noche"
    return "21-23 Noche Tardia"
df_feat["FRANJA_HORARIA"] = df_feat["HORA_NUM"].apply(franja)

print(f"Features y buckets generados. Filas: {len(df_feat):,}")


# =============================================================================
# ANÁLISIS DE FEATURES vs FRAUDE APROBADO
# =============================================================================
print("\n" + "="*65); print("ANALISIS DE FEATURES vs FRAUDE APROBADO")
print("="*65)

def analizar_feature(df_base, col_bucket):
    return (
        df_base.groupby(col_bucket, observed=True)
        .agg(Txns=("ES_FRAUDE_APROBADO","count"),
             Fraude_Ap=("ES_FRAUDE_APROBADO","sum"),
             FR_pct=("ES_FRAUDE_APROBADO", lambda x: x.mean()*100),
             Ticket_prom=("MONTO","mean"),
             Ticket_f_prom=("MONTO", lambda x: x[df_base.loc[x.index,"ES_FRAUDE_APROBADO"]==1].mean()))
        .rename(columns={"FR_pct":"FR% Ap","Ticket_prom":"Ticket prom","Ticket_f_prom":"Ticket fraude prom"})
    )

buckets_analizar = [
    ("BUCKET_N_TRX_5MIN","A1 - N trx 5min"),
    ("BUCKET_N_TRX_24H","A1 - N trx 24h"),
    ("BUCKET_GAP_MIN","A2 - Gap minutos"),
    ("ES_RAFAGA","A3 - Ráfaga (>=3 en <10min)"),
    ("BUCKET_MONTO_ACUM_24H","B - Monto acum 24h"),
    ("BUCKET_ZSCORE","B - Z-score monto"),
    ("ES_MONTO_REDONDO","B - Monto redondo"),
    ("ES_MONTO_BAJO","B - Monto bajo (<20)"),
    ("ES_PRIMERA_VEZ_COMERCIO","C - Primera vez en comercio"),
    ("BUCKET_DIAS_PRIMERA","C - Días desde primera compra"),
    ("HUBO_FRAUDE_PREVIO_24H","D - Fraude previo 24h"),
    ("HUBO_FRAUDE_PREVIO_7D","D - Fraude previo 7d"),
    ("PREV_FUE_FRAUDE","D - Anterior fue fraude"),
    ("BUCKET_RECHAZOS","F - Rechazos 24h"),
    ("BUCKET_CVV_FAIL","F - CVV fails 24h"),
    ("HUBO_CVV_FAIL_PREVIO","F - Hubo CVV fail previo"),
    ("FRANJA_HORARIA","H - Franja horaria"),
]
if "PAIS_DISTINTO_HABITUAL" in df_feat.columns:
    buckets_analizar += [
        ("PAIS_DISTINTO_HABITUAL","E - País distinto al habitual"),
        ("CAMBIO_PAIS_VS_PREV","E - Cambio país vs prev"),
        ("N_PAISES_DISTINTOS_24H","E - N países distintos 24h"),
    ]
if "IP_NUEVA_CLIENTE" in df_feat.columns:
    buckets_analizar.append(("IP_NUEVA_CLIENTE","E - IP nueva para cliente"))
if "N_CLIENTES_MISMA_IP_24H" in df_feat.columns:
    df_feat["BUCKET_CLIS_IP"] = pd.cut(
        df_feat["N_CLIENTES_MISMA_IP_24H"].clip(0,20),
        bins=[-0.001,1,2,3,5,20],
        labels=["1","2","3","4-5","6+"], include_lowest=True
    )
    buckets_analizar.append(("BUCKET_CLIS_IP","E - N clientes misma IP 24h"))

tablas_features = {}
for col_b, titulo in buckets_analizar:
    if col_b not in df_feat.columns:
        continue
    print(f"\n--- {titulo} ---")
    tf = analizar_feature(df_feat, col_b)
    print(tf.to_string())
    tablas_features[titulo] = tf


# =============================================================================
# SECCIÓN 1 — RESUMEN EJECUTIVO POR MES
# =============================================================================
print("\n" + "="*65); print("SECCION 1 — RESUMEN EJECUTIVO")
meses_disponibles = sorted(df_ap["MES"].unique())
filas_resumen = []
for mes in meses_disponibles:
    sub_mes  = df[df["MES"] == mes]
    sub_ap   = df_ap[df_ap["MES"] == mes]
    sub_f_ap = sub_ap[sub_ap["ES_FRAUDE_APROBADO"] == 1]
    sub_f_tot= sub_mes[sub_mes["ES_FRAUDE"] == 1]
    n_ap_m   = len(sub_ap); n_den_m = (sub_mes["ESTADO"] == "DENEGADA").sum()
    monto_ap_m = sub_ap["MONTO"].sum(); ticket_ap_m = sub_ap["MONTO"].mean()
    n_cli_m  = sub_ap["ID_CLIENTE"].nunique()
    n_f_tot_m= len(sub_f_tot); monto_f_tot_m = sub_f_tot["MONTO"].sum()
    n_f_ap_m = len(sub_f_ap); monto_f_ap_m = sub_f_ap["MONTO"].sum()
    ticket_f_ap_m = sub_f_ap["MONTO"].mean() if n_f_ap_m > 0 else 0
    ratio_trx_m = n_f_ap_m / n_ap_m if n_ap_m > 0 else 0
    ratio_mon_m = monto_f_ap_m / monto_ap_m if monto_ap_m > 0 else 0
    filas_resumen.append({
        "Total trx aprobadas": f"{n_ap_m:,}",
        "Total trx denegadas": f"{n_den_m:,}",
        "Total monto aprobado (S/)": f"S/ {monto_ap_m:,.2f}",
        "Ticket promedio (S/)": f"S/ {ticket_ap_m:,.2f}",
        "N clientes unicos": f"{n_cli_m:,}",
        "Trxs fraude total": f"{n_f_tot_m:,}",
        "Trxs fraude aprobado": f"{n_f_ap_m:,}",
        "Monto fraude total (S/)": f"S/ {monto_f_tot_m:,.2f}",
        "Monto fraude aprobado (S/)": f"S/ {monto_f_ap_m:,.2f}",
        "Ticket fraude aprobado": f"S/ {ticket_f_ap_m:,.2f}",
        "Ratio trx fraude aprobadas": f"{ratio_trx_m*100:.4f}%",
        "Ratio fraude soles (ap)": f"{ratio_mon_m*100:.4f}%",
    })
df_resumen_ejecutivo = pd.DataFrame(filas_resumen, index=meses_disponibles).T
df_resumen_ejecutivo.index.name = "Indicador"
print(df_resumen_ejecutivo.to_string())

n_ap = len(df_ap); monto_ap = df_ap["MONTO"].sum()
n_f_tot_g = df["ES_FRAUDE"].sum(); n_f_ap_g = df["ES_FRAUDE_APROBADO"].sum()
monto_f_ap_g = df_f_ap["MONTO"].sum() if len(df_f_ap) > 0 else 0
ratio_trx_f = n_f_ap_g / n_ap if n_ap > 0 else 0


# =============================================================================
# SECCIÓN 2 — DECILES
# =============================================================================
print("\n" + "="*65); print("SECCION 2 — DECILES")
def construir_tabla_deciles(df_b, bins_c=None, n_d=10, lb="D"):
    db = df_b.copy().dropna(subset=["MONTO"])
    if bins_c is not None:
        labels = [f"{lb}{str(i+1).zfill(2)}" for i in range(len(bins_c)-1)]
        db["DECIL"] = pd.cut(db["MONTO"], bins=bins_c, labels=labels, include_lowest=True)
        db["DECIL"] = pd.Categorical(db["DECIL"], categories=labels, ordered=True)
    else:
        labels = [f"{lb}{str(i+1).zfill(2)}" for i in range(n_d)]
        db["DECIL"] = pd.qcut(db["MONTO"], q=n_d, labels=labels, duplicates="drop")
    rows = []; fa_acum = 0; ta_acum = 0
    tot_ap = len(db[db["ESTADO"] == "APROBADA"]); tot_fa = db["ES_FRAUDE_APROBADO"].sum()
    for decil in db["DECIL"].cat.categories:
        s = db[db["DECIL"] == decil]
        if len(s) == 0: continue
        sa = s[s["ESTADO"] == "APROBADA"]; sd = s[s["ESTADO"] == "DENEGADA"]
        sfa = s[s["ES_FRAUDE_APROBADO"] == 1]; sft = s[s["ES_FRAUDE"] == 1]
        na = len(sa); ma = sa["MONTO"].sum(); nd = len(sd); md = sd["MONTO"].sum()
        nfa = len(sfa); mfa = sfa["MONTO"].sum(); nft = len(sft); mft = sft["MONTO"].sum()
        ta_acum += na; fa_acum += nfa
        rows.append({
            "Decil": decil, "Rango S/": f"{s['MONTO'].min():.0f}-{s['MONTO'].max():.0f}",
            "Trx Aprobadas": na, "Monto Aprobadas": round(ma,1),
            "Freq Aprobadas": f"{na/tot_ap*100:.2f}%" if tot_ap>0 else "0%",
            "Freq Acum Ap": f"{ta_acum/tot_ap*100:.2f}%" if tot_ap>0 else "0%",
            "Trx Denegadas": nd, "Monto Denegadas": round(md,1),
            "Total Trx": len(s), "Ticket Promedio": round(s["MONTO"].mean(),2),
            "Monto Total S/": round(s["MONTO"].sum(),1),
            "Fraude Trx Total": nft, "Fraude Monto Total": round(mft,1),
            "Fraude Trx Aprobado": nfa, "Fraude Monto Aprobado": round(mfa,1),
            "Fraude Freq Ap": f"{nfa/tot_fa*100:.4f}%" if tot_fa>0 else "0%",
            "Fraude Freq Acum": f"{fa_acum/tot_fa*100:.4f}%" if tot_fa>0 else "0%",
            "FR/Trx Ap": f"{nfa/na*100:.4f}%" if na>0 else "0%",
            "NoFR/Trx Ap": f"{(1-nfa/na)*100:.4f}%" if na>0 else "100%",
        })
    return pd.DataFrame(rows).set_index("Decil")

tabla_deciles = construir_tabla_deciles(df, n_d=10)
print(tabla_deciles.to_string())


# =============================================================================
# SECCIÓN 3 — DIMENSIONES (ajustadas para ecommerce)
# =============================================================================
print("\n" + "="*65); print("SECCION 3 — DIMENSIONES")
def res_por(df_b, col, tn=15):
    if col not in df_b.columns:
        return pd.DataFrame()
    return (df_b.groupby(col).agg(
        Txns=(col,"count"),
        Fraude_Total=("ES_FRAUDE","sum"),
        Fraude_Aprobado=("ES_FRAUDE_APROBADO","sum"),
        FR_T=("ES_FRAUDE", lambda x: x.mean()*100),
        FR_A=("ES_FRAUDE_APROBADO", lambda x: x.mean()*100),
        Monto_total=("MONTO","sum"),
        Monto_F_Ap=("MONTO", lambda x: x[df_b.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum())
    ).sort_values("Fraude_Aprobado", ascending=False).head(tn)
      .rename(columns={"FR_T":"FR% Total","FR_A":"FR% Ap","Monto_F_Ap":"Monto Fraude Ap"}))

df_por_pais   = res_por(df, "PAIS")
df_por_region = res_por(df, "REGION") if "REGION" in df.columns else pd.DataFrame()
df_por_ciudad = res_por(df, "CIUDAD") if "CIUDAD" in df.columns else pd.DataFrame()
df_por_seg    = res_por(df, "SEG_NOMBRE")
df_por_grupo  = res_por(df, "SEG_GRUPO")
df_por_tipo   = res_por(df, "TIPO_PRODUCTO")
df_por_seguro = res_por(df, "SEGURO")
df_por_cvv    = res_por(df, "COD_RED_LABEL")


# =============================================================================
# SECCIÓN 4 — MOTIVOS DE RECHAZO (nuevo bloque ecommerce)
# =============================================================================
print("\n" + "="*65); print("SECCION 4 — MOTIVOS DE RECHAZO")
if "MOTIVO_RECH" in df.columns:
    df_motivos = (
        df_den.groupby("MOTIVO_RECH")
        .agg(N_Rechazos=("MOTIVO_RECH","count"),
             Clientes_Unicos=("ID_CLIENTE","nunique"),
             Monto_Rech=("MONTO","sum"),
             Monto_Prom=("MONTO","mean"))
        .sort_values("N_Rechazos", ascending=False)
    )
    df_motivos["% del total denegado"] = (df_motivos["N_Rechazos"] / len(df_den) * 100).round(2) if len(df_den) > 0 else 0
    print(df_motivos.to_string())

    # Clientes que tuvieron rechazo y luego fraude aprobado
    cli_con_rech = set(df_den["ID_CLIENTE"].unique())
    cli_con_fraude = set(df_f_ap["ID_CLIENTE"].unique())
    interseccion = cli_con_rech & cli_con_fraude
    print(f"\nClientes con rechazo Y fraude aprobado: {len(interseccion):,} de {len(cli_con_fraude):,} clientes fraude ({len(interseccion)/max(len(cli_con_fraude),1)*100:.2f}%)")
else:
    df_motivos = pd.DataFrame()


# =============================================================================
# SECCIÓN 5 — CASCADA DE FRAUDE (clientes con N fraudes)
# =============================================================================
print("\n" + "="*65); print("SECCION 5 — CASCADA DE FRAUDE")
if len(df_f_ap) > 0:
    fraudes_por_cliente = df_f_ap.groupby("ID_CLIENTE").size().reset_index(name="N_Fraudes_Aprobados")
    fraudes_por_cliente["Bucket_N_Fraudes"] = pd.cut(
        fraudes_por_cliente["N_Fraudes_Aprobados"],
        bins=[0, 1, 2, 3, 5, 100],
        labels=["1 fraude","2 fraudes","3 fraudes","4-5 fraudes","6+ fraudes"]
    )
    tabla_cascada = (
        fraudes_por_cliente.groupby("Bucket_N_Fraudes", observed=True)
        .agg(N_Clientes=("ID_CLIENTE","nunique"),
             Total_Fraudes=("N_Fraudes_Aprobados","sum"))
    )
    tabla_cascada["% Clientes"] = (tabla_cascada["N_Clientes"] / tabla_cascada["N_Clientes"].sum() * 100).round(2)
    tabla_cascada["% Fraudes"]  = (tabla_cascada["Total_Fraudes"] / tabla_cascada["Total_Fraudes"].sum() * 100).round(2)
    print(tabla_cascada.to_string())
else:
    tabla_cascada = pd.DataFrame()


# =============================================================================
# SECCIÓN 6 — TABLAS CRUZADAS
# =============================================================================
print("\n" + "="*65); print("SECCION 6 — TABLAS CRUZADAS")
def tc(df_b, d_f, d_c, tit):
    if d_f not in df_b.columns or d_c not in df_b.columns:
        return None, None, None
    try:
        piv = df_b.pivot_table(values="ES_FRAUDE_APROBADO", index=d_f, columns=d_c, aggfunc="sum", fill_value=0)
        piv_fr = df_b.pivot_table(values="ES_FRAUDE_APROBADO", index=d_f, columns=d_c, aggfunc="mean", fill_value=0).map(lambda v: f"{v*100:.4f}%")
        piv_m = df_b[df_b["ES_FRAUDE_APROBADO"]==1].pivot_table(values="MONTO", index=d_f, columns=d_c, aggfunc="sum", fill_value=0).map(lambda v: f"S/ {v:,.2f}")
        print(f"\n--- {tit} ---"); print(piv.to_string())
        try:
            chi2, p, dof, _ = chi2_contingency(piv.values)
            print(f"Chi2={chi2:.4f} p={p:.6f} | {'SIGNIFICATIVA' if p<0.05 else 'No significativa'}")
        except: pass
        return piv, piv_fr, piv_m
    except Exception as e:
        print(f"  Error {tit}: {e}"); return None, None, None

cruces = [
    ("HUBO_CVV_FAIL_PREVIO","HUBO_FRAUDE_PREVIO_24H","CVV Fail Previo x Fraude Previo 24h"),
    ("BUCKET_N_TRX_5MIN","ES_PRIMERA_VEZ_COMERCIO","Velocidad 5min x Primera vez"),
    ("BUCKET_RECHAZOS","BUCKET_CVV_FAIL","Rechazos x CVV Fails"),
    ("BUCKET_GAP_MIN","HUBO_FRAUDE_PREVIO_24H","Gap min x Fraude previo"),
    ("FRANJA_HORARIA","ES_PRIMERA_VEZ_COMERCIO","Franja horaria x Primera vez"),
    ("BUCKET_ZSCORE","HUBO_FRAUDE_PREVIO_24H","Z-score x Fraude previo"),
]
if "PAIS_DISTINTO_HABITUAL" in df_feat.columns:
    cruces.append(("PAIS_DISTINTO_HABITUAL","HUBO_FRAUDE_PREVIO_24H","País distinto x Fraude previo"))

r_cruces = {tit: tc(df_feat, a, b, tit) for a, b, tit in cruces}


# =============================================================================
# SECCIÓN 7 — ESTADÍSTICAS DE MONTO
# =============================================================================
print("\n" + "="*65); print("SECCION 7 — ESTADISTICAS")
pcts = [10, 25, 50, 75, 90, 95, 99]
df_nofraude = df_ap[df_ap["ES_FRAUDE"] == 0]
def ss(s):
    if len(s) == 0: return [0]*13
    return [s.mean(), s.median(), s.std(), s.var(), s.min(), s.max()] + [s.quantile(p/100) for p in pcts]
stats_monto = pd.DataFrame({
    "Metrica": ["Media","Mediana","Desv.Std","Varianza","Min","Max"] + [f"P{p}" for p in pcts],
    "Total ap": ss(df_ap["MONTO"]),
    "Fraude Total": ss(df_f_tot["MONTO"]),
    "Fraude Ap": ss(df_f_ap["MONTO"]),
    "No Fraude": ss(df_nofraude["MONTO"]),
})
for c in ["Total ap","Fraude Total","Fraude Ap","No Fraude"]:
    stats_monto[c] = stats_monto[c].apply(lambda v: f"S/ {v:,.2f}")
print(stats_monto.to_string(index=False))


# =============================================================================
# EXPORTACIÓN A EXCEL
# =============================================================================
print("\nExportando a Excel...")
FH = PatternFill("solid", fgColor="1F3864"); FS = PatternFill("solid", fgColor="2E75B6")
FA = PatternFill("solid", fgColor="DEEAF1"); FY = PatternFill("solid", fgColor="FFF2CC")
FF = PatternFill("solid", fgColor="FCE4D6")
fH = Font(color="FFFFFF", bold=True, size=10); fN = Font(size=10); fI = Font(italic=True, size=9, color="1F3864")
BT = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL = Alignment(horizontal="left", vertical="center", wrap_text=True)

def h(ws, f, n, t, fl=None):
    fl = fl or FH
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=n)
    c = ws.cell(row=f, column=1, value=t); c.fill = fl; c.font = fH; c.alignment = AC; c.border = BT
def hd(ws, f):
    for r in ws.iter_rows(min_row=f, max_row=f):
        for c in r: c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
def sd(ws, fi, ff, kw=None):
    for i, r in enumerate(ws.iter_rows(min_row=fi, max_row=ff), start=1):
        fl = FA if i%2==0 else PatternFill()
        if kw and r[0].value and any(k in str(r[0].value).lower() for k in kw): fl = FF
        for c in r: c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
def interp(ws, f, n, t):
    ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=n)
    c = ws.cell(row=f, column=1, value=f"-> {t}"); c.fill = FY; c.font = fI; c.alignment = AL; c.border = BT
    ws.row_dimensions[f].height = 28
def ajc(ws):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 35)

with pd.ExcelWriter(RUTA_EXCEL_OUT, engine="openpyxl") as writer:
    # H1: Resumen
    s = "1_Resumen"; nc = len(df_resumen_ejecutivo.columns) + 1
    df_resumen_ejecutivo.to_excel(writer, sheet_name=s, startrow=3)
    ws = writer.sheets[s]
    h(ws, 1, nc, f"ECOMMERCE — {NOMBRE_COMERCIO} | {datetime.today().strftime('%d/%m/%Y')}")
    h(ws, 2, nc, "RESUMEN INDICADORES POR MES", fl=FS)
    hd(ws, 4); sd(ws, 5, ws.max_row, kw=["fraude","ratio"])
    ajc(ws)

    # H2: Deciles
    s = "2_Deciles"; nc2 = len(tabla_deciles.columns) + 2
    tabla_deciles.to_excel(writer, sheet_name=s, startrow=3)
    ws = writer.sheets[s]
    h(ws, 1, nc2, "TABLA DE DECILES POR MONTO")
    h(ws, 2, nc2, "D01=menor D10=mayor | 4 decimales en FR%", fl=FS)
    hd(ws, 4); sd(ws, 5, ws.max_row); ajc(ws)

    # H3: Dimensiones
    s = "3_Dimensiones"; w3 = writer.book.create_sheet(s); writer.sheets[s] = w3; fa = 1
    dims_export = [
        (df_por_pais, "País", "Distribución de fraude por país origen."),
        (df_por_region, "Región", "Ojo: puede tener vacíos."),
        (df_por_ciudad, "Ciudad", "Ojo: puede tener vacíos."),
        (df_por_seg, "Segmento", "Segmento individual."),
        (df_por_grupo, "Segmento Grupo", "Affluent, Mass, etc."),
        (df_por_tipo, "Tipo Producto", "TD vs TC."),
        (df_por_seguro, "Seguridad (ECI)", "Seguro vs No Seguro."),
        (df_por_cvv, "CVV/Red", "No Match = alerta."),
    ]
    for dfm, tit, itp in dims_export:
        if dfm is None or len(dfm) == 0: continue
        do = dfm.reset_index(); nd = len(do.columns)
        h(w3, fa, nd, f"{tit.upper()}"); fa += 1
        for j, col in enumerate(do.columns, start=1):
            c = w3.cell(row=fa, column=j, value=col); c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for i, row in do.iterrows():
            fl = FA if i%2==0 else PatternFill()
            for j, val in enumerate(row, start=1):
                c = w3.cell(row=fa, column=j, value=f"{val:,.4f}" if isinstance(val,float) else val)
                c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        w3.merge_cells(start_row=fa, start_column=1, end_row=fa, end_column=nd)
        c = w3.cell(row=fa, column=1, value=f"-> {itp}"); c.fill = FY; c.font = fI; c.alignment = AL; c.border = BT
        w3.row_dimensions[fa].height = 25; fa += 3
    ajc(w3)

    # H4: Motivos de Rechazo
    if not df_motivos.empty:
        s = "4_Motivos_Rechazo"; df_motivos.to_excel(writer, sheet_name=s, startrow=3)
        ws = writer.sheets[s]; nc4 = len(df_motivos.columns) + 1
        h(ws, 1, nc4, "MOTIVOS DE RECHAZO")
        h(ws, 2, nc4, "Solo transacciones denegadas | Clasificación por razón de respuesta", fl=FS)
        hd(ws, 4); sd(ws, 5, ws.max_row)
        fi = ws.max_row + 2
        interp(ws, fi, nc4, f"Clientes con rechazo Y fraude aprobado: {len(interseccion):,} de {len(cli_con_fraude):,} clientes fraude")
        ajc(ws)

    # H5: Cascada Fraude
    if not tabla_cascada.empty:
        s = "5_Cascada_Fraude"; tabla_cascada.to_excel(writer, sheet_name=s, startrow=3)
        ws = writer.sheets[s]; nc5 = len(tabla_cascada.columns) + 1
        h(ws, 1, nc5, "CASCADA DE FRAUDE — N FRAUDES POR CLIENTE")
        h(ws, 2, nc5, "Cuántos clientes tienen 1, 2, 3+ fraudes aprobados", fl=FS)
        hd(ws, 4); sd(ws, 5, ws.max_row)
        ajc(ws)

    # H6: Features Análisis
    s = "6_Features"; w6 = writer.book.create_sheet(s); writer.sheets[s] = w6; fa = 1
    for tit, tf in tablas_features.items():
        if tf is None or tf.empty: continue
        do = tf.reset_index(); nd = len(do.columns)
        h(w6, fa, nd, tit.upper()); fa += 1
        for j, col in enumerate(do.columns, start=1):
            c = w6.cell(row=fa, column=j, value=str(col)); c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for i, row in do.iterrows():
            fl = FA if i%2==0 else PatternFill()
            for j, val in enumerate(row, start=1):
                c = w6.cell(row=fa, column=j, value=f"{val:.4f}" if isinstance(val,float) else val)
                c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        fa += 2
    ajc(w6)

    # H7: Cruces
    s = "7_Cruces"; w7 = writer.book.create_sheet(s); writer.sheets[s] = w7; fa = 1
    for tit, tup in r_cruces.items():
        if tup is None or tup[0] is None: continue
        piv = tup[0]
        do = piv.reset_index(); nd = len(do.columns)
        h(w7, fa, nd, tit.upper(), fl=FS); fa += 1
        for j, col in enumerate(do.columns, start=1):
            c = w7.cell(row=fa, column=j, value=str(col)); c.fill = FS; c.font = fH; c.alignment = AC; c.border = BT
        fa += 1
        for i, row in do.iterrows():
            fl = FA if i%2==0 else PatternFill()
            for j, val in enumerate(row, start=1):
                c = w7.cell(row=fa, column=j, value=round(val,4) if isinstance(val,float) else val)
                c.fill = fl; c.font = fN; c.alignment = AC; c.border = BT
            fa += 1
        fa += 2
    ajc(w7)

    # H8: Estadísticas
    s = "8_Estadisticas"; stats_monto.to_excel(writer, sheet_name=s, index=False, startrow=3)
    ws = writer.sheets[s]; nc8 = len(stats_monto.columns)
    h(ws, 1, nc8, "ESTADISTICAS DE MONTO")
    h(ws, 2, nc8, "Total | Fraude Total | Fraude Aprobado | No Fraude", fl=FS)
    hd(ws, 4); sd(ws, 5, ws.max_row); ajc(ws)

    # H9: Datos con features (sample, para revisión)
    s = "9_Sample_Features"; sample = df_feat.sample(min(1000, len(df_feat)), random_state=42)
    cols_export = ["ID_CLIENTE","DATETIME","MONTO","ES_FRAUDE","ES_FRAUDE_APROBADO",
                   "N_TRX_5MIN","N_TRX_24H","GAP_MINUTOS","ES_RAFAGA",
                   "MONTO_ACUM_24H","ZSCORE_MONTO_CLI","ES_PRIMERA_VEZ_COMERCIO",
                   "HUBO_FRAUDE_PREVIO_24H","N_RECHAZOS_24H","N_CVV_FAIL_24H","HUBO_CVV_FAIL_PREVIO"]
    cols_export = [c for c in cols_export if c in sample.columns]
    sample[cols_export].to_excel(writer, sheet_name=s, index=False, startrow=2)
    ws = writer.sheets[s]; ncs = len(cols_export)
    h(ws, 1, ncs, "MUESTRA DE 1000 TRX CON FEATURES (para inspección visual)")
    hd(ws, 3); ajc(ws)

print(f"\nExcel exportado: {RUTA_EXCEL_OUT}")
print("Hojas: 1_Resumen | 2_Deciles | 3_Dimensiones | 4_Motivos_Rechazo")
print("       5_Cascada_Fraude | 6_Features | 7_Cruces | 8_Estadisticas | 9_Sample_Features")

# Guardar dataframe de features para uso de la app Streamlit (opcional)
RUTA_PARQUET_FEAT = RUTA_PARQUET.replace(".parquet", "_features.parquet")
try:
    df_feat.to_parquet(RUTA_PARQUET_FEAT)
    print(f"\nFeatures guardadas en: {RUTA_PARQUET_FEAT}")
    print("La app Streamlit puede cargar este archivo directamente.")
except Exception as e:
    print(f"\nNo se pudo guardar parquet de features: {e}")
