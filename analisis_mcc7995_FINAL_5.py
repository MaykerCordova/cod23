# =============================================================================
# ANÁLISIS INTEGRAL MCC 7995 — Juegos de Azar
# Feature engineering con Polars para eficiencia
# =============================================================================
import pandas as pd
import numpy as np
import polars as pl
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.stats import chi2_contingency
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# =============================================================================
# AJUSTA SOLO ESTE BLOQUE
# =============================================================================
COLS = {
    "FECHA"           : "ACF-FECHA TRX",
    "HORA"            : "ACF-HORA TRX",
    "DATETIME"        : "FECHA_HORA",
    "COMERCIO"        : "ACF-NOMBRE/LOCALIZACION COMERCIO",
    "INDICADOR"       : "ACF-INDICADOR DE FRAUDE",
    "COD_RESPUESTA"   : "ACF-COD RPTA",
    "CONDICION_RT"    : "CONDICION RT",
    "CANAL"           : "ACF-CANAL",
    "ENTRY_MODE"      : "ACF-ENTRY MODE",
    "SALDO"           : "ACF-SALDO DISPONIBLE EN MONEDA TRX",
    "ID_CLIENTE"      : "ACF-ID CLIENTE",
    "ESI_UCAP"        : "ACF-ECI/UCAF",
    "PAIS"            : "ACF-PAIS ORIGEN 87519",
    "COD_RED_COMERCIO": "ACF-COD RED COMERCIO",
    "MONTO"           : "ACF-MONTO EN MONEDA LOCAL",
    "SEGMENTO"        : "VAA-EVENTO DE COMPROMISO OTRA FUENTE",
    "TIPO_PRODUCTO"   : "ACF-TIPO PROD TC",
    "BIN"             : "ACF-BIN",
}
RUTA_PARQUET   = r"C:\ruta\al\consolidado_7995.parquet"
RUTA_EXCEL_OUT = r"C:\ruta\salida\analisis_mcc7995_resultado.xlsx"
# =============================================================================

SEG_NOMBRE = {"30":"Polo Direccion","99":"Polo Direccion","31":"Premium","32":"Preferente",
    "33":"Personal","34":"Estandar","5":"Inst. Financieras","21":"Corporativo",
    "2":"Mediano Empresas","15":"Sector Gobierno","16":"Otras Instituciones",
    "3":"Pequenas Empresas","4":"Negocios 2","7":"Negocios 3","8":"Negocios 1","13":"Microempresas"}
SEG_GRUPO = {"30":"Affluent","99":"Affluent","31":"Emerging Affluent","32":"Emerging Affluent",
    "33":"Top of Mass","34":"Mass","5":"Corporate","21":"Corporate","2":"Commercial",
    "15":"Commercial","16":"Commercial","3":"Small Business","4":"Small Business",
    "7":"Small Business","8":"Small Business","13":"Small Business"}
COD_RED_LABEL = {"S":"Estatico (TD)","D":"Dinamico (TD/TC)","E":"Estatico (TC)","N":"No Match / Sin CVV"}
C_FRAUDE="#D9534F"; C_OK="#5B9BD5"; C_088="#F0AD4E"; C_NEU="#8EA9C1"; C_ROJO2="#8B0000"

# =============================================================================
# CARGA Y PREPARACIÓN (Pandas para compatibilidad con secciones de análisis)
# =============================================================================
print("Cargando parquet...")
df_raw = pd.read_parquet(RUTA_PARQUET)
col_map = {v:k for k,v in COLS.items()}
df = df_raw.rename(columns=col_map).copy()

faltantes = [a for a in COLS if a not in df.columns]
if faltantes: print(f"  Columnas no encontradas: {faltantes}")

df["MONTO"]=pd.to_numeric(df["MONTO"],errors="coerce")
df["SALDO"]=pd.to_numeric(df["SALDO"],errors="coerce")
df["DATETIME"]=pd.to_datetime(df["DATETIME"],errors="coerce")
df["FECHA"]=df["DATETIME"].dt.normalize()
df["MES"]=df["DATETIME"].dt.to_period("M").astype(str)

for c in ["INDICADOR","COD_RESPUESTA","CONDICION_RT","CANAL",
          "ESI_UCAP","COD_RED_COMERCIO","SEGMENTO","TIPO_PRODUCTO","ENTRY_MODE"]:
    if c in df.columns: df[c]=df[c].astype(str).str.strip().str.upper()
if "BIN" in df.columns: df["BIN"]=df["BIN"].astype(str).str.split(".").str[0].str.strip()

df["ESTADO"]=df["COD_RESPUESTA"].apply(lambda x:"APROBADA" if str(x).strip() in ["00","0000","000","0"] else "DENEGADA")
df["ES_FRAUDE"]=(df["INDICADOR"]=="F").astype(int)
df["ES_FRAUDE_APROBADO"]=((df["INDICADOR"]=="F")&(df["ESTADO"]=="APROBADA")).astype(int)
df["COND_088"]=df["CONDICION_RT"].str.contains("088",na=False)

# SEGURO: 2, 5, 02, 05
df["SEGURO"]=df["ESI_UCAP"].apply(
    lambda x: "Seguro" if str(x).strip() in ["2","5","02","05"] else "No Seguro"
)
df["SEG_NOMBRE"]=df["SEGMENTO"].map(SEG_NOMBRE).fillna("Otro/Sin seg")
df["SEG_GRUPO"]=df["SEGMENTO"].map(SEG_GRUPO).fillna("Otro/Sin seg")
df["COD_RED_LABEL"]=df["COD_RED_COMERCIO"].map(COD_RED_LABEL).fillna("Otro")

df_ap=df[df["ESTADO"]=="APROBADA"].copy()
df_den=df[df["ESTADO"]=="DENEGADA"].copy()
df_f_ap=df[df["ES_FRAUDE_APROBADO"]==1].copy()
df_f_tot=df[df["ES_FRAUDE"]==1].copy()

print(f"  Total: {len(df):,} | Ap: {len(df_ap):,} | Den: {len(df_den):,}")
print(f"  Fraude total: {df['ES_FRAUDE'].sum():,} | Fraude aprobado: {df['ES_FRAUDE_APROBADO'].sum():,}")


# =============================================================================
# INGENIERÍA DE VARIABLES CON POLARS (toda la data, sin muestreo)
# =============================================================================
print("\n" + "="*65)
print("INGENIERIA DE VARIABLES — Polars (eficiente, toda la data)")
print("="*65)

# Convertir a Polars
cols_feat = ["ID_CLIENTE","DATETIME","COMERCIO","MONTO","SALDO",
             "ES_FRAUDE","ES_FRAUDE_APROBADO","ESTADO","INDICADOR",
             "SEG_NOMBRE","SEG_GRUPO","TIPO_PRODUCTO","SEGURO"]
if "BIN" in df.columns: cols_feat.append("BIN")

plf_all = pl.from_pandas(df[cols_feat].reset_index(drop=True))
plf_all = plf_all.with_columns(pl.col("DATETIME").cast(pl.Datetime("us")))
plf_all = plf_all.sort(["ID_CLIENTE","DATETIME"])

plf_ap  = plf_all.filter(pl.col("ESTADO") == "APROBADA")
plf_den = plf_all.filter(pl.col("ESTADO") == "DENEGADA")

# ── F1: RATIO MONTO / SALDO ──────────────────────────────────────────────────
print("  F1: Ratio Monto/Saldo...")
plf_ap = plf_ap.with_columns(
    pl.when(pl.col("SALDO") > 0)
    .then(pl.col("MONTO") / pl.col("SALDO"))
    .otherwise(None)
    .alias("RATIO_MONTO_SALDO")
)

# ── F2: Z-SCORE MONTO POR CLIENTE ────────────────────────────────────────────
print("  F2: Z-score monto por cliente...")
plf_ap = plf_ap.with_columns([
    pl.col("MONTO").mean().over("ID_CLIENTE").alias("_mean_cli"),
    pl.col("MONTO").std().over("ID_CLIENTE").alias("_std_cli"),
])
plf_ap = plf_ap.with_columns(
    pl.when(pl.col("_std_cli") > 0)
    .then((pl.col("MONTO") - pl.col("_mean_cli")) / pl.col("_std_cli"))
    .otherwise(0.0)
    .alias("ZSCORE_MONTO_CLI")
).drop(["_mean_cli","_std_cli"])

# ── F3: COMERCIO NUEVO ───────────────────────────────────────────────────────
print("  F3: Comercio nuevo...")
plf_ap = plf_ap.with_columns(
    pl.lit(1).cum_sum().over(["ID_CLIENTE","COMERCIO"]).alias("_rank_com")
)
plf_ap = plf_ap.with_columns(
    (pl.col("_rank_com") == 1).cast(pl.Int32).alias("ES_COMERCIO_NUEVO")
).drop("_rank_com")

# ── F4: N° COMERCIOS DISTINTOS EN EL DÍA ─────────────────────────────────────
print("  F4: Comercios distintos por dia...")
plf_ap = plf_ap.with_columns(
    pl.col("DATETIME").dt.date().alias("FECHA_DIA")
)
n_com_dia = (
    plf_ap.group_by(["ID_CLIENTE","FECHA_DIA"])
    .agg(pl.col("COMERCIO").n_unique().alias("N_COMERCIOS_DIA"))
)
plf_ap = plf_ap.join(n_com_dia, on=["ID_CLIENTE","FECHA_DIA"], how="left")

# ── F5: HORA Y FRANJA HORARIA ────────────────────────────────────────────────
print("  F5: Hora y franja horaria...")
plf_ap = plf_ap.with_columns(
    pl.col("DATETIME").dt.hour().alias("HORA_NUM")
)
plf_ap = plf_ap.with_columns(
    pl.when(pl.col("HORA_NUM") <= 5).then(pl.lit("00-05 Madrugada"))
    .when(pl.col("HORA_NUM") <= 11).then(pl.lit("06-11 Manana"))
    .when(pl.col("HORA_NUM") <= 17).then(pl.lit("12-17 Tarde"))
    .when(pl.col("HORA_NUM") <= 20).then(pl.lit("18-20 Noche"))
    .otherwise(pl.lit("21-23 Noche Tardia"))
    .alias("FRANJA_HORARIA")
)

# ── F6: DÍAS DESDE ÚLTIMA TRANSACCIÓN ────────────────────────────────────────
print("  F6: Dias desde ultima transaccion...")
plf_ap = plf_ap.with_columns(
    pl.col("DATETIME").shift(1).over("ID_CLIENTE").alias("_prev_dt")
)
plf_ap = plf_ap.with_columns(
    ((pl.col("DATETIME") - pl.col("_prev_dt")).dt.total_seconds() / 86400)
    .alias("DIAS_DESDE_ULTIMA_TRX")
).drop("_prev_dt")

# ── F7: MONTO ACUMULADO ÚLTIMAS 2 HORAS ──────────────────────────────────────
print("  F7: Monto acumulado ultimas 2 horas (rolling)...")
plf_ap_sorted = plf_ap.sort("DATETIME")
monto_rolling = (
    plf_ap_sorted
    .rolling(index_column="DATETIME", period="2h", group_by="ID_CLIENTE")
    .agg(pl.col("MONTO").sum().alias("_monto_acum_2h_total"))
)
# Restar el monto de la transaccion actual (rolling incluye la propia fila)
plf_ap = plf_ap.with_row_index("_idx")
monto_rolling = monto_rolling.with_row_index("_idx")
plf_ap = plf_ap.join(
    monto_rolling.select(["_idx","_monto_acum_2h_total"]), on="_idx", how="left"
)
plf_ap = plf_ap.with_columns(
    (pl.col("_monto_acum_2h_total") - pl.col("MONTO")).clip(0, None).alias("MONTO_ACUM_2H")
).drop(["_monto_acum_2h_total","_idx"])

# ── F8-F9: RECHAZOS PREVIOS EN 24H (merge_asof eficiente) ────────────────────
print("  F8-F9: Rechazos previos 24h (merge_asof, toda la data)...")

if len(plf_den) > 0:
    # Cumulative denial count per client
    plf_den_cum = (
        plf_den.sort(["ID_CLIENTE","DATETIME"])
        .with_columns(
            pl.lit(1).cum_sum().over("ID_CLIENTE").alias("CUM_DEN")
        )
        .select(["ID_CLIENTE","DATETIME","CUM_DEN"])
    )

    # merge_asof: para cada aprobada, encontrar la denegada mas reciente del mismo cliente
    plf_ap_rec = plf_ap.sort("DATETIME").with_row_index("_idx_ap")
    merged_now = plf_ap_rec.select(["_idx_ap","ID_CLIENTE","DATETIME"]).join_asof(
        plf_den_cum.rename({"CUM_DEN":"CUM_DEN_NOW"}),
        on="DATETIME", by="ID_CLIENTE", strategy="backward"
    )

    # Buscar denegadas 24h antes
    plf_ap_24h = plf_ap_rec.with_columns(
        (pl.col("DATETIME") - pl.duration(hours=24)).alias("DT_24H")
    ).select(["_idx_ap","ID_CLIENTE","DT_24H"])
    merged_24h = plf_ap_24h.rename({"DT_24H":"DATETIME"}).join_asof(
        plf_den_cum.rename({"CUM_DEN":"CUM_DEN_24H"}),
        on="DATETIME", by="ID_CLIENTE", strategy="backward"
    )

    # N_RECHAZOS_24H = CUM_DEN_NOW - CUM_DEN_24H
    rechazos = merged_now.select(["_idx_ap","CUM_DEN_NOW"]).join(
        merged_24h.select(["_idx_ap","CUM_DEN_24H"]), on="_idx_ap", how="left"
    ).with_columns(
        (pl.col("CUM_DEN_NOW").fill_null(0) - pl.col("CUM_DEN_24H").fill_null(0))
        .clip(0, None).cast(pl.Int32).alias("N_RECHAZOS_24H")
    )

    plf_ap = plf_ap.with_row_index("_idx_ap").join(
        rechazos.select(["_idx_ap","N_RECHAZOS_24H"]), on="_idx_ap", how="left"
    ).drop("_idx_ap")
else:
    plf_ap = plf_ap.with_columns(pl.lit(0).alias("N_RECHAZOS_24H"))

plf_ap = plf_ap.with_columns(
    (pl.col("N_RECHAZOS_24H") / (pl.col("N_RECHAZOS_24H") + 1)).alias("RATIO_FALLIDOS")
)

# ── F10: FR HISTÓRICO DEL COMERCIO ───────────────────────────────────────────
print("  F10: FR historico por comercio...")
fr_com = (
    plf_ap.group_by("COMERCIO")
    .agg(
        (pl.col("ES_FRAUDE_APROBADO").sum() / pl.col("ES_FRAUDE_APROBADO").count())
        .alias("FR_HISTORICO_COMERCIO")
    )
)
plf_ap = plf_ap.join(fr_com, on="COMERCIO", how="left")

# ── F11: CONCENTRACIÓN BIN EN COMERCIO ────────────────────────────────────────
if "BIN" in plf_ap.columns:
    print("  F11: Concentracion BIN por comercio...")
    total_com = plf_ap.group_by("COMERCIO").agg(pl.count().alias("_n_total"))
    top_bin = (
        plf_ap.group_by(["COMERCIO","BIN"]).agg(pl.count().alias("_n_bin"))
        .sort("_n_bin", descending=True)
        .group_by("COMERCIO").first()
        .select(["COMERCIO","_n_bin"])
    )
    conc = top_bin.join(total_com, on="COMERCIO").with_columns(
        (pl.col("_n_bin") / pl.col("_n_total")).alias("CONC_BIN_COMERCIO")
    ).select(["COMERCIO","CONC_BIN_COMERCIO"])
    plf_ap = plf_ap.join(conc, on="COMERCIO", how="left")
else:
    plf_ap = plf_ap.with_columns(pl.lit(None).cast(pl.Float64).alias("CONC_BIN_COMERCIO"))

# Convertir resultado a Pandas
print("  Convirtiendo resultado a Pandas...")
df_feat = plf_ap.to_pandas()

# Crear buckets para tablas cruzadas
df_feat["BUCKET_RATIO_SALDO"] = pd.cut(df_feat["RATIO_MONTO_SALDO"].clip(0,2),
    bins=[0,0.1,0.3,0.5,0.8,2], labels=["<10%","10-30%","30-50%","50-80%",">80%"], include_lowest=True)
df_feat["BUCKET_ZSCORE"] = pd.cut(df_feat["ZSCORE_MONTO_CLI"].clip(-3,3),
    bins=[-3,-2,-1,0,1,2,3], labels=["<-2SD","-2a-1SD","-1a0SD","0a1SD","1a2SD",">2SD"], include_lowest=True)
df_feat["BUCKET_N_COMERCIOS"] = pd.cut(df_feat["N_COMERCIOS_DIA"],
    bins=[0,1,2,3,5,100], labels=["1 comercio","2 com","3 com","4-5 com",">5 com"], include_lowest=True)
df_feat["BUCKET_INACTIVIDAD"] = pd.cut(df_feat["DIAS_DESDE_ULTIMA_TRX"].clip(0,60),
    bins=[-0.001,0.083,1,7,14,30,60], labels=["<2h","1dia","2-7d","8-14d","15-30d",">30d"], include_lowest=True)
df_feat["BUCKET_RECHAZOS"] = pd.cut(df_feat["N_RECHAZOS_24H"].clip(0,10),
    bins=[-0.001,0,1,2,3,10], labels=["0 rech","1 rech","2 rech","3 rech","4+ rech"], include_lowest=True)
df_feat["BUCKET_FR_COMERCIO"] = pd.cut(df_feat["FR_HISTORICO_COMERCIO"].clip(0,0.1),
    bins=[-0.001,0.001,0.005,0.01,0.05,0.1], labels=["<0.1%","0.1-0.5%","0.5-1%","1-5%",">5%"], include_lowest=True)
if df_feat["CONC_BIN_COMERCIO"].notna().any():
    df_feat["BUCKET_CONC_BIN"] = pd.cut(df_feat["CONC_BIN_COMERCIO"].clip(0,1),
        bins=[-0.001,0.25,0.5,0.75,0.9,1.0], labels=["<25%","25-50%","50-75%","75-90%",">90%"], include_lowest=True)

# Resumen de features
print("\nFeatures calculadas exitosamente sobre toda la data:")
features_num=["RATIO_MONTO_SALDO","ZSCORE_MONTO_CLI","N_COMERCIOS_DIA",
    "HORA_NUM","DIAS_DESDE_ULTIMA_TRX","MONTO_ACUM_2H","N_RECHAZOS_24H",
    "RATIO_FALLIDOS","FR_HISTORICO_COMERCIO","CONC_BIN_COMERCIO"]
for f in features_num:
    if f in df_feat.columns:
        s=df_feat[f].dropna()
        print(f"  {f:<30} n={len(s):,} mean={s.mean():.4f} median={s.median():.4f}")


# =============================================================================
# ANÁLISIS DE FEATURES — tabla por cada variable vs fraude aprobado
# =============================================================================
print("\n" + "="*65)
print("ANALISIS DE FEATURES vs FRAUDE APROBADO")
print("="*65)

def analizar_feature(df_base, col_bucket):
    return (
        df_base.groupby(col_bucket, observed=True)
        .agg(Txns=("ES_FRAUDE_APROBADO","count"),
             Fraude_Ap=("ES_FRAUDE_APROBADO","sum"),
             FR_pct=("ES_FRAUDE_APROBADO",lambda x:x.mean()*100),
             Ticket_prom=("MONTO","mean"),
             Ticket_f_prom=("MONTO",lambda x:x[df_base.loc[x.index,"ES_FRAUDE_APROBADO"]==1].mean()))
        .rename(columns={"FR_pct":"FR% Ap","Ticket_prom":"Ticket prom","Ticket_f_prom":"Ticket fraude prom"})
    )

buckets_analizar=[("BUCKET_RATIO_SALDO","Ratio Monto/Saldo"),
    ("BUCKET_ZSCORE","Z-Score Monto"),("ES_COMERCIO_NUEVO","Comercio Nuevo"),
    ("BUCKET_N_COMERCIOS","N Comercios Dia"),("FRANJA_HORARIA","Franja Horaria"),
    ("BUCKET_INACTIVIDAD","Inactividad"),("BUCKET_RECHAZOS","Rechazos 24h"),
    ("BUCKET_FR_COMERCIO","FR Historico Comercio")]
if "BUCKET_CONC_BIN" in df_feat.columns:
    buckets_analizar.append(("BUCKET_CONC_BIN","Concentracion BIN"))

tablas_features={}
for col_b,titulo in buckets_analizar:
    if col_b not in df_feat.columns: continue
    print(f"\n--- {titulo} ---")
    tf=analizar_feature(df_feat,col_b)
    print(tf.to_string())
    tablas_features[titulo]=tf

# =============================================================================
# SECCIÓN 1 — RESUMEN EJECUTIVO POR MES
# =============================================================================
print("\n" + "="*65); print("SECCION 1 — RESUMEN EJECUTIVO")
meses_disponibles=sorted(df_ap["MES"].unique())
filas_resumen=[]
for mes in meses_disponibles:
    sub_mes=df[df["MES"]==mes]; sub_ap=df_ap[df_ap["MES"]==mes]
    sub_f_ap=sub_ap[sub_ap["ES_FRAUDE_APROBADO"]==1]; sub_f_tot=sub_mes[sub_mes["ES_FRAUDE"]==1]
    n_ap_m=len(sub_ap); n_den_m=(sub_mes["ESTADO"]=="DENEGADA").sum()
    monto_ap_m=sub_ap["MONTO"].sum(); ticket_ap_m=sub_ap["MONTO"].mean()
    n_cli_m=sub_ap["ID_CLIENTE"].nunique()
    n_f_tot_m=len(sub_f_tot); monto_f_tot_m=sub_f_tot["MONTO"].sum()
    n_f_ap_m=len(sub_f_ap); monto_f_ap_m=sub_f_ap["MONTO"].sum()
    ticket_f_ap_m=sub_f_ap["MONTO"].mean() if n_f_ap_m>0 else 0
    ratio_trx_m=n_f_ap_m/n_ap_m if n_ap_m>0 else 0; ratio_mon_m=monto_f_ap_m/monto_ap_m if monto_ap_m>0 else 0
    filas_resumen.append({"Total trx aprobadas":f"{n_ap_m:,}","Total trx denegadas":f"{n_den_m:,}",
        "Total monto aprobado (S/)":f"S/ {monto_ap_m:,.2f}","Ticket promedio (S/)":f"S/ {ticket_ap_m:,.2f}",
        "N clientes unicos":f"{n_cli_m:,}","Trxs fraude total":f"{n_f_tot_m:,}",
        "Trxs fraude aprobado":f"{n_f_ap_m:,}","Monto fraude total (S/)":f"S/ {monto_f_tot_m:,.2f}",
        "Monto fraude aprobado (S/)":f"S/ {monto_f_ap_m:,.2f}","Ticket fraude aprobado":f"S/ {ticket_f_ap_m:,.2f}",
        "Ratio trx fraude aprobadas":f"{ratio_trx_m*100:.4f}%","Ratio fraude soles (ap)":f"{ratio_mon_m*100:.4f}%"})
df_resumen_ejecutivo=pd.DataFrame(filas_resumen,index=meses_disponibles).T
df_resumen_ejecutivo.index.name="Indicador"
print(df_resumen_ejecutivo.to_string())

n_ap=len(df_ap); monto_ap=df_ap["MONTO"].sum()
n_f_tot_g=df["ES_FRAUDE"].sum(); n_f_ap_g=df["ES_FRAUDE_APROBADO"].sum()
monto_f_ap_g=df_f_ap["MONTO"].sum() if len(df_f_ap)>0 else 0
ratio_trx_f=n_f_ap_g/n_ap if n_ap>0 else 0; ratio_monto_f=monto_f_ap_g/monto_ap if monto_ap>0 else 0
nivel_riesgo="ALTO >=5%" if ratio_trx_f>=0.05 else ("MEDIO [1%-5%]" if ratio_trx_f>=0.01 else "REGULAR <1%")
riesgo_texto={"REGULAR <1%":"Bajo riesgo relativo.","MEDIO [1%-5%]":"Monitoreo activo recomendado.","ALTO >=5%":"Intervencion inmediata requerida."}
df_umbral=pd.DataFrame({"Umbral":["Regular","Medio","Alto"],"Rango":["<1%","[1%-5%]",">=5%"],
    "Nivel actual":["AQUI" if ratio_trx_f<0.01 else "","AQUI" if 0.01<=ratio_trx_f<0.05 else "","AQUI" if ratio_trx_f>=0.05 else ""]})

# =============================================================================
# SECCIÓN 2 — DECILES (misma estructura en principal y apertura D10)
# =============================================================================
print("\n" + "="*65); print("SECCION 2 — DECILES")
def construir_tabla_deciles(df_b,bins_c=None,n_d=10,lb="D"):
    db=df_b.copy().dropna(subset=["MONTO"])
    if bins_c is not None:
        labels=[f"{lb}{str(i+1).zfill(2)}" for i in range(len(bins_c)-1)]
        db["DECIL"]=pd.cut(db["MONTO"],bins=bins_c,labels=labels,include_lowest=True)
        db["DECIL"]=pd.Categorical(db["DECIL"],categories=labels,ordered=True)
    else:
        labels=[f"{lb}{str(i+1).zfill(2)}" for i in range(n_d)]
        db["DECIL"]=pd.qcut(db["MONTO"],q=n_d,labels=labels,duplicates="drop")
    rows=[]; fa_acum=0; ta_acum=0; tot_ap=len(db[db["ESTADO"]=="APROBADA"]); tot_fa=db["ES_FRAUDE_APROBADO"].sum()
    for decil in db["DECIL"].cat.categories:
        s=db[db["DECIL"]==decil]
        if len(s)==0: continue
        sa=s[s["ESTADO"]=="APROBADA"]; sd=s[s["ESTADO"]=="DENEGADA"]
        sfa=s[s["ES_FRAUDE_APROBADO"]==1]; sft=s[s["ES_FRAUDE"]==1]
        na=len(sa); ma=sa["MONTO"].sum(); nd=len(sd); md=sd["MONTO"].sum()
        nfa=len(sfa); mfa=sfa["MONTO"].sum(); nft=len(sft); mft=sft["MONTO"].sum()
        ta_acum+=na; fa_acum+=nfa
        rows.append({"Decil":decil,"Rango S/":f"{s['MONTO'].min():.0f}-{s['MONTO'].max():.0f}",
            "Trx Aprobadas":na,"Monto Aprobadas":round(ma,1),
            "Freq Aprobadas":f"{na/tot_ap*100:.2f}%" if tot_ap>0 else "0%",
            "Freq Acum Ap":f"{ta_acum/tot_ap*100:.2f}%" if tot_ap>0 else "0%",
            "Trx Denegadas":nd,"Monto Denegadas":round(md,1),
            "Total Trx":len(s),"Ticket Promedio":round(s["MONTO"].mean(),2),
            "Monto Total S/":round(s["MONTO"].sum(),1),
            "Fraude Trx Total":nft,"Fraude Monto Total":round(mft,1),
            "Fraude Trx Aprobado":nfa,"Fraude Monto Aprobado":round(mfa,1),
            "Fraude Freq Ap":f"{nfa/tot_fa*100:.4f}%" if tot_fa>0 else "0%",
            "Fraude Freq Acum":f"{fa_acum/tot_fa*100:.4f}%" if tot_fa>0 else "0%",
            "FR/Trx Ap":f"{nfa/na*100:.4f}%" if na>0 else "0%",
            "NoFR/Trx Ap":f"{(1-nfa/na)*100:.4f}%" if na>0 else "100%"})
    return pd.DataFrame(rows).set_index("Decil")

tabla_deciles=construir_tabla_deciles(df,n_d=10); print(tabla_deciles.to_string())
print("\nApertura D10:")
m_d10_min=df["MONTO"].quantile(0.90); df_d10=df[df["MONTO"]>=m_d10_min].copy()
bins_d10=np.unique([df_d10["MONTO"].quantile(p/100) for p in np.linspace(0,100,11)])
if len(bins_d10)<3: bins_d10=np.unique([df_d10["MONTO"].quantile(p/100) for p in np.linspace(0,100,6)])
tabla_d10=construir_tabla_deciles(df_d10,bins_c=bins_d10,lb="SD")
print(tabla_d10.to_string())


# SECCIONES 3-7 (Velocidad, Diaria, 088, Dimensiones, Estadisticas)
print("\n" + "="*65); print("SECCION 3 — VELOCIDAD POR INTERVALO")
df_vel=df.sort_values(["ID_CLIENTE","DATETIME"]).copy()
df_vel["PREV_DT"]=df_vel.groupby("ID_CLIENTE")["DATETIME"].shift(1)
df_vel["MIN_DESDE_PREV"]=(df_vel["DATETIME"]-df_vel["PREV_DT"]).dt.total_seconds()/60
df_vel2=df_vel.dropna(subset=["MIN_DESDE_PREV"]); df_vel2=df_vel2[df_vel2["MIN_DESDE_PREV"]>=0].copy()
def bk_int(m):
    if m<=2:return "<=2 min"
    elif m<=5:return "<=5 min"
    elif m<=10:return "<=10 min"
    elif m<=15:return "<=15 min"
    elif m<=20:return "<=20 min"
    elif m<=60:return "<=1 hora"
    else:return ">1 hora"
OB=["<=2 min","<=5 min","<=10 min","<=15 min","<=20 min","<=1 hora",">1 hora"]
df_vel2["BUCKET"]=pd.Categorical(df_vel2["MIN_DESDE_PREV"].apply(bk_int),categories=OB,ordered=True)
tabla_vel=(df_vel2.groupby("BUCKET",observed=True)
    .agg(Transacciones=("MONTO","count"),Genuinas=("ES_FRAUDE",lambda x:(x==0).sum()),
         Fraude_Total=("ES_FRAUDE","sum"),Fraude_Aprobado=("ES_FRAUDE_APROBADO","sum"),
         FR_Total_pct=("ES_FRAUDE",lambda x:x.mean()*100),FR_Aprobado_pct=("ES_FRAUDE_APROBADO",lambda x:x.mean()*100),
         Monto_Genuino=("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].sum()),
         Monto_Fraude_Total=("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE"]==1].sum()),
         Monto_Fraude_Ap=("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
         Ticket_Genuino=("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].mean()),
         Ticket_Fraude=("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].mean()))
    .assign(Pct_fraudes_total=lambda x:x["Fraude_Total"]/x["Fraude_Total"].sum()*100))
print(tabla_vel.to_string())

print("\n" + "="*65); print("SECCION 4 — TRX DIARIA POR CLIENTE")
dia_cli=(df_ap.assign(FD=lambda x:x["DATETIME"].dt.date)
    .groupby(["ID_CLIENTE","FD"])
    .agg(Trx=("MONTO","count"),Monto=("MONTO","sum"),F_tot=("ES_FRAUDE","sum"),
         F_ap=("ES_FRAUDE_APROBADO","sum"),MF_tot=("MONTO",lambda x:x[df_ap.loc[x.index,"ES_FRAUDE"]==1].sum()),
         MF_ap=("MONTO",lambda x:x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()))
    .reset_index()
    .assign(tiene_f=lambda x:(x["F_ap"]>0).astype(int),Trx_g=lambda x:x["Trx"]-x["F_tot"],M_g=lambda x:x["Monto"]-x["MF_tot"]))
def bk_d(n):
    if n<=6:return str(n)
    elif n<=9:return "7-9"
    elif n<=12:return "10-12"
    else:return "13+"
OD=["1","2","3","4","5","6","7-9","10-12","13+"]
dia_cli["BK"]=pd.Categorical(dia_cli["Trx"].apply(bk_d),categories=OD,ordered=True)
tabla_diaria=(dia_cli.groupby("BK",observed=True)
    .agg(Genuina=("tiene_f",lambda x:(x==0).sum()),Fraude_Aprobado=("tiene_f","sum"),
         Total_Clientes=("tiene_f","count"),Trx_Genuina=("Trx_g","sum"),Monto_Genuina=("M_g","sum"),
         Trx_Fraude_Total=("F_tot","sum"),Trx_Fraude_Aprobado=("F_ap","sum"),
         Monto_Fraude_Total=("MF_tot","sum"),Monto_Fraude_Aprobado=("MF_ap","sum"))
    .reset_index().rename(columns={"BK":"Intervalo"})
    .assign(Impacto_pct=lambda x:x["Fraude_Aprobado"]/x["Total_Clientes"]*100,
            Tk_Fraude=lambda x:np.where(x["Trx_Fraude_Aprobado"]>0,x["Monto_Fraude_Aprobado"]/x["Trx_Fraude_Aprobado"],0),
            Tk_Genuina=lambda x:np.where(x["Trx_Genuina"]>0,x["Monto_Genuina"]/x["Trx_Genuina"],0)))
total_clf=tabla_diaria["Fraude_Aprobado"].sum()
pct_1=tabla_diaria.loc[tabla_diaria["Intervalo"]=="1","Fraude_Aprobado"].sum()/total_clf*100 if total_clf>0 else 0
td=tabla_diaria.copy()
td["Impacto_pct"]=td["Impacto_pct"].map("{:.4f}%".format)
for c in ["Monto_Genuina","Monto_Fraude_Total","Monto_Fraude_Aprobado"]:td[c]=td[c].map("S/ {:,.2f}".format)
for c in ["Tk_Fraude","Tk_Genuina"]:td[c]=td[c].map("S/ {:,.2f}".format)
tabla_diaria_display=td
print(td.to_string(index=False))

print("\n" + "="*65); print("SECCION 5 — CONDICION 088")
f088t=(df["COND_088"]&(df["ES_FRAUDE"]==1)).sum()
f088a=(df["COND_088"]&(df["ES_FRAUDE_APROBADO"]==1)).sum()
fn088t=(~df["COND_088"]&(df["ES_FRAUDE"]==1)).sum()
fn088a=(~df["COND_088"]&(df["ES_FRAUDE_APROBADO"]==1)).sum()
if n_f_tot_g>0: print(f"  TOTAL: dentro088={f088t:,} ({f088t/n_f_tot_g*100:.4f}%) fuera={fn088t:,}")
if n_f_ap_g>0:  print(f"  APROBADO: dentro088={f088a:,} ({f088a/n_f_ap_g*100:.4f}%) fuera={fn088a:,}")

print("\n" + "="*65); print("SECCION 6 — DIMENSIONES")
def res_por(df_b,col,tn=10):
    return (df_b.groupby(col).agg(Txns=(col,"count"),Fraude_Total=("ES_FRAUDE","sum"),
        Fraude_Aprobado=("ES_FRAUDE_APROBADO","sum"),FR_T=("ES_FRAUDE",lambda x:x.mean()*100),
        FR_A=("ES_FRAUDE_APROBADO",lambda x:x.mean()*100),Monto_total=("MONTO","sum"),
        Monto_F_Ap=("MONTO",lambda x:x[df_b.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()))
    .sort_values("Fraude_Aprobado",ascending=False).head(tn)
    .rename(columns={"FR_T":"FR% Total","FR_A":"FR% Ap","Monto_F_Ap":"Monto Fraude Ap"}))
df_por_canal=res_por(df,"CANAL"); df_por_seg_nombre=res_por(df,"SEG_NOMBRE")
df_por_seg_grupo=res_por(df,"SEG_GRUPO"); df_por_seguro=res_por(df,"SEGURO")
df_por_tipo=res_por(df,"TIPO_PRODUCTO"); df_por_cvv=res_por(df,"COD_RED_LABEL")
df_por_pais=res_por(df,"PAIS")
df_top_com=(df.groupby("COMERCIO").agg(FT_g=("ES_FRAUDE","sum"),MFT_g=("MONTO",lambda x:x[df.loc[x.index,"ES_FRAUDE"]==1].sum()))
    .join(df_ap.groupby("COMERCIO").agg(Tarjetas=("ID_CLIENTE","nunique"),Trx=("MONTO","count"),
        FA=("ES_FRAUDE_APROBADO","sum"),MFA=("MONTO",lambda x:x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        MT=("MONTO","sum"),FRA=("ES_FRAUDE_APROBADO",lambda x:x.mean()*100)))
    .reset_index().sort_values("MFA",ascending=False).head(10)
    .rename(columns={"COMERCIO":"Comercio","Tarjetas":"#Tarj","Trx":"#TrxAp","FT_g":"#FraudeTotal","FA":"#FraudeAp",
        "MFT_g":"Monto F Total S/","MFA":"Monto F Ap S/","MT":"Monto Total S/","FRA":"FR%(ap)"}))
df_top_com["FR%(ap)"]=df_top_com["FR%(ap)"].map("{:.4f}%".format)

print("\n" + "="*65); print("SECCION 7 — ESTADISTICAS")
pcts=[10,25,50,75,90,95,99]; df_nofraude=df_ap[df_ap["ES_FRAUDE"]==0]
def ss(s):
    if len(s)==0:return [0]*13
    return [s.mean(),s.median(),s.std(),s.var(),s.min(),s.max()]+[s.quantile(p/100) for p in pcts]
stats_monto=pd.DataFrame({"Metrica":["Media","Mediana","Desv.Std","Varianza","Min","Max"]+[f"P{p}" for p in pcts],
    "Total ap":ss(df_ap["MONTO"]),"Fraude Total":ss(df_f_tot["MONTO"]),
    "Fraude Ap":ss(df_f_ap["MONTO"]),"No Fraude":ss(df_nofraude["MONTO"])})
for c in ["Total ap","Fraude Total","Fraude Ap","No Fraude"]:
    stats_monto[c]=stats_monto[c].apply(lambda v:f"S/ {v:,.2f}")
print(stats_monto.to_string(index=False))


# SECCIÓN 8 — TABLAS CRUZADAS (base + velocidad + features)
print("\n" + "="*65); print("SECCION 8 — TABLAS CRUZADAS")
def tc(df_b,d_f,d_c,tit):
    try:
        piv=df_b.pivot_table(values="ES_FRAUDE_APROBADO",index=d_f,columns=d_c,aggfunc="sum",fill_value=0)
        piv_fr=df_b.pivot_table(values="ES_FRAUDE_APROBADO",index=d_f,columns=d_c,aggfunc="mean",fill_value=0).map(lambda v:f"{v*100:.4f}%")
        piv_m=df_b[df_b["ES_FRAUDE_APROBADO"]==1].pivot_table(values="MONTO",index=d_f,columns=d_c,aggfunc="sum",fill_value=0).map(lambda v:f"S/ {v:,.2f}")
        print(f"\n--- {tit} ---"); print("Fraudes:"); print(piv.to_string())
        chi2,p,dof,_=chi2_contingency(piv.values)
        print(f"Chi2={chi2:.4f} p={p:.6f} | {'SIGNIFICATIVA' if p<0.05 else 'No significativa'}")
        return piv,piv_fr,piv_m
    except Exception as e:
        print(f"  Error {tit}: {e}"); return None,None,None

cruces_base=[("TIPO_PRODUCTO","SEGURO","TipoProducto x Seguro"),("SEG_NOMBRE","SEGURO","Segmento x Seguro"),
    ("SEG_GRUPO","SEGURO","SegGrupo x Seguro"),("COD_RED_LABEL","SEGURO","CVV x Seguro"),
    ("CANAL","SEGURO","Canal x Seguro"),("TIPO_PRODUCTO","COD_RED_LABEL","TipoProducto x CVV")]
r_base={tit:tc(df_ap,df_,dc_,tit) for df_,dc_,tit in cruces_base}

# Velocidad × dimensiones
df_ve=df_vel2.copy()
sm=df.groupby("ID_CLIENTE")[["SEG_NOMBRE","SEG_GRUPO","TIPO_PRODUCTO"]].first()
df_ve=df_ve.join(sm,on="ID_CLIENTE",rsuffix="_v")
cruces_vel=[("BUCKET","SEG_NOMBRE","Velocidad x Segmento"),("BUCKET","TIPO_PRODUCTO","Velocidad x TipoProducto")]
r_vel={tit:tc(df_ve,df_,dc_,tit) for df_,dc_,tit in cruces_vel if dc_ in df_ve.columns}

# Features × dimensiones
cruces_feat=[("BUCKET_RATIO_SALDO","TIPO_PRODUCTO","RatioSaldo x TipoProducto"),
    ("BUCKET_ZSCORE","SEGURO","ZScore x Seguro"),("ES_COMERCIO_NUEVO","TIPO_PRODUCTO","ComercioNuevo x TipoProducto"),
    ("FRANJA_HORARIA","TIPO_PRODUCTO","FranjaHoraria x TipoProducto"),
    ("BUCKET_RECHAZOS","TIPO_PRODUCTO","Rechazos24h x TipoProducto"),
    ("BUCKET_FR_COMERCIO","SEGURO","FRHistorico x Seguro")]
r_feat={tit:tc(df_feat,df_,dc_,tit) for df_,dc_,tit in cruces_feat if df_ in df_feat.columns}

# SECCIÓN 9 — CLIENTES RECURRENTES
print("\n" + "="*65); print("SECCION 9 — CLIENTES RECURRENTES")
clf_mes={m:set(df_ap[(df_ap["MES"]==m)&(df_ap["ES_FRAUDE_APROBADO"]==1)]["ID_CLIENTE"]) for m in meses_disponibles}
tcg=df_ap["ID_CLIENTE"].nunique(); cpm=df_ap.groupby("MES")["ID_CLIENTE"].nunique()
print(f"Clientes globales: {tcg:,} | Suma por mes: {cpm.sum():,}")
dr_rows=[]
if len(meses_disponibles)>=2:
    for i in range(len(meses_disponibles)-1):
        ma=meses_disponibles[i]; mb=meses_disponibles[i+1]
        ca=clf_mes.get(ma,set()); cb=clf_mes.get(mb,set())
        rec=ca&cb; nue=cb-ca; pr=len(rec)/len(cb)*100 if len(cb)>0 else 0
        print(f"  {ma}->{mb}: fraude_A={len(ca):,} fraude_B={len(cb):,} recurrentes={len(rec):,} ({pr:.4f}%)")
        dr_rows.append({"Periodo":f"{ma}->{mb}","Fraude MesA":len(ca),"Fraude MesB":len(cb),"Recurrentes":len(rec),"% Recurrentes":f"{pr:.4f}%","Nuevos":len(nue)})
df_rec=pd.DataFrame(dr_rows)
df_cli_mes=pd.DataFrame({"Mes":meses_disponibles,"Clientes unicos":[df_ap[df_ap["MES"]==m]["ID_CLIENTE"].nunique() for m in meses_disponibles],
    "Con fraude":[len(clf_mes.get(m,set())) for m in meses_disponibles]})
df_cli_mes["% fraude"]=(df_cli_mes["Con fraude"]/df_cli_mes["Clientes unicos"]*100).map("{:.4f}%".format)

# =============================================================================
# EXPORTACIÓN A EXCEL
# =============================================================================
print("\nExportando a Excel...")
FH=PatternFill("solid",fgColor="1F3864"); FS=PatternFill("solid",fgColor="2E75B6")
FA=PatternFill("solid",fgColor="DEEAF1"); FY=PatternFill("solid",fgColor="FFF2CC")
FF=PatternFill("solid",fgColor="FCE4D6")
fH=Font(color="FFFFFF",bold=True,size=10); fN=Font(size=10); fI=Font(italic=True,size=9,color="1F3864")
BT=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
AC=Alignment(horizontal="center",vertical="center",wrap_text=True)
AL=Alignment(horizontal="left",vertical="center",wrap_text=True)

def h(ws,f,n,t,fl=None):
    fl=fl or FH; ws.merge_cells(start_row=f,start_column=1,end_row=f,end_column=n)
    c=ws.cell(row=f,column=1,value=t); c.fill=fl; c.font=fH; c.alignment=AC; c.border=BT
def hd(ws,f):
    for r in ws.iter_rows(min_row=f,max_row=f):
        for c in r: c.fill=FS; c.font=fH; c.alignment=AC; c.border=BT
def sd(ws,fi,ff,kw=None):
    for i,r in enumerate(ws.iter_rows(min_row=fi,max_row=ff),start=1):
        fl=FA if i%2==0 else PatternFill()
        if kw and r[0].value and any(k in str(r[0].value).lower() for k in kw): fl=FF
        for c in r: c.fill=fl; c.font=fN; c.alignment=AC; c.border=BT
def interp(ws,f,n,t):
    ws.merge_cells(start_row=f,start_column=1,end_row=f,end_column=n)
    c=ws.cell(row=f,column=1,value=f"-> {t}"); c.fill=FY; c.font=fI; c.alignment=AL; c.border=BT
    ws.row_dimensions[f].height=28
def ajc(ws):
    for col in ws.columns:
        ml=max((len(str(c.value)) for c in col if c.value),default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(ml+4,35)
def escribir_pivot(ws,piv,fi,tit,nc):
    h(ws,fi,nc,tit,fl=FS); fi+=1
    if piv is None: return fi+2
    do=piv.reset_index()
    for j,col in enumerate(do.columns,start=1):
        c=ws.cell(row=fi,column=j,value=str(col)); c.fill=FS; c.font=fH; c.alignment=AC; c.border=BT
    fi+=1
    for i,row in do.iterrows():
        fl=FA if i%2==0 else PatternFill()
        for j,val in enumerate(row,start=1):
            c=ws.cell(row=fi,column=j,value=round(val,4) if isinstance(val,float) else val)
            c.fill=fl; c.font=fN; c.alignment=AC; c.border=BT
        fi+=1
    return fi+2

with pd.ExcelWriter(RUTA_EXCEL_OUT,engine="openpyxl") as writer:
    # H1: Resumen
    s="1_Resumen"; nc=len(df_resumen_ejecutivo.columns)+1
    df_resumen_ejecutivo.to_excel(writer,sheet_name=s,startrow=3)
    ws=writer.sheets[s]; h(ws,1,nc,f"MCC 7995 | {datetime.today().strftime('%d/%m/%Y')}"); h(ws,2,nc,"RESUMEN INDICADORES POR MES",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row,kw=["fraude","ratio"])
    fu=ws.max_row+2; h(ws,fu,nc,"UMBRAL DE CONTROL",fl=FS); df_umbral.to_excel(writer,sheet_name=s,index=False,startrow=fu)
    fi=ws.max_row+2; interp(ws,fi,nc,f"Nivel: {nivel_riesgo}. {riesgo_texto[nivel_riesgo]}")
    if n_f_tot_g>0: interp(ws,fi+1,nc,f"Fraude total: {n_f_tot_g:,} | Aprobado: {n_f_ap_g:,} ({n_f_ap_g/n_f_tot_g*100:.2f}%)")
    ajc(ws)
    # H2: Deciles
    s="2_Deciles"; nc2=len(tabla_deciles.columns)+2; tabla_deciles.to_excel(writer,sheet_name=s,startrow=3)
    ws=writer.sheets[s]; h(ws,1,nc2,"TABLA DE DECILES POR MONTO"); h(ws,2,nc2,"D01=menor D10=mayor | 4 decimales en FR%",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row)
    dmf=tabla_deciles["Fraude Trx Aprobado"].idxmax(); dmm=tabla_deciles["Fraude Monto Aprobado"].idxmax()
    fi=ws.max_row+2; interp(ws,fi,nc2,f"Mayor cant fraude ap: {dmf} ({tabla_deciles.loc[dmf,'Fraude Trx Aprobado']:,}) Rango: {tabla_deciles.loc[dmf,'Rango S/']}")
    interp(ws,fi+1,nc2,f"Mayor monto fraude ap: {dmm} (S/ {tabla_deciles.loc[dmm,'Fraude Monto Aprobado']:,.2f})"); ajc(ws)
    # H2B: Apertura D10
    s="2B_Apertura_D10"; nc2b=len(tabla_d10.columns)+2; tabla_d10.to_excel(writer,sheet_name=s,startrow=3)
    ws=writer.sheets[s]; h(ws,1,nc2b,f"APERTURA D10 (montos >= S/ {m_d10_min:,.2f})"); h(ws,2,nc2b,"Mismas columnas que tabla principal",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row)
    if "Fraude Trx Aprobado" in tabla_d10.columns and tabla_d10["Fraude Trx Aprobado"].sum()>0:
        sm=tabla_d10["Fraude Trx Aprobado"].idxmax(); fi=ws.max_row+2
        interp(ws,fi,nc2b,f"Sub-decil top fraude ap: {sm} ({tabla_d10.loc[sm,'Fraude Trx Aprobado']:,}) Rango: {tabla_d10.loc[sm,'Rango S/']}")
    ajc(ws)
    # H3: Top Comercios
    s="3_Top_Comercios"; dtc=df_top_com.copy()
    for c in ["Monto F Total S/","Monto F Ap S/","Monto Total S/"]:
        if c in dtc.columns: dtc[c]=dtc[c].map("S/ {:,.2f}".format)
    dtc.to_excel(writer,sheet_name=s,index=False,startrow=3); ws=writer.sheets[s]; nc3=len(dtc.columns)
    h(ws,1,nc3,"TOP 10 COMERCIOS (Fraude Total=ap+den)"); h(ws,2,nc3,"Ordenado por monto fraude aprobado",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row); ajc(ws)
    # H4: Estadisticas
    s="4_Estadisticas"; stats_monto.to_excel(writer,sheet_name=s,index=False,startrow=3)
    ws=writer.sheets[s]; nc4=len(stats_monto.columns)
    h(ws,1,nc4,"ESTADISTICAS DE MONTO"); h(ws,2,nc4,"Total | Fraude Total | Fraude Aprobado | No Fraude",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row)
    if len(df_f_ap)>0:
        mfa=df_f_ap["MONTO"].mean(); mdfa=df_f_ap["MONTO"].median(); sfa=df_f_ap["MONTO"].std()
        mdn=df_nofraude["MONTO"].median(); p90f=df_f_ap["MONTO"].quantile(0.90); p90n=df_nofraude["MONTO"].quantile(0.90)
        rmm=mfa/mdfa if mdfa>0 else 0; fi=ws.max_row+2
        interp(ws,fi,nc4,f"SESGO: Media S/ {mfa:,.2f} vs mediana S/ {mdfa:,.2f} (ratio {rmm:.2f}x). "+("Montos extremos presentes." if rmm>2 else "Sesgo moderado." if rmm>1.5 else "Simetrico."))
        interp(ws,fi+1,nc4,f"P90: Fraude S/ {p90f:,.2f} vs Legitimo S/ {p90n:,.2f}. "+("Umbral efectivo aqui." if p90f>p90n else "P90 fraude no supera al legitimo."))
        interp(ws,fi+2,nc4,f"MEDIANAS: Fraude S/ {mdfa:,.2f} vs Legitimo S/ {mdn:,.2f}. "+("Fraude con montos MAYORES." if mdfa>mdn else "Fraude con montos MENORES — testeo de tarjetas."))
    ajc(ws)
    # H5: Trx Diaria
    s="5_Trx_Diaria"; tabla_diaria_display.to_excel(writer,sheet_name=s,index=False,startrow=3)
    ws=writer.sheets[s]; nc5=len(tabla_diaria_display.columns)
    h(ws,1,nc5,"TRANSACCIONALIDAD DIARIA POR CLIENTE"); h(ws,2,nc5,f"Fraude ap en {pct_1:.2f}% pares con 1 trx/dia",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row); fi=ws.max_row+2
    interp(ws,fi,nc5,f"{pct_1:.2f}% con 1 trx/dia -> ataque puntual. Regla velocidad diaria no efectiva — priorizar monto."); ajc(ws)
    # H6: Velocidad
    s="6_Velocidad"; tabla_vel.reset_index().to_excel(writer,sheet_name=s,index=False,startrow=3)
    ws=writer.sheets[s]; nc6=len(tabla_vel.columns)+1
    h(ws,1,nc6,"VELOCIDAD: INTERVALO ENTRE TRANSACCIONES"); h(ws,2,nc6,"Fraude Total | Fraude Aprobado | FR% con 4 decimales",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row)
    bm=tabla_vel["FR_Aprobado_pct"].idxmax() if len(tabla_vel)>0 else "N/A"
    pv=tabla_vel.loc[bm,"Pct_fraudes_total"] if len(tabla_vel)>0 else 0
    fi=ws.max_row+2; interp(ws,fi,nc6,f"Mayor FR aprobado: '{bm}' ({pv:.2f}% del total). <=5min -> ataques automatizados."); ajc(ws)
    # H7: Dimensiones
    s="7_Dimensiones"; w7=writer.book.create_sheet(s); writer.sheets[s]=w7; fa=1
    for dfm,tit,itp in [(df_por_canal,"Canal","Vector de ataque principal."),(df_por_seg_nombre,"Segmento Individual","Nombre de cada codigo."),
        (df_por_seg_grupo,"Segmento Grupo","Agrupacion: Affluent, Mass, etc."),(df_por_seguro,"Seguridad Comercio","No Seguro = sin 3DSecure."),
        (df_por_tipo,"Tipo Producto","TD vs TC."),( df_por_cvv,"CVV/Red","No Match = alerta."),(df_por_pais,"Pais Top10","Otros paises = riesgo.")]:
        do=dfm.reset_index(); nd=len(do.columns)
        h(w7,fa,nd,f"MCC 7995 — {tit.upper()}"); fa+=1
        for j,col in enumerate(do.columns,start=1):
            c=w7.cell(row=fa,column=j,value=col); c.fill=FS; c.font=fH; c.alignment=AC; c.border=BT
        fa+=1
        for i,row in do.iterrows():
            fl=FA if i%2==0 else PatternFill()
            for j,val in enumerate(row,start=1):
                c=w7.cell(row=fa,column=j,value=f"{val:,.4f}" if isinstance(val,float) else val)
                c.fill=fl; c.font=fN; c.alignment=AC; c.border=BT
            fa+=1
        w7.merge_cells(start_row=fa,start_column=1,end_row=fa,end_column=nd)
        c=w7.cell(row=fa,column=1,value=f"-> {itp}"); c.fill=FY; c.font=fI; c.alignment=AL; c.border=BT
        w7.row_dimensions[fa].height=25; fa+=3
    ajc(w7)
    # H8A: Cruces Base
    s="8A_Cruces_Base"; w8=writer.book.create_sheet(s); writer.sheets[s]=w8; fa=1
    for tit,(piv,_,_) in r_base.items():
        fa=escribir_pivot(w8,piv,fa,tit,12)
        try:
            chi2,p,_,_=chi2_contingency(piv.values)
            w8.merge_cells(start_row=fa,start_column=1,end_row=fa,end_column=12)
            c=w8.cell(row=fa,column=1,value=f"Chi2={chi2:.4f} p={p:.6f} {'SIGNIFICATIVA' if p<0.05 else 'No sig'}")
            c.fill=FY; c.font=fI; c.alignment=AL; c.border=BT
        except: pass
        fa+=3
    ajc(w8)
    # H8B: Cruces Velocidad
    s="8B_Cruces_Vel"; w8b=writer.book.create_sheet(s); writer.sheets[s]=w8b; fa=1
    for tit,(piv,_,_) in r_vel.items():
        fa=escribir_pivot(w8b,piv,fa,tit,15)
    ajc(w8b)
    # H8C: Cruces Features
    s="8C_Cruces_Feat"; w8c=writer.book.create_sheet(s); writer.sheets[s]=w8c; fa=1
    for tit,(piv,_,_) in r_feat.items():
        fa=escribir_pivot(w8c,piv,fa,tit,12)
    ajc(w8c)
    # H9: Features Analisis
    s="9_Features"; w9=writer.book.create_sheet(s); writer.sheets[s]=w9; fa=1
    for tit,tf in tablas_features.items():
        if tf is None: continue
        do=tf.reset_index(); nd=len(do.columns)
        h(w9,fa,nd,tit.upper()); fa+=1
        for j,col in enumerate(do.columns,start=1):
            c=w9.cell(row=fa,column=j,value=col); c.fill=FS; c.font=fH; c.alignment=AC; c.border=BT
        fa+=1
        for i,row in do.iterrows():
            fl=FA if i%2==0 else PatternFill()
            for j,val in enumerate(row,start=1):
                c=w9.cell(row=fa,column=j,value=f"{val:.4f}" if isinstance(val,float) else val)
                c.fill=fl; c.font=fN; c.alignment=AC; c.border=BT
            fa+=1
        fa+=2
    ajc(w9)
    # H10: Clientes Recurrentes
    s="10_Recurrentes"; df_cli_mes.to_excel(writer,sheet_name=s,index=False,startrow=3)
    ws=writer.sheets[s]; nc10=len(df_cli_mes.columns)
    h(ws,1,nc10,"CLIENTES UNICOS Y RECURRENCIA"); h(ws,2,nc10,"Un cliente puede aparecer en varios meses",fl=FS)
    hd(ws,4); sd(ws,5,ws.max_row)
    fi=ws.max_row+2; interp(ws,fi,nc10,f"Clientes globales: {tcg:,} | Suma por mes: {cpm.sum():,}")
    if len(df_rec)>0:
        fi2=ws.max_row+3; h(ws,fi2,nc10,"RECURRENCIA ENTRE MESES",fl=FS)
        df_rec.to_excel(writer,sheet_name=s,index=False,startrow=fi2)
        hd(ws,fi2+1); sd(ws,fi2+2,ws.max_row)
    ajc(ws)

print(f"\nExcel exportado: {RUTA_EXCEL_OUT}")
print("Hojas: 1_Resumen | 2_Deciles | 2B_Apertura_D10 | 3_Top_Comercios")
print("       4_Estadisticas | 5_Trx_Diaria | 6_Velocidad | 7_Dimensiones")
print("       8A_Cruces_Base | 8B_Cruces_Vel | 8C_Cruces_Feat")
print("       9_Features | 10_Recurrentes")
