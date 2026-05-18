# =============================================================================
# ANÁLISIS INTEGRAL MCC 7995 — Juegos de Azar
# Versión final completa con todas las correcciones
# =============================================================================

import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.stats import chi2_contingency
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# =============================================================================
# AJUSTA SOLO ESTE BLOQUE
# =============================================================================
COLS = {
    "FECHA"           : "ACF-FECHA TRX",
    "HORA"            : "ACF-HORA TRX",
    "DATETIME"        : "FECHA_HORA",
    "COMERCIO"        : "ACF-NOMBRE/LOCALIZACION COMERCIO",
    "INDICADOR"       : "ACF-INDICADOR DE FRAUDE",
    "COD_RESPUESTA"   : "ACF-COD RPTA",
    "CONDICION_RT"    : "CONDICION RT",
    "CANAL"           : "ACF-CANAL",
    "ENTRY_MODE"      : "ACF-ENTRY MODE",
    "SALDO"           : "ACF-SALDO DISPONIBLE EN MONEDA TRX",
    "ID_CLIENTE"      : "ACF-ID CLIENTE",
    "ESI_UCAP"        : "ACF-ECI/UCAF",
    "PAIS"            : "ACF-PAIS ORIGEN 87519",
    "COD_RED_COMERCIO": "ACF-COD RED COMERCIO",
    "MONTO"           : "ACF-MONTO EN MONEDA LOCAL",
    "SEGMENTO"        : "VAA-EVENTO DE COMPROMISO OTRA FUENTE",
    "TIPO_PRODUCTO"   : "ACF-TIPO PROD TC",
}
RUTA_PARQUET   = r"C:\ruta\al\consolidado_7995.parquet"
RUTA_EXCEL_OUT = r"C:\ruta\salida\analisis_mcc7995_resultado.xlsx"
# =============================================================================

SEG_NOMBRE = {
    "30":"Polo Dirección","99":"Polo Dirección",
    "31":"Premium","32":"Preferente","33":"Personal","34":"Estándar",
    "5":"Inst. Financieras","21":"Corporativo","2":"Mediano Empresas",
    "15":"Sector Gobierno","16":"Otras Instituciones",
    "3":"Pequeñas Empresas","4":"Negocios 2","7":"Negocios 3",
    "8":"Negocios 1","13":"Microempresas",
}
SEG_GRUPO = {
    "30":"Affluent","99":"Affluent",
    "31":"Emerging Affluent","32":"Emerging Affluent",
    "33":"Top of Mass","34":"Mass",
    "5":"Corporate","21":"Corporate","2":"Commercial",
    "15":"Commercial","16":"Commercial",
    "3":"Small Business","4":"Small Business",
    "7":"Small Business","8":"Small Business","13":"Small Business",
}
COD_RED_LABEL = {
    "S":"Estático (TD)","D":"Dinámico (TD/TC)",
    "E":"Estático (TC)","N":"No Match / Sin CVV",
}

C_FRAUDE="#D9534F"; C_OK="#5B9BD5"; C_088="#F0AD4E"
C_NEU="#8EA9C1";    C_ROJO2="#8B0000"

# =============================================================================
# CARGA Y PREPARACIÓN
# =============================================================================
print("Cargando parquet...")
df_raw = pd.read_parquet(RUTA_PARQUET)
col_map = {v:k for k,v in COLS.items()}
df = df_raw.rename(columns=col_map).copy()

faltantes = [a for a in COLS if a not in df.columns]
if faltantes:
    print(f"  Columnas no encontradas: {faltantes}")

df["MONTO"]    = pd.to_numeric(df["MONTO"],    errors="coerce")
df["SALDO"]    = pd.to_numeric(df["SALDO"],    errors="coerce")
df["DATETIME"] = pd.to_datetime(df["DATETIME"], errors="coerce")
df["FECHA"]    = df["DATETIME"].dt.normalize()
df["MES"]      = df["DATETIME"].dt.to_period("M").astype(str)

for c in ["INDICADOR","COD_RESPUESTA","CONDICION_RT","CANAL",
          "ESI_UCAP","COD_RED_COMERCIO","SEGMENTO","TIPO_PRODUCTO","ENTRY_MODE"]:
    if c in df.columns:
        df[c] = df[c].astype(str).str.strip().str.upper()

df["ESTADO"] = df["COD_RESPUESTA"].apply(
    lambda x: "APROBADA" if str(x).strip() in ["00","0000","000","0"] else "DENEGADA"
)
df["ES_FRAUDE"]          = (df["INDICADOR"] == "F").astype(int)
df["ES_FRAUDE_APROBADO"] = (
    (df["INDICADOR"] == "F") & (df["ESTADO"] == "APROBADA")
).astype(int)
df["COND_088"]      = df["CONDICION_RT"].str.contains("088", na=False)
df["SEGURO"]        = df["ESI_UCAP"].apply(
    lambda x: "Seguro" if str(x).strip() in ["2","5"] else "No Seguro"
)
df["SEG_NOMBRE"]    = df["SEGMENTO"].map(SEG_NOMBRE).fillna("Otro/Sin seg")
df["SEG_GRUPO"]     = df["SEGMENTO"].map(SEG_GRUPO).fillna("Otro/Sin seg")
df["COD_RED_LABEL"] = df["COD_RED_COMERCIO"].map(COD_RED_LABEL).fillna("Otro")

df_ap    = df[df["ESTADO"] == "APROBADA"].copy()
df_den   = df[df["ESTADO"] == "DENEGADA"].copy()
df_f_ap  = df[df["ES_FRAUDE_APROBADO"] == 1].copy()
df_f_tot = df[df["ES_FRAUDE"] == 1].copy()

print(f"  Total registros:    {len(df):,}")
print(f"  Aprobadas:          {len(df_ap):,}")
print(f"  Denegadas:          {len(df_den):,}")
print(f"  Fraude total:       {df['ES_FRAUDE'].sum():,}")
print(f"  Fraude aprobado:    {df['ES_FRAUDE_APROBADO'].sum():,}")
print(f"  Con condición 088:  {df['COND_088'].sum():,}")


# =============================================================================
# SECCIÓN 1 — RESUMEN EJECUTIVO POR MES
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 1 — RESUMEN EJECUTIVO POR MES")
print("="*65)

meses_disponibles = sorted(df_ap["MES"].unique())
filas_resumen = []

for mes in meses_disponibles:
    sub_mes  = df[df["MES"] == mes]
    sub_ap   = df_ap[df_ap["MES"] == mes]
    sub_f_ap = sub_ap[sub_ap["ES_FRAUDE_APROBADO"] == 1]
    sub_f_tot= sub_mes[sub_mes["ES_FRAUDE"] == 1]

    n_ap_m       = len(sub_ap)
    n_den_m      = (sub_mes["ESTADO"]=="DENEGADA").sum()
    monto_ap_m   = sub_ap["MONTO"].sum()
    ticket_ap_m  = sub_ap["MONTO"].mean()
    n_cli_m      = sub_ap["ID_CLIENTE"].nunique()
    n_f_tot_m    = len(sub_f_tot)
    monto_f_tot_m= sub_f_tot["MONTO"].sum()
    n_f_ap_m     = len(sub_f_ap)
    monto_f_ap_m = sub_f_ap["MONTO"].sum()
    ticket_f_ap_m= sub_f_ap["MONTO"].mean() if n_f_ap_m > 0 else 0
    ratio_trx_m  = n_f_ap_m / n_ap_m       if n_ap_m    > 0 else 0
    ratio_mon_m  = monto_f_ap_m / monto_ap_m if monto_ap_m > 0 else 0

    filas_resumen.append({
        "Total trx aprobadas"              : f"{n_ap_m:,}",
        "Total trx denegadas"              : f"{n_den_m:,}",
        "Total monto aprobado (S/)"        : f"S/ {monto_ap_m:,.2f}",
        "Ticket promedio por trx (S/)"     : f"S/ {ticket_ap_m:,.2f}",
        "N clientes unicos"                : f"{n_cli_m:,}",
        "Trxs fraudulentas total (F)"      : f"{n_f_tot_m:,}",
        "Trxs fraudulentas aprobadas"      : f"{n_f_ap_m:,}",
        "Monto fraude total (S/)"          : f"S/ {monto_f_tot_m:,.2f}",
        "Monto fraude aprobado (S/)"       : f"S/ {monto_f_ap_m:,.2f}",
        "Ticket promedio fraude aprobado"  : f"S/ {ticket_f_ap_m:,.2f}",
        "Ratio trx fraudulentas aprobadas" : f"{ratio_trx_m*100:.4f}%",
        "Ratio fraude en soles (aprobado)" : f"{ratio_mon_m*100:.4f}%",
    })

df_resumen_ejecutivo = pd.DataFrame(filas_resumen, index=meses_disponibles).T
df_resumen_ejecutivo.index.name = "Indicador"
print(df_resumen_ejecutivo.to_string())

# Métricas globales
n_ap          = len(df_ap)
n_den         = len(df_den)
monto_ap      = df_ap["MONTO"].sum()
n_f_tot_g     = df["ES_FRAUDE"].sum()
n_f_ap_g      = df["ES_FRAUDE_APROBADO"].sum()
monto_f_ap_g  = df_f_ap["MONTO"].sum() if len(df_f_ap) > 0 else 0
monto_f_tot_g = df_f_tot["MONTO"].sum() if len(df_f_tot) > 0 else 0
ticket_ap_g   = df_ap["MONTO"].mean()
ticket_f_ap_g = df_f_ap["MONTO"].mean() if len(df_f_ap) > 0 else 0
n_clientes    = df_ap["ID_CLIENTE"].nunique()
ratio_trx_f   = n_f_ap_g / n_ap     if n_ap     > 0 else 0
ratio_monto_f = monto_f_ap_g / monto_ap if monto_ap > 0 else 0

nivel_riesgo = (
    "ALTO >=5%"      if ratio_trx_f >= 0.05 else
    "MEDIO [1%-5%]"  if ratio_trx_f >= 0.01 else
    "REGULAR <1%"
)
riesgo_texto = {
    "REGULAR <1%"   : "Fraud rate en nivel REGULAR (<1%). Bajo riesgo relativo.",
    "MEDIO [1%-5%]" : "Fraud rate en nivel MEDIO (1%-5%). Monitoreo activo recomendado.",
    "ALTO >=5%"     : "ALERTA: Fraud rate supera el 5%. Intervencion inmediata requerida.",
}
df_umbral = pd.DataFrame({
    "Umbral":["Regular","Medio","Alto"],
    "Rango":["< 1%","[1%-5%]",">= 5%"],
    "Nivel actual":[
        "AQUI" if ratio_trx_f < 0.01 else "",
        "AQUI" if 0.01 <= ratio_trx_f < 0.05 else "",
        "AQUI" if ratio_trx_f >= 0.05 else "",
    ]
})


# =============================================================================
# SECCIÓN 2 — TABLA DE DECILES
# =============================================================================
print("\n" + "="*65)
print("SECCION 2 — TABLA DE DECILES POR MONTO")
print("="*65)

def construir_tabla_deciles(df_base, n_deciles=10, label="D"):
    df_base = df_base.copy().dropna(subset=["MONTO"])
    df_base["DECIL"] = pd.qcut(
        df_base["MONTO"], q=n_deciles,
        labels=[f"{label}{str(i+1).zfill(2)}" for i in range(n_deciles)],
        duplicates="drop"
    )
    rows = []
    fraude_ap_acum = 0
    trx_ap_acum    = 0
    total_ap       = len(df_base[df_base["ESTADO"]=="APROBADA"])
    total_f_ap     = df_base["ES_FRAUDE_APROBADO"].sum()

    for decil in df_base["DECIL"].cat.categories:
        sub      = df_base[df_base["DECIL"]==decil]
        sub_ap   = sub[sub["ESTADO"]=="APROBADA"]
        sub_den  = sub[sub["ESTADO"]=="DENEGADA"]
        sub_f_ap = sub[sub["ES_FRAUDE_APROBADO"]==1]
        sub_f_tot= sub[sub["ES_FRAUDE"]==1]

        n_ap_d   = len(sub_ap);   m_ap_d  = sub_ap["MONTO"].sum()
        n_den_d  = len(sub_den);  m_den_d = sub_den["MONTO"].sum()
        n_f_ap_d = len(sub_f_ap); m_f_ap_d= sub_f_ap["MONTO"].sum()
        n_f_tot_d= len(sub_f_tot);m_f_tot_d=sub_f_tot["MONTO"].sum()

        trx_ap_acum    += n_ap_d
        fraude_ap_acum += n_f_ap_d
        freq_ap      = n_ap_d / total_ap      if total_ap   > 0 else 0
        freq_ap_acum = trx_ap_acum / total_ap if total_ap   > 0 else 0
        freq_f_ap    = n_f_ap_d / total_f_ap  if total_f_ap > 0 else 0
        freq_f_acum  = fraude_ap_acum / total_f_ap if total_f_ap > 0 else 0

        rows.append({
            "Decil"                   : decil,
            "Rango S/"                : f"{sub['MONTO'].min():.0f}-{sub['MONTO'].max():.0f}",
            "Trx Aprobadas"           : n_ap_d,
            "Monto Aprobadas"         : round(m_ap_d,1),
            "Freq Aprobadas"          : f"{freq_ap*100:.2f}%",
            "Freq Acum Aprobadas"     : f"{freq_ap_acum*100:.2f}%",
            "Trx Denegadas"           : n_den_d,
            "Monto Denegadas"         : round(m_den_d,1),
            "Total Trx"               : len(sub),
            "Ticket Promedio"         : round(sub["MONTO"].mean(),2),
            "Monto Total S/"          : round(sub["MONTO"].sum(),1),
            "Fraude Trx Total"        : n_f_tot_d,
            "Fraude Monto Total"      : round(m_f_tot_d,1),
            "Fraude Trx Aprobado"     : n_f_ap_d,
            "Fraude Monto Aprobado"   : round(m_f_ap_d,1),
            "Fraude Freq Aprobado"    : f"{freq_f_ap*100:.4f}%",
            "Fraude Freq Acumulada"   : f"{freq_f_acum*100:.4f}%",
            "Fraude/Trx Aprobadas"    : f"{n_f_ap_d/n_ap_d*100:.4f}%" if n_ap_d>0 else "0.0000%",
            "No Fraude/Trx Aprobadas" : f"{(1-n_f_ap_d/n_ap_d)*100:.4f}%" if n_ap_d>0 else "100.0000%",
        })
    return pd.DataFrame(rows).set_index("Decil")

tabla_deciles = construir_tabla_deciles(df, n_deciles=10)
print(tabla_deciles.to_string())

# Apertura D10 con cortes fijos
print("\nApertura Decil 10 (top monto) — cortes fijos:")
monto_d10_min = df["MONTO"].quantile(0.90)
df_d10 = df[df["MONTO"] >= monto_d10_min].copy()

bins_d10 = np.unique([df_d10["MONTO"].quantile(p/100) for p in np.linspace(0,100,11)])
if len(bins_d10) < 3:
    bins_d10 = np.unique([df_d10["MONTO"].quantile(p/100) for p in np.linspace(0,100,6)])

labels_d10 = [f"SD{str(i+1).zfill(2)}" for i in range(len(bins_d10)-1)]
df_d10["DECIL_D10"] = pd.cut(df_d10["MONTO"], bins=bins_d10,
                              labels=labels_d10, include_lowest=True)

tabla_d10 = (
    df_d10.groupby("DECIL_D10", observed=True)
    .agg(
        Trx_Aprobadas    =("ESTADO",              lambda x: (x=="APROBADA").sum()),
        Monto_Aprobado   =("MONTO",               lambda x: x[df_d10.loc[x.index,"ESTADO"]=="APROBADA"].sum()),
        Trx_Denegadas    =("ESTADO",              lambda x: (x=="DENEGADA").sum()),
        Total_Trx        =("MONTO",               "count"),
        Ticket_Promedio  =("MONTO",               "mean"),
        Monto_Total      =("MONTO",               "sum"),
        Fraude_Trx_Total =("ES_FRAUDE",            "sum"),
        Fraude_Monto_Tot =("MONTO",               lambda x: x[df_d10.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Fraude_Trx_Ap    =("ES_FRAUDE_APROBADO",   "sum"),
        Fraude_Monto_Ap  =("MONTO",               lambda x: x[df_d10.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
    )
    .assign(
        FR_Ap_pct=lambda x: (x["Fraude_Trx_Ap"]/x["Trx_Aprobadas"]*100).map("{:.4f}%".format),
    )
)
print(tabla_d10.to_string())

# Grafico deciles
deciles_idx    = tabla_deciles.index.tolist()
n_f_ap_d_list  = tabla_deciles["Fraude Trx Aprobado"].tolist()
n_f_tot_d_list = tabla_deciles["Fraude Trx Total"].tolist()
fr_ap_d = [
    tabla_deciles.loc[d,"Fraude Trx Aprobado"]/tabla_deciles.loc[d,"Trx Aprobadas"]*100
    if tabla_deciles.loc[d,"Trx Aprobadas"]>0 else 0
    for d in deciles_idx
]
fig = make_subplots(rows=1, cols=2,
    subplot_titles=["Fraudes por decil (total vs aprobado)","FR% aprobado por decil"])
fig.add_trace(go.Bar(x=deciles_idx, y=n_f_tot_d_list, name="Fraude Total",
    marker_color=C_ROJO2, opacity=0.6,
    text=[f"{v:,}" for v in n_f_tot_d_list], textposition="outside"), row=1, col=1)
fig.add_trace(go.Bar(x=deciles_idx, y=n_f_ap_d_list, name="Fraude Aprobado",
    marker_color=C_FRAUDE,
    text=[f"{v:,}" for v in n_f_ap_d_list], textposition="outside"), row=1, col=1)
fig.add_trace(go.Bar(x=deciles_idx, y=fr_ap_d, name="FR%",
    marker_color=C_088,
    text=[f"{v:.4f}%" for v in fr_ap_d], textposition="outside"), row=1, col=2)
fig.update_layout(title_text="MCC 7995 — Deciles por monto",
    barmode="group", height=450, plot_bgcolor="white")
fig.write_html("g1_deciles.html"); fig.show()


# =============================================================================
# SECCIÓN 3 — VELOCIDAD POR INTERVALO DE TIEMPO
# =============================================================================
print("\n" + "="*65)
print("SECCION 3 — VELOCIDAD: INTERVALO ENTRE TRANSACCIONES")
print("="*65)

df_vel = df.sort_values(["ID_CLIENTE","DATETIME"]).copy()
df_vel["PREV_DT"] = df_vel.groupby("ID_CLIENTE")["DATETIME"].shift(1)
df_vel["MIN_DESDE_PREV"] = (
    df_vel["DATETIME"] - df_vel["PREV_DT"]
).dt.total_seconds() / 60
df_vel2 = df_vel.dropna(subset=["MIN_DESDE_PREV"]).copy()
df_vel2 = df_vel2[df_vel2["MIN_DESDE_PREV"] >= 0]

def bucket_intervalo(m):
    if   m <=  2: return "<=2 min"
    elif m <=  5: return "<=5 min"
    elif m <= 10: return "<=10 min"
    elif m <= 15: return "<=15 min"
    elif m <= 20: return "<=20 min"
    elif m <= 60: return "<=1 hora"
    else:         return ">1 hora"

ORDEN_BUCKETS = ["<=2 min","<=5 min","<=10 min","<=15 min","<=20 min","<=1 hora",">1 hora"]
df_vel2["BUCKET"] = pd.Categorical(
    df_vel2["MIN_DESDE_PREV"].apply(bucket_intervalo),
    categories=ORDEN_BUCKETS, ordered=True
)

tabla_vel = (
    df_vel2.groupby("BUCKET", observed=True)
    .agg(
        Transacciones         =("MONTO",              "count"),
        Genuinas              =("ES_FRAUDE",           lambda x: (x==0).sum()),
        Fraude_Total          =("ES_FRAUDE",           "sum"),
        Fraude_Aprobado       =("ES_FRAUDE_APROBADO",  "sum"),
        FR_Total_pct          =("ES_FRAUDE",           lambda x: x.mean()*100),
        FR_Aprobado_pct       =("ES_FRAUDE_APROBADO",  lambda x: x.mean()*100),
        Monto_Genuino         =("MONTO",               lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].sum()),
        Monto_Fraude_Total    =("MONTO",               lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_Fraude_Aprobado =("MONTO",               lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        Ticket_Genuino_Prom   =("MONTO",               lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].mean()),
        Ticket_Fraude_Prom    =("MONTO",               lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].mean()),
    )
    .assign(Pct_fraudes_total=lambda x: x["Fraude_Total"]/x["Fraude_Total"].sum()*100)
)
print(tabla_vel.to_string())

bkts = tabla_vel.index.tolist()
fig2 = make_subplots(rows=1, cols=2,
    subplot_titles=["Genuinas vs Fraude por intervalo","FR% por intervalo"])
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["Genuinas"].tolist(), name="Genuinas",
    marker_color=C_OK,
    text=[f"{v:,}" for v in tabla_vel["Genuinas"]], textposition="outside"), row=1, col=1)
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["Fraude_Total"].tolist(), name="Fraude Total",
    marker_color=C_ROJO2, opacity=0.6,
    text=[f"{v:,}" for v in tabla_vel["Fraude_Total"]], textposition="outside"), row=1, col=1)
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["Fraude_Aprobado"].tolist(), name="Fraude Aprobado",
    marker_color=C_FRAUDE,
    text=[f"{v:,}" for v in tabla_vel["Fraude_Aprobado"]], textposition="outside"), row=1, col=1)
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["FR_Aprobado_pct"].round(4).tolist(), name="FR% Ap",
    marker_color=C_088,
    text=[f"{v:.4f}%" for v in tabla_vel["FR_Aprobado_pct"]], textposition="outside"), row=1, col=2)
fig2.update_layout(title_text="MCC 7995 — Fraude por intervalo",
    barmode="group", height=450, plot_bgcolor="white")
fig2.write_html("g2_velocidad_intervalo.html"); fig2.show()


# =============================================================================
# SECCIÓN 4 — TRANSACCIONALIDAD DIARIA POR CLIENTE
# =============================================================================
print("\n" + "="*65)
print("SECCION 4 — TRANSACCIONALIDAD DIARIA POR CLIENTE")
print("="*65)

dia_cliente = (
    df_ap
    .assign(FECHA_DIA=lambda x: x["DATETIME"].dt.date)
    .groupby(["ID_CLIENTE","FECHA_DIA"])
    .agg(
        Trx_dia              =("MONTO",              "count"),
        Monto_dia            =("MONTO",              "sum"),
        Fraudes_tot_dia      =("ES_FRAUDE",           "sum"),
        Fraudes_ap_dia       =("ES_FRAUDE_APROBADO",  "sum"),
        Monto_fraude_tot_dia =("MONTO",              lambda x: x[df_ap.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_fraude_ap_dia  =("MONTO",              lambda x: x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
    )
    .reset_index()
    .assign(
        tiene_fraude  =lambda x: (x["Fraudes_ap_dia"] > 0).astype(int),
        Trx_genuina   =lambda x: x["Trx_dia"] - x["Fraudes_tot_dia"],
        Monto_genuino =lambda x: x["Monto_dia"] - x["Monto_fraude_tot_dia"],
    )
)

def bucket_diario(n):
    if   n <= 6:  return str(n)
    elif n <= 9:  return "7-9"
    elif n <= 12: return "10-12"
    else:         return "13+"

ORDEN_DIA = ["1","2","3","4","5","6","7-9","10-12","13+"]
dia_cliente["BUCKET_DIA"] = pd.Categorical(
    dia_cliente["Trx_dia"].apply(bucket_diario),
    categories=ORDEN_DIA, ordered=True
)

tabla_diaria = (
    dia_cliente.groupby("BUCKET_DIA", observed=True)
    .agg(
        Genuina               =("tiene_fraude",        lambda x: (x==0).sum()),
        Fraude_Aprobado       =("tiene_fraude",        "sum"),
        Total_Clientes        =("tiene_fraude",        "count"),
        Trx_Genuina           =("Trx_genuina",         "sum"),
        Monto_Genuina         =("Monto_genuino",       "sum"),
        Trx_Fraude_Total      =("Fraudes_tot_dia",     "sum"),
        Trx_Fraude_Aprobado   =("Fraudes_ap_dia",      "sum"),
        Monto_Fraude_Total    =("Monto_fraude_tot_dia","sum"),
        Monto_Fraude_Aprobado =("Monto_fraude_ap_dia", "sum"),
    )
    .reset_index()
    .rename(columns={"BUCKET_DIA":"Intervalo"})
    .assign(
        Impacto_fraude_pct  =lambda x: x["Fraude_Aprobado"]/x["Total_Clientes"]*100,
        Ticket_Fraude_Prom  =lambda x: np.where(x["Trx_Fraude_Aprobado"]>0,
                                x["Monto_Fraude_Aprobado"]/x["Trx_Fraude_Aprobado"],0),
        Ticket_Genuina_Prom =lambda x: np.where(x["Trx_Genuina"]>0,
                                x["Monto_Genuina"]/x["Trx_Genuina"],0),
    )
)

total_cli_fraude = tabla_diaria["Fraude_Aprobado"].sum()
pct_1trx = (
    tabla_diaria.loc[tabla_diaria["Intervalo"]=="1","Fraude_Aprobado"].sum()
    / total_cli_fraude * 100 if total_cli_fraude > 0 else 0
)
print(f"  El fraude aprobado en un {pct_1trx:.2f}% corresponde a clientes con 1 trx/dia")

tabla_diaria_display = tabla_diaria.copy()
tabla_diaria_display["Impacto_fraude_pct"]    = tabla_diaria_display["Impacto_fraude_pct"].map("{:.4f}%".format)
tabla_diaria_display["Monto_Genuina"]         = tabla_diaria_display["Monto_Genuina"].map("S/ {:,.2f}".format)
tabla_diaria_display["Monto_Fraude_Total"]    = tabla_diaria_display["Monto_Fraude_Total"].map("S/ {:,.2f}".format)
tabla_diaria_display["Monto_Fraude_Aprobado"] = tabla_diaria_display["Monto_Fraude_Aprobado"].map("S/ {:,.2f}".format)
tabla_diaria_display["Ticket_Fraude_Prom"]    = tabla_diaria_display["Ticket_Fraude_Prom"].map("S/ {:,.2f}".format)
tabla_diaria_display["Ticket_Genuina_Prom"]   = tabla_diaria_display["Ticket_Genuina_Prom"].map("S/ {:,.2f}".format)
tabla_diaria_display.columns = [
    "Intervalo","Genuina","Fraude Aprobado","Total Clientes",
    "Trx Genuina","Monto Genuina",
    "Trx Fraude Total","Trx Fraude Aprobado",
    "Monto Fraude Total","Monto Fraude Aprobado",
    "Impacto Fraude %","Ticket Fraude Prom","Ticket Genuina Prom"
]
print(tabla_diaria_display.to_string(index=False))

buckets_d = tabla_diaria["Intervalo"].astype(str).tolist()
fr_d_pct  = (tabla_diaria["Fraude_Aprobado"]/tabla_diaria["Total_Clientes"]*100).round(4).tolist()
fig_d = make_subplots(rows=1, cols=2,
    subplot_titles=["Clientes por nivel trx/dia","FR% por nivel trx/dia"])
fig_d.add_trace(go.Bar(x=buckets_d, y=tabla_diaria["Genuina"].tolist(), name="Genuina",
    marker_color=C_OK,
    text=tabla_diaria["Genuina"].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=1)
fig_d.add_trace(go.Bar(x=buckets_d, y=tabla_diaria["Fraude_Aprobado"].tolist(), name="Fraude Aprobado",
    marker_color=C_FRAUDE,
    text=tabla_diaria["Fraude_Aprobado"].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=1)
fig_d.add_trace(go.Bar(x=buckets_d, y=fr_d_pct, name="FR%",
    marker_color=C_088,
    text=[f"{v:.4f}%" for v in fr_d_pct], textposition="outside"), row=1, col=2)
fig_d.update_layout(title_text="MCC 7995 — Transaccionalidad diaria",
    barmode="group", height=450, plot_bgcolor="white")
fig_d.write_html("g3_trx_diaria.html"); fig_d.show()


# =============================================================================
# SECCIÓN 5 — CONDICIÓN 088
# =============================================================================
print("\n" + "="*65)
print("SECCION 5 — CONDICION 088")
print("="*65)

f_en_088_tot  = (df["COND_088"] & (df["ES_FRAUDE"]==1)).sum()
f_fuera_088_tot=(~df["COND_088"] & (df["ES_FRAUDE"]==1)).sum()
f_en_088_ap   = (df["COND_088"] & (df["ES_FRAUDE_APROBADO"]==1)).sum()
f_fuera_088_ap=(~df["COND_088"] & (df["ES_FRAUDE_APROBADO"]==1)).sum()

print(f"  FRAUDE TOTAL:    dentro={f_en_088_tot:,} ({f_en_088_tot/n_f_tot_g*100:.4f}%)  fuera={f_fuera_088_tot:,}" if n_f_tot_g>0 else "")
print(f"  FRAUDE APROBADO: dentro={f_en_088_ap:,} ({f_en_088_ap/n_f_ap_g*100:.4f}%)   fuera={f_fuera_088_ap:,}" if n_f_ap_g>0 else "")

comp_088 = pd.DataFrame({
    "Con 088":{
        "Transacciones"           : df["COND_088"].sum(),
        "Fraude Total"            : f_en_088_tot,
        "Fraude Aprobado"         : f_en_088_ap,
        "FR% Aprobado"            : f"{df.loc[df['COND_088'],'ES_FRAUDE_APROBADO'].mean()*100:.4f}%",
        "Monto prom (S/)"         : round(df.loc[df["COND_088"],"MONTO"].mean(),2),
        "Monto mediano (S/)"      : round(df.loc[df["COND_088"],"MONTO"].median(),2),
        "Monto fraude aprobado S/": round(df.loc[df["COND_088"]&(df["ES_FRAUDE_APROBADO"]==1),"MONTO"].sum(),2),
    },
    "Sin 088":{
        "Transacciones"           : (~df["COND_088"]).sum(),
        "Fraude Total"            : f_fuera_088_tot,
        "Fraude Aprobado"         : f_fuera_088_ap,
        "FR% Aprobado"            : f"{df.loc[~df['COND_088'],'ES_FRAUDE_APROBADO'].mean()*100:.4f}%",
        "Monto prom (S/)"         : round(df.loc[~df["COND_088"],"MONTO"].mean(),2),
        "Monto mediano (S/)"      : round(df.loc[~df["COND_088"],"MONTO"].median(),2),
        "Monto fraude aprobado S/": round(df.loc[~df["COND_088"]&(df["ES_FRAUDE_APROBADO"]==1),"MONTO"].sum(),2),
    },
}).T
print(comp_088.to_string())

# =============================================================================
# SECCIÓN 6 — DISTRIBUCIÓN POR DIMENSIONES
# =============================================================================
print("\n" + "="*65)
print("SECCION 6 — DISTRIBUCION DEL FRAUDE POR DIMENSIONES")
print("="*65)

def resumen_por(df_base, col, top_n=10):
    return (
        df_base.groupby(col)
        .agg(
            Txns              =(col,                 "count"),
            Fraude_Total      =("ES_FRAUDE",          "sum"),
            Fraude_Aprobado   =("ES_FRAUDE_APROBADO", "sum"),
            FR_Total_pct      =("ES_FRAUDE",          lambda x: x.mean()*100),
            FR_Aprobado_pct   =("ES_FRAUDE_APROBADO", lambda x: x.mean()*100),
            Monto_total       =("MONTO",              "sum"),
            Monto_Fraude_Ap   =("MONTO",              lambda x: x[df_base.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        )
        .sort_values("Fraude_Aprobado", ascending=False)
        .head(top_n)
        .rename(columns={
            "FR_Total_pct"  :"FR% Total",
            "FR_Aprobado_pct":"FR% Aprobado",
            "Monto_Fraude_Ap":"Monto Fraude Aprobado",
        })
    )

df_por_canal     = resumen_por(df, "CANAL")
df_por_seg_nombre= resumen_por(df, "SEG_NOMBRE")  # nombre individual
df_por_seg_grupo = resumen_por(df, "SEG_GRUPO")   # agrupacion
df_por_seguro    = resumen_por(df, "SEGURO")
df_por_tipo      = resumen_por(df, "TIPO_PRODUCTO")
df_por_cvv       = resumen_por(df, "COD_RED_LABEL")
df_por_pais      = resumen_por(df, "PAIS")

for nombre, tabla in [
    ("Canal",                     df_por_canal),
    ("Segmento INDIVIDUAL",       df_por_seg_nombre),
    ("Segmento GRUPO",            df_por_seg_grupo),
    ("Seguro/No Seguro",          df_por_seguro),
    ("Tipo Producto",             df_por_tipo),
    ("CVV/Red Comercio",          df_por_cvv),
    ("Pais (Top 10)",             df_por_pais),
]:
    print(f"\n— Por {nombre}:")
    print(tabla.to_string())

# Top Comercios — fraude total desde df completo, aprobado desde df_ap
print("\n— Top 10 comercios:")
df_top_comercios = (
    df.groupby("COMERCIO")
    .agg(
        Fraude_Total_g      =("ES_FRAUDE",           "sum"),
        Monto_Fraude_Tot_g  =("MONTO",              lambda x: x[df.loc[x.index,"ES_FRAUDE"]==1].sum()),
    )
    .join(
        df_ap.groupby("COMERCIO")
        .agg(
            Tarjetas        =("ID_CLIENTE",           "nunique"),
            Trx             =("MONTO",                "count"),
            Fraude_Aprobado =("ES_FRAUDE_APROBADO",   "sum"),
            Monto_Fraude_Ap =("MONTO",               lambda x: x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
            Monto_total     =("MONTO",                "sum"),
            FR_Ap_pct       =("ES_FRAUDE_APROBADO",   lambda x: x.mean()*100),
        )
    )
    .reset_index()
    .sort_values("Monto_Fraude_Ap", ascending=False)
    .head(10)
    .rename(columns={
        "COMERCIO"          :"Comercio",
        "Tarjetas"          :"# Tarjetas",
        "Trx"               :"# Trx Aprobadas",
        "Fraude_Total_g"    :"# Fraude Total (ap+den)",
        "Fraude_Aprobado"   :"# Fraude Aprobado",
        "Monto_Fraude_Tot_g":"Monto Fraude Total S/",
        "Monto_Fraude_Ap"   :"Monto Fraude Aprobado S/",
        "Monto_total"       :"Monto Total S/",
        "FR_Ap_pct"         :"Fraud Rate % (ap)",
    })
)
df_top_comercios["Fraud Rate % (ap)"] = df_top_comercios["Fraud Rate % (ap)"].map("{:.4f}%".format)
print(df_top_comercios.to_string(index=False))

# Grafico dimensiones (seg_nombre + canal + seguro + cvv)
dims_graf = [("SEG_NOMBRE","Segmento Individual"),("SEG_GRUPO","Segmento Grupo"),
             ("CANAL","Canal"),("SEGURO","Seguro/No Seguro")]
fig3 = make_subplots(rows=2, cols=2, subplot_titles=[d[1] for d in dims_graf])
for i,(col,titulo) in enumerate(dims_graf):
    r,c = divmod(i,2)
    sub = resumen_por(df, col, top_n=8).reset_index()
    fig3.add_trace(go.Bar(x=sub[col], y=sub["Fraude_Total"], name="Fraude Total",
        marker_color=C_ROJO2, opacity=0.6,
        text=sub["Fraude_Total"].apply(lambda v: f"{v:,}"),
        textposition="outside", showlegend=(i==0)), row=r+1, col=c+1)
    fig3.add_trace(go.Bar(x=sub[col], y=sub["Fraude_Aprobado"], name="Fraude Aprobado",
        marker_color=C_FRAUDE,
        text=sub["Fraude_Aprobado"].apply(lambda v: f"{v:,}"),
        textposition="outside", showlegend=(i==0)), row=r+1, col=c+1)
fig3.update_layout(title_text="MCC 7995 — Fraude por dimension",
    barmode="group", height=700, plot_bgcolor="white")
fig3.write_html("g4_fraude_dimension.html"); fig3.show()


# =============================================================================
# SECCIÓN 7 — ESTADÍSTICAS DE MONTO CON INTERPRETACIONES
# =============================================================================
print("\n" + "="*65)
print("SECCION 7 — ESTADISTICAS DE MONTO")
print("="*65)

pcts        = [10,25,50,75,90,95,99]
df_nofraude = df_ap[df_ap["ES_FRAUDE"]==0]

def stats_serie(s):
    if len(s)==0: return [0]*13
    return ([s.mean(),s.median(),s.std(),s.var(),s.min(),s.max()] +
            [s.quantile(p/100) for p in pcts])

stats_monto = pd.DataFrame({
    "Metrica"         :["Media","Mediana","Desv. Estandar","Varianza","Minimo","Maximo"]+
                       [f"Percentil {p}" for p in pcts],
    "Total aprobadas" : stats_serie(df_ap["MONTO"]),
    "Fraude Total"    : stats_serie(df_f_tot["MONTO"]),
    "Fraude Aprobado" : stats_serie(df_f_ap["MONTO"]),
    "No Fraude"       : stats_serie(df_nofraude["MONTO"]),
})
for col in ["Total aprobadas","Fraude Total","Fraude Aprobado","No Fraude"]:
    stats_monto[col] = stats_monto[col].apply(lambda v: f"S/ {v:,.2f}")
print(stats_monto.to_string(index=False))

# Interpretaciones automaticas
def interpretar_stats(serie_f_ap, serie_nof):
    if len(serie_f_ap)==0:
        print("  Sin fraudes aprobados para interpretar.")
        return
    media_f  = serie_f_ap.mean();   median_f = serie_f_ap.median()
    std_f    = serie_f_ap.std();    media_n  = serie_nof.mean()
    median_n = serie_nof.median()
    p75_f    = serie_f_ap.quantile(0.75); p90_f = serie_f_ap.quantile(0.90)
    p95_f    = serie_f_ap.quantile(0.95); p99_f = serie_f_ap.quantile(0.99)
    p75_n    = serie_nof.quantile(0.75);  p90_n = serie_nof.quantile(0.90)
    p95_n    = serie_nof.quantile(0.95)

    ratio_mm = media_f/median_f if median_f>0 else 0
    print("\nINTERPRETACIONES AUTOMATICAS:")

    # Sesgo media vs mediana
    if ratio_mm > 2:
        print(f"  SESGO ALTO: Media S/ {media_f:,.2f} es {ratio_mm:.1f}x mayor que mediana"
              f" S/ {median_f:,.2f}. Hay transacciones extremas que jalan el promedio."
              f" El fraude TIPICO es S/ {median_f:,.2f}, no la media.")
    elif ratio_mm > 1.5:
        print(f"  SESGO MODERADO: Media S/ {media_f:,.2f} vs mediana S/ {median_f:,.2f}"
              f" (ratio {ratio_mm:.2f}x). Algunos montos altos distorsionan el promedio.")
    else:
        print(f"  DISTRIBUCION SIMETRICA: Media S/ {media_f:,.2f} ~= mediana"
              f" S/ {median_f:,.2f}. Fraude homogeneo en monto.")

    # Desviacion estandar
    if media_f>0:
        cv = std_f/media_f
        if cv>1:
            print(f"  ALTA VARIABILIDAD: Desv. std S/ {std_f:,.2f} (CV={cv:.2f}x la media)."
                  f" No hay un patron unico de monto en el fraude.")
        else:
            print(f"  VARIABILIDAD MODERADA: Desv. std S/ {std_f:,.2f} (CV={cv:.2f}x)."
                  f" Montos relativamente consistentes.")

    # Comparacion mediana fraude vs legitimo
    if median_f > median_n:
        print(f"  MONTOS MAYORES: Mediana fraude S/ {median_f:,.2f} > legitimo"
              f" S/ {median_n:,.2f}. Los atacantes apuestan mas que el cliente tipico.")
    else:
        print(f"  MONTOS MENORES: Mediana fraude S/ {median_f:,.2f} <= legitimo"
              f" S/ {median_n:,.2f}. Posible testeo de tarjetas con montos bajos.")

    # Percentiles - umbrales de regla
    print(f"\n  ANALISIS DE PERCENTILES (fraude aprobado vs legitimo):")
    for pf, pn, p_label in [(p75_f,p75_n,"P75"),(p90_f,p90_n,"P90"),(p95_f,p95_n,"P95")]:
        diff = pf - pn
        if diff > 0:
            print(f"  {p_label}: Fraude S/ {pf:,.2f} | Legitimo S/ {pn:,.2f}"
                  f" | Diferencia +S/ {diff:,.2f} — umbral efectivo aqui.")
        else:
            print(f"  {p_label}: Fraude S/ {pf:,.2f} | Legitimo S/ {pn:,.2f}"
                  f" — fraude NO supera al legitimo en este percentil.")
    print(f"  P99: Fraude S/ {p99_f:,.2f} — cola extrema del fraude.")

interpretar_stats(df_f_ap["MONTO"], df_nofraude["MONTO"])

# Evolucion mensual
evol = (
    df.groupby("MES")
    .agg(
        Aprobadas      =("ESTADO",             lambda x: (x=="APROBADA").sum()),
        Denegadas      =("ESTADO",             lambda x: (x=="DENEGADA").sum()),
        Fraude_Total   =("ES_FRAUDE",           "sum"),
        Fraude_Aprobado=("ES_FRAUDE_APROBADO",  "sum"),
        Con_088        =("COND_088",            "sum"),
        Monto_total    =("MONTO",               "sum"),
    ).reset_index()
)
fig4 = make_subplots(rows=1, cols=2,
    subplot_titles=["Aprobadas / Denegadas","Fraude Total vs Aprobado vs 088"])
for serie,color,nombre in [("Aprobadas",C_OK,"Aprobadas"),("Denegadas",C_NEU,"Denegadas")]:
    fig4.add_trace(go.Bar(x=evol["MES"], y=evol[serie], name=nombre,
        marker_color=color,
        text=evol[serie].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=1)
for serie,color,nombre in [("Fraude_Total",C_ROJO2,"Fraude Total"),
                            ("Fraude_Aprobado",C_FRAUDE,"Fraude Aprobado"),
                            ("Con_088",C_088,"Con 088")]:
    fig4.add_trace(go.Bar(x=evol["MES"], y=evol[serie], name=nombre,
        marker_color=color,
        text=evol[serie].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=2)
fig4.update_layout(title_text="MCC 7995 — Evolucion mensual",
    barmode="group", height=450, plot_bgcolor="white")
fig4.write_html("g5_evolucion_mensual.html"); fig4.show()


# =============================================================================
# SECCIÓN 8 — TABLAS CRUZADAS CON CHI-CUADRADO
# =============================================================================
print("\n" + "="*65)
print("SECCION 8 — TABLAS CRUZADAS DE DOS DIMENSIONES")
print("="*65)

def tabla_cruzada(df_base, dim_filas, dim_cols, titulo):
    print(f"\n--- {titulo} ---")
    piv_cnt = df_base.pivot_table(
        values="ES_FRAUDE_APROBADO", index=dim_filas, columns=dim_cols,
        aggfunc="sum", fill_value=0
    )
    piv_fr = df_base.pivot_table(
        values="ES_FRAUDE_APROBADO", index=dim_filas, columns=dim_cols,
        aggfunc="mean", fill_value=0
    ).applymap(lambda v: f"{v*100:.4f}%")
    piv_monto = df_base[df_base["ES_FRAUDE_APROBADO"]==1].pivot_table(
        values="MONTO", index=dim_filas, columns=dim_cols,
        aggfunc="sum", fill_value=0
    ).applymap(lambda v: f"S/ {v:,.2f}")

    print("Fraudes Aprobados (cantidad):"); print(piv_cnt.to_string())
    print("Fraud Rate % por celda:");       print(piv_fr.to_string())
    print("Monto Fraude Aprobado S/:");     print(piv_monto.to_string())

    try:
        chi2,p_val,dof,_ = chi2_contingency(piv_cnt.values)
        sig = "SIGNIFICATIVA" if p_val<0.05 else "No significativa"
        print(f"Chi2={chi2:.4f} | p={p_val:.6f} | gl={dof} | {sig}")
        if p_val < 0.05:
            print(f"  -> {dim_filas} x {dim_cols} SI tiene asociacion con el fraude.")
        else:
            print(f"  -> No hay asociacion estadistica significativa.")
    except Exception as e:
        print(f"  Chi2 no calculable: {e}")
    return piv_cnt, piv_fr, piv_monto

cruces = [
    ("TIPO_PRODUCTO","SEGURO",       "Tipo Producto x Seguro/No Seguro"),
    ("SEG_NOMBRE",   "SEGURO",       "Segmento Individual x Seguro/No Seguro"),
    ("SEG_GRUPO",    "SEGURO",       "Segmento Grupo x Seguro/No Seguro"),
    ("COD_RED_LABEL","SEGURO",       "CVV/Red x Seguro/No Seguro"),
    ("CANAL",        "SEGURO",       "Canal x Seguro/No Seguro"),
    ("TIPO_PRODUCTO","COD_RED_LABEL","Tipo Producto x CVV/Red"),
]
resultados_cruce = {}
for dim_f,dim_c,titulo in cruces:
    piv_cnt,piv_fr,piv_monto = tabla_cruzada(df_ap, dim_f, dim_c, titulo)
    resultados_cruce[titulo] = (piv_cnt, piv_fr, piv_monto)

# =============================================================================
# SECCIÓN 9 — CLIENTES ÚNICOS Y RECURRENCIA DE FRAUDE
# =============================================================================
print("\n" + "="*65)
print("SECCION 9 — CLIENTES UNICOS Y RECURRENCIA DE FRAUDE")
print("="*65)

print("NOTA: La suma de clientes unicos por mes supera al total global")
print("porque un cliente puede comprar en varios meses (se cuenta una vez por mes).")

clientes_por_mes = df_ap.groupby("MES")["ID_CLIENTE"].nunique()
total_cli_global = df_ap["ID_CLIENTE"].nunique()
print(f"\nClientes unicos por mes:")
print(clientes_por_mes.to_string())
print(f"\nSuma por mes:          {clientes_por_mes.sum():,}")
print(f"Clientes globales:     {total_cli_global:,}")
print(f"Solapamiento:          {clientes_por_mes.sum()-total_cli_global:,} apariciones multi-mes")

cli_fraude_por_mes = {
    mes: set(df_ap[(df_ap["MES"]==mes)&(df_ap["ES_FRAUDE_APROBADO"]==1)]["ID_CLIENTE"])
    for mes in meses_disponibles
}
print("\nClientes con fraude aprobado por mes:")
for mes,clientes in cli_fraude_por_mes.items():
    print(f"  {mes}: {len(clientes):,} clientes")

df_recurrencia_rows = []
if len(meses_disponibles) >= 2:
    print("\nRecurrencia entre meses consecutivos:")
    for i in range(len(meses_disponibles)-1):
        mes_a = meses_disponibles[i]
        mes_b = meses_disponibles[i+1]
        cli_f_a = cli_fraude_por_mes.get(mes_a, set())
        cli_f_b = cli_fraude_por_mes.get(mes_b, set())
        recurrentes = cli_f_a & cli_f_b
        nuevos_b    = cli_f_b - cli_f_a
        pct_rec = len(recurrentes)/len(cli_f_b)*100 if len(cli_f_b)>0 else 0
        print(f"  {mes_a}->{mes_b}: fraude_A={len(cli_f_a):,} fraude_B={len(cli_f_b):,}"
              f" recurrentes={len(recurrentes):,} ({pct_rec:.4f}%) nuevos={len(nuevos_b):,}")
        df_recurrencia_rows.append({
            "Periodo"             : f"{mes_a} -> {mes_b}",
            "Clientes fraude MesA": len(cli_f_a),
            "Clientes fraude MesB": len(cli_f_b),
            "Recurrentes"         : len(recurrentes),
            "% Recurrentes"       : f"{pct_rec:.4f}%",
            "Nuevos en MesB"      : len(nuevos_b),
        })

    if len(meses_disponibles) >= 3:
        rec_todos = cli_fraude_por_mes[meses_disponibles[0]]
        for m in meses_disponibles[1:]:
            rec_todos = rec_todos & cli_fraude_por_mes.get(m, set())
        print(f"\nClientes con fraude en TODOS los meses: {len(rec_todos):,}")
        if len(rec_todos) > 0:
            print(f"  IDs (primeros 10): {list(rec_todos)[:10]}")

df_recurrencia_tabla = pd.DataFrame(df_recurrencia_rows)

# Tabla resumen por mes para Excel
df_clientes_mes = pd.DataFrame({
    "Mes"                  : meses_disponibles,
    "Clientes unicos"      : [df_ap[df_ap["MES"]==m]["ID_CLIENTE"].nunique() for m in meses_disponibles],
    "Clientes con fraude"  : [len(cli_fraude_por_mes.get(m,set())) for m in meses_disponibles],
    "% con fraude"         : [
        f"{len(cli_fraude_por_mes.get(m,set()))/df_ap[df_ap['MES']==m]['ID_CLIENTE'].nunique()*100:.4f}%"
        if df_ap[df_ap["MES"]==m]["ID_CLIENTE"].nunique()>0 else "0.0000%"
        for m in meses_disponibles
    ],
})


# =============================================================================
# EXPORTACIÓN A EXCEL — TODAS LAS HOJAS
# =============================================================================
print("\nExportando a Excel...")

FILL_HEADER  = PatternFill("solid", fgColor="1F3864")
FILL_SUBHEAD = PatternFill("solid", fgColor="2E75B6")
FILL_FILA_A  = PatternFill("solid", fgColor="DEEAF1")
FILL_AMARILLO= PatternFill("solid", fgColor="FFF2CC")
FILL_FRAUDE  = PatternFill("solid", fgColor="FCE4D6")
FONT_HEADER  = Font(color="FFFFFF", bold=True, size=10)
FONT_NORMAL  = Font(size=10)
FONT_INTERP  = Font(italic=True, size=9, color="1F3864")
BORDER_THIN  = Border(left=Side(style="thin"),right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
ALIGN_CENTER = Alignment(horizontal="center",vertical="center",wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",  vertical="center",wrap_text=True)

def estilizar_header(ws, fila, n_cols, texto, fill=None):
    fill = fill or FILL_HEADER
    ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=n_cols)
    c = ws.cell(row=fila,column=1,value=texto)
    c.fill=fill; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def estilizar_encabezados_df(ws, fila):
    for row in ws.iter_rows(min_row=fila,max_row=fila):
        for c in row:
            c.fill=FILL_SUBHEAD; c.font=FONT_HEADER
            c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def estilizar_datos(ws, fi, ff, kw=None):
    for i,row in enumerate(ws.iter_rows(min_row=fi,max_row=ff),start=1):
        fill = FILL_FILA_A if i%2==0 else PatternFill()
        if kw and row[0].value and any(k in str(row[0].value).lower() for k in kw):
            fill = FILL_FRAUDE
        for c in row:
            c.fill=fill; c.font=FONT_NORMAL
            c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def agregar_interp(ws, fila, n_cols, texto):
    ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=n_cols)
    c = ws.cell(row=fila,column=1,value=f"-> {texto}")
    c.fill=FILL_AMARILLO; c.font=FONT_INTERP
    c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
    ws.row_dimensions[fila].height=30

def ajustar_col(ws):
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value),default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(ml+4,35)

with pd.ExcelWriter(RUTA_EXCEL_OUT, engine="openpyxl") as writer:

    # HOJA 1: RESUMEN EJECUTIVO
    sh = "1_Resumen_Ejecutivo"
    nc = len(df_resumen_ejecutivo.columns)+1
    df_resumen_ejecutivo.to_excel(writer,sheet_name=sh,startrow=3)
    ws = writer.sheets[sh]
    estilizar_header(ws,1,nc,f"MCC 7995 — JUEGOS DE AZAR | {datetime.today().strftime('%d/%m/%Y')}")
    estilizar_header(ws,2,nc,"RESUMEN DE INDICADORES DE FRAUDE POR MES",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4)
    estilizar_datos(ws,5,ws.max_row,kw=["fraude","ratio"])
    fu = ws.max_row+2
    estilizar_header(ws,fu,nc,"UMBRAL DE CONTROL — Ratio Trx Fraudulentas",fill=FILL_SUBHEAD)
    df_umbral.to_excel(writer,sheet_name=sh,index=False,startrow=fu)
    fi = ws.max_row+2
    agregar_interp(ws,fi,nc,f"Nivel de riesgo: {nivel_riesgo}. {riesgo_texto[nivel_riesgo]}")
    agregar_interp(ws,fi+1,nc,
        f"Fraude total: {n_f_tot_g:,} | Fraude aprobado: {n_f_ap_g:,} "
        f"({n_f_ap_g/n_f_tot_g*100:.2f}% del total). "
        "La diferencia son fraudes en transacciones denegadas." if n_f_tot_g>0 else "Sin fraudes.")
    ajustar_col(ws)

    # HOJA 2: DECILES
    sh = "2_Deciles"
    nc2 = len(tabla_deciles.columns)+2
    tabla_deciles.to_excel(writer,sheet_name=sh,startrow=3)
    ws = writer.sheets[sh]
    estilizar_header(ws,1,nc2,"MCC 7995 — TABLA DE DECILES POR MONTO")
    estilizar_header(ws,2,nc2,"D01=menor monto D10=mayor monto | Fraude Total=ap+den | Fraude Aprobado=solo aprobadas",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    d_max_f = tabla_deciles["Fraude Trx Aprobado"].idxmax()
    d_max_m = tabla_deciles["Fraude Monto Aprobado"].idxmax()
    fi = ws.max_row+2
    agregar_interp(ws,fi,nc2,f"Mayor concentracion fraude aprobado en cantidad: {d_max_f} ({tabla_deciles.loc[d_max_f,'Fraude Trx Aprobado']:,}) Rango: {tabla_deciles.loc[d_max_f,'Rango S/']}")
    agregar_interp(ws,fi+1,nc2,f"Mayor monto fraude aprobado: {d_max_m} (S/ {tabla_deciles.loc[d_max_m,'Fraude Monto Aprobado']:,.2f}) Rango: {tabla_deciles.loc[d_max_m,'Rango S/']}")
    ajustar_col(ws)

    # HOJA 2B: APERTURA D10
    sh = "2B_Apertura_D10"
    nc2b = len(tabla_d10.columns)+2
    tabla_d10.to_excel(writer,sheet_name=sh,startrow=3)
    ws = writer.sheets[sh]
    estilizar_header(ws,1,nc2b,f"MCC 7995 — APERTURA DECIL 10 (montos >= S/ {monto_d10_min:,.2f})")
    estilizar_header(ws,2,nc2b,"Cortes fijos por percentil dentro del D10",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    fi = ws.max_row+2
    if "Fraude_Trx_Ap" in tabla_d10.columns:
        d10_max = tabla_d10["Fraude_Trx_Ap"].idxmax()
        agregar_interp(ws,fi,nc2b,f"Sub-decil con mas fraudes aprobados en D10: {d10_max} ({tabla_d10.loc[d10_max,'Fraude_Trx_Ap']:,}). Este rango es candidato principal para umbral de monto.")
    ajustar_col(ws)

    # HOJA 3: TOP COMERCIOS
    sh = "3_Top_Comercios"
    df_tc = df_top_comercios.copy()
    df_tc["Monto Fraude Total S/"]    = df_tc["Monto Fraude Total S/"].map("S/ {:,.2f}".format)
    df_tc["Monto Fraude Aprobado S/"] = df_tc["Monto Fraude Aprobado S/"].map("S/ {:,.2f}".format)
    df_tc["Monto Total S/"]           = df_tc["Monto Total S/"].map("S/ {:,.2f}".format)
    df_tc.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws = writer.sheets[sh]
    estilizar_header(ws,1,len(df_tc.columns),"MCC 7995 — TOP 10 COMERCIOS CON MAYOR MONTO FRAUDE APROBADO")
    estilizar_header(ws,2,len(df_tc.columns),"Fraude Total=ap+den desde todo el dataset | Fraude Aprobado=solo aprobadas",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    top1 = df_tc.iloc[0]["Comercio"] if len(df_tc)>0 else "N/A"
    monto_top1 = df_tc.iloc[0]["Monto Fraude Aprobado S/"] if len(df_tc)>0 else "S/ 0"
    fi = ws.max_row+2
    agregar_interp(ws,fi,len(df_tc.columns),f"Comercio con mayor monto fraude aprobado: '{top1}' ({monto_top1}). Si concentra >50% del total es candidato para regla especifica.")
    ajustar_col(ws)

    # HOJA 4: ESTADISTICAS MONTO
    sh = "4_Estadisticas_Monto"
    stats_monto.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws = writer.sheets[sh]
    estilizar_header(ws,1,len(stats_monto.columns),"MCC 7995 — ESTADISTICAS DE MONTO: TOTAL vs APROBADO vs NO FRAUDE")
    estilizar_header(ws,2,len(stats_monto.columns),"Fraude Total=F en cualquier estado | Fraude Aprobado=F que pasaron el control | No Fraude=aprobadas sin F",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    if len(df_f_ap)>0:
        media_f_ap=df_f_ap["MONTO"].mean(); med_f_ap=df_f_ap["MONTO"].median()
        std_f_ap=df_f_ap["MONTO"].std();    med_nof=df_nofraude["MONTO"].median()
        p90_f_ap=df_f_ap["MONTO"].quantile(0.90); p90_nof=df_nofraude["MONTO"].quantile(0.90)
        p95_f_ap=df_f_ap["MONTO"].quantile(0.95); p95_nof=df_nofraude["MONTO"].quantile(0.95)
        ratio_mm=media_f_ap/med_f_ap if med_f_ap>0 else 0
        fi = ws.max_row+2
        agregar_interp(ws,fi,len(stats_monto.columns),
            f"SESGO: Media S/ {media_f_ap:,.2f} vs mediana S/ {med_f_ap:,.2f} (ratio {ratio_mm:.2f}x). "
            +("Alta presencia de montos extremos — el fraude tipico es la mediana." if ratio_mm>2
              else "Sesgo moderado." if ratio_mm>1.5 else "Distribucion simetrica."))
        agregar_interp(ws,fi+1,len(stats_monto.columns),
            f"DISPERSION: Desv. std S/ {std_f_ap:,.2f}. "
            +(f"Alta variabilidad (CV={std_f_ap/media_f_ap:.2f}x) — montos heterogeneos." if media_f_ap>0 and std_f_ap/media_f_ap>1
              else "Variabilidad moderada."))
        agregar_interp(ws,fi+2,len(stats_monto.columns),
            f"P90: Fraude S/ {p90_f_ap:,.2f} vs Legitimo S/ {p90_nof:,.2f}. "
            +("Umbral en este rango captura el 10% de fraudes mas altos." if p90_f_ap>p90_nof
              else "P90 del fraude no supera al legitimo — umbral de monto no efectivo aqui."))
        agregar_interp(ws,fi+3,len(stats_monto.columns),
            f"COMPARACION: Mediana fraude S/ {med_f_ap:,.2f} vs legitimo S/ {med_nof:,.2f}. "
            +("Fraude con montos MAYORES — atacantes apuestan mas que el cliente tipico." if med_f_ap>med_nof
              else "Fraude con montos MENORES/IGUALES — posible testeo de tarjetas."))
    ajustar_col(ws)

    # HOJA 5: TRX DIARIA CLIENTE
    sh = "5_Trx_Diaria_Cliente"
    tabla_diaria_display.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws = writer.sheets[sh]
    estilizar_header(ws,1,len(tabla_diaria_display.columns),"MCC 7995 — TRANSACCIONALIDAD DIARIA POR CLIENTE")
    estilizar_header(ws,2,len(tabla_diaria_display.columns),
        f"Fraude aprobado en {pct_1trx:.2f}% de pares cliente-dia con 1 sola transaccion",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    fi = ws.max_row+2
    agregar_interp(ws,fi,len(tabla_diaria_display.columns),
        f"El {pct_1trx:.2f}% de clientes fraudulentos hace 1 trx/dia -> ataque de monto alto en una sola operacion. Regla de velocidad diaria no seria efectiva — priorizar umbral de monto.")
    ajustar_col(ws)

    # HOJA 6: VELOCIDAD INTERVALO
    sh = "6_Velocidad_Intervalo"
    tabla_vel.reset_index().to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws = writer.sheets[sh]
    nc6 = len(tabla_vel.columns)+1
    estilizar_header(ws,1,nc6,"MCC 7995 — VELOCIDAD: INTERVALO ENTRE TRANSACCIONES DEL MISMO CLIENTE")
    estilizar_header(ws,2,nc6,"En que ventana de tiempo se concentran los ataques?",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    bk_max = tabla_vel["FR_Aprobado_pct"].idxmax() if len(tabla_vel)>0 else "N/A"
    pct_v  = tabla_vel.loc[bk_max,"Pct_fraudes_total"] if len(tabla_vel)>0 else 0
    fi = ws.max_row+2
    agregar_interp(ws,fi,nc6,f"Intervalo con mayor FR aprobado: '{bk_max}' ({pct_v:.2f}% del total). Si <=5 min -> ataques automatizados.")
    ajustar_col(ws)

    # HOJA 7: POR DIMENSION
    sh = "7_Por_Dimension"
    ws7 = writer.book.create_sheet(sh); writer.sheets[sh] = ws7
    fa = 1
    dims_exp = [
        (df_por_canal,      "Canal",              "Canal con mayor fraude indica vector de ataque principal."),
        (df_por_seg_nombre, "Segmento Individual","Nombre individual de cada codigo de segmento."),
        (df_por_seg_grupo,  "Segmento Grupo",     "Agrupacion de segmentos (Affluent, Mass, Corporate, etc.)."),
        (df_por_seguro,     "Seguridad Comercio", "No Seguro indica ausencia de 3DSecure."),
        (df_por_tipo,       "Tipo Producto",       "TD vs TC puede tener distinto perfil de fraude."),
        (df_por_cvv,        "CVV/Red Comercio",    "Alta concentracion en No Match o Estatico es alerta."),
        (df_por_pais,       "Pais Origen Top10",   "Trxs desde paises distintos al Peru son alto riesgo."),
    ]
    for df_dim,titulo,interp in dims_exp:
        df_out = df_dim.reset_index(); nd = len(df_out.columns)
        estilizar_header(ws7,fa,nd,f"MCC 7995 — {titulo.upper()}"); fa+=1
        for j,col in enumerate(df_out.columns,start=1):
            c=ws7.cell(row=fa,column=j,value=col)
            c.fill=FILL_SUBHEAD; c.font=FONT_HEADER
            c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
        fa+=1
        for i,row in df_out.iterrows():
            fill=FILL_FILA_A if i%2==0 else PatternFill()
            for j,val in enumerate(row,start=1):
                c=ws7.cell(row=fa,column=j,
                    value=f"{val:,.4f}" if isinstance(val,float) else val)
                c.fill=fill; c.font=FONT_NORMAL
                c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
            fa+=1
        ws7.merge_cells(start_row=fa,start_column=1,end_row=fa,end_column=nd)
        c=ws7.cell(row=fa,column=1,value=f"-> {interp}")
        c.fill=FILL_AMARILLO; c.font=FONT_INTERP
        c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
        ws7.row_dimensions[fa].height=25; fa+=3
    ajustar_col(ws7)

    # HOJA 8: TABLAS CRUZADAS
    sh = "8_Tablas_Cruzadas"
    ws8 = writer.book.create_sheet(sh); writer.sheets[sh] = ws8
    fa = 1
    for dim_f,dim_c,titulo in cruces:
        estilizar_header(ws8,fa,10,f"MCC 7995 — {titulo.upper()}"); fa+=1
        piv = df_ap.pivot_table(
            values="ES_FRAUDE_APROBADO",index=dim_f,columns=dim_c,
            aggfunc="sum",fill_value=0
        ).reset_index()
        for j,col in enumerate(piv.columns,start=1):
            c=ws8.cell(row=fa,column=j,value=str(col))
            c.fill=FILL_SUBHEAD; c.font=FONT_HEADER
            c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
        fa+=1
        for i,row in piv.iterrows():
            fill=FILL_FILA_A if i%2==0 else PatternFill()
            for j,val in enumerate(row,start=1):
                c=ws8.cell(row=fa,column=j,value=val)
                c.fill=fill; c.font=FONT_NORMAL
                c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
            fa+=1
        try:
            contingencia = df_ap.pivot_table(
                values="ES_FRAUDE_APROBADO",index=dim_f,columns=dim_c,
                aggfunc="sum",fill_value=0
            ).values
            chi2,p_val,dof,_ = chi2_contingency(contingencia)
            sig = "SIGNIFICATIVA" if p_val<0.05 else "No significativa"
            ws8.merge_cells(start_row=fa,start_column=1,end_row=fa,end_column=10)
            c=ws8.cell(row=fa,column=1,
                value=f"Chi2={chi2:.4f} | p-valor={p_val:.6f} | gl={dof} | {sig}")
            c.fill=FILL_AMARILLO; c.font=FONT_INTERP
            c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
        except: pass
        fa+=3
    ajustar_col(ws8)

    # HOJA 9: CLIENTES RECURRENTES
    sh = "9_Clientes_Recurrentes"
    df_clientes_mes.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws = writer.sheets[sh]
    nc9 = len(df_clientes_mes.columns)
    estilizar_header(ws,1,nc9,"MCC 7995 — CLIENTES UNICOS Y RECURRENCIA DE FRAUDE")
    estilizar_header(ws,2,nc9,"Un cliente puede aparecer en varios meses — no suman al total global",fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws,4); estilizar_datos(ws,5,ws.max_row)
    fi = ws.max_row+2
    agregar_interp(ws,fi,nc9,
        f"Clientes unicos globales: {total_cli_global:,}. "
        f"La suma por mes ({clientes_por_mes.sum():,}) es mayor por clientes que compraron en varios meses.")
    if len(df_recurrencia_tabla)>0:
        fi2 = ws.max_row+3
        estilizar_header(ws,fi2,nc9,"RECURRENCIA DE FRAUDE ENTRE MESES CONSECUTIVOS",fill=FILL_SUBHEAD)
        df_recurrencia_tabla.to_excel(writer,sheet_name=sh,index=False,startrow=fi2)
        estilizar_encabezados_df(ws,fi2+1)
        estilizar_datos(ws,fi2+2,ws.max_row)
    ajustar_col(ws)

print(f"\nExcel exportado: {RUTA_EXCEL_OUT}")
print("Hojas: 1_Resumen_Ejecutivo | 2_Deciles | 2B_Apertura_D10 | 3_Top_Comercios")
print("       4_Estadisticas_Monto | 5_Trx_Diaria_Cliente | 6_Velocidad_Intervalo")
print("       7_Por_Dimension | 8_Tablas_Cruzadas | 9_Clientes_Recurrentes")
print("\nGraficos: g1_deciles | g2_velocidad_intervalo | g3_trx_diaria")
print("          g4_fraude_dimension | g5_evolucion_mensual")
