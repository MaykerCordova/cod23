"""
╔══════════════════════════════════════════════════════════════════╗
║         BITÁCORA AUTOMÁTICA - BANCA PRIVADA  v2                  ║
║         Buzón: Prevención de Fraudes                             ║
║                                                                  ║
║  Semáforo de tiempos de respuesta:                               ║
║    🟢 Verde    : < 20 minutos                                    ║
║    🟡 Amarillo : 20 – 29 minutos                                 ║
║    🔴 Rojo     : >= 30 minutos                                   ║
║                                                                  ║
║  Mejoras v2:                                                     ║
║    1. Ignora stores de archivo histórico ("Archivar en línea")   ║
║    2. Restrict nativo de Outlook por asunto en Enviados          ║
║    3. Filtro de fecha desde FECHA_INICIO                         ║
╚══════════════════════════════════════════════════════════════════╝
"""

import win32com.client
import pandas as pd
from datetime import datetime
import os
import warnings

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════════════════════════
#  CONFIGURACIÓN  ← Ajusta estos valores antes de correr
# ══════════════════════════════════════════════════════════════════

MAILBOX_NAME   = "prevencionfraudeperu@scotiabank.com.pe"
FOLDER_BANCA   = "BANCA PRIVADA"
FOLDER_SENT    = "Elementos enviados"

KEYWORD_ASUNTO = "BANCA PRIVADA"   # Filtro en asunto (case-insensitive)
KEYWORD_RESP   = "respuesta"       # Palabra en cuerpo para respuesta válida
                                   # (deja en "" para no filtrar por cuerpo)

FECHA_INICIO   = "2026-02-01"      # Solo procesa correos desde esta fecha

# Umbrales semáforo (minutos)
LIMITE_VERDE    = 20   # < 20 min → Verde
LIMITE_AMARILLO = 30   # 20-29 min → Amarillo  |  >= 30 → Rojo

# Rutas de salida
OUTPUT_DIR    = r"C:\Users\s7689735\The Bank of Nova Scotia\cordova pintado, Jose Mayker - BP\data"
BITACORA_FILE = os.path.join(OUTPUT_DIR, "bitacora_banca_privada.xlsx")
REPORTE_FILE  = os.path.join(OUTPUT_DIR, "reporte_banca_privada.xlsx")
CHART_FILE    = os.path.join(OUTPUT_DIR, "grafico_semaforo.png")

# Correo del reporte
DESTINATARIO_REPORTE = ""    # Ej: "jefa.mitigacion@scotiabank.com.pe"
ENVIAR_CORREO        = False # Cambia a True para activar envio automatico


# ══════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════

def semaforo(minutos):
    if pd.isna(minutos):
        return "Sin respuesta"
    if minutos < LIMITE_VERDE:
        return "Verde"
    elif minutos < LIMITE_AMARILLO:
        return "Amarillo"
    else:
        return "Rojo"

def safe_datetime(com_date):
    if com_date is None:
        return None
    try:
        if hasattr(com_date, 'year'):
            if hasattr(com_date, 'tzinfo') and com_date.tzinfo is not None:
                return com_date.replace(tzinfo=None)
            return com_date
        return None
    except Exception:
        return None

def get_body_text(item, n_chars=500):
    try:
        body = str(item.Body or "")
        return body[:n_chars].replace('\r', '').replace('\n', ' ').strip()
    except Exception:
        return ""

def _log(msg, nivel=0):
    print("  " * nivel + msg)

def _es_archivo(display_name):
    """Detecta si un store es el archivo historico de Exchange."""
    nombre = display_name.lower()
    return "archivar" in nombre or "archive" in nombre or "archiv" in nombre


# ══════════════════════════════════════════════════════════════════
#  CONEXION OUTLOOK
# ══════════════════════════════════════════════════════════════════

def _get_outlook_ns():
    outlook = win32com.client.Dispatch("Outlook.Application")
    return outlook.GetNamespace("MAPI"), outlook

def _buscar_en_store(store, nombre_carpeta):
    """Busca carpeta en el store. IGNORA stores de archivo historico."""
    # MEJORA 1: saltar archivo historico
    if _es_archivo(store.DisplayName):
        return None
    try:
        root = store.GetRootFolder()
        for folder in root.Folders:
            if folder.Name.strip().lower() == nombre_carpeta.strip().lower():
                return folder
    except Exception:
        pass
    return None

def get_folder(mailbox_name, folder_name):
    ns, _ = _get_outlook_ns()

    # Intento 1: store cuyo nombre contiene mailbox_name
    for store in ns.Stores:
        if _es_archivo(store.DisplayName):
            continue
        if mailbox_name.lower() in store.DisplayName.lower():
            folder = _buscar_en_store(store, folder_name)
            if folder:
                _log(f"Carpeta '{folder_name}' encontrada en '{store.DisplayName}'", 1)
                return folder

    # Intento 2: buscar en todos (excepto archivo)
    _log(f"Buzon '{mailbox_name}' no encontrado exactamente. Buscando en todos los stores...", 1)
    for store in ns.Stores:
        if _es_archivo(store.DisplayName):
            continue
        folder = _buscar_en_store(store, folder_name)
        if folder:
            _log(f"Carpeta '{folder_name}' encontrada en '{store.DisplayName}'", 1)
            return folder

    raise ValueError(
        f"No se encontro la carpeta '{folder_name}'.\n"
        f"Verifica que el buzon '{mailbox_name}' este abierto en Outlook y la carpeta exista."
    )


# ══════════════════════════════════════════════════════════════════
#  LECTURA DE CORREOS RECIBIDOS
# ══════════════════════════════════════════════════════════════════

def leer_carpeta_banca_privada(folder):
    """Lee correos de BANCA PRIVADA desde FECHA_INICIO."""
    registros = []
    fecha_filtro = datetime.strptime(FECHA_INICIO, "%Y-%m-%d")

    items = folder.Items
    items.Sort("[ReceivedTime]", True)
    _log(f"Leyendo {items.Count} items de la carpeta...", 1)

    for item in items:
        try:
            if not hasattr(item, 'ReceivedTime'):
                continue

            fecha_recepcion = safe_datetime(item.ReceivedTime)
            if fecha_recepcion is None:
                continue

            # MEJORA 3: filtro de fecha
            if fecha_recepcion < fecha_filtro:
                continue

            asunto = str(item.Subject or "")
            if KEYWORD_ASUNTO and KEYWORD_ASUNTO.lower() not in asunto.lower():
                continue

            conv_id = str(item.ConversationID or "")
            remitente = ""
            remitente_email = ""
            try:
                remitente = str(item.SenderName or "")
                remitente_email = str(item.SenderEmailAddress or "")
            except Exception:
                pass

            registros.append({
                "ConversationID"  : conv_id,
                "EntryID"         : str(item.EntryID or ""),
                "Asunto"          : asunto,
                "Remitente"       : remitente,
                "Remitente_Email" : remitente_email,
                "Fecha_Recepcion" : fecha_recepcion,
                "Body_Preview"    : get_body_text(item, 200),
            })

        except Exception as e:
            _log(f"[WARN] Error procesando item: {e}", 2)
            continue

    _log(f"✅ {len(registros)} correos encontrados desde {FECHA_INICIO}", 1)
    return registros


# ══════════════════════════════════════════════════════════════════
#  INDEXAR ENVIADOS
# ══════════════════════════════════════════════════════════════════

def construir_indice_enviados(sent_folder):
    """
    Indexa Elementos enviados con doble filtro:
      - MEJORA 2: Restrict nativo por asunto (de 295k items a decenas)
      - MEJORA 3: Filtro manual de fecha como segunda capa
    """
    indice = {}
    fecha_filtro = datetime.strptime(FECHA_INICIO, "%Y-%m-%d")

    # MEJORA 2: Restrict nativo de Outlook — solo BANCA PRIVADA
    filtro = f"@SQL=\"urn:schemas:httpmail:subject\" LIKE '%{KEYWORD_ASUNTO}%'"
    try:
        items = sent_folder.Items.Restrict(filtro)
        _log(f"Filtro Restrict aplicado. Items relevantes: {items.Count}", 1)
    except Exception as e:
        _log(f"[WARN] Restrict fallo ({e}), iterando sin filtro...", 1)
        items = sent_folder.Items

    total = items.Count
    _log(f"Indexando {total} items de Enviados...", 1)

    count_banca = 0
    errores = 0

    # Iterar por indice explicito — mas estable con COM
    for i in range(1, total + 1):
        try:
            item = items[i]
            if not hasattr(item, 'SentOn'):
                continue

            sent_time = safe_datetime(item.SentOn)
            if sent_time is None:
                continue

            # MEJORA 3: filtro de fecha segunda capa
            if sent_time < fecha_filtro:
                continue

            conv_id = str(item.ConversationID or "")
            if not conv_id:
                continue

            body = get_body_text(item, 500)
            es_valida = KEYWORD_RESP.lower() in body.lower() if KEYWORD_RESP else True

            remitente = ""
            try:
                remitente = str(item.SenderName or "")
            except Exception:
                pass

            if conv_id not in indice:
                indice[conv_id] = []
            indice[conv_id].append({
                "sent_time": sent_time,
                "remitente": remitente,
                "es_valida": es_valida,
                "asunto"   : str(item.Subject or ""),
            })
            count_banca += 1

        except Exception:
            errores += 1
            continue

    for cid in indice:
        indice[cid].sort(key=lambda x: x["sent_time"])

    _log(f"✅ {count_banca} enviados indexados ({len(indice)} hilos) | Errores ignorados: {errores}", 1)
    return indice


# ══════════════════════════════════════════════════════════════════
#  CRUCE Y CALCULO DE TIEMPOS
# ══════════════════════════════════════════════════════════════════

def cruzar_y_calcular(registros, indice_enviados):
    resultados = []
    sin_resp = 0

    for r in registros:
        conv_id        = r["ConversationID"]
        fecha_rec      = r["Fecha_Recepcion"]
        primer_resp_dt = None
        resp_remitente = None
        es_valida      = False

        if conv_id in indice_enviados:
            posteriores = [
                x for x in indice_enviados[conv_id]
                if x["sent_time"] > fecha_rec
            ]
            if posteriores:
                if KEYWORD_RESP:
                    validas = [x for x in posteriores if x["es_valida"]]
                    primera = validas[0] if validas else posteriores[0]
                else:
                    primera = posteriores[0]
                primer_resp_dt = primera["sent_time"]
                resp_remitente = primera["remitente"]
                es_valida      = primera["es_valida"]

        if primer_resp_dt:
            minutos = round((primer_resp_dt - fecha_rec).total_seconds() / 60, 1)
        else:
            minutos = None
            sin_resp += 1

        sem = semaforo(minutos)

        resultados.append({
            "ConversationID"  : conv_id,
            "Asunto"          : r["Asunto"],
            "Remitente"       : r["Remitente"],
            "Remitente_Email" : r["Remitente_Email"],
            "Fecha_Recepcion" : fecha_rec,
            "Fecha_Respuesta" : primer_resp_dt,
            "Respondido_Por"  : resp_remitente or "",
            "Min_Respuesta"   : minutos,
            "Semaforo"        : sem,
            "Resp_Valida"     : "Si" if es_valida else ("No" if primer_resp_dt else ""),
            "Fecha_Proceso"   : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    _log(f"✅ Cruce completado. Sin respuesta: {sin_resp}/{len(registros)}", 1)
    return pd.DataFrame(resultados)


# ══════════════════════════════════════════════════════════════════
#  GUARDAR INCREMENTAL (UPSERT por ConversationID)
# ══════════════════════════════════════════════════════════════════

def guardar_incremental(df_nuevo):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    if os.path.exists(BITACORA_FILE):
        df_existente  = pd.read_excel(BITACORA_FILE)
        ids_nuevos    = set(df_nuevo["ConversationID"].tolist())
        df_no_tocados = df_existente[~df_existente["ConversationID"].isin(ids_nuevos)]
        df_final      = pd.concat([df_no_tocados, df_nuevo], ignore_index=True)
        n_act = len(df_existente[df_existente["ConversationID"].isin(ids_nuevos)])
        _log(f"Nuevos: {len(ids_nuevos) - n_act} | Actualizados: {n_act} | Previos: {len(df_existente)}", 1)
    else:
        df_final = df_nuevo
        _log(f"Archivo nuevo creado con {len(df_final)} registros", 1)

    df_final = df_final.sort_values("Fecha_Recepcion", ascending=False)
    df_final.to_excel(BITACORA_FILE, index=False, sheet_name="Bitacora")
    _log(f"✅ Bitacora guardada → {BITACORA_FILE} ({len(df_final)} registros)", 1)
    return df_final


# ══════════════════════════════════════════════════════════════════
#  GRAFICO
# ══════════════════════════════════════════════════════════════════

COLORES = {
    "Verde"         : "#27ae60",
    "Amarillo"      : "#f39c12",
    "Rojo"          : "#e74c3c",
    "Sin respuesta" : "#bdc3c7",
}

def generar_grafico(df):
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
    fig.patch.set_facecolor("#f8f9fa")
    fig.suptitle(
        f"BANCA PRIVADA - Tiempos de Respuesta\n"
        f"Actualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        fontsize=13, fontweight='bold', color="#2c3e50", y=1.02
    )

    # Panel 1: Distribucion semaforo
    conteo = df["Semaforo"].value_counts()
    colores_pie = [COLORES.get(k, "#95a5a6") for k in conteo.index]
    wedges, texts, autotexts = ax1.pie(
        conteo.values, labels=None, colors=colores_pie,
        autopct="%1.1f%%", startangle=90, pctdistance=0.75,
        wedgeprops={"edgecolor": "white", "linewidth": 2}
    )
    for at in autotexts:
        at.set_fontsize(10); at.set_fontweight("bold"); at.set_color("white")
    leyenda = [
        mpatches.Patch(color=COLORES.get(k, "#95a5a6"), label=f"{k}  ({v})")
        for k, v in conteo.items()
    ]
    ax1.legend(handles=leyenda, loc="lower center", bbox_to_anchor=(0.5, -0.15),
               fontsize=9, frameon=False)
    ax1.set_title("Distribucion por Semaforo", fontweight="bold", color="#2c3e50", pad=12)
    ax1.set_facecolor("#f8f9fa")

    # Panel 2: Promedio diario
    df_con_resp = df[df["Min_Respuesta"].notna()].copy()
    if len(df_con_resp) > 0:
        df_con_resp["Fecha"] = pd.to_datetime(df_con_resp["Fecha_Recepcion"]).dt.date
        df_diario = (
            df_con_resp.groupby("Fecha")["Min_Respuesta"]
            .mean().reset_index().sort_values("Fecha").tail(14)
        )
        bar_colors = [
            "#27ae60" if v < LIMITE_VERDE else
            "#f39c12" if v < LIMITE_AMARILLO else
            "#e74c3c"
            for v in df_diario["Min_Respuesta"]
        ]
        bars = ax2.bar(range(len(df_diario)), df_diario["Min_Respuesta"],
                       color=bar_colors, edgecolor="white", linewidth=1.5, width=0.6)
        for bar, val in zip(bars, df_diario["Min_Respuesta"]):
            ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
                     f"{val:.0f}m", ha="center", va="bottom", fontsize=8, color="#2c3e50")
        ax2.axhline(LIMITE_VERDE, color="#27ae60", linestyle="--", linewidth=1.5,
                    label=f"Limite Verde ({LIMITE_VERDE} min)")
        ax2.axhline(LIMITE_AMARILLO, color="#e74c3c", linestyle="--", linewidth=1.5,
                    label=f"Limite Rojo ({LIMITE_AMARILLO} min)")
        ax2.set_xticks(range(len(df_diario)))
        ax2.set_xticklabels([str(d) for d in df_diario["Fecha"]],
                            rotation=40, ha="right", fontsize=8)
        ax2.set_ylabel("Minutos promedio", fontsize=9)
        ax2.legend(fontsize=8, frameon=False)
        ax2.spines[["top", "right"]].set_visible(False)
    else:
        ax2.text(0.5, 0.5, "Sin datos con respuesta", ha="center", va="center",
                 transform=ax2.transAxes, fontsize=12, color="#7f8c8d")

    ax2.set_title("Tiempo Promedio por Dia (ultimos 14 dias)",
                  fontweight="bold", color="#2c3e50", pad=12)
    ax2.set_facecolor("#f8f9fa")
    plt.tight_layout()
    plt.savefig(CHART_FILE, dpi=150, bbox_inches="tight", facecolor="#f8f9fa")
    plt.close()
    _log(f"✅ Grafico guardado → {CHART_FILE}", 1)


# ══════════════════════════════════════════════════════════════════
#  REPORTE EXCEL FORMATEADO
# ══════════════════════════════════════════════════════════════════

def _calcular_resumen(df):
    total    = len(df)
    con_r    = df["Min_Respuesta"].notna().sum()
    sin_r    = total - con_r
    prom     = df["Min_Respuesta"].mean() if con_r > 0 else None
    mediana  = df["Min_Respuesta"].median() if con_r > 0 else None
    verde    = (df["Semaforo"] == "Verde").sum()
    amarillo = (df["Semaforo"] == "Amarillo").sum()
    rojo     = (df["Semaforo"] == "Rojo").sum()
    pct_v    = round(verde / con_r * 100, 1) if con_r > 0 else 0
    pct_a    = round(amarillo / con_r * 100, 1) if con_r > 0 else 0
    pct_r    = round(rojo / con_r * 100, 1) if con_r > 0 else 0
    return dict(total=total, con_resp=con_r, sin_resp=sin_r,
                prom=prom, mediana=mediana,
                verde=verde, amarillo=amarillo, rojo=rojo,
                pct_verde=pct_v, pct_amarillo=pct_a, pct_rojo=pct_r)

def generar_reporte_excel(df):
    r = _calcular_resumen(df)

    _log("", 0)
    _log("RESUMEN:", 1)
    _log(f"   Total correos       : {r['total']}", 1)
    _log(f"   Con respuesta       : {r['con_resp']}", 1)
    _log(f"   Sin respuesta       : {r['sin_resp']}", 1)
    if r['prom']:
        _log(f"   Promedio (min)      : {r['prom']:.1f}", 1)
        _log(f"   Mediana  (min)      : {r['mediana']:.1f}", 1)
    _log(f"   Verde  (<20m)       : {r['verde']}  ({r['pct_verde']}%)", 1)
    _log(f"   Amarillo (20-29m)   : {r['amarillo']}  ({r['pct_amarillo']}%)", 1)
    _log(f"   Rojo   (>=30m)      : {r['rojo']}  ({r['pct_rojo']}%)", 1)

    columnas_export = [
        "Asunto", "Remitente", "Remitente_Email",
        "Fecha_Recepcion", "Fecha_Respuesta", "Respondido_Por",
        "Min_Respuesta", "Semaforo", "Resp_Valida",
        "ConversationID", "Fecha_Proceso"
    ]
    df_exp = df[[c for c in columnas_export if c in df.columns]].copy()

    with pd.ExcelWriter(REPORTE_FILE, engine="openpyxl") as writer:
        df_exp.to_excel(writer, index=False, sheet_name="Detalle")
        ws = writer.sheets["Detalle"]

        HDR_FILL = PatternFill("solid", fgColor="1F3864")
        HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
        FILL_V   = PatternFill("solid", fgColor="C6EFCE")
        FILL_A   = PatternFill("solid", fgColor="FFEB9C")
        FILL_R   = PatternFill("solid", fgColor="FFC7CE")
        FILL_S   = PatternFill("solid", fgColor="F2F2F2")
        borde    = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

        for col_idx in range(1, len(df_exp.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center"); cell.border = borde

        try:
            sem_col_idx = list(df_exp.columns).index("Semaforo") + 1
        except ValueError:
            sem_col_idx = None

        for row_idx in range(2, len(df_exp) + 2):
            val = str(ws.cell(row=row_idx, column=sem_col_idx).value or "") if sem_col_idx else ""
            if "Verde" in val:
                fill = FILL_V
            elif "Amarillo" in val:
                fill = FILL_A
            elif "Rojo" in val:
                fill = FILL_R
            else:
                fill = FILL_S
            for col_idx in range(1, len(df_exp.columns) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = fill; c.border = borde
                c.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A2"
        for col_idx, col_name in enumerate(df_exp.columns, 1):
            try:
                max_len = max(len(str(col_name)),
                              df_exp[col_name].dropna().astype(str).str.len().max())
            except Exception:
                max_len = 15
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

        # Hoja Resumen
        resumen_rows = [
            ["Metrica", "Valor", "%"],
            ["Total correos",              r["total"],    ""],
            ["Con respuesta",              r["con_resp"], ""],
            ["Sin respuesta",              r["sin_resp"], ""],
            ["Promedio respuesta (min)",   round(r["prom"], 1) if r["prom"] else "N/A", ""],
            ["Mediana respuesta (min)",    round(r["mediana"], 1) if r["mediana"] else "N/A", ""],
            ["--- SEMAFORO ---",           "", ""],
            ["Verde (< 20 min)",           r["verde"],    f"{r['pct_verde']}%"],
            ["Amarillo (20-29 min)",       r["amarillo"], f"{r['pct_amarillo']}%"],
            ["Rojo (>= 30 min)",           r["rojo"],     f"{r['pct_rojo']}%"],
        ]
        pd.DataFrame(resumen_rows[1:], columns=resumen_rows[0]).to_excel(
            writer, index=False, sheet_name="Resumen")

        ws2 = writer.sheets["Resumen"]
        for col_idx in range(1, 4):
            ws2.cell(row=1, column=col_idx).fill = HDR_FILL
            ws2.cell(row=1, column=col_idx).font = HDR_FONT
            ws2.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

        if os.path.exists(CHART_FILE):
            try:
                img = XLImage(CHART_FILE)
                img.anchor = "E2"
                ws2.add_image(img)
            except Exception as e:
                _log(f"[WARN] No se pudo insertar grafico: {e}", 2)

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 12

    _log(f"✅ Reporte Excel generado → {REPORTE_FILE}", 1)


# ══════════════════════════════════════════════════════════════════
#  ENVIO DE REPORTE POR CORREO
# ══════════════════════════════════════════════════════════════════

def enviar_reporte(df, destinatario):
    r = _calcular_resumen(df)
    _, outlook = _get_outlook_ns()
    mail = outlook.CreateItem(0)
    prom_str = f"{r['prom']:.1f} min" if r["prom"] else "N/A"

    mail.To = destinatario
    mail.Subject = (
        f"[BANCA PRIVADA] Reporte de Tiempos de Respuesta - "
        f"{datetime.now().strftime('%d/%m/%Y')}"
    )
    mail.HTMLBody = f"""
    <html><body style="font-family:Calibri,Arial,sans-serif;color:#2c3e50;margin:20px;">
    <h2 style="color:#1a5276;border-bottom:2px solid #1a5276;padding-bottom:8px;">
      Reporte BANCA PRIVADA - {datetime.now().strftime('%d/%m/%Y %H:%M')}
    </h2>
    <table cellpadding="10" cellspacing="0"
           style="border-collapse:collapse;width:420px;font-size:13px;">
      <tr style="background:#1F3864;color:white;">
        <th style="text-align:left;">Metrica</th>
        <th style="text-align:center;width:80px;">Valor</th>
        <th style="text-align:center;width:60px;">%</th>
      </tr>
      <tr style="background:#EAF2FF;">
        <td>Total correos</td><td align="center"><b>{r['total']}</b></td><td align="center">-</td>
      </tr>
      <tr><td>Con respuesta</td><td align="center">{r['con_resp']}</td><td align="center">-</td></tr>
      <tr style="background:#F0F0F0;">
        <td>Promedio de respuesta</td><td align="center"><b>{prom_str}</b></td><td align="center">-</td>
      </tr>
      <tr style="background:#C6EFCE;">
        <td>Verde (&lt;{LIMITE_VERDE} min)</td>
        <td align="center"><b>{r['verde']}</b></td><td align="center">{r['pct_verde']}%</td>
      </tr>
      <tr style="background:#FFEB9C;">
        <td>Amarillo ({LIMITE_VERDE}-{LIMITE_AMARILLO-1} min)</td>
        <td align="center"><b>{r['amarillo']}</b></td><td align="center">{r['pct_amarillo']}%</td>
      </tr>
      <tr style="background:#FFC7CE;">
        <td>Rojo (&gt;={LIMITE_AMARILLO} min)</td>
        <td align="center"><b>{r['rojo']}</b></td><td align="center">{r['pct_rojo']}%</td>
      </tr>
    </table>
    <br>
    <p style="font-size:11px;color:#7f8c8d;">
      Adjuntos: reporte Excel + grafico de semaforo.<br>
      Generado automaticamente - No responder.
    </p>
    </body></html>
    """
    for archivo in [REPORTE_FILE, CHART_FILE]:
        if os.path.exists(archivo):
            mail.Attachments.Add(os.path.abspath(archivo))
    mail.Send()
    _log(f"✅ Reporte enviado a: {destinatario}", 1)


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    sep = "=" * 60
    print(f"\n{sep}")
    print(f"  BITACORA BANCA PRIVADA  v2")
    print(f"  Inicio : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Desde  : {FECHA_INICIO}")
    print(f"{sep}\n")

    print("[1/6] Conectando a Outlook...")
    try:
        carpeta_bp  = get_folder(MAILBOX_NAME, FOLDER_BANCA)
        carpeta_env = get_folder(MAILBOX_NAME, FOLDER_SENT)
    except ValueError as e:
        print(f"\n ERROR: {e}")
        return

    print("\n[2/6] Leyendo carpeta BANCA PRIVADA...")
    registros = leer_carpeta_banca_privada(carpeta_bp)
    if not registros:
        print("\n No se encontraron correos. Revisa la carpeta y la FECHA_INICIO.")
        return

    print("\n[3/6] Indexando Elementos enviados...")
    indice_env = construir_indice_enviados(carpeta_env)

    print("\n[4/6] Calculando tiempos de respuesta...")
    df = cruzar_y_calcular(registros, indice_env)

    print("\n[5/6] Guardando bitacora incremental...")
    df_final = guardar_incremental(df)

    print("\n[6/6] Generando grafico y reporte Excel...")
    generar_grafico(df_final)
    generar_reporte_excel(df_final)

    if ENVIAR_CORREO and DESTINATARIO_REPORTE:
        print("\n[+] Enviando reporte por correo...")
        try:
            enviar_reporte(df_final, DESTINATARIO_REPORTE)
        except Exception as e:
            _log(f"Error al enviar correo: {e}", 1)
    elif ENVIAR_CORREO and not DESTINATARIO_REPORTE:
        _log("ENVIAR_CORREO=True pero DESTINATARIO_REPORTE esta vacio.", 1)

    print(f"\n{sep}")
    print(f"  PROCESO COMPLETADO EXITOSAMENTE")
    print(f"  Fin: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{sep}\n")


if __name__ == "__main__":
    main()
