# =============================================================================
# ANÁLISIS INTEGRAL MCC 7995 — Juegos de Azar
# Versión final completa con ingeniería de variables
# =============================================================================
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.stats import chi2_contingency
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# =============================================================================
# AJUSTA SOLO ESTE BLOQUE
# =============================================================================
COLS = {
    "FECHA"           : "ACF-FECHA TRX",
    "HORA"            : "ACF-HORA TRX",
    "DATETIME"        : "FECHA_HORA",
    "COMERCIO"        : "ACF-NOMBRE/LOCALIZACION COMERCIO",
    "INDICADOR"       : "ACF-INDICADOR DE FRAUDE",
    "COD_RESPUESTA"   : "ACF-COD RPTA",
    "CONDICION_RT"    : "CONDICION RT",
    "CANAL"           : "ACF-CANAL",
    "ENTRY_MODE"      : "ACF-ENTRY MODE",
    "SALDO"           : "ACF-SALDO DISPONIBLE EN MONEDA TRX",
    "ID_CLIENTE"      : "ACF-ID CLIENTE",
    "ESI_UCAP"        : "ACF-ECI/UCAF",
    "PAIS"            : "ACF-PAIS ORIGEN 87519",
    "COD_RED_COMERCIO": "ACF-COD RED COMERCIO",
    "MONTO"           : "ACF-MONTO EN MONEDA LOCAL",
    "SEGMENTO"        : "VAA-EVENTO DE COMPROMISO OTRA FUENTE",
    "TIPO_PRODUCTO"   : "ACF-TIPO PROD TC",
    "BIN"             : "ACF-BIN",   # si no existe se omite
}
RUTA_PARQUET   = r"C:\ruta\al\consolidado_7995.parquet"
RUTA_EXCEL_OUT = r"C:\ruta\salida\analisis_mcc7995_resultado.xlsx"
# =============================================================================

SEG_NOMBRE = {
    "30":"Polo Direccion","99":"Polo Direccion",
    "31":"Premium","32":"Preferente","33":"Personal","34":"Estandar",
    "5":"Inst. Financieras","21":"Corporativo","2":"Mediano Empresas",
    "15":"Sector Gobierno","16":"Otras Instituciones",
    "3":"Pequenas Empresas","4":"Negocios 2","7":"Negocios 3",
    "8":"Negocios 1","13":"Microempresas",
}
SEG_GRUPO = {
    "30":"Affluent","99":"Affluent",
    "31":"Emerging Affluent","32":"Emerging Affluent",
    "33":"Top of Mass","34":"Mass",
    "5":"Corporate","21":"Corporate","2":"Commercial",
    "15":"Commercial","16":"Commercial",
    "3":"Small Business","4":"Small Business",
    "7":"Small Business","8":"Small Business","13":"Small Business",
}
COD_RED_LABEL = {
    "S":"Estatico (TD)","D":"Dinamico (TD/TC)",
    "E":"Estatico (TC)","N":"No Match / Sin CVV",
}
C_FRAUDE="#D9534F"; C_OK="#5B9BD5"; C_088="#F0AD4E"
C_NEU="#8EA9C1";    C_ROJO2="#8B0000"

# =============================================================================
# CARGA Y PREPARACIÓN BASE
# =============================================================================
print("Cargando parquet...")
df_raw = pd.read_parquet(RUTA_PARQUET)
col_map = {v:k for k,v in COLS.items()}
df = df_raw.rename(columns=col_map).copy()

faltantes = [a for a in COLS if a not in df.columns]
if faltantes:
    print(f"  Columnas no encontradas (se omiten): {faltantes}")

df["MONTO"]    = pd.to_numeric(df["MONTO"],    errors="coerce")
df["SALDO"]    = pd.to_numeric(df["SALDO"],    errors="coerce")
df["DATETIME"] = pd.to_datetime(df["DATETIME"], errors="coerce")
df["FECHA"]    = df["DATETIME"].dt.normalize()
df["MES"]      = df["DATETIME"].dt.to_period("M").astype(str)

for c in ["INDICADOR","COD_RESPUESTA","CONDICION_RT","CANAL",
          "ESI_UCAP","COD_RED_COMERCIO","SEGMENTO","TIPO_PRODUCTO","ENTRY_MODE"]:
    if c in df.columns:
        df[c] = df[c].astype(str).str.strip().str.upper()

if "BIN" in df.columns:
    df["BIN"] = df["BIN"].astype(str).str.split(".").str[0].str.strip()

df["ESTADO"] = df["COD_RESPUESTA"].apply(
    lambda x: "APROBADA" if str(x).strip() in ["00","0000","000","0"] else "DENEGADA"
)
df["ES_FRAUDE"]          = (df["INDICADOR"] == "F").astype(int)
df["ES_FRAUDE_APROBADO"] = (
    (df["INDICADOR"] == "F") & (df["ESTADO"] == "APROBADA")
).astype(int)
df["COND_088"]      = df["CONDICION_RT"].str.contains("088", na=False)
df["SEGURO"]        = df["ESI_UCAP"].apply(
    lambda x: "Seguro" if str(x).strip() in ["2","5"] else "No Seguro"
)
df["SEG_NOMBRE"]    = df["SEGMENTO"].map(SEG_NOMBRE).fillna("Otro/Sin seg")
df["SEG_GRUPO"]     = df["SEGMENTO"].map(SEG_GRUPO).fillna("Otro/Sin seg")
df["COD_RED_LABEL"] = df["COD_RED_COMERCIO"].map(COD_RED_LABEL).fillna("Otro")

df_ap    = df[df["ESTADO"] == "APROBADA"].copy()
df_den   = df[df["ESTADO"] == "DENEGADA"].copy()
df_f_ap  = df[df["ES_FRAUDE_APROBADO"] == 1].copy()
df_f_tot = df[df["ES_FRAUDE"] == 1].copy()

print(f"  Total:           {len(df):,}")
print(f"  Aprobadas:       {len(df_ap):,}")
print(f"  Denegadas:       {len(df_den):,}")
print(f"  Fraude total:    {df['ES_FRAUDE'].sum():,}")
print(f"  Fraude aprobado: {df['ES_FRAUDE_APROBADO'].sum():,}")


# =============================================================================
# INGENIERÍA DE VARIABLES
# =============================================================================
print("\n" + "="*65)
print("INGENIERIA DE VARIABLES — MCC 7995")
print("="*65)

df_feat = df_ap.sort_values(["ID_CLIENTE","DATETIME"]).copy()

# ── FEATURE 1: RATIO MONTO / SALDO ───────────────────────────────────────────
print("  Calculando: Ratio Monto/Saldo...")
df_feat["RATIO_MONTO_SALDO"] = np.where(
    df_feat["SALDO"] > 0,
    df_feat["MONTO"] / df_feat["SALDO"],
    np.nan
)
df_feat["BUCKET_RATIO_SALDO"] = pd.cut(
    df_feat["RATIO_MONTO_SALDO"].clip(0,2),
    bins=[0, 0.1, 0.3, 0.5, 0.8, 2],
    labels=["<10%","10-30%","30-50%","50-80%",">80%"],
    include_lowest=True
)

# ── FEATURE 2: Z-SCORE MONTO POR CLIENTE ─────────────────────────────────────
print("  Calculando: Z-score monto por cliente...")
stats_cli = (
    df_feat.groupby("ID_CLIENTE")["MONTO"]
    .agg(mean_cli="mean", std_cli="std")
    .reset_index()
)
df_feat = df_feat.merge(stats_cli, on="ID_CLIENTE", how="left")
df_feat["ZSCORE_MONTO_CLI"] = np.where(
    df_feat["std_cli"] > 0,
    (df_feat["MONTO"] - df_feat["mean_cli"]) / df_feat["std_cli"],
    0
)
df_feat["BUCKET_ZSCORE"] = pd.cut(
    df_feat["ZSCORE_MONTO_CLI"].clip(-3,3),
    bins=[-3,-2,-1,0,1,2,3],
    labels=["<-2 SD","-2 a -1 SD","-1 a 0 SD","0 a 1 SD","1 a 2 SD",">2 SD"],
    include_lowest=True
)

# ── FEATURE 3: COMERCIO NUEVO (primera vez cliente en ese comercio) ───────────
print("  Calculando: Comercio nuevo...")
df_feat = df_feat.sort_values(["ID_CLIENTE","COMERCIO","DATETIME"])
df_feat["RANK_COMERCIO"] = df_feat.groupby(["ID_CLIENTE","COMERCIO"]).cumcount()
df_feat["ES_COMERCIO_NUEVO"] = (df_feat["RANK_COMERCIO"] == 0).astype(int)

# ── FEATURE 4: N° COMERCIOS DISTINTOS EN EL DÍA ──────────────────────────────
print("  Calculando: Comercios distintos por dia...")
df_feat["FECHA_DIA"] = df_feat["DATETIME"].dt.date
comercios_dia = (
    df_feat.groupby(["ID_CLIENTE","FECHA_DIA"])["COMERCIO"]
    .nunique()
    .reset_index()
    .rename(columns={"COMERCIO":"N_COMERCIOS_DIA"})
)
df_feat = df_feat.merge(comercios_dia, on=["ID_CLIENTE","FECHA_DIA"], how="left")
df_feat["BUCKET_N_COMERCIOS"] = pd.cut(
    df_feat["N_COMERCIOS_DIA"],
    bins=[0,1,2,3,5,100],
    labels=["1 comercio","2 comercios","3 comercios","4-5 comercios",">5 comercios"],
    include_lowest=True
)

# ── FEATURE 5: HORA Y FRANJA HORARIA ─────────────────────────────────────────
print("  Calculando: Hora y franja horaria...")
df_feat["HORA_NUM"] = df_feat["DATETIME"].dt.hour
df_feat["FRANJA_HORARIA"] = pd.cut(
    df_feat["HORA_NUM"],
    bins=[-1,5,11,17,20,23],
    labels=["Madrugada 00-05","Manana 06-11","Tarde 12-17","Noche 18-20","Noche Tardia 21-23"],
    include_lowest=True
)

# ── FEATURE 6: DÍAS DESDE ÚLTIMA TRANSACCIÓN ─────────────────────────────────
print("  Calculando: Dias desde ultima transaccion...")
df_feat = df_feat.sort_values(["ID_CLIENTE","DATETIME"])
df_feat["PREV_DATETIME"] = df_feat.groupby("ID_CLIENTE")["DATETIME"].shift(1)
df_feat["DIAS_DESDE_ULTIMA_TRX"] = (
    df_feat["DATETIME"] - df_feat["PREV_DATETIME"]
).dt.total_seconds() / 86400
df_feat["BUCKET_INACTIVIDAD"] = pd.cut(
    df_feat["DIAS_DESDE_ULTIMA_TRX"].clip(0,60),
    bins=[-0.001,0.083,1,7,14,30,60],
    labels=["<2 horas","1 dia","2-7 dias","8-14 dias","15-30 dias",">30 dias"],
    include_lowest=True
)

# ── FEATURE 7: MONTO ACUMULADO ÚLTIMAS 2 HORAS ───────────────────────────────
print("  Calculando: Monto acumulado ultimas 2 horas...")
df_feat = df_feat.sort_values(["ID_CLIENTE","DATETIME"])

def monto_acum_ventana(grupo, ventana_horas=2):
    resultados = []
    tiempos = grupo["DATETIME"].values
    montos  = grupo["MONTO"].values
    ventana = np.timedelta64(int(ventana_horas*3600), 's')
    for i in range(len(tiempos)):
        mask = (tiempos >= tiempos[i] - ventana) & (tiempos < tiempos[i])
        resultados.append(montos[mask].sum())
    return pd.Series(resultados, index=grupo.index)

df_feat["MONTO_ACUM_2H"] = (
    df_feat.groupby("ID_CLIENTE", group_keys=False)
    .apply(monto_acum_ventana)
)
df_feat["BUCKET_MONTO_ACUM_2H"] = pd.cut(
    df_feat["MONTO_ACUM_2H"].clip(0, df_feat["MONTO_ACUM_2H"].quantile(0.99)),
    bins=5,
    labels=["Q1 Bajo","Q2","Q3","Q4","Q5 Alto"],
    include_lowest=True
)

# ── FEATURE 8: N° RECHAZOS PREVIOS EN 24H ────────────────────────────────────
print("  Calculando: Rechazos previos en 24h...")
df_den_sorted = df_den.sort_values(["ID_CLIENTE","DATETIME"]).copy()

def rechazos_24h(df_ap_grp, df_den_all):
    resultados = []
    for _, row in df_ap_grp.iterrows():
        cli   = row["ID_CLIENTE"]
        t_ref = row["DATETIME"]
        t_min = t_ref - pd.Timedelta(hours=24)
        den_cli = df_den_all[
            (df_den_all["ID_CLIENTE"]==cli) &
            (df_den_all["DATETIME"] >= t_min) &
            (df_den_all["DATETIME"] <  t_ref)
        ]
        resultados.append(len(den_cli))
    return resultados

# Para eficiencia, calculamos solo sobre sample si dataset muy grande
if len(df_feat) <= 200000:
    df_feat["N_RECHAZOS_24H"] = rechazos_24h(df_feat, df_den_sorted)
else:
    print("  (Dataset grande: N_RECHAZOS_24H calculado sobre muestra de 200k)")
    idx_sample = df_feat.index[:200000]
    df_feat["N_RECHAZOS_24H"] = 0
    df_feat.loc[idx_sample, "N_RECHAZOS_24H"] = rechazos_24h(
        df_feat.loc[idx_sample], df_den_sorted
    )

# ── FEATURE 9: RATIO FALLIDOS / (FALLIDOS + EXITOSOS) ────────────────────────
print("  Calculando: Ratio fallidos...")
df_feat["RATIO_FALLIDOS"] = (
    df_feat["N_RECHAZOS_24H"] /
    (df_feat["N_RECHAZOS_24H"] + 1)
)
df_feat["BUCKET_RECHAZOS"] = pd.cut(
    df_feat["N_RECHAZOS_24H"].clip(0,10),
    bins=[-0.001,0,1,2,3,10],
    labels=["0 rechazos","1 rechazo","2 rechazos","3 rechazos","4+ rechazos"],
    include_lowest=True
)

# ── FEATURE 10: FRAUD RATE HISTÓRICO DEL COMERCIO ────────────────────────────
print("  Calculando: FR historico por comercio...")
fr_comercio = (
    df_ap.groupby("COMERCIO")
    .agg(
        total_trx_c  =("MONTO",              "count"),
        fraude_ap_c  =("ES_FRAUDE_APROBADO", "sum"),
    )
    .assign(FR_HISTORICO_COMERCIO=lambda x: x["fraude_ap_c"]/x["total_trx_c"])
    .reset_index()[["COMERCIO","FR_HISTORICO_COMERCIO"]]
)
df_feat = df_feat.merge(fr_comercio, on="COMERCIO", how="left")
df_feat["BUCKET_FR_COMERCIO"] = pd.cut(
    df_feat["FR_HISTORICO_COMERCIO"].clip(0,0.1),
    bins=[-0.001,0.001,0.005,0.01,0.05,0.1],
    labels=["<0.1%","0.1-0.5%","0.5-1%","1-5%",">5%"],
    include_lowest=True
)

# ── FEATURE 11: CONCENTRACIÓN BIN EN COMERCIO ────────────────────────────────
if "BIN" in df_feat.columns:
    print("  Calculando: Concentracion BIN por comercio...")
    bin_comercio = (
        df_ap.groupby(["COMERCIO","BIN"])["MONTO"].count()
        .reset_index()
        .rename(columns={"MONTO":"n_bin"})
    )
    total_comercio = df_ap.groupby("COMERCIO")["MONTO"].count().reset_index().rename(columns={"MONTO":"n_total"})
    bin_top = (
        bin_comercio
        .sort_values("n_bin", ascending=False)
        .groupby("COMERCIO").first()
        .reset_index()
        .rename(columns={"n_bin":"n_top_bin"})
    )
    conc_bin = bin_top.merge(total_comercio, on="COMERCIO")
    conc_bin["CONC_BIN_COMERCIO"] = conc_bin["n_top_bin"] / conc_bin["n_total"]
    df_feat = df_feat.merge(conc_bin[["COMERCIO","CONC_BIN_COMERCIO"]], on="COMERCIO", how="left")
    df_feat["BUCKET_CONC_BIN"] = pd.cut(
        df_feat["CONC_BIN_COMERCIO"].clip(0,1),
        bins=[-0.001,0.25,0.5,0.75,0.9,1.0],
        labels=["<25%","25-50%","50-75%","75-90%",">90%"],
        include_lowest=True
    )
else:
    print("  (BIN no disponible: CONC_BIN_COMERCIO omitido)")
    df_feat["CONC_BIN_COMERCIO"] = np.nan

print("\nFeatures calculadas. Resumen:")
features_num = ["RATIO_MONTO_SALDO","ZSCORE_MONTO_CLI","N_COMERCIOS_DIA",
                "HORA_NUM","DIAS_DESDE_ULTIMA_TRX","MONTO_ACUM_2H",
                "N_RECHAZOS_24H","RATIO_FALLIDOS","FR_HISTORICO_COMERCIO"]
for f in features_num:
    if f in df_feat.columns:
        print(f"  {f:<30} mean={df_feat[f].mean():.4f}  median={df_feat[f].median():.4f}  nulls={df_feat[f].isna().sum():,}")


# =============================================================================
# ANÁLISIS DE FEATURES — Fraude vs No Fraude por cada variable
# =============================================================================
print("\n" + "="*65)
print("ANALISIS DE FEATURES — Fraude Aprobado vs No Fraude")
print("="*65)

def analizar_feature_bucket(df_base, col_bucket, col_fraude="ES_FRAUDE_APROBADO"):
    """Tabla resumen de una variable categorica vs fraude aprobado."""
    return (
        df_base.groupby(col_bucket, observed=True)
        .agg(
            Txns          =(col_fraude,  "count"),
            Fraude_Ap     =(col_fraude,  "sum"),
            FR_pct        =(col_fraude,  lambda x: x.mean()*100),
            Monto_prom    =("MONTO",     "mean"),
            Monto_f_prom  =("MONTO",     lambda x: x[df_base.loc[x.index,col_fraude]==1].mean()),
        )
        .rename(columns={"FR_pct":"FR% Aprobado","Monto_prom":"Ticket prom","Monto_f_prom":"Ticket fraude prom"})
    )

buckets_analizar = [
    ("BUCKET_RATIO_SALDO",   "Ratio Monto/Saldo"),
    ("BUCKET_ZSCORE",        "Z-Score Monto por Cliente"),
    ("ES_COMERCIO_NUEVO",    "Comercio Nuevo (0=recurrente, 1=nuevo)"),
    ("BUCKET_N_COMERCIOS",   "N Comercios Distintos en el Dia"),
    ("FRANJA_HORARIA",       "Franja Horaria"),
    ("BUCKET_INACTIVIDAD",   "Inactividad desde ultima trx"),
    ("BUCKET_MONTO_ACUM_2H", "Monto Acumulado ultimas 2h"),
    ("BUCKET_RECHAZOS",      "N Rechazos previos 24h"),
    ("BUCKET_FR_COMERCIO",   "FR Historico del Comercio"),
]
if "BUCKET_CONC_BIN" in df_feat.columns:
    buckets_analizar.append(("BUCKET_CONC_BIN","Concentracion BIN en Comercio"))

tablas_features = {}
for col_b, titulo in buckets_analizar:
    if col_b not in df_feat.columns:
        continue
    print(f"\n--- {titulo} ---")
    tabla_f = analizar_feature_bucket(df_feat, col_b)
    print(tabla_f.to_string())
    tablas_features[titulo] = tabla_f

# =============================================================================
# SECCIONES 1-7 (igual que versión anterior)
# =============================================================================
print("\n" + "="*65)
print("SECCION 1 — RESUMEN EJECUTIVO POR MES")
print("="*65)

meses_disponibles = sorted(df_ap["MES"].unique())
filas_resumen = []
for mes in meses_disponibles:
    sub_mes  = df[df["MES"]==mes]
    sub_ap   = df_ap[df_ap["MES"]==mes]
    sub_f_ap = sub_ap[sub_ap["ES_FRAUDE_APROBADO"]==1]
    sub_f_tot= sub_mes[sub_mes["ES_FRAUDE"]==1]
    n_ap_m=len(sub_ap); n_den_m=(sub_mes["ESTADO"]=="DENEGADA").sum()
    monto_ap_m=sub_ap["MONTO"].sum(); ticket_ap_m=sub_ap["MONTO"].mean()
    n_cli_m=sub_ap["ID_CLIENTE"].nunique()
    n_f_tot_m=len(sub_f_tot); monto_f_tot_m=sub_f_tot["MONTO"].sum()
    n_f_ap_m=len(sub_f_ap); monto_f_ap_m=sub_f_ap["MONTO"].sum()
    ticket_f_ap_m=sub_f_ap["MONTO"].mean() if n_f_ap_m>0 else 0
    ratio_trx_m=n_f_ap_m/n_ap_m if n_ap_m>0 else 0
    ratio_mon_m=monto_f_ap_m/monto_ap_m if monto_ap_m>0 else 0
    filas_resumen.append({
        "Total trx aprobadas"             :f"{n_ap_m:,}",
        "Total trx denegadas"             :f"{n_den_m:,}",
        "Total monto aprobado (S/)"       :f"S/ {monto_ap_m:,.2f}",
        "Ticket promedio por trx (S/)"    :f"S/ {ticket_ap_m:,.2f}",
        "N clientes unicos"               :f"{n_cli_m:,}",
        "Trxs fraudulentas total (F)"     :f"{n_f_tot_m:,}",
        "Trxs fraudulentas aprobadas"     :f"{n_f_ap_m:,}",
        "Monto fraude total (S/)"         :f"S/ {monto_f_tot_m:,.2f}",
        "Monto fraude aprobado (S/)"      :f"S/ {monto_f_ap_m:,.2f}",
        "Ticket promedio fraude aprobado" :f"S/ {ticket_f_ap_m:,.2f}",
        "Ratio trx fraudulentas aprobadas":f"{ratio_trx_m*100:.4f}%",
        "Ratio fraude en soles (aprobado)":f"{ratio_mon_m*100:.4f}%",
    })

df_resumen_ejecutivo = pd.DataFrame(filas_resumen, index=meses_disponibles).T
df_resumen_ejecutivo.index.name = "Indicador"
print(df_resumen_ejecutivo.to_string())

n_ap=len(df_ap); n_den=len(df_den); monto_ap=df_ap["MONTO"].sum()
n_f_tot_g=df["ES_FRAUDE"].sum(); n_f_ap_g=df["ES_FRAUDE_APROBADO"].sum()
monto_f_ap_g=df_f_ap["MONTO"].sum() if len(df_f_ap)>0 else 0
monto_f_tot_g=df_f_tot["MONTO"].sum() if len(df_f_tot)>0 else 0
ratio_trx_f=n_f_ap_g/n_ap if n_ap>0 else 0
ratio_monto_f=monto_f_ap_g/monto_ap if monto_ap>0 else 0
nivel_riesgo=("ALTO >=5%" if ratio_trx_f>=0.05 else "MEDIO [1%-5%]" if ratio_trx_f>=0.01 else "REGULAR <1%")
riesgo_texto={
    "REGULAR <1%"  :"Fraud rate en nivel REGULAR (<1%). Bajo riesgo relativo.",
    "MEDIO [1%-5%]":"Fraud rate en nivel MEDIO (1%-5%). Monitoreo activo recomendado.",
    "ALTO >=5%"    :"ALERTA: Fraud rate supera el 5%. Intervencion inmediata requerida.",
}
df_umbral=pd.DataFrame({
    "Umbral":["Regular","Medio","Alto"],"Rango":["< 1%","[1%-5%]",">= 5%"],
    "Nivel actual":["AQUI" if ratio_trx_f<0.01 else "","AQUI" if 0.01<=ratio_trx_f<0.05 else "","AQUI" if ratio_trx_f>=0.05 else ""],
})


# =============================================================================
# SECCIÓN 2 — TABLA DE DECILES (misma estructura en principal Y apertura D10)
# =============================================================================
print("\n" + "="*65)
print("SECCION 2 — TABLA DE DECILES")
print("="*65)

def construir_tabla_deciles(df_base, bins_custom=None, n_deciles=10, label="D"):
    df_base = df_base.copy().dropna(subset=["MONTO"])
    if bins_custom is not None:
        labels = [f"{label}{str(i+1).zfill(2)}" for i in range(len(bins_custom)-1)]
        df_base["DECIL"] = pd.cut(df_base["MONTO"], bins=bins_custom,
                                   labels=labels, include_lowest=True)
        df_base["DECIL"] = pd.Categorical(df_base["DECIL"], categories=labels, ordered=True)
    else:
        labels = [f"{label}{str(i+1).zfill(2)}" for i in range(n_deciles)]
        df_base["DECIL"] = pd.qcut(df_base["MONTO"], q=n_deciles,
                                    labels=labels, duplicates="drop")
    rows=[]; fraude_ap_acum=0; trx_ap_acum=0
    total_ap=len(df_base[df_base["ESTADO"]=="APROBADA"])
    total_f_ap=df_base["ES_FRAUDE_APROBADO"].sum()

    for decil in df_base["DECIL"].cat.categories:
        sub=df_base[df_base["DECIL"]==decil]
        if len(sub)==0: continue
        sub_ap=sub[sub["ESTADO"]=="APROBADA"]; sub_den=sub[sub["ESTADO"]=="DENEGADA"]
        sub_f_ap=sub[sub["ES_FRAUDE_APROBADO"]==1]; sub_f_tot=sub[sub["ES_FRAUDE"]==1]
        n_ap_d=len(sub_ap); m_ap_d=sub_ap["MONTO"].sum()
        n_den_d=len(sub_den); m_den_d=sub_den["MONTO"].sum()
        n_f_ap_d=len(sub_f_ap); m_f_ap_d=sub_f_ap["MONTO"].sum()
        n_f_tot_d=len(sub_f_tot); m_f_tot_d=sub_f_tot["MONTO"].sum()
        trx_ap_acum+=n_ap_d; fraude_ap_acum+=n_f_ap_d
        freq_ap=n_ap_d/total_ap if total_ap>0 else 0
        freq_ap_acum=trx_ap_acum/total_ap if total_ap>0 else 0
        freq_f_ap=n_f_ap_d/total_f_ap if total_f_ap>0 else 0
        freq_f_acum=fraude_ap_acum/total_f_ap if total_f_ap>0 else 0
        rows.append({
            "Decil"                  :decil,
            "Rango S/"               :f"{sub['MONTO'].min():.0f}-{sub['MONTO'].max():.0f}",
            "Trx Aprobadas"          :n_ap_d,
            "Monto Aprobadas"        :round(m_ap_d,1),
            "Freq Aprobadas"         :f"{freq_ap*100:.2f}%",
            "Freq Acum Aprobadas"    :f"{freq_ap_acum*100:.2f}%",
            "Trx Denegadas"          :n_den_d,
            "Monto Denegadas"        :round(m_den_d,1),
            "Total Trx"              :len(sub),
            "Ticket Promedio"        :round(sub["MONTO"].mean(),2),
            "Monto Total S/"         :round(sub["MONTO"].sum(),1),
            "Fraude Trx Total"       :n_f_tot_d,
            "Fraude Monto Total"     :round(m_f_tot_d,1),
            "Fraude Trx Aprobado"    :n_f_ap_d,
            "Fraude Monto Aprobado"  :round(m_f_ap_d,1),
            "Fraude Freq Aprobado"   :f"{freq_f_ap*100:.4f}%",
            "Fraude Freq Acumulada"  :f"{freq_f_acum*100:.4f}%",
            "Fraude/Trx Aprobadas"   :f"{n_f_ap_d/n_ap_d*100:.4f}%" if n_ap_d>0 else "0.0000%",
            "No Fraude/Trx Aprobadas":f"{(1-n_f_ap_d/n_ap_d)*100:.4f}%" if n_ap_d>0 else "100.0000%",
        })
    return pd.DataFrame(rows).set_index("Decil")

tabla_deciles = construir_tabla_deciles(df, n_deciles=10)
print(tabla_deciles.to_string())

# Apertura D10 — mismas columnas que tabla principal
print("\nApertura Decil 10:")
monto_d10_min = df["MONTO"].quantile(0.90)
df_d10 = df[df["MONTO"] >= monto_d10_min].copy()
bins_d10 = np.unique([df_d10["MONTO"].quantile(p/100) for p in np.linspace(0,100,11)])
if len(bins_d10) < 3:
    bins_d10 = np.unique([df_d10["MONTO"].quantile(p/100) for p in np.linspace(0,100,6)])
tabla_d10 = construir_tabla_deciles(df_d10, bins_custom=bins_d10, label="SD")
print(tabla_d10.to_string())

# Grafico deciles
deciles_idx=tabla_deciles.index.tolist()
n_f_ap_d_list=tabla_deciles["Fraude Trx Aprobado"].tolist()
n_f_tot_d_list=tabla_deciles["Fraude Trx Total"].tolist()
fr_ap_d=[tabla_deciles.loc[d,"Fraude Trx Aprobado"]/tabla_deciles.loc[d,"Trx Aprobadas"]*100
         if tabla_deciles.loc[d,"Trx Aprobadas"]>0 else 0 for d in deciles_idx]
fig=make_subplots(rows=1,cols=2,subplot_titles=["Fraudes por decil","FR% aprobado"])
fig.add_trace(go.Bar(x=deciles_idx,y=n_f_tot_d_list,name="Fraude Total",marker_color=C_ROJO2,opacity=0.6,text=[f"{v:,}" for v in n_f_tot_d_list],textposition="outside"),row=1,col=1)
fig.add_trace(go.Bar(x=deciles_idx,y=n_f_ap_d_list,name="Fraude Aprobado",marker_color=C_FRAUDE,text=[f"{v:,}" for v in n_f_ap_d_list],textposition="outside"),row=1,col=1)
fig.add_trace(go.Bar(x=deciles_idx,y=fr_ap_d,name="FR%",marker_color=C_088,text=[f"{v:.4f}%" for v in fr_ap_d],textposition="outside"),row=1,col=2)
fig.update_layout(title_text="MCC 7995 — Deciles",barmode="group",height=450,plot_bgcolor="white")
fig.write_html("g1_deciles.html"); fig.show()


# =============================================================================
# SECCIONES 3-7 (Velocidad, Diaria, 088, Dimensiones, Estadísticas)
# =============================================================================

# ── VELOCIDAD INTERVALO ───────────────────────────────────────────────────────
print("\n" + "="*65); print("SECCION 3 — VELOCIDAD POR INTERVALO")
df_vel=df.sort_values(["ID_CLIENTE","DATETIME"]).copy()
df_vel["PREV_DT"]=df_vel.groupby("ID_CLIENTE")["DATETIME"].shift(1)
df_vel["MIN_DESDE_PREV"]=(df_vel["DATETIME"]-df_vel["PREV_DT"]).dt.total_seconds()/60
df_vel2=df_vel.dropna(subset=["MIN_DESDE_PREV"]).copy()
df_vel2=df_vel2[df_vel2["MIN_DESDE_PREV"]>=0]
def bucket_int(m):
    if m<=2:return "<=2 min"
    elif m<=5:return "<=5 min"
    elif m<=10:return "<=10 min"
    elif m<=15:return "<=15 min"
    elif m<=20:return "<=20 min"
    elif m<=60:return "<=1 hora"
    else:return ">1 hora"
ORDEN_B=["<=2 min","<=5 min","<=10 min","<=15 min","<=20 min","<=1 hora",">1 hora"]
df_vel2["BUCKET"]=pd.Categorical(df_vel2["MIN_DESDE_PREV"].apply(bucket_int),categories=ORDEN_B,ordered=True)
tabla_vel=(
    df_vel2.groupby("BUCKET",observed=True)
    .agg(
        Transacciones        =("MONTO","count"),
        Genuinas             =("ES_FRAUDE",lambda x:(x==0).sum()),
        Fraude_Total         =("ES_FRAUDE","sum"),
        Fraude_Aprobado      =("ES_FRAUDE_APROBADO","sum"),
        FR_Total_pct         =("ES_FRAUDE",lambda x:x.mean()*100),
        FR_Aprobado_pct      =("ES_FRAUDE_APROBADO",lambda x:x.mean()*100),
        Monto_Genuino        =("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].sum()),
        Monto_Fraude_Total   =("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_Fraude_Aprobado=("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        Ticket_Genuino_Prom  =("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].mean()),
        Ticket_Fraude_Prom   =("MONTO",lambda x:x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].mean()),
    )
    .assign(Pct_fraudes_total=lambda x:x["Fraude_Total"]/x["Fraude_Total"].sum()*100)
)
print(tabla_vel.to_string())

# ── TRANSACCIONALIDAD DIARIA ──────────────────────────────────────────────────
print("\n" + "="*65); print("SECCION 4 — TRANSACCIONALIDAD DIARIA")
dia_cliente=(
    df_ap.assign(FECHA_DIA=lambda x:x["DATETIME"].dt.date)
    .groupby(["ID_CLIENTE","FECHA_DIA"])
    .agg(
        Trx_dia             =("MONTO","count"),
        Monto_dia           =("MONTO","sum"),
        Fraudes_tot_dia     =("ES_FRAUDE","sum"),
        Fraudes_ap_dia      =("ES_FRAUDE_APROBADO","sum"),
        Monto_fraude_tot_dia=("MONTO",lambda x:x[df_ap.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_fraude_ap_dia =("MONTO",lambda x:x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
    )
    .reset_index()
    .assign(
        tiene_fraude =lambda x:(x["Fraudes_ap_dia"]>0).astype(int),
        Trx_genuina  =lambda x:x["Trx_dia"]-x["Fraudes_tot_dia"],
        Monto_genuino=lambda x:x["Monto_dia"]-x["Monto_fraude_tot_dia"],
    )
)
def buck_dia(n):
    if n<=6:return str(n)
    elif n<=9:return "7-9"
    elif n<=12:return "10-12"
    else:return "13+"
ORDEN_DIA=["1","2","3","4","5","6","7-9","10-12","13+"]
dia_cliente["BUCKET_DIA"]=pd.Categorical(dia_cliente["Trx_dia"].apply(buck_dia),categories=ORDEN_DIA,ordered=True)
tabla_diaria=(
    dia_cliente.groupby("BUCKET_DIA",observed=True)
    .agg(
        Genuina              =("tiene_fraude",lambda x:(x==0).sum()),
        Fraude_Aprobado      =("tiene_fraude","sum"),
        Total_Clientes       =("tiene_fraude","count"),
        Trx_Genuina          =("Trx_genuina","sum"),
        Monto_Genuina        =("Monto_genuino","sum"),
        Trx_Fraude_Total     =("Fraudes_tot_dia","sum"),
        Trx_Fraude_Aprobado  =("Fraudes_ap_dia","sum"),
        Monto_Fraude_Total   =("Monto_fraude_tot_dia","sum"),
        Monto_Fraude_Aprobado=("Monto_fraude_ap_dia","sum"),
    )
    .reset_index().rename(columns={"BUCKET_DIA":"Intervalo"})
    .assign(
        Impacto_fraude_pct =lambda x:x["Fraude_Aprobado"]/x["Total_Clientes"]*100,
        Ticket_Fraude_Prom =lambda x:np.where(x["Trx_Fraude_Aprobado"]>0,x["Monto_Fraude_Aprobado"]/x["Trx_Fraude_Aprobado"],0),
        Ticket_Genuina_Prom=lambda x:np.where(x["Trx_Genuina"]>0,x["Monto_Genuina"]/x["Trx_Genuina"],0),
    )
)
total_cli_fraude=tabla_diaria["Fraude_Aprobado"].sum()
pct_1trx=tabla_diaria.loc[tabla_diaria["Intervalo"]=="1","Fraude_Aprobado"].sum()/total_cli_fraude*100 if total_cli_fraude>0 else 0
td=tabla_diaria.copy()
td["Impacto_fraude_pct"]=td["Impacto_fraude_pct"].map("{:.4f}%".format)
td["Monto_Genuina"]=td["Monto_Genuina"].map("S/ {:,.2f}".format)
td["Monto_Fraude_Total"]=td["Monto_Fraude_Total"].map("S/ {:,.2f}".format)
td["Monto_Fraude_Aprobado"]=td["Monto_Fraude_Aprobado"].map("S/ {:,.2f}".format)
td["Ticket_Fraude_Prom"]=td["Ticket_Fraude_Prom"].map("S/ {:,.2f}".format)
td["Ticket_Genuina_Prom"]=td["Ticket_Genuina_Prom"].map("S/ {:,.2f}".format)
tabla_diaria_display=td.rename(columns={"Impacto_fraude_pct":"Impacto Fraude %","Ticket_Fraude_Prom":"Ticket Fraude Prom","Ticket_Genuina_Prom":"Ticket Genuina Prom","Fraude_Aprobado":"Fraude Aprobado","Total_Clientes":"Total Clientes","Trx_Genuina":"Trx Genuina","Monto_Genuina":"Monto Genuina","Trx_Fraude_Total":"Trx Fraude Total","Trx_Fraude_Aprobado":"Trx Fraude Aprobado","Monto_Fraude_Total":"Monto Fraude Total","Monto_Fraude_Aprobado":"Monto Fraude Aprobado"})
print(tabla_diaria_display.to_string(index=False))

# ── CONDICIÓN 088 ─────────────────────────────────────────────────────────────
print("\n" + "="*65); print("SECCION 5 — CONDICION 088")
f_en_088_tot=(df["COND_088"]&(df["ES_FRAUDE"]==1)).sum()
f_fuera_088_tot=(~df["COND_088"]&(df["ES_FRAUDE"]==1)).sum()
f_en_088_ap=(df["COND_088"]&(df["ES_FRAUDE_APROBADO"]==1)).sum()
f_fuera_088_ap=(~df["COND_088"]&(df["ES_FRAUDE_APROBADO"]==1)).sum()
print(f"  FRAUDE TOTAL:    dentro={f_en_088_tot:,} ({f_en_088_tot/n_f_tot_g*100:.4f}%)  fuera={f_fuera_088_tot:,}" if n_f_tot_g>0 else "")
print(f"  FRAUDE APROBADO: dentro={f_en_088_ap:,} ({f_en_088_ap/n_f_ap_g*100:.4f}%)   fuera={f_fuera_088_ap:,}" if n_f_ap_g>0 else "")

# ── DIMENSIONES ───────────────────────────────────────────────────────────────
print("\n" + "="*65); print("SECCION 6 — DIMENSIONES")
def resumen_por(df_base,col,top_n=10):
    return (
        df_base.groupby(col)
        .agg(Txns=(col,"count"),Fraude_Total=("ES_FRAUDE","sum"),Fraude_Aprobado=("ES_FRAUDE_APROBADO","sum"),
             FR_Total_pct=("ES_FRAUDE",lambda x:x.mean()*100),FR_Aprobado_pct=("ES_FRAUDE_APROBADO",lambda x:x.mean()*100),
             Monto_total=("MONTO","sum"),Monto_Fraude_Ap=("MONTO",lambda x:x[df_base.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()))
        .sort_values("Fraude_Aprobado",ascending=False).head(top_n)
        .rename(columns={"FR_Total_pct":"FR% Total","FR_Aprobado_pct":"FR% Aprobado","Monto_Fraude_Ap":"Monto Fraude Ap"})
    )
df_por_canal=resumen_por(df,"CANAL"); df_por_seg_nombre=resumen_por(df,"SEG_NOMBRE")
df_por_seg_grupo=resumen_por(df,"SEG_GRUPO"); df_por_seguro=resumen_por(df,"SEGURO")
df_por_tipo=resumen_por(df,"TIPO_PRODUCTO"); df_por_cvv=resumen_por(df,"COD_RED_LABEL")
df_por_pais=resumen_por(df,"PAIS")

# Top Comercios desde df completo para fraude total
df_top_comercios=(
    df.groupby("COMERCIO")
    .agg(Fraude_Total_g=("ES_FRAUDE","sum"),Monto_Fraude_Tot_g=("MONTO",lambda x:x[df.loc[x.index,"ES_FRAUDE"]==1].sum()))
    .join(df_ap.groupby("COMERCIO").agg(Tarjetas=("ID_CLIENTE","nunique"),Trx=("MONTO","count"),
        Fraude_Aprobado=("ES_FRAUDE_APROBADO","sum"),Monto_Fraude_Ap=("MONTO",lambda x:x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        Monto_total=("MONTO","sum"),FR_Ap_pct=("ES_FRAUDE_APROBADO",lambda x:x.mean()*100)))
    .reset_index().sort_values("Monto_Fraude_Ap",ascending=False).head(10)
    .rename(columns={"COMERCIO":"Comercio","Tarjetas":"# Tarjetas","Trx":"# Trx Ap",
        "Fraude_Total_g":"# Fraude Total","Fraude_Aprobado":"# Fraude Ap",
        "Monto_Fraude_Tot_g":"Monto F Total S/","Monto_Fraude_Ap":"Monto F Ap S/",
        "Monto_total":"Monto Total S/","FR_Ap_pct":"FR% (ap)"})
)
df_top_comercios["FR% (ap)"]=df_top_comercios["FR% (ap)"].map("{:.4f}%".format)

# ── ESTADÍSTICAS ──────────────────────────────────────────────────────────────
print("\n" + "="*65); print("SECCION 7 — ESTADISTICAS DE MONTO")
pcts=[10,25,50,75,90,95,99]; df_nofraude=df_ap[df_ap["ES_FRAUDE"]==0]
def stats_s(s):
    if len(s)==0:return [0]*13
    return [s.mean(),s.median(),s.std(),s.var(),s.min(),s.max()]+[s.quantile(p/100) for p in pcts]
stats_monto=pd.DataFrame({
    "Metrica":["Media","Mediana","Desv. Std","Varianza","Minimo","Maximo"]+[f"P{p}" for p in pcts],
    "Total aprobadas":stats_s(df_ap["MONTO"]),"Fraude Total":stats_s(df_f_tot["MONTO"]),
    "Fraude Aprobado":stats_s(df_f_ap["MONTO"]),"No Fraude":stats_s(df_nofraude["MONTO"]),
})
for c in ["Total aprobadas","Fraude Total","Fraude Aprobado","No Fraude"]:
    stats_monto[c]=stats_monto[c].apply(lambda v:f"S/ {v:,.2f}")
print(stats_monto.to_string(index=False))

if len(df_f_ap)>0:
    media_f=df_f_ap["MONTO"].mean(); median_f=df_f_ap["MONTO"].median()
    std_f=df_f_ap["MONTO"].std();    median_n=df_nofraude["MONTO"].median()
    ratio_mm=media_f/median_f if median_f>0 else 0
    p90_f=df_f_ap["MONTO"].quantile(0.90); p90_n=df_nofraude["MONTO"].quantile(0.90)
    print(f"\n  SESGO: Media S/ {media_f:,.2f} vs mediana S/ {median_f:,.2f} (ratio {ratio_mm:.2f}x). "
          +("Alta presencia de montos extremos." if ratio_mm>2 else "Sesgo moderado." if ratio_mm>1.5 else "Distribucion simetrica."))
    print(f"  COMPARACION: Mediana fraude S/ {median_f:,.2f} vs legitimo S/ {median_n:,.2f}. "
          +("Fraude con montos MAYORES." if median_f>median_n else "Fraude con montos MENORES — posible testeo de tarjetas."))
    print(f"  P90: Fraude S/ {p90_f:,.2f} vs Legitimo S/ {p90_n:,.2f}. "
          +("Umbral efectivo en este rango." if p90_f>p90_n else "P90 fraude no supera al legitimo."))


# =============================================================================
# SECCIÓN 8 — TABLAS CRUZADAS (incluye velocidad × segmento y features nuevas)
# =============================================================================
print("\n" + "="*65); print("SECCION 8 — TABLAS CRUZADAS")

def tabla_cruzada_completa(df_base, dim_filas, dim_cols, titulo):
    print(f"\n--- {titulo} ---")
    try:
        piv_cnt=df_base.pivot_table(values="ES_FRAUDE_APROBADO",index=dim_filas,columns=dim_cols,aggfunc="sum",fill_value=0)
        piv_fr=df_base.pivot_table(values="ES_FRAUDE_APROBADO",index=dim_filas,columns=dim_cols,aggfunc="mean",fill_value=0).map(lambda v:f"{v*100:.4f}%")
        piv_monto=df_base[df_base["ES_FRAUDE_APROBADO"]==1].pivot_table(values="MONTO",index=dim_filas,columns=dim_cols,aggfunc="sum",fill_value=0).map(lambda v:f"S/ {v:,.2f}")
        print("Fraudes Aprobados:"); print(piv_cnt.to_string())
        print("FR% Aprobado:");      print(piv_fr.to_string())
        print("Monto Fraude S/:");   print(piv_monto.to_string())
        chi2,p_val,dof,_=chi2_contingency(piv_cnt.values)
        sig="SIGNIFICATIVA" if p_val<0.05 else "No significativa"
        print(f"Chi2={chi2:.4f} p={p_val:.6f} gl={dof} | {sig}")
        return piv_cnt, piv_fr, piv_monto
    except Exception as e:
        print(f"  Error en cruce {dim_filas} x {dim_cols}: {e}")
        return None, None, None

# Cruces base
cruces_base = [
    ("TIPO_PRODUCTO","SEGURO",        "Tipo Producto x Seguro/No Seguro"),
    ("SEG_NOMBRE",   "SEGURO",        "Segmento Individual x Seguro/No Seguro"),
    ("SEG_GRUPO",    "SEGURO",        "Segmento Grupo x Seguro/No Seguro"),
    ("COD_RED_LABEL","SEGURO",        "CVV/Red x Seguro/No Seguro"),
    ("CANAL",        "SEGURO",        "Canal x Seguro/No Seguro"),
    ("TIPO_PRODUCTO","COD_RED_LABEL", "Tipo Producto x CVV/Red"),
]
# Agrega aqui nuevos cruces base:
# ("COLUMNA_FILAS","COLUMNA_COLS","Titulo"),

resultados_cruce_base = {}
for dim_f,dim_c,titulo in cruces_base:
    piv_c,piv_f,piv_m = tabla_cruzada_completa(df_ap, dim_f, dim_c, titulo)
    resultados_cruce_base[titulo] = (piv_c, piv_f, piv_m)

# Cruces velocidad × segmento/tipo producto
print("\n=== CRUCE VELOCIDAD BUCKET x DIMENSIONES ===")
df_vel2_enr = df_vel2.copy()
seg_map = df.groupby("ID_CLIENTE")[["SEG_NOMBRE","SEG_GRUPO","TIPO_PRODUCTO"]].first()
df_vel2_enr = df_vel2_enr.join(seg_map, on="ID_CLIENTE", rsuffix="_v")

cruces_velocidad = [
    ("BUCKET","SEG_NOMBRE",    "Velocidad Intervalo x Segmento Individual"),
    ("BUCKET","SEG_GRUPO",     "Velocidad Intervalo x Segmento Grupo"),
    ("BUCKET","TIPO_PRODUCTO", "Velocidad Intervalo x Tipo Producto"),
]
resultados_cruce_vel = {}
for dim_f,dim_c,titulo in cruces_velocidad:
    if dim_c not in df_vel2_enr.columns:
        print(f"  {dim_c} no disponible en df_vel2_enr, omitiendo")
        continue
    piv_c,piv_f,piv_m = tabla_cruzada_completa(df_vel2_enr, dim_f, dim_c, titulo)
    resultados_cruce_vel[titulo] = (piv_c, piv_f, piv_m)

# Cruces con features nuevas
print("\n=== CRUCE FEATURES NUEVAS x FRAUDE ===")
cruces_features = [
    ("BUCKET_RATIO_SALDO",   "TIPO_PRODUCTO","Ratio Monto/Saldo x Tipo Producto"),
    ("BUCKET_ZSCORE",        "SEGURO",       "Z-Score Monto x Seguro/No Seguro"),
    ("ES_COMERCIO_NUEVO",    "TIPO_PRODUCTO","Comercio Nuevo x Tipo Producto"),
    ("BUCKET_N_COMERCIOS",   "SEGURO",       "N Comercios Dia x Seguro/No Seguro"),
    ("FRANJA_HORARIA",       "TIPO_PRODUCTO","Franja Horaria x Tipo Producto"),
    ("BUCKET_INACTIVIDAD",   "SEGURO",       "Inactividad x Seguro/No Seguro"),
    ("BUCKET_RECHAZOS",      "TIPO_PRODUCTO","Rechazos 24h x Tipo Producto"),
    ("BUCKET_FR_COMERCIO",   "SEGURO",       "FR Historico Comercio x Seguro/No Seguro"),
]
if "BUCKET_CONC_BIN" in df_feat.columns:
    cruces_features.append(("BUCKET_CONC_BIN","TIPO_PRODUCTO","Concentracion BIN x Tipo Producto"))

resultados_cruce_feat = {}
for dim_f,dim_c,titulo in cruces_features:
    if dim_f not in df_feat.columns:
        continue
    piv_c,piv_f,piv_m = tabla_cruzada_completa(df_feat, dim_f, dim_c, titulo)
    resultados_cruce_feat[titulo] = (piv_c, piv_f, piv_m)

# =============================================================================
# SECCIÓN 9 — CLIENTES ÚNICOS Y RECURRENCIA
# =============================================================================
print("\n" + "="*65); print("SECCION 9 — CLIENTES RECURRENTES")
cli_fraude_por_mes={mes:set(df_ap[(df_ap["MES"]==mes)&(df_ap["ES_FRAUDE_APROBADO"]==1)]["ID_CLIENTE"]) for mes in meses_disponibles}
total_cli_global=df_ap["ID_CLIENTE"].nunique()
clientes_por_mes=df_ap.groupby("MES")["ID_CLIENTE"].nunique()
print(f"Clientes globales: {total_cli_global:,}  |  Suma por mes: {clientes_por_mes.sum():,}")
df_recurrencia_rows=[]
if len(meses_disponibles)>=2:
    for i in range(len(meses_disponibles)-1):
        ma=meses_disponibles[i]; mb=meses_disponibles[i+1]
        cli_a=cli_fraude_por_mes.get(ma,set()); cli_b=cli_fraude_por_mes.get(mb,set())
        rec=cli_a&cli_b; nuevos=cli_b-cli_a
        pct_r=len(rec)/len(cli_b)*100 if len(cli_b)>0 else 0
        print(f"  {ma}->{mb}: fraude_A={len(cli_a):,} fraude_B={len(cli_b):,} recurrentes={len(rec):,} ({pct_r:.4f}%) nuevos={len(nuevos):,}")
        df_recurrencia_rows.append({"Periodo":f"{ma}->{mb}","Fraude MesA":len(cli_a),"Fraude MesB":len(cli_b),"Recurrentes":len(rec),"% Recurrentes":f"{pct_r:.4f}%","Nuevos en MesB":len(nuevos)})
df_recurrencia_tabla=pd.DataFrame(df_recurrencia_rows)
df_clientes_mes=pd.DataFrame({"Mes":meses_disponibles,"Clientes unicos":[df_ap[df_ap["MES"]==m]["ID_CLIENTE"].nunique() for m in meses_disponibles],"Clientes con fraude":[len(cli_fraude_por_mes.get(m,set())) for m in meses_disponibles]})
df_clientes_mes["% con fraude"]=(df_clientes_mes["Clientes con fraude"]/df_clientes_mes["Clientes unicos"]*100).map("{:.4f}%".format)


# =============================================================================
# EXPORTACIÓN A EXCEL — TODAS LAS HOJAS
# =============================================================================
print("\nExportando a Excel...")
FILL_HEADER=PatternFill("solid",fgColor="1F3864"); FILL_SUBHEAD=PatternFill("solid",fgColor="2E75B6")
FILL_FILA_A=PatternFill("solid",fgColor="DEEAF1"); FILL_AMARILLO=PatternFill("solid",fgColor="FFF2CC")
FILL_FRAUDE=PatternFill("solid",fgColor="FCE4D6"); FILL_VERDE=PatternFill("solid",fgColor="E2EFDA")
FONT_HEADER=Font(color="FFFFFF",bold=True,size=10); FONT_NORMAL=Font(size=10)
FONT_BOLD=Font(bold=True,size=10); FONT_INTERP=Font(italic=True,size=9,color="1F3864")
BORDER_THIN=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
ALIGN_CENTER=Alignment(horizontal="center",vertical="center",wrap_text=True)
ALIGN_LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)

def eh(ws,fila,nc,texto,fill=None):
    fill=fill or FILL_HEADER
    ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=nc)
    c=ws.cell(row=fila,column=1,value=texto)
    c.fill=fill; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def eed(ws,fila):
    for row in ws.iter_rows(min_row=fila,max_row=fila):
        for c in row: c.fill=FILL_SUBHEAD; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def ed(ws,fi,ff,kw=None):
    for i,row in enumerate(ws.iter_rows(min_row=fi,max_row=ff),start=1):
        fill=FILL_FILA_A if i%2==0 else PatternFill()
        if kw and row[0].value and any(k in str(row[0].value).lower() for k in kw): fill=FILL_FRAUDE
        for c in row: c.fill=fill; c.font=FONT_NORMAL; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def ai(ws,fila,nc,texto):
    ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=nc)
    c=ws.cell(row=fila,column=1,value=f"-> {texto}")
    c.fill=FILL_AMARILLO; c.font=FONT_INTERP; c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
    ws.row_dimensions[fila].height=28

def ac(ws):
    for col in ws.columns:
        ml=max((len(str(c.value)) for c in col if c.value),default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(ml+4,35)

def escribir_pivot_en_hoja(ws, piv, fila_inicio, titulo, n_cols_ws):
    """Escribe una tabla pivot en una hoja existente."""
    eh(ws, fila_inicio, n_cols_ws, titulo, fill=FILL_SUBHEAD)
    fila_inicio += 1
    if piv is None: return fila_inicio + 2
    df_out = piv.reset_index()
    for j,col in enumerate(df_out.columns,start=1):
        c=ws.cell(row=fila_inicio,column=j,value=str(col))
        c.fill=FILL_SUBHEAD; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
    fila_inicio += 1
    for i,row in df_out.iterrows():
        fill=FILL_FILA_A if i%2==0 else PatternFill()
        for j,val in enumerate(row,start=1):
            c=ws.cell(row=fila_inicio,column=j,value=val if not isinstance(val,float) else round(val,4))
            c.fill=fill; c.font=FONT_NORMAL; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
        fila_inicio += 1
    return fila_inicio + 2

with pd.ExcelWriter(RUTA_EXCEL_OUT, engine="openpyxl") as writer:

    # HOJA 1: RESUMEN EJECUTIVO
    sh="1_Resumen_Ejecutivo"; nc=len(df_resumen_ejecutivo.columns)+1
    df_resumen_ejecutivo.to_excel(writer,sheet_name=sh,startrow=3)
    ws=writer.sheets[sh]
    eh(ws,1,nc,f"MCC 7995 — JUEGOS DE AZAR | {datetime.today().strftime('%d/%m/%Y')}")
    eh(ws,2,nc,"RESUMEN DE INDICADORES DE FRAUDE POR MES",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row,kw=["fraude","ratio"])
    fu=ws.max_row+2; eh(ws,fu,nc,"UMBRAL DE CONTROL",fill=FILL_SUBHEAD)
    df_umbral.to_excel(writer,sheet_name=sh,index=False,startrow=fu)
    fi=ws.max_row+2
    ai(ws,fi,nc,f"Nivel de riesgo: {nivel_riesgo}. {riesgo_texto[nivel_riesgo]}")
    ai(ws,fi+1,nc,f"Fraude total: {n_f_tot_g:,} | Fraude aprobado: {n_f_ap_g:,} ({n_f_ap_g/n_f_tot_g*100:.2f}% del total)." if n_f_tot_g>0 else "Sin fraudes.")
    ac(ws)

    # HOJA 2: DECILES
    sh="2_Deciles"; nc2=len(tabla_deciles.columns)+2
    tabla_deciles.to_excel(writer,sheet_name=sh,startrow=3)
    ws=writer.sheets[sh]
    eh(ws,1,nc2,"MCC 7995 — TABLA DE DECILES POR MONTO")
    eh(ws,2,nc2,"D01=menor monto D10=mayor | Fraude Total=ap+den | Fraude Aprobado=solo aprobadas",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    d_max_f=tabla_deciles["Fraude Trx Aprobado"].idxmax()
    d_max_m=tabla_deciles["Fraude Monto Aprobado"].idxmax()
    fi=ws.max_row+2
    ai(ws,fi,nc2,f"Mayor concentracion fraudes aprobados: {d_max_f} ({tabla_deciles.loc[d_max_f,'Fraude Trx Aprobado']:,}) Rango: {tabla_deciles.loc[d_max_f,'Rango S/']}")
    ai(ws,fi+1,nc2,f"Mayor monto fraude aprobado: {d_max_m} (S/ {tabla_deciles.loc[d_max_m,'Fraude Monto Aprobado']:,.2f}) Rango: {tabla_deciles.loc[d_max_m,'Rango S/']}")
    ac(ws)

    # HOJA 2B: APERTURA D10 (mismas columnas que hoja 2)
    sh="2B_Apertura_D10"; nc2b=len(tabla_d10.columns)+2
    tabla_d10.to_excel(writer,sheet_name=sh,startrow=3)
    ws=writer.sheets[sh]
    eh(ws,1,nc2b,f"MCC 7995 — APERTURA DECIL 10 (montos >= S/ {monto_d10_min:,.2f})")
    eh(ws,2,nc2b,"Misma estructura que tabla principal — cortes fijos por percentil dentro del D10",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    fi=ws.max_row+2
    if "Fraude Trx Aprobado" in tabla_d10.columns and tabla_d10["Fraude Trx Aprobado"].sum()>0:
        sd_max=tabla_d10["Fraude Trx Aprobado"].idxmax()
        ai(ws,fi,nc2b,f"Sub-decil con mas fraudes aprobados: {sd_max} ({tabla_d10.loc[sd_max,'Fraude Trx Aprobado']:,}) Rango: {tabla_d10.loc[sd_max,'Rango S/']}")
    ac(ws)

    # HOJA 3: TOP COMERCIOS
    sh="3_Top_Comercios"
    df_tc=df_top_comercios.copy()
    for col_fmt in ["Monto F Total S/","Monto F Ap S/","Monto Total S/"]:
        if col_fmt in df_tc.columns:
            df_tc[col_fmt]=df_tc[col_fmt].map("S/ {:,.2f}".format)
    df_tc.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws=writer.sheets[sh]
    nc3=len(df_tc.columns)
    eh(ws,1,nc3,"MCC 7995 — TOP 10 COMERCIOS (Fraude Total=ap+den | Fraude Ap=solo aprobadas)")
    eh(ws,2,nc3,"Ordenado por Monto Fraude Aprobado",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    top1=df_tc.iloc[0]["Comercio"] if len(df_tc)>0 else "N/A"
    fi=ws.max_row+2
    ai(ws,fi,nc3,f"Comercio top: '{top1}'. Si concentra >50% del monto fraude -> candidato para regla especifica.")
    ac(ws)

    # HOJA 4: ESTADÍSTICAS
    sh="4_Estadisticas_Monto"
    stats_monto.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws=writer.sheets[sh]; nc4=len(stats_monto.columns)
    eh(ws,1,nc4,"MCC 7995 — ESTADISTICAS DE MONTO")
    eh(ws,2,nc4,"Fraude Total=F en cualquier estado | Fraude Aprobado=F aprobadas | No Fraude=aprobadas sin F",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    if len(df_f_ap)>0:
        media_fa=df_f_ap["MONTO"].mean(); med_fa=df_f_ap["MONTO"].median()
        std_fa=df_f_ap["MONTO"].std(); med_n=df_nofraude["MONTO"].median()
        p90_fa=df_f_ap["MONTO"].quantile(0.90); p90_n=df_nofraude["MONTO"].quantile(0.90)
        rmm=media_fa/med_fa if med_fa>0 else 0
        fi=ws.max_row+2
        ai(ws,fi,nc4,f"SESGO: Media S/ {media_fa:,.2f} vs mediana S/ {med_fa:,.2f} (ratio {rmm:.2f}x). "+("Alta presencia de montos extremos." if rmm>2 else "Sesgo moderado." if rmm>1.5 else "Distribucion simetrica."))
        ai(ws,fi+1,nc4,f"DISPERSION: Desv. std S/ {std_fa:,.2f}. "+("Alta variabilidad (CV={:.2f}x).".format(std_fa/media_fa) if media_fa>0 and std_fa/media_fa>1 else "Variabilidad moderada."))
        ai(ws,fi+2,nc4,f"P90: Fraude S/ {p90_fa:,.2f} vs Legitimo S/ {p90_n:,.2f}. "+("Umbral efectivo en este rango." if p90_fa>p90_n else "P90 fraude no supera al legitimo."))
        ai(ws,fi+3,nc4,f"COMPARACION MEDIANAS: Fraude S/ {med_fa:,.2f} vs Legitimo S/ {med_n:,.2f}. "+("Fraude con montos MAYORES." if med_fa>med_n else "Fraude con montos MENORES — posible testeo de tarjetas."))
    ac(ws)

    # HOJA 5: TRX DIARIA
    sh="5_Trx_Diaria_Cliente"
    tabla_diaria_display.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws=writer.sheets[sh]; nc5=len(tabla_diaria_display.columns)
    eh(ws,1,nc5,"MCC 7995 — TRANSACCIONALIDAD DIARIA POR CLIENTE")
    eh(ws,2,nc5,f"Fraude aprobado en {pct_1trx:.2f}% de pares cliente-dia con 1 sola trx",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    fi=ws.max_row+2
    ai(ws,fi,nc5,f"El {pct_1trx:.2f}% de clientes fraudulentos hace 1 trx/dia -> ataque puntual de monto alto. Regla de velocidad diaria no seria efectiva — priorizar umbral de monto.")
    ac(ws)

    # HOJA 6: VELOCIDAD INTERVALO
    sh="6_Velocidad_Intervalo"
    tabla_vel.reset_index().to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws=writer.sheets[sh]; nc6=len(tabla_vel.columns)+1
    eh(ws,1,nc6,"MCC 7995 — VELOCIDAD: INTERVALO ENTRE TRANSACCIONES DEL MISMO CLIENTE")
    eh(ws,2,nc6,"Fraude Total=cualquier estado | Fraude Aprobado=solo aprobadas",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    bk_max=tabla_vel["FR_Aprobado_pct"].idxmax() if len(tabla_vel)>0 else "N/A"
    pct_v=tabla_vel.loc[bk_max,"Pct_fraudes_total"] if len(tabla_vel)>0 else 0
    fi=ws.max_row+2
    ai(ws,fi,nc6,f"Intervalo mayor FR aprobado: '{bk_max}' ({pct_v:.2f}% del total fraudes). Si <=5 min -> ataques automatizados.")
    ac(ws)

    # HOJA 7: POR DIMENSION
    sh="7_Por_Dimension"; ws7=writer.book.create_sheet(sh); writer.sheets[sh]=ws7; fa=1
    dims_exp=[(df_por_canal,"Canal","Canal con mayor fraude indica vector principal."),
              (df_por_seg_nombre,"Segmento Individual","Nombre individual de cada codigo."),
              (df_por_seg_grupo,"Segmento Grupo","Agrupacion: Affluent, Mass, Corporate..."),
              (df_por_seguro,"Seguridad Comercio","No Seguro = ausencia de 3DSecure."),
              (df_por_tipo,"Tipo Producto","TD vs TC pueden tener distinto perfil."),
              (df_por_cvv,"CVV/Red Comercio","No Match o Estatico = alerta."),
              (df_por_pais,"Pais Origen Top10","Trxs desde otros paises = alto riesgo.")]
    for df_dim,titulo,interp in dims_exp:
        df_out=df_dim.reset_index(); nd=len(df_out.columns)
        eh(ws7,fa,nd,f"MCC 7995 — {titulo.upper()}"); fa+=1
        for j,col in enumerate(df_out.columns,start=1):
            c=ws7.cell(row=fa,column=j,value=col); c.fill=FILL_SUBHEAD; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
        fa+=1
        for i,row in df_out.iterrows():
            fill=FILL_FILA_A if i%2==0 else PatternFill()
            for j,val in enumerate(row,start=1):
                c=ws7.cell(row=fa,column=j,value=f"{val:,.4f}" if isinstance(val,float) else val)
                c.fill=fill; c.font=FONT_NORMAL; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
            fa+=1
        ws7.merge_cells(start_row=fa,start_column=1,end_row=fa,end_column=nd)
        c=ws7.cell(row=fa,column=1,value=f"-> {interp}")
        c.fill=FILL_AMARILLO; c.font=FONT_INTERP; c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
        ws7.row_dimensions[fa].height=25; fa+=3
    ac(ws7)

    # HOJA 8A: TABLAS CRUZADAS BASE
    sh="8A_Cruces_Base"; ws8=writer.book.create_sheet(sh); writer.sheets[sh]=ws8; fa=1
    for dim_f,dim_c,titulo in cruces_base:
        piv_c=resultados_cruce_base.get(titulo,(None,None,None))[0]
        fa=escribir_pivot_en_hoja(ws8,piv_c,fa,titulo,12)
        try:
            chi2,p_val,dof,_=chi2_contingency(piv_c.values) if piv_c is not None else (0,1,0,0)
            ws8.merge_cells(start_row=fa,start_column=1,end_row=fa,end_column=12)
            c=ws8.cell(row=fa,column=1,value=f"Chi2={chi2:.4f} p={p_val:.6f} | {'SIGNIFICATIVA' if p_val<0.05 else 'No significativa'}")
            c.fill=FILL_AMARILLO; c.font=FONT_INTERP; c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
        except: pass
        fa+=3
    ac(ws8)

    # HOJA 8B: CRUCES VELOCIDAD × SEGMENTO
    sh="8B_Cruces_Velocidad"; ws8b=writer.book.create_sheet(sh); writer.sheets[sh]=ws8b; fa=1
    for titulo,(piv_c,piv_f,piv_m) in resultados_cruce_vel.items():
        fa=escribir_pivot_en_hoja(ws8b,piv_c,fa,titulo,15)
    ac(ws8b)

    # HOJA 8C: CRUCES FEATURES NUEVAS
    sh="8C_Cruces_Features"; ws8c=writer.book.create_sheet(sh); writer.sheets[sh]=ws8c; fa=1
    for titulo,(piv_c,piv_f,piv_m) in resultados_cruce_feat.items():
        fa=escribir_pivot_en_hoja(ws8c,piv_c,fa,titulo,12)
    ac(ws8c)

    # HOJA 9: FEATURES (tablas de analisis)
    sh="9_Features_Analisis"; ws9=writer.book.create_sheet(sh); writer.sheets[sh]=ws9; fa=1
    for titulo,tabla_f in tablas_features.items():
        if tabla_f is None: continue
        df_out=tabla_f.reset_index(); nd=len(df_out.columns)
        eh(ws9,fa,nd,f"MCC 7995 — {titulo.upper()}"); fa+=1
        for j,col in enumerate(df_out.columns,start=1):
            c=ws9.cell(row=fa,column=j,value=col); c.fill=FILL_SUBHEAD; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
        fa+=1
        for i,row in df_out.iterrows():
            fill=FILL_FILA_A if i%2==0 else PatternFill()
            for j,val in enumerate(row,start=1):
                c=ws9.cell(row=fa,column=j,value=f"{val:.4f}" if isinstance(val,float) else val)
                c.fill=fill; c.font=FONT_NORMAL; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
            fa+=1
        fa+=2
    ac(ws9)

    # HOJA 10: CLIENTES RECURRENTES
    sh="10_Clientes_Recurrentes"
    df_clientes_mes.to_excel(writer,sheet_name=sh,index=False,startrow=3)
    ws=writer.sheets[sh]; nc10=len(df_clientes_mes.columns)
    eh(ws,1,nc10,"MCC 7995 — CLIENTES UNICOS Y RECURRENCIA DE FRAUDE")
    eh(ws,2,nc10,"Un cliente puede aparecer en varios meses — no suman al total global",fill=FILL_SUBHEAD)
    eed(ws,4); ed(ws,5,ws.max_row)
    fi=ws.max_row+2
    ai(ws,fi,nc10,f"Clientes globales: {total_cli_global:,} | Suma por mes: {clientes_por_mes.sum():,} | Diferencia (multi-mes): {clientes_por_mes.sum()-total_cli_global:,}")
    if len(df_recurrencia_tabla)>0:
        fi2=ws.max_row+3
        eh(ws,fi2,nc10,"RECURRENCIA ENTRE MESES CONSECUTIVOS",fill=FILL_SUBHEAD)
        df_recurrencia_tabla.to_excel(writer,sheet_name=sh,index=False,startrow=fi2)
        eed(ws,fi2+1); ed(ws,fi2+2,ws.max_row)
    ac(ws)

print(f"\nExcel exportado: {RUTA_EXCEL_OUT}")
print("Hojas generadas:")
print("  1_Resumen_Ejecutivo | 2_Deciles | 2B_Apertura_D10 | 3_Top_Comercios")
print("  4_Estadisticas_Monto | 5_Trx_Diaria_Cliente | 6_Velocidad_Intervalo")
print("  7_Por_Dimension | 8A_Cruces_Base | 8B_Cruces_Velocidad | 8C_Cruces_Features")
print("  9_Features_Analisis | 10_Clientes_Recurrentes")
