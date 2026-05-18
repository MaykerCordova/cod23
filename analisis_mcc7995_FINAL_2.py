# =============================================================================
# ANÁLISIS INTEGRAL MCC 7995 — Juegos de Azar
# Fuente: Parquet consolidado desde Monitor (aprobadas + denegadas)
# Versión final — todas las variables consistentes
# =============================================================================

import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# =============================================================================
# ██████  AJUSTA SOLO ESTE BLOQUE
# =============================================================================
COLS = {
    "FECHA"           : "ACF-FECHA TRX",
    "HORA"            : "ACF-HORA TRX",
    "DATETIME"        : "FECHA_HORA",
    "COMERCIO"        : "ACF-NOMBRE/LOCALIZACION COMERCIO",
    "INDICADOR"       : "ACF-INDICADOR DE FRAUDE",
    "COD_RESPUESTA"   : "ACF-COD RPTA",
    "CONDICION_RT"    : "CONDICION RT",
    "CANAL"           : "ACF-CANAL",
    "ENTRY_MODE"      : "ACF-ENTRY MODE",
    "SALDO"           : "ACF-SALDO DISPONIBLE EN MONEDA TRX",
    "ID_CLIENTE"      : "ACF-ID CLIENTE",
    "ESI_UCAP"        : "ACF-ECI/UCAF",
    "PAIS"            : "ACF-PAIS ORIGEN 87519",
    "COD_RED_COMERCIO": "ACF-COD RED COMERCIO",
    "MONTO"           : "ACF-MONTO EN MONEDA LOCAL",
    "SEGMENTO"        : "VAA-EVENTO DE COMPROMISO OTRA FUENTE",
    "TIPO_PRODUCTO"   : "ACF-TIPO PROD TC",
}

RUTA_PARQUET   = r"C:\ruta\al\consolidado_7995.parquet"
RUTA_EXCEL_OUT = r"C:\ruta\salida\analisis_mcc7995_resultado.xlsx"
# =============================================================================

SEG_NOMBRE = {
    "30":"Polo Dirección","99":"Polo Dirección",
    "31":"Premium","32":"Preferente","33":"Personal","34":"Estándar",
    "5":"Inst. Financieras","21":"Corporativo","2":"Mediano Empresas",
    "15":"Sector Gobierno","16":"Otras Instituciones",
    "3":"Pequeñas Empresas","4":"Negocios 2","7":"Negocios 3",
    "8":"Negocios 1","13":"Microempresas",
}
SEG_GRUPO = {
    "30":"Affluent","99":"Affluent",
    "31":"Emerging Affluent","32":"Emerging Affluent",
    "33":"Top of Mass","34":"Mass",
    "5":"Corporate","21":"Corporate","2":"Commercial",
    "15":"Commercial","16":"Commercial",
    "3":"Small Business","4":"Small Business","7":"Small Business",
    "8":"Small Business","13":"Small Business",
}
COD_RED_LABEL = {
    "S":"Estático (TD)","D":"Dinámico (TD/TC)",
    "E":"Estático (TC)","N":"No Match / Sin CVV",
}

C_FRAUDE = "#D9534F"
C_OK     = "#5B9BD5"
C_088    = "#F0AD4E"
C_NEU    = "#8EA9C1"
C_ROJO2  = "#8B0000"

# =============================================================================
# CARGA Y PREPARACIÓN
# =============================================================================
print("Cargando parquet...")
df_raw = pd.read_parquet(RUTA_PARQUET)
col_map = {v: k for k, v in COLS.items()}
df = df_raw.rename(columns=col_map).copy()

faltantes = [a for a in COLS if a not in df.columns]
if faltantes:
    print(f"  ⚠️  Columnas no encontradas: {faltantes}")

df["MONTO"] = pd.to_numeric(df["MONTO"], errors="coerce")
df["SALDO"] = pd.to_numeric(df["SALDO"], errors="coerce")

for c in ["INDICADOR","COD_RESPUESTA","CONDICION_RT","CANAL",
          "ESI_UCAP","COD_RED_COMERCIO","SEGMENTO","TIPO_PRODUCTO","ENTRY_MODE"]:
    if c in df.columns:
        df[c] = df[c].astype(str).str.strip().str.upper()

df["DATETIME"] = pd.to_datetime(df["DATETIME"], errors="coerce")
df["FECHA"]    = df["DATETIME"].dt.normalize()
df["MES"]      = df["DATETIME"].dt.to_period("M").astype(str)

# Estado desde código de respuesta
df["ESTADO"] = df["COD_RESPUESTA"].apply(
    lambda x: "APROBADA" if str(x).strip() in ["00","0000","000","0"] else "DENEGADA"
)

# Variables de fraude
df["ES_FRAUDE"]          = (df["INDICADOR"] == "F").astype(int)
df["ES_FRAUDE_APROBADO"] = (
    (df["INDICADOR"] == "F") & (df["ESTADO"] == "APROBADA")
).astype(int)

df["COND_088"]      = df["CONDICION_RT"].str.contains("088", na=False)
df["SEGURO"]        = df["ESI_UCAP"].apply(lambda x: "Seguro" if str(x).strip() in ["2","5"] else "No Seguro")
df["SEG_NOMBRE"]    = df["SEGMENTO"].map(SEG_NOMBRE).fillna("Otro")
df["SEG_GRUPO"]     = df["SEGMENTO"].map(SEG_GRUPO).fillna("Otro")
df["COD_RED_LABEL"] = df["COD_RED_COMERCIO"].map(COD_RED_LABEL).fillna("Otro")

df_ap  = df[df["ESTADO"] == "APROBADA"].copy()
df_den = df[df["ESTADO"] == "DENEGADA"].copy()
df_f_ap  = df[df["ES_FRAUDE_APROBADO"] == 1].copy()
df_f_tot = df[df["ES_FRAUDE"] == 1].copy()

print(f"  Registros totales:      {len(df):,}")
print(f"  Rango fechas:           {df['FECHA'].min().date()} → {df['FECHA'].max().date()}")
print(f"  Aprobadas:              {len(df_ap):,}")
print(f"  Denegadas:              {len(df_den):,}")
print(f"  Fraudes total (F):      {df['ES_FRAUDE'].sum():,}")
print(f"  Fraudes aprobados:      {df['ES_FRAUDE_APROBADO'].sum():,}")
print(f"  Con condición 088:      {df['COND_088'].sum():,}")

# =============================================================================
# SECCIÓN 1 — RESUMEN EJECUTIVO POR MES
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 1 — RESUMEN EJECUTIVO POR MES")
print("="*65)

meses_disponibles = sorted(df_ap["MES"].unique())
filas_resumen = []

for mes in meses_disponibles:
    sub_mes   = df[df["MES"] == mes]
    sub_ap    = df_ap[df_ap["MES"] == mes]
    sub_f_ap  = sub_ap[sub_ap["ES_FRAUDE_APROBADO"] == 1]
    sub_f_tot = sub_mes[sub_mes["ES_FRAUDE"] == 1]

    n_ap_m       = len(sub_ap)
    n_den_m      = (sub_mes["ESTADO"] == "DENEGADA").sum()
    monto_ap_m   = sub_ap["MONTO"].sum()
    ticket_ap_m  = sub_ap["MONTO"].mean()
    n_cli_m      = sub_ap["ID_CLIENTE"].nunique()
    n_f_tot_m    = len(sub_f_tot)
    monto_f_tot_m= sub_f_tot["MONTO"].sum()
    n_f_ap_m     = len(sub_f_ap)
    monto_f_ap_m = sub_f_ap["MONTO"].sum()
    ticket_f_ap_m= sub_f_ap["MONTO"].mean() if n_f_ap_m > 0 else 0
    ratio_trx_m  = n_f_ap_m / n_ap_m    if n_ap_m    > 0 else 0
    ratio_mon_m  = monto_f_ap_m / monto_ap_m if monto_ap_m > 0 else 0

    filas_resumen.append({
        "Total trx aprobadas"              : f"{n_ap_m:,}",
        "Total trx denegadas"              : f"{n_den_m:,}",
        "Total monto aprobado (S/)"        : f"S/ {monto_ap_m:,.2f}",
        "Ticket promedio por trx (S/)"     : f"S/ {ticket_ap_m:,.2f}",
        "N° clientes únicos"               : f"{n_cli_m:,}",
        "Trxs fraudulentas total (F)"      : f"{n_f_tot_m:,}",
        "Trxs fraudulentas aprobadas"      : f"{n_f_ap_m:,}",
        "Monto fraude total (S/)"          : f"S/ {monto_f_tot_m:,.2f}",
        "Monto fraude aprobado (S/)"       : f"S/ {monto_f_ap_m:,.2f}",
        "Ticket promedio fraude aprobado"  : f"S/ {ticket_f_ap_m:,.2f}",
        "Ratio trx fraudulentas aprobadas" : f"{ratio_trx_m*100:.2f}%",
        "Ratio fraude en soles (aprobado)" : f"{ratio_mon_m*100:.2f}%",
    })

df_resumen_ejecutivo = pd.DataFrame(filas_resumen, index=meses_disponibles).T
df_resumen_ejecutivo.index.name = "Indicador"
print(df_resumen_ejecutivo.to_string())

# Métricas globales
n_ap         = len(df_ap)
n_den        = len(df_den)
monto_ap     = df_ap["MONTO"].sum()
n_f_tot_g    = df["ES_FRAUDE"].sum()
n_f_ap_g     = df["ES_FRAUDE_APROBADO"].sum()
monto_f_ap_g = df_f_ap["MONTO"].sum() if len(df_f_ap) > 0 else 0
monto_f_tot_g= df_f_tot["MONTO"].sum() if len(df_f_tot) > 0 else 0
ticket_ap_g  = df_ap["MONTO"].mean()
ticket_f_ap_g= df_f_ap["MONTO"].mean() if len(df_f_ap) > 0 else 0
n_clientes   = df_ap["ID_CLIENTE"].nunique()
ratio_trx_f  = n_f_ap_g / n_ap    if n_ap    > 0 else 0
ratio_monto_f= monto_f_ap_g / monto_ap if monto_ap > 0 else 0

nivel_riesgo = (
    "ALTO ≥5%"      if ratio_trx_f >= 0.05 else
    "MEDIO [1%-5%]" if ratio_trx_f >= 0.01 else
    "REGULAR <1%"
)
riesgo_texto = {
    "REGULAR <1%"  : "El fraud rate se encuentra en nivel REGULAR (<1%). El segmento presenta bajo riesgo relativo.",
    "MEDIO [1%-5%]": "⚠ El fraud rate está en nivel MEDIO (1%–5%). Se recomienda monitoreo activo y revisión de reglas.",
    "ALTO ≥5%"     : "🔴 ALERTA: El fraud rate supera el 5%. El segmento requiere intervención inmediata.",
}
df_umbral = pd.DataFrame({
    "Umbral" : ["Regular",  "Medio",       "Alto"],
    "Rango"  : ["< 1%",     "[1% – 5%]",   "≥ 5%"],
    "Nivel actual": [
        "◀ AQUÍ" if ratio_trx_f < 0.01 else "",
        "◀ AQUÍ" if 0.01 <= ratio_trx_f < 0.05 else "",
        "◀ AQUÍ" if ratio_trx_f >= 0.05 else "",
    ]
})

# =============================================================================
# SECCIÓN 2 — TABLA DE DECILES
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 2 — TABLA DE DECILES POR MONTO")
print("="*65)

def construir_tabla_deciles(df_base, n_deciles=10, label="D"):
    df_base = df_base.copy().dropna(subset=["MONTO"])
    df_base["DECIL"] = pd.qcut(
        df_base["MONTO"], q=n_deciles,
        labels=[f"{label}{str(i+1).zfill(2)}" for i in range(n_deciles)],
        duplicates="drop"
    )
    rows = []
    fraude_ap_acum = 0
    trx_ap_acum    = 0
    total_ap       = len(df_base[df_base["ESTADO"] == "APROBADA"])
    total_f_ap     = df_base["ES_FRAUDE_APROBADO"].sum()

    for decil in df_base["DECIL"].cat.categories:
        sub     = df_base[df_base["DECIL"] == decil]
        sub_ap  = sub[sub["ESTADO"] == "APROBADA"]
        sub_den = sub[sub["ESTADO"] == "DENEGADA"]
        sub_f_ap  = sub[sub["ES_FRAUDE_APROBADO"] == 1]
        sub_f_tot = sub[sub["ES_FRAUDE"] == 1]

        n_ap_d    = len(sub_ap)
        m_ap_d    = sub_ap["MONTO"].sum()
        trx_ap_acum += n_ap_d
        freq_ap      = n_ap_d / total_ap if total_ap > 0 else 0
        freq_ap_acum = trx_ap_acum / total_ap if total_ap > 0 else 0

        n_den_d  = len(sub_den)
        m_den_d  = sub_den["MONTO"].sum()

        n_f_ap_d  = len(sub_f_ap)
        m_f_ap_d  = sub_f_ap["MONTO"].sum()
        n_f_tot_d = len(sub_f_tot)
        m_f_tot_d = sub_f_tot["MONTO"].sum()

        fraude_ap_acum += n_f_ap_d
        freq_f_ap      = n_f_ap_d / total_f_ap if total_f_ap > 0 else 0
        freq_f_ap_acum = fraude_ap_acum / total_f_ap if total_f_ap > 0 else 0

        rows.append({
            "Decil"                   : decil,
            "Rango S/"                : f"{sub['MONTO'].min():.0f}–{sub['MONTO'].max():.0f}",
            "Trx Aprobadas"           : n_ap_d,
            "Monto Aprobadas"         : round(m_ap_d, 1),
            "Freq Aprobadas"          : f"{freq_ap*100:.0f}%",
            "Freq Acum Aprobadas"     : f"{freq_ap_acum*100:.0f}%",
            "Trx Denegadas"           : n_den_d,
            "Monto Denegadas"         : round(m_den_d, 1),
            "Total Trx"               : len(sub),
            "Ticket Promedio"         : round(sub["MONTO"].mean(), 2),
            "Monto Total S/"          : round(sub["MONTO"].sum(), 1),
            "Fraude Trx Total"        : n_f_tot_d,
            "Fraude Monto Total"      : round(m_f_tot_d, 1),
            "Fraude Trx Aprobado"     : n_f_ap_d,
            "Fraude Monto Aprobado"   : round(m_f_ap_d, 1),
            "Fraude Freq Aprobado"    : f"{freq_f_ap*100:.0f}%",
            "Fraude Freq Acumulada"   : f"{freq_f_ap_acum*100:.0f}%",
            "Fraude/Trx Aprobadas"    : f"{n_f_ap_d/n_ap_d*100:.0f}%" if n_ap_d > 0 else "0%",
            "No Fraude/Trx Aprobadas" : f"{(1-n_f_ap_d/n_ap_d)*100:.0f}%" if n_ap_d > 0 else "100%",
        })
    return pd.DataFrame(rows).set_index("Decil")

tabla_deciles = construir_tabla_deciles(df, n_deciles=10)
print("\nTabla de deciles (D01=menor monto, D10=mayor monto):")
print(tabla_deciles.to_string())

print("\nApertura del Decil 10 (top monto):")
monto_d10_min = df["MONTO"].quantile(0.90)
df_d10 = df[df["MONTO"] >= monto_d10_min].copy()
tabla_d10 = None
for n in [10, 8, 6, 5]:
    try:
        tabla_d10 = construir_tabla_deciles(df_d10, n_deciles=n, label="D")
        print(f"  (Apertura con {n} grupos)")
        break
    except ValueError:
        continue
if tabla_d10 is None:
    bins = sorted(set([df_d10["MONTO"].quantile(p/100) for p in range(0,101,10)]))
    df_d10["DECIL"] = pd.cut(df_d10["MONTO"], bins=bins, include_lowest=True)
    tabla_d10 = df_d10.groupby("DECIL", observed=True).agg(
        Trx=("MONTO","count"),
        Fraude_Total=("ES_FRAUDE","sum"),
        Fraude_Aprobado=("ES_FRAUDE_APROBADO","sum"),
        Monto_total=("MONTO","sum"), Ticket=("MONTO","mean")
    )
print(tabla_d10.to_string())

# Gráfico deciles
deciles_idx = tabla_deciles.index.tolist()
n_f_ap_d_list  = tabla_deciles["Fraude Trx Aprobado"].tolist()
n_f_tot_d_list = tabla_deciles["Fraude Trx Total"].tolist()
fr_ap_d = [
    tabla_deciles.loc[d,"Fraude Trx Aprobado"] / tabla_deciles.loc[d,"Trx Aprobadas"] * 100
    if tabla_deciles.loc[d,"Trx Aprobadas"] > 0 else 0
    for d in deciles_idx
]
fig = make_subplots(rows=1, cols=2,
    subplot_titles=["Fraudes por decil (total vs aprobado)","Fraud Rate % (aprobado) por decil"])
fig.add_trace(go.Bar(x=deciles_idx, y=n_f_tot_d_list, name="Fraude Total",
    marker_color=C_ROJO2, opacity=0.6,
    text=[f"{v:,}" for v in n_f_tot_d_list], textposition="outside"), row=1, col=1)
fig.add_trace(go.Bar(x=deciles_idx, y=n_f_ap_d_list, name="Fraude Aprobado",
    marker_color=C_FRAUDE,
    text=[f"{v:,}" for v in n_f_ap_d_list], textposition="outside"), row=1, col=1)
fig.add_trace(go.Bar(x=deciles_idx, y=fr_ap_d, name="FR% Aprobado",
    marker_color=C_088,
    text=[f"{v:.1f}%" for v in fr_ap_d], textposition="outside"), row=1, col=2)
fig.update_layout(title_text="MCC 7995 — Deciles por monto",
    barmode="group", showlegend=True, height=450, plot_bgcolor="white")
fig.write_html("g1_deciles.html"); fig.show()

# =============================================================================
# SECCIÓN 3 — VELOCIDAD POR INTERVALO DE TIEMPO
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 3 — VELOCIDAD: INTERVALO ENTRE TRANSACCIONES DEL MISMO CLIENTE")
print("="*65)

df_vel = df.sort_values(["ID_CLIENTE","DATETIME"]).copy()
df_vel["PREV_DT"] = df_vel.groupby("ID_CLIENTE")["DATETIME"].shift(1)
df_vel["MIN_DESDE_PREV"] = (df_vel["DATETIME"] - df_vel["PREV_DT"]).dt.total_seconds() / 60
df_vel2 = df_vel.dropna(subset=["MIN_DESDE_PREV"]).copy()
df_vel2 = df_vel2[df_vel2["MIN_DESDE_PREV"] >= 0]

def bucket_intervalo(m):
    if   m <=  2: return "≤2 min"
    elif m <=  5: return "≤5 min"
    elif m <= 10: return "≤10 min"
    elif m <= 15: return "≤15 min"
    elif m <= 20: return "≤20 min"
    elif m <= 60: return "≤1 hora"
    else:         return ">1 hora"

ORDEN_BUCKETS = ["≤2 min","≤5 min","≤10 min","≤15 min","≤20 min","≤1 hora",">1 hora"]
df_vel2["BUCKET"] = pd.Categorical(
    df_vel2["MIN_DESDE_PREV"].apply(bucket_intervalo),
    categories=ORDEN_BUCKETS, ordered=True
)

tabla_vel = (
    df_vel2.groupby("BUCKET", observed=True)
    .agg(
        Transacciones         =("MONTO",               "count"),
        Genuinas              =("ES_FRAUDE",            lambda x: (x==0).sum()),
        Fraude_Total          =("ES_FRAUDE",            "sum"),
        Fraude_Aprobado       =("ES_FRAUDE_APROBADO",   "sum"),
        FR_Total_pct          =("ES_FRAUDE",            lambda x: x.mean()*100),
        FR_Aprobado_pct       =("ES_FRAUDE_APROBADO",   lambda x: x.mean()*100),
        Monto_Genuino         =("MONTO",                lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].sum()),
        Monto_Fraude_Total    =("MONTO",                lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_Fraude_Aprobado =("MONTO",                lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        Ticket_Genuino_Prom   =("MONTO",                lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE"]==0].mean()),
        Ticket_Fraude_Prom    =("MONTO",                lambda x: x[df_vel2.loc[x.index,"ES_FRAUDE_APROBADO"]==1].mean()),
    )
    .assign(Pct_fraudes_total=lambda x: x["Fraude_Total"] / x["Fraude_Total"].sum() * 100)
)
print(tabla_vel.to_string())

bkts = tabla_vel.index.tolist()
fig2 = make_subplots(rows=1, cols=2,
    subplot_titles=["Genuinas vs Fraude (total/aprobado) por intervalo",
                    "Fraud Rate % por intervalo"])
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["Genuinas"].tolist(), name="Genuinas",
    marker_color=C_OK,
    text=[f"{v:,}" for v in tabla_vel["Genuinas"]], textposition="outside"), row=1, col=1)
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["Fraude_Total"].tolist(), name="Fraude Total",
    marker_color=C_ROJO2, opacity=0.6,
    text=[f"{v:,}" for v in tabla_vel["Fraude_Total"]], textposition="outside"), row=1, col=1)
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["Fraude_Aprobado"].tolist(), name="Fraude Aprobado",
    marker_color=C_FRAUDE,
    text=[f"{v:,}" for v in tabla_vel["Fraude_Aprobado"]], textposition="outside"), row=1, col=1)
fig2.add_trace(go.Bar(x=bkts, y=tabla_vel["FR_Aprobado_pct"].round(2).tolist(), name="FR% Aprobado",
    marker_color=C_088,
    text=[f"{v:.1f}%" for v in tabla_vel["FR_Aprobado_pct"]], textposition="outside"), row=1, col=2)
fig2.update_layout(title_text="MCC 7995 — Fraude por intervalo entre transacciones",
    barmode="group", height=450, plot_bgcolor="white")
fig2.write_html("g2_velocidad_intervalo.html"); fig2.show()

# =============================================================================
# SECCIÓN 4 — TRANSACCIONALIDAD DIARIA POR CLIENTE
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 4 — TRANSACCIONALIDAD DIARIA POR CLIENTE")
print("="*65)

dia_cliente = (
    df_ap
    .assign(FECHA_DIA=lambda x: x["DATETIME"].dt.date)
    .groupby(["ID_CLIENTE","FECHA_DIA"])
    .agg(
        Trx_dia               =("MONTO",               "count"),
        Monto_dia             =("MONTO",               "sum"),
        Fraudes_tot_dia       =("ES_FRAUDE",            "sum"),
        Fraudes_ap_dia        =("ES_FRAUDE_APROBADO",   "sum"),
        Monto_fraude_tot_dia  =("MONTO",               lambda x: x[df_ap.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_fraude_ap_dia   =("MONTO",               lambda x: x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
    )
    .reset_index()
    .assign(
        tiene_fraude  = lambda x: (x["Fraudes_ap_dia"] > 0).astype(int),
        Trx_genuina   = lambda x: x["Trx_dia"] - x["Fraudes_tot_dia"],
        Monto_genuino = lambda x: x["Monto_dia"] - x["Monto_fraude_tot_dia"],
    )
)

def bucket_diario(n):
    if   n <= 6:  return str(n)
    elif n <= 9:  return "7-9"
    elif n <= 12: return "10-12"
    else:         return "13+"

ORDEN_DIA = ["1","2","3","4","5","6","7-9","10-12","13+"]
dia_cliente["BUCKET_DIA"] = pd.Categorical(
    dia_cliente["Trx_dia"].apply(bucket_diario), categories=ORDEN_DIA, ordered=True
)

tabla_diaria = (
    dia_cliente.groupby("BUCKET_DIA", observed=True)
    .agg(
        Genuina               =("tiene_fraude",        lambda x: (x==0).sum()),
        Fraude_Aprobado       =("tiene_fraude",        "sum"),
        Total_Clientes        =("tiene_fraude",        "count"),
        Trx_Genuina           =("Trx_genuina",         "sum"),
        Monto_Genuina         =("Monto_genuino",       "sum"),
        Trx_Fraude_Total      =("Fraudes_tot_dia",     "sum"),
        Trx_Fraude_Aprobado   =("Fraudes_ap_dia",      "sum"),
        Monto_Fraude_Total    =("Monto_fraude_tot_dia","sum"),
        Monto_Fraude_Aprobado =("Monto_fraude_ap_dia", "sum"),
    )
    .reset_index()
    .rename(columns={"BUCKET_DIA":"Intervalo"})
    .assign(
        Impacto_fraude_pct  = lambda x: x["Fraude_Aprobado"] / x["Total_Clientes"] * 100,
        Ticket_Fraude_Prom  = lambda x: np.where(x["Trx_Fraude_Aprobado"] > 0,
                                x["Monto_Fraude_Aprobado"] / x["Trx_Fraude_Aprobado"], 0),
        Ticket_Genuina_Prom = lambda x: np.where(x["Trx_Genuina"] > 0,
                                x["Monto_Genuina"] / x["Trx_Genuina"], 0),
    )
)

total_cli_fraude = tabla_diaria["Fraude_Aprobado"].sum()
pct_1trx = (
    tabla_diaria.loc[tabla_diaria["Intervalo"]=="1","Fraude_Aprobado"].sum()
    / total_cli_fraude * 100 if total_cli_fraude > 0 else 0
)
print(f"\n  ➤ El fraude aprobado en un {pct_1trx:.0f}% corresponde a clientes con 1 trx/día")

tabla_diaria_display = tabla_diaria.copy()
tabla_diaria_display["Impacto_fraude_pct"]    = tabla_diaria_display["Impacto_fraude_pct"].map("{:.0f}%".format)
tabla_diaria_display["Monto_Genuina"]         = tabla_diaria_display["Monto_Genuina"].map("S/ {:,.2f}".format)
tabla_diaria_display["Monto_Fraude_Total"]    = tabla_diaria_display["Monto_Fraude_Total"].map("S/ {:,.2f}".format)
tabla_diaria_display["Monto_Fraude_Aprobado"] = tabla_diaria_display["Monto_Fraude_Aprobado"].map("S/ {:,.2f}".format)
tabla_diaria_display["Ticket_Fraude_Prom"]    = tabla_diaria_display["Ticket_Fraude_Prom"].map("S/ {:,.2f}".format)
tabla_diaria_display["Ticket_Genuina_Prom"]   = tabla_diaria_display["Ticket_Genuina_Prom"].map("S/ {:,.2f}".format)
tabla_diaria_display.columns = [
    "Intervalo","Genuina","Fraude Aprobado","Total Clientes",
    "Trx Genuina","Monto Genuina",
    "Trx Fraude Total","Trx Fraude Aprobado",
    "Monto Fraude Total","Monto Fraude Aprobado",
    "Impacto Fraude %","Ticket Fraude Prom","Ticket Genuina Prom"
]
print(tabla_diaria_display.to_string(index=False))

buckets_d = tabla_diaria["Intervalo"].astype(str).tolist()
fr_d_pct  = (tabla_diaria["Fraude_Aprobado"] / tabla_diaria["Total_Clientes"] * 100).round(1).tolist()
fig_d = make_subplots(rows=1, cols=2,
    subplot_titles=["Clientes por nivel trx/día","Fraud Rate % por nivel trx/día"])
fig_d.add_trace(go.Bar(x=buckets_d, y=tabla_diaria["Genuina"].tolist(), name="Genuina",
    marker_color=C_OK,
    text=tabla_diaria["Genuina"].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=1)
fig_d.add_trace(go.Bar(x=buckets_d, y=tabla_diaria["Fraude_Aprobado"].tolist(), name="Fraude Aprobado",
    marker_color=C_FRAUDE,
    text=tabla_diaria["Fraude_Aprobado"].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=1)
fig_d.add_trace(go.Bar(x=buckets_d, y=fr_d_pct, name="FR%",
    marker_color=C_088,
    text=[f"{v:.1f}%" for v in fr_d_pct], textposition="outside"), row=1, col=2)
fig_d.update_layout(title_text="MCC 7995 — Transaccionalidad diaria por cliente",
    barmode="group", height=450, plot_bgcolor="white")
fig_d.write_html("g3_trx_diaria.html"); fig_d.show()

# =============================================================================
# SECCIÓN 5 — CONDICIÓN 088
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 5 — CONDICIÓN 088")
print("="*65)

f_en_088_tot  = (df["COND_088"] & (df["ES_FRAUDE"]==1)).sum()
f_fuera_088_tot=(~df["COND_088"] & (df["ES_FRAUDE"]==1)).sum()
f_en_088_ap   = (df["COND_088"] & (df["ES_FRAUDE_APROBADO"]==1)).sum()
f_fuera_088_ap=(~df["COND_088"] & (df["ES_FRAUDE_APROBADO"]==1)).sum()
tot_f_g       = df["ES_FRAUDE"].sum()
tot_f_ap_g    = df["ES_FRAUDE_APROBADO"].sum()

print(f"  FRAUDE TOTAL:    dentro 088={f_en_088_tot:,} ({f_en_088_tot/tot_f_g*100:.1f}%)  fuera={f_fuera_088_tot:,} ({f_fuera_088_tot/tot_f_g*100:.1f}%)" if tot_f_g > 0 else "")
print(f"  FRAUDE APROBADO: dentro 088={f_en_088_ap:,} ({f_en_088_ap/tot_f_ap_g*100:.1f}%)  fuera={f_fuera_088_ap:,} ({f_fuera_088_ap/tot_f_ap_g*100:.1f}%)" if tot_f_ap_g > 0 else "")

comp_088 = pd.DataFrame({
    "Con 088": {
        "Transacciones"           : df["COND_088"].sum(),
        "Fraude Total"            : f_en_088_tot,
        "Fraude Aprobado"         : f_en_088_ap,
        "FR% Aprobado"            : df.loc[df["COND_088"],"ES_FRAUDE_APROBADO"].mean()*100,
        "Monto prom (S/)"         : df.loc[df["COND_088"],"MONTO"].mean(),
        "Monto mediano (S/)"      : df.loc[df["COND_088"],"MONTO"].median(),
        "Monto fraude aprobado S/": df.loc[df["COND_088"] & (df["ES_FRAUDE_APROBADO"]==1),"MONTO"].sum(),
    },
    "Sin 088": {
        "Transacciones"           : (~df["COND_088"]).sum(),
        "Fraude Total"            : f_fuera_088_tot,
        "Fraude Aprobado"         : f_fuera_088_ap,
        "FR% Aprobado"            : df.loc[~df["COND_088"],"ES_FRAUDE_APROBADO"].mean()*100,
        "Monto prom (S/)"         : df.loc[~df["COND_088"],"MONTO"].mean(),
        "Monto mediano (S/)"      : df.loc[~df["COND_088"],"MONTO"].median(),
        "Monto fraude aprobado S/": df.loc[~df["COND_088"] & (df["ES_FRAUDE_APROBADO"]==1),"MONTO"].sum(),
    },
}).T
print(comp_088.to_string())

# =============================================================================
# SECCIÓN 6 — DISTRIBUCIÓN POR DIMENSIONES
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 6 — DISTRIBUCIÓN DEL FRAUDE POR DIMENSIONES")
print("="*65)

def resumen_por(df_base, col, top_n=10):
    return (
        df_base.groupby(col)
        .agg(
            Txns               =(col,                  "count"),
            Fraude_Total       =("ES_FRAUDE",           "sum"),
            Fraude_Aprobado    =("ES_FRAUDE_APROBADO",  "sum"),
            FR_Total_pct       =("ES_FRAUDE",           lambda x: x.mean()*100),
            FR_Aprobado_pct    =("ES_FRAUDE_APROBADO",  lambda x: x.mean()*100),
            Monto_total        =("MONTO",               "sum"),
            Monto_Fraude_Ap    =("MONTO",               lambda x: x[df_base.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        )
        .sort_values("Fraude_Aprobado", ascending=False)
        .head(top_n)
        .rename(columns={
            "FR_Total_pct"  : "FR% Total",
            "FR_Aprobado_pct":"FR% Aprobado",
            "Monto_Fraude_Ap":"Monto Fraude Aprobado",
        })
    )

df_por_canal    = resumen_por(df, "CANAL")
df_por_segmento = resumen_por(df, "SEG_NOMBRE")
df_por_seguro   = resumen_por(df, "SEGURO")
df_por_tipo     = resumen_por(df, "TIPO_PRODUCTO")
df_por_cvv      = resumen_por(df, "COD_RED_LABEL")
df_por_pais     = resumen_por(df, "PAIS")

for nombre, tabla in [
    ("Canal",            df_por_canal),
    ("Segmento",         df_por_segmento),
    ("Seguro/No Seguro", df_por_seguro),
    ("Tipo Producto",    df_por_tipo),
    ("CVV/Red Comercio", df_por_cvv),
    ("País (Top 10)",    df_por_pais),
]:
    print(f"\n— Por {nombre}:")
    print(tabla.to_string())

print("\n— Top 10 comercios por fraude aprobado:")
df_top_comercios = (
    df_ap.groupby("COMERCIO")
    .agg(
        Tarjetas            =("ID_CLIENTE",           "nunique"),
        Trx                 =("MONTO",                "count"),
        Fraude_Total        =("ES_FRAUDE",             "sum"),
        Fraude_Aprobado     =("ES_FRAUDE_APROBADO",    "sum"),
        Monto_Fraude_Total  =("MONTO",                lambda x: x[df_ap.loc[x.index,"ES_FRAUDE"]==1].sum()),
        Monto_Fraude_Ap     =("MONTO",                lambda x: x[df_ap.loc[x.index,"ES_FRAUDE_APROBADO"]==1].sum()),
        Monto_total         =("MONTO",                "sum"),
        FR_Ap_pct           =("ES_FRAUDE_APROBADO",   lambda x: x.mean()*100),
    )
    .sort_values("Monto_Fraude_Ap", ascending=False)
    .head(10)
    .reset_index()
    .rename(columns={
        "COMERCIO"          : "Comercio",
        "Tarjetas"          : "# Tarjetas",
        "Trx"               : "# Trx",
        "Fraude_Total"      : "# Fraude Total",
        "Fraude_Aprobado"   : "# Fraude Aprobado",
        "Monto_Fraude_Total": "Monto Fraude Total S/",
        "Monto_Fraude_Ap"   : "Monto Fraude Aprobado S/",
        "Monto_total"       : "Monto Total S/",
        "FR_Ap_pct"         : "Fraud Rate % (ap)",
    })
)
print(df_top_comercios.to_string(index=False))

# Gráfico dimensiones
dims_graf = [("SEG_NOMBRE","Segmento"),("SEGURO","Seguro/No Seguro"),
             ("CANAL","Canal"),("COD_RED_LABEL","CVV/Red")]
fig3 = make_subplots(rows=2, cols=2, subplot_titles=[d[1] for d in dims_graf])
for i,(col,titulo) in enumerate(dims_graf):
    r,c = divmod(i,2)
    sub = resumen_por(df, col, top_n=8).reset_index()
    fig3.add_trace(go.Bar(x=sub[col], y=sub["Fraude_Total"], name="Fraude Total",
        marker_color=C_ROJO2, opacity=0.6,
        text=sub["Fraude_Total"].apply(lambda v: f"{v:,}"),
        textposition="outside", showlegend=(i==0)), row=r+1, col=c+1)
    fig3.add_trace(go.Bar(x=sub[col], y=sub["Fraude_Aprobado"], name="Fraude Aprobado",
        marker_color=C_FRAUDE,
        text=sub["Fraude_Aprobado"].apply(lambda v: f"{v:,}"),
        textposition="outside", showlegend=(i==0)), row=r+1, col=c+1)
fig3.update_layout(title_text="MCC 7995 — Fraude por dimensión",
    barmode="group", height=700, plot_bgcolor="white")
fig3.write_html("g4_fraude_dimension.html"); fig3.show()

# =============================================================================
# SECCIÓN 7 — ESTADÍSTICAS DE MONTO
# =============================================================================
print("\n" + "="*65)
print("SECCIÓN 7 — ESTADÍSTICAS DE MONTO")
print("="*65)

pcts       = [10, 25, 50, 75, 90, 95, 99]
df_nofraude= df_ap[df_ap["ES_FRAUDE"] == 0]

def stats_serie(s):
    if len(s) == 0:
        return [0]*13
    return ([s.mean(), s.median(), s.std(), s.var(), s.min(), s.max()] +
            [s.quantile(p/100) for p in pcts])

stats_monto = pd.DataFrame({
    "Métrica"         : (["Media","Mediana","Desv. Estándar","Varianza","Mínimo","Máximo"] +
                         [f"Percentil {p}" for p in pcts]),
    "Total aprobadas" : stats_serie(df_ap["MONTO"]),
    "Fraude Total"    : stats_serie(df_f_tot["MONTO"]),
    "Fraude Aprobado" : stats_serie(df_f_ap["MONTO"]),
    "No Fraude"       : stats_serie(df_nofraude["MONTO"]),
})
for col in ["Total aprobadas","Fraude Total","Fraude Aprobado","No Fraude"]:
    stats_monto[col] = stats_monto[col].apply(lambda v: f"S/ {v:,.2f}")
print(stats_monto.to_string(index=False))

# =============================================================================
# EVOLUCIÓN MENSUAL
# =============================================================================
evol = (
    df.groupby("MES")
    .agg(
        Aprobadas      =("ESTADO",             lambda x: (x=="APROBADA").sum()),
        Denegadas      =("ESTADO",             lambda x: (x=="DENEGADA").sum()),
        Fraude_Total   =("ES_FRAUDE",          "sum"),
        Fraude_Aprobado=("ES_FRAUDE_APROBADO", "sum"),
        Con_088        =("COND_088",           "sum"),
        Monto_total    =("MONTO",              "sum"),
    )
    .reset_index()
)
fig4 = make_subplots(rows=1, cols=2,
    subplot_titles=["Aprobadas / Denegadas / Fraude","Fraude Total vs Aprobado vs 088"])
for serie,color,nombre in [("Aprobadas",C_OK,"Aprobadas"),
                            ("Denegadas",C_NEU,"Denegadas")]:
    fig4.add_trace(go.Bar(x=evol["MES"], y=evol[serie], name=nombre,
        marker_color=color,
        text=evol[serie].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=1)
for serie,color,nombre in [("Fraude_Total",C_ROJO2,"Fraude Total"),
                            ("Fraude_Aprobado",C_FRAUDE,"Fraude Aprobado"),
                            ("Con_088",C_088,"Con 088")]:
    fig4.add_trace(go.Bar(x=evol["MES"], y=evol[serie], name=nombre,
        marker_color=color,
        text=evol[serie].apply(lambda v: f"{v:,}"), textposition="outside"), row=1, col=2)
fig4.update_layout(title_text="MCC 7995 — Evolución mensual",
    barmode="group", height=450, plot_bgcolor="white")
fig4.write_html("g5_evolucion_mensual.html"); fig4.show()

# =============================================================================
# EXPORTACIÓN A EXCEL
# =============================================================================
print("\nExportando a Excel...")

FILL_HEADER  = PatternFill("solid", fgColor="1F3864")
FILL_SUBHEAD = PatternFill("solid", fgColor="2E75B6")
FILL_FILA_A  = PatternFill("solid", fgColor="DEEAF1")
FILL_AMARILLO= PatternFill("solid", fgColor="FFF2CC")
FILL_FRAUDE  = PatternFill("solid", fgColor="FCE4D6")
FONT_HEADER  = Font(color="FFFFFF", bold=True, size=10)
FONT_NORMAL  = Font(size=10)
FONT_INTERP  = Font(italic=True, size=9, color="1F3864")
BORDER_THIN  = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"),  bottom=Side(style="thin"))
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def estilizar_header(ws, fila, n_cols, texto, fill=None):
    fill = fill or FILL_HEADER
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.fill=fill; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def estilizar_encabezados_df(ws, fila):
    for row in ws.iter_rows(min_row=fila, max_row=fila):
        for c in row:
            c.fill=FILL_SUBHEAD; c.font=FONT_HEADER; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def estilizar_datos(ws, fila_inicio, fila_fin, fraude_keywords=None):
    for i, row in enumerate(ws.iter_rows(min_row=fila_inicio, max_row=fila_fin), start=1):
        fill = FILL_FILA_A if i%2==0 else PatternFill()
        if fraude_keywords and row[0].value:
            if any(k in str(row[0].value).lower() for k in fraude_keywords):
                fill = FILL_FRAUDE
        for c in row:
            c.fill=fill; c.font=FONT_NORMAL; c.alignment=ALIGN_CENTER; c.border=BORDER_THIN

def agregar_interpretacion(ws, fila, n_cols, texto):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=n_cols)
    c = ws.cell(row=fila, column=1, value=f"📌 {texto}")
    c.fill=FILL_AMARILLO; c.font=FONT_INTERP; c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
    ws.row_dimensions[fila].height = 30

def ajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+4, 35)

with pd.ExcelWriter(RUTA_EXCEL_OUT, engine="openpyxl") as writer:

    # ── HOJA 1: RESUMEN EJECUTIVO ─────────────────────────────────────────────
    sheet = "1_Resumen_Ejecutivo"
    n_cols_res = len(df_resumen_ejecutivo.columns) + 1
    df_resumen_ejecutivo.to_excel(writer, sheet_name=sheet, startrow=3)
    ws = writer.sheets[sheet]
    estilizar_header(ws, 1, n_cols_res,
        f"MCC 7995 — JUEGOS DE AZAR | Análisis: {datetime.today().strftime('%d/%m/%Y')}")
    estilizar_header(ws, 2, n_cols_res,
        "RESUMEN DE INDICADORES DE FRAUDE POR MES", fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws, 4)
    estilizar_datos(ws, 5, ws.max_row, fraude_keywords=["fraude","ratio"])
    fila_u = ws.max_row + 2
    estilizar_header(ws, fila_u, n_cols_res,
        "UMBRAL DE CONTROL — Ratio Trx Fraudulentas Aprobadas", fill=FILL_SUBHEAD)
    df_umbral.to_excel(writer, sheet_name=sheet, index=False, startrow=fila_u)
    fi = ws.max_row + 2
    agregar_interpretacion(ws, fi, n_cols_res,
        f"Nivel de riesgo actual: {nivel_riesgo}. {riesgo_texto[nivel_riesgo]}")
    agregar_interpretacion(ws, fi+1, n_cols_res,
        f"Fraude total (F): {n_f_tot_g:,} trxs — Fraude aprobado: {n_f_ap_g:,} trxs "
        f"({n_f_ap_g/n_f_tot_g*100:.1f}% del total). "
        "La diferencia son fraudes en transacciones denegadas." if n_f_tot_g > 0 else "Sin fraudes.")
    ajustar_columnas(ws)

    # ── HOJA 2: DECILES ───────────────────────────────────────────────────────
    sheet = "2_Deciles"
    n_cols_dec = len(tabla_deciles.columns) + 2
    tabla_deciles.to_excel(writer, sheet_name=sheet, startrow=3)
    ws = writer.sheets[sheet]
    estilizar_header(ws, 1, n_cols_dec, "MCC 7995 — TABLA DE DECILES POR MONTO")
    estilizar_header(ws, 2, n_cols_dec,
        "D01=menor monto → D10=mayor monto. Fraude Total incluye aprobadas+denegadas.",
        fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws, 4)
    estilizar_datos(ws, 5, ws.max_row)
    decil_max_f = tabla_deciles["Fraude Trx Aprobado"].idxmax()
    decil_max_m = tabla_deciles["Fraude Monto Aprobado"].idxmax()
    fi = ws.max_row + 2
    agregar_interpretacion(ws, fi, n_cols_dec,
        f"Mayor concentración de fraude aprobado en cantidad: {decil_max_f} "
        f"({tabla_deciles.loc[decil_max_f,'Fraude Trx Aprobado']:,} fraudes) — "
        f"Rango: {tabla_deciles.loc[decil_max_f,'Rango S/']}")
    agregar_interpretacion(ws, fi+1, n_cols_dec,
        f"Mayor concentración de fraude aprobado en monto: {decil_max_m} "
        f"(S/ {tabla_deciles.loc[decil_max_m,'Fraude Monto Aprobado']:,.2f}) — "
        f"Rango: {tabla_deciles.loc[decil_max_m,'Rango S/']}")
    ajustar_columnas(ws)

    # ── HOJA 3: TOP COMERCIOS ─────────────────────────────────────────────────
    sheet = "3_Top_Comercios"
    df_tc = df_top_comercios.copy()
    df_tc["Monto Fraude Total S/"]    = df_tc["Monto Fraude Total S/"].map("S/ {:,.2f}".format)
    df_tc["Monto Fraude Aprobado S/"] = df_tc["Monto Fraude Aprobado S/"].map("S/ {:,.2f}".format)
    df_tc["Monto Total S/"]           = df_tc["Monto Total S/"].map("S/ {:,.2f}".format)
    df_tc["Fraud Rate % (ap)"]        = df_tc["Fraud Rate % (ap)"].map("{:.2f}%".format)
    df_tc.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]
    estilizar_header(ws, 1, len(df_tc.columns),
        "MCC 7995 — TOP 10 COMERCIOS CON MAYOR MONTO DE FRAUDE APROBADO")
    estilizar_header(ws, 2, len(df_tc.columns),
        "Ordenado por monto de fraude aprobado acumulado", fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws, 4)
    estilizar_datos(ws, 5, ws.max_row)
    top1       = df_tc.iloc[0]["Comercio"]               if len(df_tc) > 0 else "N/A"
    monto_top1 = df_tc.iloc[0]["Monto Fraude Aprobado S/"] if len(df_tc) > 0 else "S/ 0"
    fi = ws.max_row + 2
    agregar_interpretacion(ws, fi, len(df_tc.columns),
        f"Comercio con mayor monto de fraude aprobado: '{top1}' ({monto_top1}). "
        "Si concentra >50% del monto total de fraude, es candidato prioritario para regla específica.")
    ajustar_columnas(ws)

    # ── HOJA 4: ESTADÍSTICAS MONTO ────────────────────────────────────────────
    sheet = "4_Estadisticas_Monto"
    stats_monto.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]
    estilizar_header(ws, 1, len(stats_monto.columns),
        "MCC 7995 — ESTADÍSTICAS DE MONTO: TOTAL vs APROBADO vs NO FRAUDE")
    estilizar_header(ws, 2, len(stats_monto.columns),
        "Fraude Total = F en cualquier estado | Fraude Aprobado = F que pasaron el control",
        fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws, 4)
    estilizar_datos(ws, 5, ws.max_row)
    med_f_ap  = df_f_ap["MONTO"].median()  if len(df_f_ap)  > 0 else 0
    med_f_tot = df_f_tot["MONTO"].median() if len(df_f_tot) > 0 else 0
    med_nof   = df_nofraude["MONTO"].median()
    p90_f_ap  = df_f_ap["MONTO"].quantile(0.90) if len(df_f_ap) > 0 else 0
    p90_nof   = df_nofraude["MONTO"].quantile(0.90)
    fi = ws.max_row + 2
    agregar_interpretacion(ws, fi, len(stats_monto.columns),
        f"Mediana fraude aprobado S/ {med_f_ap:,.2f} | fraude total S/ {med_f_tot:,.2f} | "
        f"legítimo S/ {med_nof:,.2f}.")
    agregar_interpretacion(ws, fi+1, len(stats_monto.columns),
        f"P90 fraude aprobado S/ {p90_f_ap:,.2f} vs legítimo S/ {p90_nof:,.2f}. "
        "Si el P90 del fraude aprobado es mayor, los ataques exitosos son de montos altos — "
        "un umbral de monto puede ser efectivo como control.")
    ajustar_columnas(ws)

    # ── HOJA 5: TRANSACCIONALIDAD DIARIA ─────────────────────────────────────
    sheet = "5_Trx_Diaria_Cliente"
    tabla_diaria_display.to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]
    estilizar_header(ws, 1, len(tabla_diaria_display.columns),
        "MCC 7995 — TRANSACCIONALIDAD DIARIA POR CLIENTE")
    estilizar_header(ws, 2, len(tabla_diaria_display.columns),
        f"Fraude aprobado en un {pct_1trx:.0f}% corresponde a clientes con 1 sola transacción por día",
        fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws, 4)
    estilizar_datos(ws, 5, ws.max_row)
    fi = ws.max_row + 2
    agregar_interpretacion(ws, fi, len(tabla_diaria_display.columns),
        f"El {pct_1trx:.0f}% de clientes fraudulentos aprobados hace 1 trx/día → "
        "ataque de monto alto en una sola operación. "
        "Una regla de velocidad diaria no sería efectiva — priorizar umbral de monto.")
    ajustar_columnas(ws)

    # ── HOJA 6: VELOCIDAD POR INTERVALO ──────────────────────────────────────
    sheet = "6_Velocidad_Intervalo"
    tabla_vel.reset_index().to_excel(writer, sheet_name=sheet, index=False, startrow=3)
    ws = writer.sheets[sheet]
    estilizar_header(ws, 1, len(tabla_vel.columns)+1,
        "MCC 7995 — VELOCIDAD: INTERVALO ENTRE TRANSACCIONES DEL MISMO CLIENTE")
    estilizar_header(ws, 2, len(tabla_vel.columns)+1,
        "¿En qué ventana de tiempo se concentran los ataques fraudulentos?",
        fill=FILL_SUBHEAD)
    estilizar_encabezados_df(ws, 4)
    estilizar_datos(ws, 5, ws.max_row)
    bucket_max = tabla_vel["FR_Aprobado_pct"].idxmax() if len(tabla_vel) > 0 else "N/A"
    pct_vel    = tabla_vel.loc[bucket_max,"Pct_fraudes_total"] if len(tabla_vel) > 0 else 0
    fi = ws.max_row + 2
    agregar_interpretacion(ws, fi, len(tabla_vel.columns)+1,
        f"Intervalo con mayor fraud rate aprobado: '{bucket_max}' ({pct_vel:.1f}% del total de fraudes). "
        "Si es ≤5 min → indica ataques automatizados sobre el mismo cliente.")
    agregar_interpretacion(ws, fi+1, len(tabla_vel.columns)+1,
        "Una regla de velocidad efectiva bloquea al cliente si hace N transacciones "
        "en el intervalo de mayor concentración de fraude detectado en esta tabla.")
    ajustar_columnas(ws)

    # ── HOJA 7: POR DIMENSIONES ───────────────────────────────────────────────
    sheet = "7_Por_Dimension"
    ws7 = writer.book.create_sheet(sheet)
    writer.sheets[sheet] = ws7
    fila_actual = 1

    dims_export = [
        (df_por_canal,    "Canal",
         "El canal con mayor fraude aprobado indica el vector de ataque principal."),
        (df_por_segmento, "Segmento Cliente",
         "Si el fraude aprobado se concentra en un segmento, la regla puede ser más quirúrgica."),
        (df_por_seguro,   "Seguridad Comercio",
         "'No Seguro' indica ausencia de 3DSecure — vector frecuente en e-commerce."),
        (df_por_tipo,     "Tipo Producto",
         "TD vs TC puede tener distinto perfil de fraude y distintas reglas aplicables."),
        (df_por_cvv,      "CVV/Red Comercio",
         "Alta concentración en No Match o Estático es señal de alerta."),
        (df_por_pais,     "País Origen (Top 10)",
         "Transacciones desde países distintos al Perú son señal de alto riesgo."),
    ]

    for df_dim, titulo, interp in dims_export:
        df_out = df_dim.reset_index()
        n_cols_d = len(df_out.columns)
        estilizar_header(ws7, fila_actual, n_cols_d, f"MCC 7995 — {titulo.upper()}")
        fila_actual += 1
        for j, col in enumerate(df_out.columns, start=1):
            c = ws7.cell(row=fila_actual, column=j, value=col)
            c.fill=FILL_SUBHEAD; c.font=FONT_HEADER
            c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
        fila_actual += 1
        for i, row in df_out.iterrows():
            fill = FILL_FILA_A if i%2==0 else PatternFill()
            for j, val in enumerate(row, start=1):
                c = ws7.cell(row=fila_actual, column=j,
                             value=f"{val:,.2f}" if isinstance(val, float) else val)
                c.fill=fill; c.font=FONT_NORMAL
                c.alignment=ALIGN_CENTER; c.border=BORDER_THIN
            fila_actual += 1
        ws7.merge_cells(start_row=fila_actual, start_column=1,
                        end_row=fila_actual, end_column=n_cols_d)
        c = ws7.cell(row=fila_actual, column=1, value=f"📌 {interp}")
        c.fill=FILL_AMARILLO; c.font=FONT_INTERP
        c.alignment=ALIGN_LEFT; c.border=BORDER_THIN
        ws7.row_dimensions[fila_actual].height = 30
        fila_actual += 3

    ajustar_columnas(ws7)

print(f"\n✅ Excel exportado: {RUTA_EXCEL_OUT}")
print("   Hojas: 1_Resumen_Ejecutivo | 2_Deciles | 3_Top_Comercios")
print("          4_Estadisticas_Monto | 5_Trx_Diaria_Cliente")
print("          6_Velocidad_Intervalo | 7_Por_Dimension")
print("\n   Gráficos HTML:")
print("   g1_deciles | g2_velocidad_intervalo | g3_trx_diaria")
print("   g4_fraude_dimension | g5_evolucion_mensual")
